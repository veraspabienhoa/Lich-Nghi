# V86.13 - Khóa đăng ký riêng từng vai trò + Admin miễn mọi giới hạn lịch nghỉ + V86.12 features (2026-08-20)
import streamlit as st
import pandas as pd
from datetime import date, timedelta, datetime, timezone
import calendar
import requests
import os
import io
import gspread
from google.oauth2.service_account import Credentials
import streamlit.components.v1 as components
import time
import smtplib
import unicodedata
import hashlib
import base64
import secrets
import hmac
import json
import zipfile
import re
import numbers
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from urllib.parse import urljoin

# --- CLOUD RUN + POSTGRESQL SHARED DATA LAYER (V75) ---
try:
    import vera_postgres as vpg
except Exception:
    vpg = None


# --- CẤU HÌNH MÚI GIỜ VIỆT NAM ---
VN_TZ = timezone(timedelta(hours=7))

def get_vn_today():
    return datetime.now(VN_TZ).date()


# ==========================================================
# V79 - TIMESOFT: TỰ ĐĂNG NHẬP + TỰ PHÁT HIỆN API BÁO CÁO
# ==========================================================
# Tài khoản/mật khẩu KHÔNG ghi trong app.py. Khai báo tại Streamlit Secrets:
# [TIMESOFT]
# base_url = "https://vera.timesoft.vn"
# username = "..."
# password = "..."
#
# V82: Cloud Run ưu tiên biến môi trường/Secret Manager; máy local vẫn dùng Streamlit Secrets.
def _timesoft_setting(secret_key, env_key, default=""):
    env_val = str(os.getenv(env_key, "") or "").strip()
    if env_val:
        return env_val
    try:
        return str(st.secrets["TIMESOFT"].get(secret_key, default) or default).strip()
    except Exception:
        return str(default or "").strip()

TIMESOFT_BASE_URL = _timesoft_setting("base_url", "TIMESOFT_BASE_URL", "https://vera.timesoft.vn")
TIMESOFT_USERNAME = _timesoft_setting("username", "TIMESOFT_USERNAME", "")
TIMESOFT_PASSWORD = _timesoft_setting("password", "TIMESOFT_PASSWORD", "")
TIMESOFT_BASE_URL = str(TIMESOFT_BASE_URL or "https://vera.timesoft.vn").rstrip("/")
TIMESOFT_REPORT_PAGES = {
    "summary_invoice": {
        "label": "Báo cáo tổng hợp doanh thu",
        "path": "/Report/ReportSummaryInvoice/Index",
        "tokens": ["reportsummaryinvoice", "summaryinvoice", "invoice", "doanh thu", "hóa đơn", "hoa don"],
    },
    "employee_checkin": {
        "label": "Báo cáo chấm công nhân viên",
        "path": "/Report/ReportEmployeeCheckin/Index",
        "tokens": ["reportemployeecheckin", "employeecheckin", "checkin", "chấm công", "cham cong"],
    },
}
TIMESOFT_CAPTURE_MAX_CHARS = 250_000
TIMESOFT_PREVIEW_MAX_CHARS = 6_000

# V81 - API đã được xác định từ request thực tế của TimeSoft.
TIMESOFT_DIRECT_APIS = {
    "summary_invoice": {
        "url": "/Report/ReportSummaryInvoice/SearchFullText",
        "referer": "/Report/ReportSummaryInvoice/Index",
    },
    "employee_checkin": {
        "url": "/Report/ReportEmployeeCheckin/SearchElastic",
        "referer": "/Report/ReportEmployeeCheckin/Index",
    },
}
TIMESOFT_HTTP_SESSION_TTL_SECONDS = 15 * 60
TIMESOFT_CHECKIN_PAGE_SIZE = 100
TIMESOFT_MAX_CHECKIN_PAGES = 500

# V83 - Chu kỳ HIỂN THỊ khớp Cloud Scheduler. app.py không tự tạo timer nền.
TIMESOFT_BACKGROUND_INTERVAL_MINUTES = 30


def timesoft_is_configured():
    return bool(str(TIMESOFT_BASE_URL).strip() and str(TIMESOFT_USERNAME).strip() and str(TIMESOFT_PASSWORD).strip())


def _timesoft_safe_request_headers(headers):
    """Chỉ giữ header kỹ thuật an toàn; tuyệt đối không trả Cookie/Authorization ra UI/log."""
    allow = {"accept", "content-type", "x-requested-with", "referer", "origin"}
    safe = {}
    for k, v in (headers or {}).items():
        lk = str(k).lower().strip()
        if lk in allow:
            safe[lk] = str(v)
    return safe


def _timesoft_body_preview(body):
    if body is None:
        return ""
    txt = str(body)
    return txt[:TIMESOFT_PREVIEW_MAX_CHARS]


def _timesoft_json_kind(body_text):
    text0 = str(body_text or "").strip()
    if not text0:
        return ""
    try:
        obj = json.loads(text0)
        if isinstance(obj, dict):
            return "dict:" + ",".join(list(map(str, obj.keys()))[:20])
        if isinstance(obj, list):
            return f"list:{len(obj)}"
    except Exception:
        pass
    return ""


def _timesoft_candidate_score(report_key, item):
    """Chấm điểm request để tự chọn API dữ liệu thật thay vì đoán endpoint."""
    info = TIMESOFT_REPORT_PAGES.get(report_key, {})
    url = str(item.get("url", "")).lower()
    body = str(item.get("body_preview", "")).lower()
    post_data = str(item.get("post_data", "")).lower()
    content_type = str(item.get("content_type", "")).lower()
    score = 0

    if item.get("resource_type") in {"xhr", "fetch"}:
        score += 25
    if int(item.get("status", 0) or 0) == 200:
        score += 8
    if "json" in content_type or item.get("json_kind"):
        score += 18
    if str(item.get("method", "")).upper() == "POST":
        score += 4

    for token in info.get("tokens", []):
        token_l = str(token).lower()
        if token_l and token_l in url:
            score += 24
        if token_l and token_l in body:
            score += 8
        if token_l and token_l in post_data:
            score += 5

    generic_data_tokens = [
        "data", "rows", "items", "result", "total", "amount", "invoice",
        "employee", "checkin", "checkout", "time", "date", "customer",
    ]
    for token in generic_data_tokens:
        if token in body:
            score += 1

    # Giảm điểm các request tài nguyên/telemetry không phải dữ liệu báo cáo.
    if any(x in url for x in ["signalr", "notification", "favicon", "analytics", "chat", "socket"]):
        score -= 20
    return score


def _timesoft_pick_best_candidates(captured_by_report):
    result = {}
    for report_key, items in (captured_by_report or {}).items():
        scored = []
        for item in items:
            clean = dict(item)
            clean["score"] = _timesoft_candidate_score(report_key, clean)
            scored.append(clean)
        scored.sort(key=lambda x: (int(x.get("score", 0)), int(x.get("status", 0) == 200)), reverse=True)
        result[report_key] = {
            "report": TIMESOFT_REPORT_PAGES.get(report_key, {}).get("label", report_key),
            "page_url": urljoin(TIMESOFT_BASE_URL + "/", TIMESOFT_REPORT_PAGES.get(report_key, {}).get("path", "").lstrip("/")),
            "best": scored[0] if scored else None,
            "candidates": scored[:20],
        }
    return result


def _timesoft_visible_input(page, selectors):
    for selector in selectors:
        try:
            loc = page.locator(selector)
            count = min(loc.count(), 12)
            for i in range(count):
                node = loc.nth(i)
                if node.is_visible():
                    return node
        except Exception:
            continue
    return None


def _timesoft_login_error_text(page):
    """Lấy thông báo lỗi đăng nhập đang hiển thị nhưng không lấy giá trị input/credential."""
    selectors = [
        '.validation-summary-errors', '.field-validation-error', '.alert-danger',
        '.alert-warning', '.error', '.error-message', '.text-danger',
        '[role="alert"]', '.toast-message', '.notifyjs-bootstrap-error',
    ]
    messages = []
    for selector in selectors:
        try:
            loc = page.locator(selector)
            for i in range(min(loc.count(), 8)):
                node = loc.nth(i)
                if not node.is_visible():
                    continue
                txt = re.sub(r"\s+", " ", str(node.inner_text() or "")).strip()
                if txt and txt not in messages:
                    messages.append(txt[:300])
        except Exception:
            continue
    return " | ".join(messages[:4])


def _timesoft_login_with_playwright(page, verify_url=None):
    """Đăng nhập TimeSoft và xác minh bằng cách mở lại trang báo cáo thật.

    V80: không kết luận thất bại chỉ vì URL Login chưa đổi ngay sau submit.
    Một số bản TimeSoft xử lý đăng nhập bằng JavaScript/AJAX rồi mới tạo session.
    Sau khi submit, hàm chủ động mở verify_url; chỉ khi verify_url lại redirect về
    /User/Login và vẫn có ô password thì mới xem là đăng nhập thất bại.
    Credential/cookie/authorization tuyệt đối không được ghi ra log/UI.
    """
    password_box = _timesoft_visible_input(page, [
        'input[type="password"]',
        'input[name*="password" i]',
        'input[id*="password" i]',
        'input[name*="pass" i]',
        'input[id*="pass" i]',
    ])
    if password_box is None:
        # Không có ô password: có thể session đã đăng nhập sẵn.
        return True, "Session TimeSoft đã đăng nhập."

    username_box = _timesoft_visible_input(page, [
        'input[name="UserName"]',
        'input[name="Username"]',
        'input[name*="username" i]',
        'input[id*="username" i]',
        'input[name*="user" i]',
        'input[id*="user" i]',
        'input[name*="account" i]',
        'input[id*="account" i]',
        'input[name*="login" i]',
        'input[id*="login" i]',
        'input[type="email"]',
        'input[type="text"]',
    ])
    if username_box is None:
        return False, "Không nhận diện được ô tài khoản trên trang đăng nhập TimeSoft."

    try:
        # click + fill giúp kích hoạt đúng event input/change của framework phía trang web.
        username_box.click()
        username_box.fill(str(TIMESOFT_USERNAME))
        password_box.click()
        password_box.fill(str(TIMESOFT_PASSWORD))
        try:
            password_box.press("Tab")
        except Exception:
            pass
        page.wait_for_timeout(250)
    except Exception as e:
        return False, f"Không nhập được thông tin đăng nhập TimeSoft: {type(e).__name__}"

    submit = None
    submit_selectors = [
        'button[type="submit"]',
        'input[type="submit"]',
        'input[type="button"][value*="đăng" i]',
        'input[type="button"][value*="login" i]',
        'button:has-text("Đăng nhập")',
        'button:has-text("Đăng Nhập")',
        'button:has-text("Login")',
        'a:has-text("Đăng nhập")',
        'a:has-text("Đăng Nhập")',
        '[onclick*="login" i]',
        '[id*="login" i]',
    ]
    for selector in submit_selectors:
        try:
            loc = page.locator(selector)
            for i in range(min(loc.count(), 10)):
                node = loc.nth(i)
                if node.is_visible() and node.is_enabled():
                    submit = node
                    break
            if submit is not None:
                break
        except Exception:
            continue

    try:
        if submit is not None:
            submit.click(timeout=8_000)
        else:
            # Fallback an toàn: Enter trên ô password thường kích hoạt submit/JS login.
            password_box.press("Enter")

        # Không phụ thuộc hoàn toàn vào networkidle vì TimeSoft có thể polling.
        try:
            page.wait_for_load_state("domcontentloaded", timeout=10_000)
        except Exception:
            pass
        page.wait_for_timeout(2_000)
    except Exception as e:
        return False, f"Không gửi được biểu mẫu đăng nhập TimeSoft: {type(e).__name__}"

    # Xác minh session bằng chính trang report, tránh false-negative nếu Login dùng AJAX.
    target_url = verify_url or urljoin(
        TIMESOFT_BASE_URL + "/",
        TIMESOFT_REPORT_PAGES["summary_invoice"]["path"].lstrip("/"),
    )
    try:
        page.goto(target_url, wait_until="domcontentloaded", timeout=35_000)
        try:
            page.wait_for_load_state("networkidle", timeout=8_000)
        except Exception:
            pass
        page.wait_for_timeout(1_000)
    except Exception as e:
        return False, f"Đã gửi đăng nhập nhưng không mở được trang kiểm tra TimeSoft: {type(e).__name__}"

    final_url = str(page.url or "")
    try:
        still_password = _timesoft_visible_input(page, [
            'input[type="password"]',
            'input[name*="password" i]',
            'input[id*="password" i]',
        ])
    except Exception:
        still_password = None

    is_login_url = "/user/login" in final_url.lower()
    if is_login_url and still_password is not None:
        err = _timesoft_login_error_text(page)
        suffix = f" Thông báo từ TimeSoft: {err}" if err else ""
        return False, (
            "TimeSoft vẫn chuyển về trang đăng nhập sau khi xác minh session. "
            "Tài khoản đã được nạp từ Secrets nhưng form/luồng đăng nhập tự động chưa hoàn tất."
            + suffix
        )

    return True, f"Đăng nhập TimeSoft thành công · đã xác minh bằng trang báo cáo ({final_url})."


def timesoft_auto_discover_apis():
    """Tự đăng nhập và bắt XHR/fetch thật của 2 trang báo cáo để xác định API chính xác.

    Không đoán endpoint. Hàm chỉ chọn từ request thực tế do chính trang TimeSoft phát sinh.
    Yêu cầu môi trường có package playwright và Chromium.
    """
    if not timesoft_is_configured():
        return False, "Chưa khai báo đầy đủ [TIMESOFT] trong Streamlit Secrets.", {}

    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return False, (
            "Chưa có Playwright. Thêm `playwright` vào requirements.txt và cài Chromium "
            "(`playwright install chromium`) để bật Auto Discovery."
        ), {}

    captured = {key: [] for key in TIMESOFT_REPORT_PAGES}
    active_report = {"key": None}
    browser = None

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu"],
            )
            context = browser.new_context(
                ignore_https_errors=False,
                viewport={"width": 1440, "height": 1000},
                locale="vi-VN",
            )
            page = context.new_page()

            def on_response(response):
                report_key = active_report.get("key")
                if report_key not in captured:
                    return
                try:
                    req = response.request
                    resource_type = str(req.resource_type or "").lower()
                    if resource_type not in {"xhr", "fetch"}:
                        return
                    headers = _timesoft_safe_request_headers(req.headers)
                    content_type = str(response.headers.get("content-type", "") or "")
                    body_text = ""
                    if any(x in content_type.lower() for x in ["json", "text", "javascript", "html"]):
                        try:
                            body_text = response.text() or ""
                        except Exception:
                            body_text = ""
                    if len(body_text) > TIMESOFT_CAPTURE_MAX_CHARS:
                        body_text = body_text[:TIMESOFT_CAPTURE_MAX_CHARS]
                    post_data = req.post_data or ""
                    # Không capture request login vì active_report=None ở bước đăng nhập.
                    captured[report_key].append({
                        "method": str(req.method or "GET").upper(),
                        "url": str(req.url or ""),
                        "resource_type": resource_type,
                        "status": int(response.status or 0),
                        "content_type": content_type,
                        "post_data": str(post_data)[:50_000],
                        "request_headers": headers,
                        "json_kind": _timesoft_json_kind(body_text),
                        "body_preview": _timesoft_body_preview(body_text),
                    })
                except Exception:
                    return

            page.on("response", on_response)

            # Mở trang report trước: TimeSoft sẽ tự redirect đến Login nếu session chưa có.
            first_url = urljoin(
                TIMESOFT_BASE_URL + "/",
                TIMESOFT_REPORT_PAGES["summary_invoice"]["path"].lstrip("/"),
            )
            page.goto(first_url, wait_until="domcontentloaded", timeout=35_000)
            page.wait_for_timeout(1_000)
            ok_login, login_msg = _timesoft_login_with_playwright(page, verify_url=first_url)
            if not ok_login:
                context.close()
                browser.close()
                return False, login_msg, {}

            # Duyệt từng report. Request XHR/fetch được bắt trong on_response ở trên.
            for report_key, info in TIMESOFT_REPORT_PAGES.items():
                active_report["key"] = report_key
                report_url = urljoin(TIMESOFT_BASE_URL + "/", info["path"].lstrip("/"))
                page.goto(report_url, wait_until="domcontentloaded", timeout=35_000)
                # Chờ AJAX của report. networkidle có thể timeout nếu trang dùng polling, nên có fallback.
                try:
                    page.wait_for_load_state("networkidle", timeout=12_000)
                except Exception:
                    pass
                page.wait_for_timeout(4_000)

            active_report["key"] = None
            context.close()
            browser.close()
            browser = None

        picked = _timesoft_pick_best_candidates(captured)
        missing = [k for k, v in picked.items() if not v.get("best")]
        if missing:
            names = ", ".join(TIMESOFT_REPORT_PAGES[k]["label"] for k in missing)
            return False, f"Đã đăng nhập nhưng chưa bắt được XHR/fetch dữ liệu của: {names}.", picked
        return True, "Đã tự đăng nhập và xác định request API thực tế của cả 2 báo cáo TimeSoft.", picked

    except Exception as e:
        try:
            if browser is not None:
                browser.close()
        except Exception:
            pass
        return False, f"TimeSoft Auto Discovery lỗi: {type(e).__name__}: {e}", _timesoft_pick_best_candidates(captured)


def _timesoft_discovery_rows(discovery):
    rows = []
    for report_key, report_info in (discovery or {}).items():
        best = (report_info or {}).get("best") or {}
        rows.append({
            "Báo cáo": (report_info or {}).get("report", report_key),
            "Method": best.get("method", ""),
            "API URL": best.get("url", ""),
            "HTTP": best.get("status", ""),
            "Kiểu": best.get("json_kind", "") or best.get("content_type", ""),
            "Điểm nhận diện": best.get("score", ""),
        })
    return pd.DataFrame(rows)


def _timesoft_sanitized_discovery(discovery):
    """JSON kỹ thuật an toàn hơn: chỉ giữ request cần thiết, không xuất response/PII/token."""
    safe = {}
    allowed = {
        "method", "url", "resource_type", "status", "content_type",
        "post_data", "request_headers", "json_kind", "score",
    }
    for report_key, info in (discovery or {}).items():
        best = (info or {}).get("best") or {}
        safe_best = {k: v for k, v in best.items() if k in allowed}
        # Tuyệt đối không xuất Cookie/Authorization nếu về sau header capture thay đổi.
        headers = dict(safe_best.get("request_headers") or {})
        for bad in ["cookie", "authorization", "proxy-authorization"]:
            headers.pop(bad, None)
        safe_best["request_headers"] = headers
        safe[report_key] = {
            "report": (info or {}).get("report", report_key),
            "page_url": (info or {}).get("page_url", ""),
            "best": safe_best or None,
        }
    return safe


def _timesoft_date_range_text(start_date, end_date):
    if isinstance(start_date, datetime):
        start_date = start_date.date()
    if isinstance(end_date, datetime):
        end_date = end_date.date()
    if not isinstance(start_date, date) or not isinstance(end_date, date):
        raise ValueError("Khoảng ngày TimeSoft không hợp lệ.")
    if start_date > end_date:
        start_date, end_date = end_date, start_date
    return f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"


def _timesoft_requests_session_from_browser_cookies(cookies, user_agent=""):
    session = requests.Session()
    session.headers.update({
        "Accept-Language": "vi-VN,vi;q=0.9,en;q=0.7",
        "User-Agent": str(user_agent or "Mozilla/5.0"),
    })
    for cookie in cookies or []:
        name = str(cookie.get("name", "") or "")
        value = str(cookie.get("value", "") or "")
        if not name:
            continue
        kwargs = {}
        domain = str(cookie.get("domain", "") or "").strip()
        path = str(cookie.get("path", "") or "/").strip() or "/"
        if domain:
            kwargs["domain"] = domain
        if path:
            kwargs["path"] = path
        try:
            session.cookies.set(name, value, **kwargs)
        except Exception:
            session.cookies.set(name, value)
    return session


def _timesoft_new_authenticated_http_session():
    """Dùng Playwright chỉ để đăng nhập, sau đó chuyển cookie sang requests.Session."""
    if not timesoft_is_configured():
        return False, "Chưa khai báo đầy đủ [TIMESOFT] trong Streamlit Secrets.", None
    try:
        from playwright.sync_api import sync_playwright
    except Exception:
        return False, "Chưa cài Playwright/Chromium cho TimeSoft.", None

    browser = None
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=["--no-sandbox", "--disable-dev-shm-usage", "--disable-gpu"],
            )
            context = browser.new_context(
                ignore_https_errors=False,
                viewport={"width": 1440, "height": 1000},
                locale="vi-VN",
            )
            page = context.new_page()
            verify_url = urljoin(
                TIMESOFT_BASE_URL + "/",
                TIMESOFT_REPORT_PAGES["summary_invoice"]["path"].lstrip("/"),
            )
            page.goto(verify_url, wait_until="domcontentloaded", timeout=35_000)
            page.wait_for_timeout(600)
            ok_login, login_msg = _timesoft_login_with_playwright(page, verify_url=verify_url)
            if not ok_login:
                context.close()
                browser.close()
                browser = None
                return False, login_msg, None

            cookies = context.cookies()
            try:
                user_agent = page.evaluate("navigator.userAgent")
            except Exception:
                user_agent = "Mozilla/5.0"
            context.close()
            browser.close()
            browser = None

        http_session = _timesoft_requests_session_from_browser_cookies(cookies, user_agent)
        return True, "Đã tạo phiên TimeSoft và chuyển session sang API trực tiếp.", http_session
    except Exception as e:
        try:
            if browser is not None:
                browser.close()
        except Exception:
            pass
        return False, f"Không tạo được phiên API TimeSoft: {type(e).__name__}: {e}", None


def _timesoft_get_http_session(force_login=False):
    key_session = "_timesoft_http_session_v81"
    key_time = "_timesoft_http_session_time_v81"
    now_ts = time.time()
    if not force_login:
        existing = st.session_state.get(key_session)
        created_at = float(st.session_state.get(key_time, 0) or 0)
        if existing is not None and (now_ts - created_at) < TIMESOFT_HTTP_SESSION_TTL_SECONDS:
            return True, "Đang dùng lại phiên TimeSoft còn hiệu lực.", existing

    ok, msg, session = _timesoft_new_authenticated_http_session()
    if ok and session is not None:
        st.session_state[key_session] = session
        st.session_state[key_time] = now_ts
    return ok, msg, session


def _timesoft_post_json(session, api_path, referer_path, payload, timeout=60):
    url = urljoin(TIMESOFT_BASE_URL + "/", str(api_path).lstrip("/"))
    referer = urljoin(TIMESOFT_BASE_URL + "/", str(referer_path).lstrip("/"))
    headers = {
        "Accept": "application/json, text/plain, */*",
        "Content-Type": "application/json;charset=UTF-8",
        "Referer": referer,
        "Origin": TIMESOFT_BASE_URL,
        "X-Requested-With": "XMLHttpRequest",
    }
    try:
        resp = session.post(url, json=payload, headers=headers, timeout=timeout, allow_redirects=True)
    except Exception as e:
        return False, f"Không gọi được TimeSoft API: {type(e).__name__}: {e}", None, False

    final_url = str(resp.url or "")
    ctype = str(resp.headers.get("content-type", "") or "").lower()
    if resp.status_code in {401, 403} or "/user/login" in final_url.lower():
        return False, "Phiên đăng nhập TimeSoft đã hết hạn.", None, True
    if resp.status_code != 200:
        return False, f"TimeSoft API HTTP {resp.status_code}.", None, False
    if "json" not in ctype:
        # Login page đôi lúc trả HTTP 200 HTML.
        body_start = str(resp.text or "")[:1000].lower()
        if "login" in final_url.lower() or "password" in body_start or "đăng nhập" in body_start:
            return False, "Phiên đăng nhập TimeSoft không còn hợp lệ.", None, True
    try:
        data = resp.json()
    except Exception:
        return False, "TimeSoft API không trả JSON hợp lệ.", None, False

    return_code = str((data or {}).get("ReturnCode", "") or "")
    if return_code and return_code != "000000":
        msg = str((data or {}).get("Message", "") or "").strip()
        return False, f"TimeSoft trả mã {return_code}: {msg or 'Không rõ lỗi'}", data, False
    return True, "OK", data, False


def _timesoft_summary_invoice_payload(start_date, end_date):
    return {
        "objectSearch": {
            "TypeData": 0,
            "TypeInvoice": 1,
            "TachInvoiceSpa": 0,
            "CreateTimeRange": _timesoft_date_range_text(start_date, end_date),
        },
        "pageSize": 0,
    }


def _timesoft_employee_checkin_payload(start_date, end_date, page_index=1, page_size=TIMESOFT_CHECKIN_PAGE_SIZE):
    return {
        "objectSearch": {
            "CreateDateRange": _timesoft_date_range_text(start_date, end_date),
        },
        "pageIndex": int(page_index),
        "pageSize": int(page_size),
    }


def _timesoft_fetch_summary_invoice(session, start_date, end_date):
    cfg = TIMESOFT_DIRECT_APIS["summary_invoice"]
    payload = _timesoft_summary_invoice_payload(start_date, end_date)
    ok, msg, data, auth_failed = _timesoft_post_json(
        session, cfg["url"], cfg["referer"], payload
    )
    if not ok:
        return False, msg, {}, pd.DataFrame(), auth_failed
    rows = (data or {}).get("Data") or []
    df = pd.json_normalize(rows, sep=".") if isinstance(rows, list) and rows else pd.DataFrame()
    summary_keys = [
        "Total", "TotalMoney", "TotalDiscount", "TotalQuantity", "TotalInvoice",
        "TotalActualRevenu", "TotalPromotion", "TotalPayment", "TotalActualTerm",
        "TotalDebt", "ReportByYear", "Message",
    ]
    meta = {k: (data or {}).get(k) for k in summary_keys if k in (data or {})}
    return True, "Đã lấy báo cáo tổng hợp doanh thu.", meta, df, False


def _timesoft_fetch_employee_checkin(session, start_date, end_date):
    cfg = TIMESOFT_DIRECT_APIS["employee_checkin"]
    all_rows = []
    total_expected = None
    for page_index in range(1, TIMESOFT_MAX_CHECKIN_PAGES + 1):
        payload = _timesoft_employee_checkin_payload(
            start_date, end_date, page_index=page_index, page_size=TIMESOFT_CHECKIN_PAGE_SIZE
        )
        ok, msg, data, auth_failed = _timesoft_post_json(
            session, cfg["url"], cfg["referer"], payload
        )
        if not ok:
            return False, msg, pd.DataFrame(), {"Total": total_expected}, auth_failed
        if total_expected is None:
            try:
                total_expected = int((data or {}).get("Total") or 0)
            except Exception:
                total_expected = 0
        page_rows = (data or {}).get("Data") or []
        if not isinstance(page_rows, list) or not page_rows:
            break
        all_rows.extend(page_rows)
        if total_expected and len(all_rows) >= total_expected:
            break

    df = pd.json_normalize(all_rows, sep=".") if all_rows else pd.DataFrame()
    meta = {"Total": total_expected if total_expected is not None else len(all_rows)}
    return True, "Đã lấy báo cáo chấm công nhân viên.", df, meta, False


def timesoft_direct_sync(start_date, end_date, force_login=False):
    """Đồng bộ trực tiếp 2 API. Nếu session hết hạn, tự đăng nhập lại và thử thêm 1 lần."""
    if not timesoft_is_configured():
        return False, "Chưa cấu hình TimeSoft trong Secrets.", {}

    last_error = ""
    for attempt in range(2):
        force_now = bool(force_login or attempt > 0)
        ok_session, session_msg, session = _timesoft_get_http_session(force_login=force_now)
        if not ok_session or session is None:
            return False, session_msg, {}

        ok_inv, msg_inv, inv_meta, inv_df, inv_auth = _timesoft_fetch_summary_invoice(
            session, start_date, end_date
        )
        if not ok_inv:
            last_error = msg_inv
            if inv_auth and attempt == 0:
                continue
            return False, msg_inv, {}

        ok_chk, msg_chk, chk_df, chk_meta, chk_auth = _timesoft_fetch_employee_checkin(
            session, start_date, end_date
        )
        if not ok_chk:
            last_error = msg_chk
            if chk_auth and attempt == 0:
                continue
            return False, msg_chk, {}

        result = {
            "start_date": start_date,
            "end_date": end_date,
            "synced_at": datetime.now(VN_TZ),
            "session_message": session_msg,
            "summary_invoice_meta": inv_meta,
            "summary_invoice_df": inv_df,
            "employee_checkin_meta": chk_meta,
            "employee_checkin_df": chk_df,
        }
        return True, (
            f"Lấy dữ liệu TimeSoft thành công: {len(inv_df)} dòng doanh thu chi tiết, "
            f"{len(chk_df)} dòng chấm công."
        ), result

    return False, last_error or "Không đồng bộ được TimeSoft.", {}


# ==========================================================
# V82 - SNAPSHOT ĐỒNG BỘ NỀN CLOUD RUN / POSTGRESQL
# ==========================================================
def _timesoft_bg_key(prefix, target_date=None):
    d = target_date if isinstance(target_date, date) else get_vn_today()
    return f"{prefix}_{d.strftime('%Y%m%d')}"


def _timesoft_read_background_snapshot(target_date=None):
    """Đọc snapshot do Cloud Run Job ghi vào PostgreSQL; không gọi TimeSoft."""
    if vpg is None or not vpg.is_enabled():
        return {}
    d = target_date if isinstance(target_date, date) else get_vn_today()
    try:
        return {
            "date": d,
            "status": vpg.read_dataset("timesoft_background_status", allow_stale=True),
            "summary_invoice": vpg.read_dataset(_timesoft_bg_key("timesoft_summary_invoice", d), allow_stale=True),
            "summary_totals": vpg.read_dataset(_timesoft_bg_key("timesoft_summary_totals", d), allow_stale=True),
            "employee_checkin": vpg.read_dataset(_timesoft_bg_key("timesoft_employee_checkin", d), allow_stale=True),
        }
    except Exception:
        return {}


def _timesoft_background_status_row():
    snap = _timesoft_read_background_snapshot(get_vn_today())
    df = snap.get("status") if isinstance(snap, dict) else None
    if isinstance(df, pd.DataFrame) and not df.empty:
        return df.iloc[0].to_dict()
    return {}


def _timesoft_checkin_display_df(df):
    if not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame()
    preferred = [
        "WorkDateStr",
        "employeeInfo.EmployeeCode",
        "employeeInfo.Name",
        "WorkTimeName",
        "StartWorkTime",
        "EndWorkTime",
        "MachineTimeCheckInStr",
        "MachineTimeCheckOutStr",
        "GoWorkTypeName",
        "LastCheckInTypeName",
        "TotalMinuteInGoLate",
        "TotalMinuteBackHomeEarly",
        "TotalMinuteInDay",
        "TotalCheckInOneDay",
    ]
    cols = [c for c in preferred if c in df.columns]
    out = df[cols].copy() if cols else df.copy()
    rename = {
        "WorkDateStr": "Ngày",
        "employeeInfo.EmployeeCode": "Mã NV",
        "employeeInfo.Name": "Nhân viên",
        "WorkTimeName": "Ca",
        "StartWorkTime": "Giờ bắt đầu ca",
        "EndWorkTime": "Giờ kết thúc ca",
        "MachineTimeCheckInStr": "Check-in",
        "MachineTimeCheckOutStr": "Check-out",
        "GoWorkTypeName": "Trạng thái vào",
        "LastCheckInTypeName": "Trạng thái ra",
        "TotalMinuteInGoLate": "Phút đi trễ",
        "TotalMinuteBackHomeEarly": "Phút về sớm",
        "TotalMinuteInDay": "Tổng phút trong ngày",
        "TotalCheckInOneDay": "Số lần check-in",
    }
    return out.rename(columns=rename)


def render_timesoft_background_snapshot_today(show_status=True):
    """Hiển thị snapshot TimeSoft nền hôm nay. Chỉ Admin được phép gọi/nhìn thấy."""
    if str(st.session_state.get("current_role", "")).strip().lower() != "admin":
        return

    if vpg is None or not vpg.is_enabled():
        st.warning(
            "PostgreSQL chưa được bật nên chưa thể đọc snapshot nền. "
            "Trên Cloud Run cần bật PostgreSQL/Cloud SQL và để Cloud Scheduler chạy Job."
        )
        return

    bg_status = _timesoft_background_status_row()
    if show_status:
        if bg_status:
            bg_ok = str(bg_status.get("status", "")).lower() == "success"
            last_sync = str(bg_status.get("synced_at_vn", "") or bg_status.get("synced_at", ""))
            (st.success if bg_ok else st.warning)(
                f"{'✅' if bg_ok else '⚠️'} Cloud Run Job: {bg_status.get('status', 'unknown')} · "
                f"lần chạy gần nhất {last_sync or 'chưa rõ'}"
            )
            b1, b2, b3, b4 = st.columns(4)
            b1.metric("Chu kỳ", f"{TIMESOFT_BACKGROUND_INTERVAL_MINUTES} phút")
            b2.metric("Doanh thu · dòng", int(float(bg_status.get("invoice_rows", 0) or 0)))
            b3.metric("Chấm công · dòng", int(float(bg_status.get("checkin_rows", 0) or 0)))
            b4.metric("Thời gian chạy", f"{float(bg_status.get('duration_seconds', 0) or 0):.1f}s")
        else:
            st.info(
                "PostgreSQL đã bật nhưng chưa có snapshot TimeSoft nền. "
                "Sau khi Cloud Scheduler gọi Cloud Run Job thành công lần đầu, dữ liệu sẽ xuất hiện tại đây."
            )

    bg_date = get_vn_today()
    bg_snap = _timesoft_read_background_snapshot(bg_date)

    bg_tot = bg_snap.get("summary_totals") if isinstance(bg_snap, dict) else None
    if isinstance(bg_tot, pd.DataFrame) and not bg_tot.empty:
        tr = bg_tot.iloc[0].to_dict()
        t1, t2, t3 = st.columns(3)
        t1.metric("Tổng tiền", f"{float(tr.get('TotalMoney', 0) or 0):,.0f} đ".replace(",", "."))
        t2.metric("Giảm giá", f"{float(tr.get('TotalDiscount', 0) or 0):,.0f} đ".replace(",", "."))
        t3.metric("Doanh thu thực", f"{float(tr.get('TotalActualRevenu', 0) or 0):,.0f} đ".replace(",", "."))

    bg_inv = bg_snap.get("summary_invoice") if isinstance(bg_snap, dict) else None
    if isinstance(bg_inv, pd.DataFrame) and not bg_inv.empty:
        st.caption("Doanh thu nền hôm nay")
        st.dataframe(bg_inv, width="stretch", hide_index=True, height=300)
    else:
        st.info("Chưa có dữ liệu doanh thu nền hôm nay.")

    bg_chk = bg_snap.get("employee_checkin") if isinstance(bg_snap, dict) else None
    bg_chk_view = _timesoft_checkin_display_df(bg_chk)
    if isinstance(bg_chk_view, pd.DataFrame) and not bg_chk_view.empty:
        st.caption("Chấm công nền hôm nay")
        st.dataframe(bg_chk_view, width="stretch", hide_index=True, height=420)
    else:
        st.info("Chưa có dữ liệu chấm công nền hôm nay.")


def _timesoft_export_workbook(sync_result):
    output = io.BytesIO()
    inv_df = sync_result.get("summary_invoice_df")
    chk_df = sync_result.get("employee_checkin_df")
    inv_meta = sync_result.get("summary_invoice_meta") or {}
    chk_meta = sync_result.get("employee_checkin_meta") or {}
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_rows = []
        for k, v in inv_meta.items():
            summary_rows.append({"Nhóm": "Doanh thu", "Chỉ số": k, "Giá trị": v})
        for k, v in chk_meta.items():
            summary_rows.append({"Nhóm": "Chấm công", "Chỉ số": k, "Giá trị": v})
        pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name="TongHop")
        (inv_df if isinstance(inv_df, pd.DataFrame) else pd.DataFrame()).to_excel(
            writer, index=False, sheet_name="DoanhThu"
        )
        (chk_df if isinstance(chk_df, pd.DataFrame) else pd.DataFrame()).to_excel(
            writer, index=False, sheet_name="ChamCong"
        )
    return output.getvalue()


# --- CHUẨN HÓA TÊN / TÀI KHOẢN ---
def normalize_name(name):
    """Đồng nhất cách gõ Thúy/Thuý để tránh lỗi so sánh dữ liệu nghiệp vụ."""
    return str(name).replace("Thuý", "Thúy").replace("thuý", "thúy").strip()

def remove_vietnamese_accents(value):
    """Bỏ dấu tiếng Việt nhưng vẫn giữ nguyên số 0 đầu chuỗi và ký tự khác."""
    text = unicodedata.normalize("NFD", str(value))
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return text.replace("đ", "d").replace("Đ", "D")

def normalize_login_name(value):
    """Tên đăng nhập: không phân biệt dấu, HOA/thường; không ép kiểu số."""
    return " ".join(remove_vietnamese_accents(str(value)).strip().split()).casefold()

def clean_employee_match_name(value):
    """
    Tên dùng để ĐỐI CHIẾU dữ liệu nhân viên giữa Bảng tour / TimeSoft / Google Sheet.
    Dấu * ở cuối tên trong Bảng tour chỉ là ký hiệu vận hành, không phải một phần tên nhân viên.
    Ví dụ: ``Cẩm Nhung *`` == ``Cẩm Nhung``.
    Không dùng hàm này để thay đổi tên đăng nhập hiển thị/lưu hồ sơ.
    """
    text = normalize_name(value)
    text = re.sub(r"\s*\*+\s*$", "", str(text)).strip()
    return " ".join(text.split())

def normalize_employee_match_name(value):
    return normalize_login_name(clean_employee_match_name(value))

def sort_employee_names(values):
    """Sắp xếp tên nhân viên A→Z, bỏ trùng theo chuẩn không dấu/không phân biệt hoa thường."""
    by_key = {}
    for value in values or []:
        name = str(value or "").strip()
        if not name:
            continue
        key = normalize_login_name(name)
        if key and key not in by_key:
            by_key[key] = name
    return sorted(by_key.values(), key=lambda x: normalize_login_name(x))

def password_matches(input_password, stored_password):
    """Mật khẩu được so sánh đúng ký tự, có phân biệt HOA/thường và chấp nhận ký tự đặc biệt/0 đầu."""
    return hmac.compare_digest(str(input_password), str(stored_password))

def is_locked_value(value):
    return str(value).strip().casefold() in {"1", "true", "yes", "y", "khóa", "khoa", "locked", "x"}

def clean_leave_reason_display(value):
    """Làm sạch tên lý do nghỉ để hiển thị/dropdown luôn khớp chính xác."""
    text = str(value or "").replace("🔴", "").strip()
    return " ".join(text.split())

def normalize_leave_reason(value):
    """Chuẩn hóa loại nghỉ để so sánh trùng dữ liệu ổn định."""
    return clean_leave_reason_display(value).casefold()

PROGRESSIVE_PENALTY_REASONS = {
    normalize_login_name("Nghỉ không phép"): "Nghỉ không phép",
    normalize_login_name("Đi trễ không phép"): "Đi trễ không phép",
    normalize_login_name("Về sớm không phép"): "Về sớm không phép",
    # Chấp nhận tên cũ/biến thể nếu danh mục đang dùng "Ra sớm không phép".
    normalize_login_name("Ra sớm không phép"): "Về sớm không phép",
}

# Các loại này tuyệt đối KHÔNG tham gia xếp hạng Người Thứ 1/2/3... dù tên danh mục
# sau này có bị chỉnh gần giống một nhóm phạt lũy tiến.
PROGRESSIVE_PENALTY_EXCLUDED_REASONS = {
    normalize_login_name("Ra ngoài vào muộn dưới 30 phút"),
    normalize_login_name("Ra ngoài vào muộn dưới 60 phút"),
    normalize_login_name("Ra ngoài vào muộn dưới 120 phút"),
}

def get_progressive_penalty_reason(value):
    """Trả về tên chuẩn nếu loại nghỉ thuộc nhóm phạt lũy tiến, ngược lại trả về None."""
    key = normalize_login_name(str(value).replace("🔴", "").strip())
    if key in PROGRESSIVE_PENALTY_EXCLUDED_REASONS:
        return None
    return PROGRESSIVE_PENALTY_REASONS.get(key)

def is_progressive_penalty_reason(value):
    return get_progressive_penalty_reason(value) is not None

def is_nghi_khong_phep_reason(value):
    """Giữ tương thích với code cũ: chỉ kiểm tra riêng Nghỉ không phép."""
    return get_progressive_penalty_reason(value) == "Nghỉ không phép"

def _fallback_admin_remember_token():
    """Token bền vững cho tài khoản admin dự phòng; không lưu mật khẩu trong trình duyệt."""
    try:
        secret = str(st.secrets.get("vera_persistent_login_secret", "VERA-SPA-PERSISTENT-LOGIN-2026"))
    except Exception:
        secret = "VERA-SPA-PERSISTENT-LOGIN-2026"
    digest = hmac.new(secret.encode("utf-8"), b"fallback-admin", hashlib.sha256).hexdigest()
    return "vera_admin_" + digest

def _is_valid_fallback_admin_token(token):
    if not token:
        return False
    return hmac.compare_digest(str(token), _fallback_admin_remember_token())

# --- DANH SÁCH NGÂN HÀNG VIỆT NAM (VietQR, tự làm mới mỗi 24 giờ) ---
VIETQR_BANKS_API = "https://api.vietqr.io/v2/banks"

# Danh sách dự phòng khi API bên ngoài tạm thời không truy cập được.
FALLBACK_VN_BANKS = [
    ("Vietcombank", "Ngân hàng TMCP Ngoại thương Việt Nam"),
    ("VietinBank", "Ngân hàng TMCP Công thương Việt Nam"),
    ("BIDV", "Ngân hàng TMCP Đầu tư và Phát triển Việt Nam"),
    ("Agribank", "Ngân hàng Nông nghiệp và Phát triển Nông thôn Việt Nam"),
    ("MBBank", "Ngân hàng TMCP Quân đội"),
    ("Techcombank", "Ngân hàng TMCP Kỹ thương Việt Nam"),
    ("ACB", "Ngân hàng TMCP Á Châu"),
    ("VPBank", "Ngân hàng TMCP Việt Nam Thịnh Vượng"),
    ("TPBank", "Ngân hàng TMCP Tiên Phong"),
    ("Sacombank", "Ngân hàng TMCP Sài Gòn Thương Tín"),
    ("HDBank", "Ngân hàng TMCP Phát triển Thành phố Hồ Chí Minh"),
    ("VIB", "Ngân hàng TMCP Quốc tế Việt Nam"),
    ("SHB", "Ngân hàng TMCP Sài Gòn - Hà Nội"),
    ("Eximbank", "Ngân hàng TMCP Xuất Nhập khẩu Việt Nam"),
    ("MSB", "Ngân hàng TMCP Hàng Hải Việt Nam"),
    ("OCB", "Ngân hàng TMCP Phương Đông"),
    ("PVcomBank", "Ngân hàng TMCP Đại Chúng Việt Nam"),
    ("LPBank", "Ngân hàng TMCP Lộc Phát Việt Nam"),
    ("SeABank", "Ngân hàng TMCP Đông Nam Á"),
    ("ABBANK", "Ngân hàng TMCP An Bình"),
    ("BacABank", "Ngân hàng TMCP Bắc Á"),
    ("NamABank", "Ngân hàng TMCP Nam Á"),
    ("NCB", "Ngân hàng TMCP Quốc Dân"),
    ("VietABank", "Ngân hàng TMCP Việt Á"),
    ("VietBank", "Ngân hàng TMCP Việt Nam Thương Tín"),
    ("BaoVietBank", "Ngân hàng TMCP Bảo Việt"),
    ("KienLongBank", "Ngân hàng TMCP Kiên Long"),
    ("PGBank", "Ngân hàng TMCP Thịnh vượng và Phát triển"),
    ("SaigonBank", "Ngân hàng TMCP Sài Gòn Công Thương"),
    ("SCB", "Ngân hàng TMCP Sài Gòn"),
    ("COOPBANK", "Ngân hàng Hợp tác xã Việt Nam"),
    ("ShinhanBank", "Ngân hàng TNHH MTV Shinhan Việt Nam"),
    ("Woori", "Ngân hàng TNHH MTV Woori Việt Nam"),
    ("HSBC", "Ngân hàng TNHH MTV HSBC (Việt Nam)"),
    ("StandardChartered", "Ngân hàng TNHH MTV Standard Chartered Bank Việt Nam"),
    ("PublicBank", "Ngân hàng TNHH MTV Public Việt Nam"),
    ("CIMB", "Ngân hàng TNHH MTV CIMB Việt Nam"),
    ("HongLeong", "Ngân hàng TNHH MTV Hong Leong Việt Nam"),
    ("MBV", "Ngân hàng TNHH MTV Việt Nam Hiện Đại"),
    ("Vikki", "Ngân hàng TNHH MTV Số Vikki"),
    ("GPBank", "Ngân hàng Thương mại TNHH MTV Dầu Khí Toàn Cầu"),
    ("CBBank", "Ngân hàng Thương mại TNHH MTV Xây dựng Việt Nam"),
    ("VRB", "Ngân hàng Liên doanh Việt - Nga"),
    ("IndovinaBank", "Ngân hàng TNHH Indovina"),
    ("KBank", "Ngân hàng Đại chúng TNHH Kasikornbank"),
    ("VBSP", "Ngân hàng Chính sách Xã hội"),
]

@st.cache_data(ttl=86400, show_spinner=False)
def load_vietnam_banks():
    """Lấy danh sách ngân hàng từ VietQR; trả về list dict gồm label/value."""
    try:
        r = requests.get(VIETQR_BANKS_API, timeout=12)
        r.raise_for_status()
        payload = r.json()
        rows = payload.get("data", []) if isinstance(payload, dict) else []
        banks = []
        seen = set()
        for item in rows:
            if not isinstance(item, dict):
                continue
            full_name = str(item.get("name", "")).replace("\n", " ").strip()
            short_name = str(item.get("shortName", item.get("short_name", ""))).strip()
            code = str(item.get("code", "")).strip()
            if not full_name:
                continue
            # Chỉ giữ ngân hàng / chi nhánh ngân hàng / ngân hàng số, loại ví & công ty tài chính.
            name_fold = remove_vietnamese_accents(full_name).casefold()
            if not ("ngan hang" in name_fold or "bank" in name_fold):
                continue
            unique_key = (remove_vietnamese_accents(full_name).casefold(), short_name.casefold())
            if unique_key in seen:
                continue
            seen.add(unique_key)
            short_display = short_name or code
            label = f"{short_display} — {full_name}" if short_display else full_name
            banks.append({"label": label, "value": full_name, "short": short_display})
        if banks:
            return sorted(banks, key=lambda x: remove_vietnamese_accents(x["label"]).casefold())
    except Exception:
        pass

    return [
        {"label": f"{short} — {name}", "value": name, "short": short}
        for short, name in FALLBACK_VN_BANKS
    ]

def bank_selectbox(label, key, current_value=""):
    """Dropdown ngân hàng có ô gõ tìm kiếm tích hợp của Streamlit."""
    banks = load_vietnam_banks()
    current = str(current_value or "").strip()

    # Nếu dữ liệu cũ chưa khớp tên mới từ API, vẫn giữ làm lựa chọn đầu tiên.
    if current and not any(normalize_login_name(x["value"]) == normalize_login_name(current) for x in banks):
        banks = [{"label": f"{current} (đang lưu)", "value": current, "short": ""}] + banks

    placeholder = "-- Chọn ngân hàng --"
    labels = [placeholder] + [x["label"] for x in banks]
    label_to_value = {x["label"]: x["value"] for x in banks}
    index = 0
    if current:
        for i, item in enumerate(banks, start=1):
            if normalize_login_name(item["value"]) == normalize_login_name(current) or normalize_login_name(item.get("short", "")) == normalize_login_name(current):
                index = i
                break

    selected = st.selectbox(
        label,
        labels,
        index=index,
        key=key,
        filter_mode="contains",
        placeholder="Gõ tên hoặc tên viết tắt ngân hàng để tìm...",
        help="Mở danh sách rồi gõ tên ngân hàng hoặc tên viết tắt, ví dụ: VCB, Vietcombank, ACB, MB..."
    )
    return "" if selected == placeholder else label_to_value.get(selected, selected)

# --- ĐỊA CHỈ HÀNH CHÍNH VIỆT NAM (SAU SÁP NHẬP 07/2025: 34 TỈNH/THÀNH, 2 CẤP) ---
VN_ADMIN_API_V2 = "https://provinces.open-api.vn/api/v2/"
FALLBACK_VN_PROVINCES_2025 = [
    "An Giang", "Bắc Ninh", "Cà Mau", "Cần Thơ", "Cao Bằng", "Đà Nẵng",
    "Đắk Lắk", "Điện Biên", "Đồng Nai", "Đồng Tháp", "Gia Lai", "Hà Nội",
    "Hà Tĩnh", "Hải Phòng", "Hồ Chí Minh", "Huế", "Hưng Yên", "Khánh Hòa",
    "Lai Châu", "Lâm Đồng", "Lạng Sơn", "Lào Cai", "Nghệ An", "Ninh Bình",
    "Phú Thọ", "Quảng Ngãi", "Quảng Ninh", "Quảng Trị", "Sơn La", "Tây Ninh",
    "Thái Nguyên", "Thanh Hóa", "Tuyên Quang", "Vĩnh Long"
]

@st.cache_data(ttl=604800, show_spinner=False)
def load_vietnam_admin_divisions():
    """
    Lấy dữ liệu hành chính Việt Nam sau 01/07/2025 từ Province Open API v2.
    Trả về list: [{code, name, wards:[{code,name}]}]. Cache 7 ngày để không gọi API liên tục.
    """
    try:
        r = requests.get(VN_ADMIN_API_V2, params={"depth": 2}, timeout=15)
        r.raise_for_status()
        payload = r.json()
        result = []
        if isinstance(payload, list):
            for item in payload:
                if not isinstance(item, dict):
                    continue
                p_name = str(item.get("name", "")).strip()
                p_name = re.sub(r"^(Tỉnh|Thành phố)\s+", "", p_name, flags=re.IGNORECASE).strip()
                p_code = item.get("code", "")
                wards_raw = item.get("wards") or item.get("communes") or item.get("children") or []
                wards = []
                if isinstance(wards_raw, list):
                    for w in wards_raw:
                        if isinstance(w, dict) and str(w.get("name", "")).strip():
                            wards.append({"code": w.get("code", ""), "name": str(w.get("name", "")).strip()})
                if p_name:
                    result.append({"code": p_code, "name": p_name, "wards": wards})
        if result:
            return sorted(result, key=lambda x: remove_vietnamese_accents(x["name"]).casefold()), ""
    except Exception as e:
        return ([{"code": "", "name": x, "wards": []} for x in FALLBACK_VN_PROVINCES_2025],
                f"Không tải được danh mục Phường/Xã trực tuyến: {e}")
    return ([{"code": "", "name": x, "wards": []} for x in FALLBACK_VN_PROVINCES_2025],
            "Không nhận được dữ liệu hành chính trực tuyến.")

def _address_component_match(text, candidate):
    t = normalize_login_name(text)
    c = normalize_login_name(candidate)
    return bool(c and c in t)

def parse_combined_vietnam_address(address, divisions=None):
    """Tách gần đúng địa chỉ cũ thành địa chỉ chi tiết + Phường/Xã + Tỉnh/Thành."""
    raw = str(address or "").strip()
    if not raw:
        return "", "", ""
    divisions = divisions or load_vietnam_admin_divisions()[0]
    province = ""
    ward = ""
    province_obj = None
    # Ưu tiên tên dài để tránh trùng một phần.
    for p in sorted(divisions, key=lambda x: len(str(x.get("name", ""))), reverse=True):
        if _address_component_match(raw, p.get("name", "")):
            province = str(p.get("name", "")).strip()
            province_obj = p
            break
    if province_obj:
        for w in sorted(province_obj.get("wards", []), key=lambda x: len(str(x.get("name", ""))), reverse=True):
            if _address_component_match(raw, w.get("name", "")):
                ward = str(w.get("name", "")).strip()
                break
    parts = [x.strip() for x in raw.split(",") if x.strip()]
    detail_parts = []
    for part in parts:
        n = normalize_login_name(part)
        if province and n == normalize_login_name(province):
            continue
        if ward and n == normalize_login_name(ward):
            continue
        detail_parts.append(part)
    detail = ", ".join(detail_parts).strip()
    if not detail and raw and not province and not ward:
        detail = raw
    return detail, ward, province

def combine_vietnam_address(detail, ward, province):
    parts = [str(x).strip().strip(",") for x in (detail, ward, province) if str(x).strip().strip(",")]
    # Loại trùng nếu người dùng gõ lại tên Phường/Xã/Tỉnh trong ô chi tiết.
    out = []
    seen = set()
    for x in parts:
        k = normalize_login_name(x)
        if k not in seen:
            out.append(x)
            seen.add(k)
    return ", ".join(out)

def vietnam_address_inputs(prefix, current_address="", show_preview=True):
    """
    Render 3 box: Tỉnh/Thành phố -> Phường/Xã -> Địa chỉ chi tiết.
    Không đặt trong st.form vì Phường/Xã phải đổi ngay khi Tỉnh/Thành thay đổi.
    Kết quả trả về là CHUỖI ĐÃ GHÉP để lưu vào đúng 1 cột Địa chỉ.
    """
    divisions, api_err = load_vietnam_admin_divisions()
    parsed_detail, parsed_ward, parsed_province = parse_combined_vietnam_address(current_address, divisions)
    province_names = [str(p.get("name", "")).strip() for p in divisions if str(p.get("name", "")).strip()]
    province_options = [""] + province_names
    p_key = f"{prefix}_province"
    w_key = f"{prefix}_ward"
    d_key = f"{prefix}_detail"
    manual_w_key = f"{prefix}_ward_manual"

    if p_key not in st.session_state:
        st.session_state[p_key] = parsed_province if parsed_province in province_names else ""
    if d_key not in st.session_state:
        st.session_state[d_key] = parsed_detail

    province = st.selectbox(
        "Tỉnh/Thành phố", province_options, key=p_key, filter_mode="contains",
        placeholder="Gõ để tìm Tỉnh/Thành phố..."
    )
    province_obj = next((p for p in divisions if str(p.get("name", "")).strip() == province), None)
    ward_names = [str(w.get("name", "")).strip() for w in (province_obj or {}).get("wards", []) if str(w.get("name", "")).strip()]

    if ward_names:
        ward_options = [""] + sorted(ward_names, key=lambda x: remove_vietnamese_accents(x).casefold())
        existing_ward = st.session_state.get(w_key, "")
        if existing_ward not in ward_options:
            st.session_state[w_key] = parsed_ward if parsed_ward in ward_options else ""
        ward = st.selectbox(
            "Phường/Xã", ward_options, key=w_key, filter_mode="contains",
            placeholder="Gõ để tìm Phường/Xã..."
        )
    else:
        # API tạm lỗi hoặc tỉnh chưa có danh mục: vẫn cho nhập tay để công việc không bị chặn.
        if manual_w_key not in st.session_state:
            st.session_state[manual_w_key] = parsed_ward
        ward = st.text_input("Phường/Xã", key=manual_w_key, placeholder="Nhập Phường/Xã")
        if api_err:
            st.caption("⚠️ Danh mục Phường/Xã trực tuyến đang tạm không khả dụng; có thể nhập tay.")

    detail = st.text_input(
        "Địa chỉ chi tiết", key=d_key,
        placeholder="Số nhà, tên đường, khu phố/thôn/ấp..."
    )
    combined = combine_vietnam_address(detail, ward, province)
    if show_preview:
        st.caption(f"📍 Địa chỉ sẽ lưu: {combined or '(chưa nhập)'}")
    return combined

def employee_registration_window(today=None):
    """Nhân viên được thao tác từ hôm nay đến hết tháng kế tiếp."""
    today = today or get_vn_today()
    if today.month == 12:
        next_month_first = date(today.year + 1, 1, 1)
    else:
        next_month_first = date(today.year, today.month + 1, 1)
    max_date = next_month_first.replace(day=calendar.monthrange(next_month_first.year, next_month_first.month)[1])
    return today, max_date

def normalize_schedule_date(value):
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value.strftime('%d/%m/%Y')
    try:
        parsed = pd.to_datetime(str(value).strip(), dayfirst=True, errors='raise')
        return parsed.strftime('%d/%m/%Y')
    except Exception:
        return str(value).strip()

def render_leave_filter_label_css():
    """Tăng cỡ chữ + Bold cho nhãn bộ lọc thời gian / nhân viên ở các trang lịch nghỉ."""
    st.markdown(
        """
        <style>
        div[data-testid="stSelectbox"] > label,
        div[data-testid="stTextInput"] > label {
            font-size: 18px !important;
            font-weight: 700 !important;
            line-height: 1.25 !important;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def schedule_key(row):
    reason_col = 'Lý do nghỉ' if 'Lý do nghỉ' in row else 'Loại nghỉ'
    return (
        normalize_schedule_date(row.get('Ngày', '')),
        normalize_login_name(row.get('Tên nhân viên', '')),
        remove_vietnamese_accents(str(row.get(reason_col, '')).strip()).casefold(),
    )

# --- HÀM ĐỊNH DẠNG BẢNG HIỂN THỊ TRỰC QUAN ---
def format_display_df(df):
    d = df.copy()
    def fmt_num(val):
        if pd.isna(val) or val == "": return ""
        try:
            v = float(val)
            if v == 0: return ""
            return str(int(v)) if v.is_integer() else str(v)
        except: return str(val)
    
    for col in ['Số ngày tính', 'Số ngày phép cộng dồn']:
        if col in d.columns:
            d[col] = d[col].apply(fmt_num)
            
    if 'Ngày' in d.columns:
        d['Ngày'] = pd.to_datetime(d['Ngày'], errors='coerce').dt.strftime('%d/%m/%Y').fillna(d['Ngày'])
        
    return d

# --- HÀM GỬI EMAIL BÁO CÁO ---
def send_email_report(sender_email, sender_password, to_email, emp_name, df_emp, total_phat, start_str, end_str):
    try:
        subject = f"Báo cáo chi tiết lịch nghỉ và vi phạm - {emp_name} ({start_str} đến {end_str})"
        
        # Định dạng lại bảng để hiển thị đẹp trong email
        df_display = format_display_df(df_emp[['Ngày', 'Lý do nghỉ', 'Chi tiết', 'Số ngày tính', 'Phạt vi phạm']])
        df_display['Phạt vi phạm'] = df_display['Phạt vi phạm'].apply(lambda x: f"{float(x):,.0f}" if float(x) > 0 else "")
        
        # Thêm style CSS cho bảng HTML
        html_table = df_display.to_html(index=False, justify='center')
        html_table = html_table.replace('<table border="1" class="dataframe">', '<table style="width:100%; border-collapse: collapse; border: 1px solid #D9D9D9; font-family: Arial, sans-serif;">')
        html_table = html_table.replace('<th>', '<th style="background-color: #f2f2f2; border: 1px solid #D9D9D9; padding: 8px; text-align: center; white-space: normal; overflow-wrap: anywhere; word-break: break-word;">')
        html_table = html_table.replace('<td>', '<td style="border: 1px solid #D9D9D9; padding: 8px; text-align: center;">')
        
        html_content = f"""
        <html>
        <body>
            <p>Chào <b>{emp_name}</b>,</p>
            <p>Hệ thống quản lý Vera Spa gửi bạn chi tiết lịch nghỉ và vi phạm trong giai đoạn từ <b>{start_str}</b> đến <b>{end_str}</b>:</p>
            <br>
            {html_table}
            <br>
            <h3 style="color: red;">Tổng tiền phạt vi phạm: {total_phat:,.0f} VNĐ</h3>
            <p><i>Vui lòng kiểm tra lại thông tin. Nếu có bất kỳ sai sót nào, xin vui lòng phản hồi lại trong thời gian sớm nhất.</i></p>
            <br>
            <p>Trân trọng,</p>
            <p><b>VERA SPA</b></p>
        </body>
        </html>
        """
        
        msg = MIMEMultipart()
        msg['From'] = f"Vera Spa <{sender_email}>"
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(html_content, 'html'))
        
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(msg)
        server.quit()
        
        return True, "Thành công"
    except Exception as e:
        return False, str(e)


# --- THEO DÕI SỐ NGƯỜI ĐANG TRUY CẬP & TRẠNG THÁI HỆ THỐNG ---
@st.cache_resource
def get_active_users():
    return {}

@st.cache_resource
def get_system_status():
    # Giữ để tương thích các phiên bản cũ; V86.13 dùng khóa đăng ký theo từng vai trò lưu Google Sheet.
    return {"lock_nv": False}

REGISTRATION_LOCK_WORKSHEET = "KhoaQuyenDangKy"
REGISTRATION_LOCK_ROLES = ["quanly", "letan", "leader", "nhanvien", "locker", "tapvu"]
REGISTRATION_LOCK_LABELS = {
    "quanly": "Quản lý",
    "letan": "Lễ tân",
    "leader": "Leader",
    "nhanvien": "Nhân viên",
    "locker": "Locker",
    "tapvu": "Tạp vụ",
}
REGISTRATION_LOCK_HEADERS = ["Vai trò", "Khóa đăng ký", "Ngày cập nhật", "Giờ cập nhật", "Người cập nhật"]


def _registration_lock_ws():
    client = get_gspread_client()
    if not client:
        return None
    ss = client.open_by_key(SHEET_MAT_KHAU_ID)
    try:
        ws = ss.worksheet(REGISTRATION_LOCK_WORKSHEET)
    except Exception:
        ws = ss.add_worksheet(title=REGISTRATION_LOCK_WORKSHEET, rows=30, cols=5)
    return ws


@st.cache_data(ttl=15, show_spinner=False)
def load_registration_role_locks():
    locks = {role: False for role in REGISTRATION_LOCK_ROLES}
    try:
        ws = _registration_lock_ws()
        if ws is None:
            return locks
        vals = ws.get_all_values()
        if not vals:
            ws.update("A1:E1", [REGISTRATION_LOCK_HEADERS], value_input_option="USER_ENTERED")
            return locks
        header = [normalize_login_name(x) for x in vals[0]]
        try:
            role_idx = header.index(normalize_login_name("Vai trò"))
        except Exception:
            role_idx = 0
        try:
            lock_idx = header.index(normalize_login_name("Khóa đăng ký"))
        except Exception:
            lock_idx = 1
        for row in vals[1:]:
            role = str(row[role_idx] if role_idx < len(row) else "").strip().lower()
            if role in locks:
                value = row[lock_idx] if lock_idx < len(row) else ""
                locks[role] = is_locked_value(value)
        return locks
    except Exception:
        return locks


def set_registration_role_lock(role, locked, actor=""):
    role = str(role or "").strip().lower()
    if role == "admin":
        return False, "Admin luôn được mở quyền đăng ký và không thể bị khóa."
    if role not in REGISTRATION_LOCK_ROLES:
        return False, f"Vai trò không hợp lệ: {role}"
    try:
        ws = _registration_lock_ws()
        if ws is None:
            return False, "Không kết nối được Google Sheets."
        vals = ws.get_all_values()
        if not vals:
            ws.update("A1:E1", [REGISTRATION_LOCK_HEADERS], value_input_option="USER_ENTERED")
            vals = [REGISTRATION_LOCK_HEADERS]

        row_num = None
        for i, row in enumerate(vals[1:], start=2):
            if row and str(row[0]).strip().lower() == role:
                row_num = i
                break

        now = datetime.now(VN_TZ)
        payload = [[
            role,
            "1" if bool(locked) else "0",
            now.strftime("%d/%m/%Y"),
            now.strftime("%H:%M:%S"),
            str(actor or st.session_state.get("current_user", "")).strip(),
        ]]
        if row_num is None:
            ws.append_row(payload[0], value_input_option="USER_ENTERED")
        else:
            ws.update(f"A{row_num}:E{row_num}", payload, value_input_option="USER_ENTERED")

        try:
            load_registration_role_locks.clear()
        except Exception:
            pass
        return True, (
            f"Đã {'KHÓA' if locked else 'MỞ'} quyền đăng ký của vai trò "
            f"{REGISTRATION_LOCK_LABELS.get(role, role)}."
        )
    except Exception as e:
        return False, f"Lỗi lưu khóa quyền đăng ký: {e}"


def is_registration_role_locked(role):
    role = str(role or "").strip().lower()
    if role == "admin":
        return False
    return bool(load_registration_role_locks().get(role, False))


active_users = get_active_users()
system_status = get_system_status()

# Cập nhật thời gian hoạt động của user hiện tại
if st.session_state.get("logged_in") and st.session_state.get("current_user"):
    active_users[st.session_state.current_user] = time.time()

# Dọn dẹp user đã ngưng hoạt động > 5 phút (300 giây)
current_t = time.time()
for u in list(active_users.keys()):
    if current_t - active_users[u] > 300: 
        del active_users[u]

online_users_count = len(active_users)
online_users_list = list(active_users.keys())

# --- CẤU HÌNH TRANG ---
st.set_page_config(page_title="Vera Spa Tam Hiệp Đồng Nai", page_icon="📅", layout="wide", initial_sidebar_state="auto")

# --- JAVASCRIPT: NHỚ ĐĂNG NHẬP + ĐÓNG DROPDOWN KHI BẤM RA NGOÀI ---
components.html("""
<script>
(function () {
    try {
        const parentWin = window.parent;
        const parentDoc = parentWin.document;
        const url = new URL(parentWin.location.href);
        const STORAGE_KEY = 'vera_remember_token';

        // Khi server yêu cầu quên đăng nhập: xóa token khỏi trình duyệt.
        if (url.searchParams.get('forget_login') === '1') {
            parentWin.localStorage.removeItem(STORAGE_KEY);
            url.searchParams.delete('forget_login');
            url.searchParams.delete('remember_token');
            parentWin.history.replaceState({}, '', url.toString());
        } else {
            const tokenInUrl = url.searchParams.get('remember_token');
            const savedToken = parentWin.localStorage.getItem(STORAGE_KEY);

            // Token mới sau khi đăng nhập -> lưu trong localStorage (không lưu mật khẩu).
            if (tokenInUrl) {
                parentWin.localStorage.setItem(STORAGE_KEY, tokenInUrl);
                // Không để bearer token nằm lâu trên thanh địa chỉ sau khi đã lưu cục bộ.
                url.searchParams.delete('remember_token');
                parentWin.history.replaceState({}, '', url.toString());
            } else if (savedToken) {
                // Lần mở app sau: đưa token trở lại URL để server xác thực.
                url.searchParams.set('remember_token', savedToken);
                parentWin.location.replace(url.toString());
                return;
            }
        }

        parentDoc.addEventListener('keydown', function(event) {
            if ((event.key === 'c' || event.key === 'C')) {
                const tag = (event.target.tagName || '').toLowerCase();
                if (tag !== 'input' && tag !== 'textarea') event.stopPropagation();
            }
        }, true);

        // Bấm ra khoảng trống: blur ô select đang hoạt động để popover đóng lại.
        parentDoc.addEventListener('pointerdown', function(event) {
            const insideSelect = event.target.closest && event.target.closest('[data-baseweb="select"], [data-baseweb="popover"]');
            if (!insideSelect) {
                const active = parentDoc.activeElement;
                if (active && active.closest && active.closest('[data-baseweb="select"]')) active.blur();
            }
        }, true);
    } catch (e) {
        console.debug('Vera helper:', e);
    }
})();
</script>
""", height=0, width=0)


# --- ÉP CSS GIAO DIỆN CỐ ĐỊNH (TỐI ƯU HIỆU NĂNG) ---
st.markdown("""
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Cinzel+Decorative:wght@400;700&display=swap');
        @import url('https://fonts.googleapis.com/css2?family=Roboto:wght@400;700&display=swap');
        @import url('https://fonts.googleapis.com/css2?family=Arial:wght@400;700&display=swap');
        
        /* Cấu hình Giao diện toàn trang */
        html, body, [class*="st-"], .stMarkdown, .stText, div, span, p {
            font-family: 'Roboto', sans-serif !important;
            color: #333333 !important;
        }
        
        /* --- FIX LỖI MŨI TÊN (KHÔI PHỤC FONT ICON CỦA STREAMLIT) --- */
        span.material-symbols-rounded, 
        [data-testid="stIconMaterial"], 
        .stIcon, 
        span[class*="stIcon"] {
            font-family: "Material Symbols Rounded" !important;
        }
        
        p, .stText, [data-testid="stMarkdownContainer"] {
            font-size: 16px !important;
        }
        
        .block-container { padding-top: 0.85rem; padding-bottom: 0.75rem; max-width: 1500px; }
        div[data-testid="stVerticalBlock"] > div { gap: 0.12rem !important; }
        div.stButton, div[data-testid="stDownloadButton"], div[data-testid="stFormSubmitButton"] { margin: 0 !important; padding: 0 !important; }
        button { margin-top: 1px !important; min-height: 40px; padding-top: 0.32rem !important; padding-bottom: 0.32rem !important; transition: background-color .16s ease, color .16s ease, border-color .16s ease, transform .12s ease !important; }
        div.stButton > button:hover,
        div[data-testid="stDownloadButton"] > button:hover,
        div[data-testid="stFormSubmitButton"] > button:hover {
            background-color: #c27ba0 !important;
            color: #ffffff !important;
            border-color: #a85f86 !important;
            transform: translateY(-1px);
        }
        input, textarea { font-size: 16px !important; }
        [data-testid="stDataFrame"], [data-testid="stDataEditor"] { width: 100% !important; }

        /* --- WRAPTEXT TIÊU ĐỀ BẢNG ---
           Khi Admin thu hẹp cột, tiêu đề được phép xuống dòng thay vì bị cắt chữ. */
        table th,
        [data-testid="stTable"] th,
        [data-testid="stDataFrame"] [role="columnheader"],
        [data-testid="stDataEditor"] [role="columnheader"],
        [data-testid="stDataFrame"] [role="columnheader"] > div,
        [data-testid="stDataEditor"] [role="columnheader"] > div {
            white-space: normal !important;
            overflow-wrap: anywhere !important;
            word-break: break-word !important;
            text-overflow: clip !important;
            line-height: 1.15 !important;
            height: auto !important;
            min-height: 38px !important;
        }

        /* Các nhãn header DOM của Glide/Streamlit ở những phiên bản có expose header. */
        [data-testid="stDataFrame"] [class*="header"],
        [data-testid="stDataEditor"] [class*="header"] {
            white-space: normal !important;
            overflow-wrap: anywhere !important;
            word-break: break-word !important;
        }

        /* Hiệu ứng hover cho TOÀN BỘ dropdown/select/multiselect */
        div[data-baseweb="select"],
        [data-testid="stSelectbox"],
        [data-testid="stMultiSelect"] {
            transition: transform .14s ease !important;
        }
        div[data-baseweb="select"] > div,
        [data-testid="stSelectbox"] div[data-baseweb="select"] > div,
        [data-testid="stMultiSelect"] div[data-baseweb="select"] > div {
            transition: background-color .16s ease, border-color .16s ease, box-shadow .16s ease !important;
        }
        div[data-baseweb="select"]:hover > div,
        [data-testid="stSelectbox"]:hover div[data-baseweb="select"] > div,
        [data-testid="stMultiSelect"]:hover div[data-baseweb="select"] > div {
            background-color: #f7e8ef !important;
            border-color: #c27ba0 !important;
            box-shadow: 0 0 0 1px #c27ba0 inset, 0 2px 8px rgba(194, 123, 160, .18) !important;
        }
        div[data-baseweb="select"]:hover svg,
        [data-testid="stSelectbox"]:hover svg,
        [data-testid="stMultiSelect"]:hover svg {
            fill: #a85f86 !important;
            color: #a85f86 !important;
        }
        div[data-baseweb="popover"] [role="option"],
        div[data-baseweb="menu"] [role="option"],
        ul[role="listbox"] li,
        ul[role="listbox"] [role="option"] {
            transition: background-color .14s ease, color .14s ease, padding-left .14s ease !important;
        }
        div[data-baseweb="popover"] [role="option"]:hover,
        div[data-baseweb="menu"] [role="option"]:hover,
        ul[role="listbox"] li:hover,
        ul[role="listbox"] [role="option"]:hover {
            background-color: #f3dce8 !important;
            color: #7d3159 !important;
            padding-left: 14px !important;
        }

        /* V65 - BORDER GIAO DIỆN.
           Dropdown/select box giữ RGB(242,242,242) = #F2F2F2.
           Riêng TOÀN BỘ đường viền bảng dùng RGB(217,217,217) = #D9D9D9. */
        [data-testid="stSelectbox"] div[data-baseweb="select"] > div,
        [data-testid="stMultiSelect"] div[data-baseweb="select"] > div,
        div[data-baseweb="select"] > div {
            border-color: #F2F2F2 !important;
            box-shadow: 0 0 0 1px #F2F2F2 inset !important;
        }
        [data-testid="stSelectbox"]:hover div[data-baseweb="select"] > div,
        [data-testid="stMultiSelect"]:hover div[data-baseweb="select"] > div,
        div[data-baseweb="select"]:hover > div,
        [data-testid="stSelectbox"] div[data-baseweb="select"] > div:focus-within,
        [data-testid="stMultiSelect"] div[data-baseweb="select"] > div:focus-within {
            border-color: #F2F2F2 !important;
            box-shadow: 0 0 0 1px #F2F2F2 inset !important;
        }
        div[data-baseweb="popover"],
        div[data-baseweb="menu"],
        ul[role="listbox"] {
            border-color: #F2F2F2 !important;
        }

        /* V65 - HTML tables và các bảng Streamlit/Data Editor: border #D9D9D9. */
        table, table th, table td {
            border-color: #D9D9D9 !important;
        }
        [data-testid="stDataFrame"],
        [data-testid="stDataEditor"],
        [data-testid="stTable"] {
            border-color: #D9D9D9 !important;
        }
        [data-testid="stDataFrame"] [role="grid"],
        [data-testid="stDataEditor"] [role="grid"],
        [data-testid="stDataFrame"] [role="gridcell"],
        [data-testid="stDataEditor"] [role="gridcell"],
        [data-testid="stDataFrame"] [role="columnheader"],
        [data-testid="stDataEditor"] [role="columnheader"] {
            border-color: #D9D9D9 !important;
        }

        /* V62 - NỀN TIÊU ĐỀ/NHÃN TOÀN HỆ THỐNG: RGB(217,217,217) = #D9D9D9.
           Áp dụng cho heading và toàn bộ nhãn widget như "Chọn ngày nghỉ", "Chọn nhân viên", ... */
        /* V71 - BỘ CỠ CHỮ MẶC ĐỊNH TOÀN HỆ THỐNG.
           Tiêu đề lớn 28px → tiêu đề con 22px → tiêu đề nhỏ 18px → label/nội dung 16px → bảng 13px. */
        h1, .custom-main-title { font-size: 28px !important; line-height: 1.22 !important; }
        h2, h3 { font-size: 22px !important; line-height: 1.25 !important; }
        h4, h5, h6, [data-testid="stExpander"] details summary p { font-size: 18px !important; line-height: 1.28 !important; }
        p, .stText, [data-testid="stMarkdownContainer"],
        [data-testid="stWidgetLabel"], [data-testid="stWidgetLabel"] p,
        input, textarea, button, [data-baseweb="select"] { font-size: 16px !important; }
        table, table th, table td,
        [data-testid="stDataFrame"], [data-testid="stDataEditor"], [data-testid="stTable"] { font-size: 13px !important; }

        h1, h2, h3, h4, h5, h6 {
            background: #D9D9D9 !important;
            border-left: 5px solid #A6A6A6 !important;
            border-radius: 7px !important;
            padding: 0.38rem 0.65rem !important;
            margin-top: 0.28rem !important;
            margin-bottom: 0.4rem !important;
            color: #222222 !important;
        }
        .custom-main-title {
            background: #D9D9D9 !important;
            border-left: 5px solid #A6A6A6 !important;
            border-radius: 7px !important;
            padding: 0.45rem 0.7rem !important;
            color: #222222 !important;
        }

        /* Nhãn của input/select/date/multiselect/radio/checkbox/file uploader/data editor... */
        [data-testid="stWidgetLabel"],
        label[data-testid="stWidgetLabel"],
        div[data-testid="stWidgetLabel"] {
            display: block !important;
            width: 100% !important;
            max-width: 100% !important;
            box-sizing: border-box !important;
            background-color: #D9D9D9 !important;
            border-radius: 5px !important;
            padding: 0.28rem 0.5rem !important;
            margin-bottom: 0.22rem !important;
            color: #222222 !important;
        }
        [data-testid="stWidgetLabel"] p,
        [data-testid="stWidgetLabel"] span,
        [data-testid="stWidgetLabel"] div {
            background: transparent !important;
            color: #222222 !important;
        }

        /* Một số phiên bản Streamlit dùng label/legend riêng cho widget. */
        [data-testid="stSelectbox"] > label,
        [data-testid="stMultiSelect"] > label,
        [data-testid="stDateInput"] > label,
        [data-testid="stTextInput"] > label,
        [data-testid="stTextArea"] > label,
        [data-testid="stNumberInput"] > label,
        [data-testid="stFileUploader"] > label,
        [data-testid="stRadio"] > label,
        [data-testid="stCheckbox"] > label,
        [data-testid="stSlider"] > label,
        [data-testid="stTimeInput"] > label,
        [data-testid="stColorPicker"] > label,
        [data-testid="stDataEditor"] > label,
        [data-testid="stDataFrame"] > label,
        fieldset > legend {
            display: block !important;
            width: 100% !important;
            box-sizing: border-box !important;
            background-color: #D9D9D9 !important;
            border-radius: 5px !important;
            padding: 0.28rem 0.5rem !important;
            color: #222222 !important;
        }

        /* Dòng tiêu đề của Expander cũng dùng cùng nền xám. */
        [data-testid="stExpander"] details > summary {
            background-color: #D9D9D9 !important;
            border-radius: 6px !important;
            padding-left: 0.55rem !important;
            padding-right: 0.55rem !important;
        }

        @media (max-width: 768px) {
            .block-container {
                padding-top: 0.6rem !important;
                padding-left: 0.45rem !important;
                padding-right: 0.45rem !important;
            }
            .custom-main-title { font-size: 28px !important; line-height: 1.22 !important; margin-bottom: 8px !important; }
            .custom-main-title > div { float: none !important; text-align: left !important; margin-top: 6px !important; }
            p, .stText, [data-testid="stMarkdownContainer"] { font-size: 16px !important; }
            div[data-testid="stVerticalBlock"] > div { gap: 0.08rem !important; }
            button { min-height: 42px !important; font-size: 16px !important; margin-top: 0 !important; padding-top: 0.25rem !important; padding-bottom: 0.25rem !important; }
            div[data-baseweb="popover"] { max-width: calc(100vw - 12px) !important; }
            [data-testid="stDataFrame"], [data-testid="stDataEditor"] { font-size: 13px !important; }
            [data-testid="stTabs"] button { white-space: nowrap !important; }
        }
        
        /* Loại bỏ thanh cuộn dropdown */
        div[data-baseweb="popover"] > div,
        div[data-baseweb="select"] ul[role="listbox"],
        div[data-testid="stSelectboxVirtualDropdown"] {
            max-height: 85vh !important; 
        }
        
        .custom-main-title {
            font-family: 'Roboto', sans-serif !important;
            font-size: 28px; font-weight: bold; margin-bottom: 20px; color: #333 !important;
        }
        
        /* GIẢM SIZE CHỮ: ĐĂNG KÝ - THAY ĐỔI LỊCH NGHỈ */
        [data-testid="stExpander"] details summary p {
            font-size: 18px !important;
            font-weight: 700 !important;
            color: #d32f2f !important;
            text-transform: uppercase;
        }
    </style>
""", unsafe_allow_html=True)

# --- KẾT NỐI GSPREAD ---
SHEET_MAT_KHAU_ID = "1DGXy3kPyMPwtz-3CnG8i6BiQbXFDApasoXVFzSmUe24"
SHEET_DU_PHONG_ID = "1Kz0aw-JatptAN9G7YSwZ6rJO09urOPaD-rS-18eZSY0"
SHEET_LICH_NGHI_2_ID = "1bLxn-L5gXui8pCL1b9TxshCNcykM7jg0J49Dkr5b4DI"
SHEET_CHINH_ID = "1xTjmi6BaQFSqsgn9-EM7MjVS2n2FNuxT"
BANG_TOUR_FILE_ID = "151d1ueCwH2KXX-HPQF1uj340uWSCS2dW"
PAYROLL_SOURCE_SHEET_ID = "1WtYsbEAlifL1PZ-nSGBojgL4Bnur-1vF"

# V84 - Điều khiển Auto Update phạt. Trạng thái lưu trên Google Sheet để mọi instance dùng chung.
AUTO_PENALTY_CONFIG_WORKSHEET = "CauHinhAutoPhat"
AUTO_PENALTY_MINUTES = 5
AUTO_PENALTY_CONFIG_HEADERS = [
    "Key", "Trạng thái", "Ngưỡng phút", "Ngày cập nhật", "Giờ cập nhật", "Người cập nhật"
]
AUTO_PENALTY_CONFIG_KEY = "AUTO_PENALTY"
AUTO_PENALTY_RUNNING = "RUNNING"
AUTO_PENALTY_PAUSED = "PAUSED"
PAYROLL_SOURCE_WORKSHEET = "Báo cáo doanh thu hóa đơn"
PAYROLL_STORAGE_WORKSHEET = "BangLuong"
PAYROLL_CONFIG_WORKSHEET = "CauHinhLuong"
UI_LAYOUT_WORKSHEET = "CauHinhCot"
UI_THEME_WORKSHEET = "CauHinhGiaoDien"
# V85.2 - Thứ tự MENU CHỨC NĂNG riêng cho Admin, lưu Google Sheet.
ADMIN_MENU_CONFIG_WORKSHEET = "CauHinhMenuAdmin"
ADMIN_MENU_CONFIG_HEADERS = ["ConfigKey", "Thứ tự JSON", "Cập nhật lúc", "Người cập nhật"]
ADMIN_MENU_CONFIG_KEY = "admin_menu_order"
# V86.4 - Tài liệu Hướng dẫn sử dụng lưu bền vững trong Google Sheet dưới dạng base64 chunks.
USAGE_GUIDE_WORKSHEET = "HuongDanSuDung"
USAGE_GUIDE_MAX_BYTES = 4 * 1024 * 1024
USAGE_GUIDE_CHUNK_SIZE = 32000
USAGE_GUIDE_META_HEADERS = [
    "Tên tài liệu", "Phiên bản", "Tên file", "MIME", "Dung lượng", "SHA256",
    "Cập nhật lúc", "Người cập nhật", "Ghi chú", "Số chunk"
]
TICHLUY_WORKSHEET = "TichLuy"
VIOLATION_DEBT_WORKSHEET = "NoViPham"
VIOLATION_DEBT_HEADERS = [
    "STT", "Tên nhân viên", "Số tiền", "Nội dung", "Loại",
    "Kỳ phát sinh từ", "Kỳ phát sinh đến", "Bắt đầu trừ từ",
    "Trạng thái", "Mã nguồn", "Ngày cập nhật", "Giờ cập nhật",
    "Người cập nhật", "Kỳ đã khấu trừ"
]
VIOLATION_DEBT_OPEN_STATUS = "Chưa hoàn thành"
VIOLATION_DEBT_DONE_STATUS = "Đã hoàn thành"
VIOLATION_DEBT_CONTENT = "Chưa hoàn thành nghĩa vụ Vi phạm"
# V41: kỳ lương cố định 01-15 / 16-cuối tháng + loại quanly khỏi TichLuy
TICHLUY_TARGET_DEFAULT = 5000000
TICHLUY_PERIOD_DEFAULT = 500000
TICHLUY_HEADERS = [
    "STT", "Tên nhân viên", "Ngày bắt đầu làm", "Mục tiêu tích lũy", "Đã tích lũy",
    "Còn lại", "Kỳ gần nhất", "Số tiền kỳ gần nhất", "Chi tiết các kỳ"
]

# Các bộ phận không tham gia đăng ký lịch nghỉ / Tích lũy.
# Leader được xem như Nhân viên trong các nghiệp vụ nghỉ phép, Tích lũy và Bảng lương.
LEAVE_EXCLUDED_ROLES = {"letan", "locker", "tapvu"}
TICHLUY_EXCLUDED_ROLES = {"admin", "letan", "quanly", "locker", "tapvu"}
FRONTDESK_ROLES = {"letan", "quanly"}
EMPLOYEE_LIKE_ROLES = {"nhanvien", "leader"}
PAYROLL_ELIGIBLE_ROLES = {"nhanvien", "leader"}
ALL_ACCOUNT_ROLES = ["nhanvien", "leader", "quanly", "letan", "locker", "tapvu", "admin"]
FRONTDESK_MANAGEABLE_ROLES = {"nhanvien", "locker", "tapvu"}

# Phân quyền theo từng chức năng: Admin có thể cấu hình theo vai trò và ghi đè theo từng tài khoản.
FEATURE_PERMISSION_WORKSHEET = "PhanQuyenChucNang"
FEATURE_PERMISSION_HEADERS = ["Phạm vi", "Đối tượng", "Chức năng", "Cho phép", "Ngày cập nhật", "Giờ cập nhật", "Người cập nhật"]
FEATURE_DEFINITIONS = {
    "tour": "🧭 Bảng tour",
    "payroll": "💰 Bảng lương",
    "payroll_history": "🗂 Lịch sử bảng lương",
    "payroll_email": "📧 Gửi email bảng lương",
    "leave": "📅 Đăng ký nghỉ phép",
    "leave_manage": "✏️ Quản lý lịch nghỉ",
    "shift": "⏰ Thiết lập ca làm việc",
    "staff_list": "👥 Danh sách nhân sự",
    "employee_add": "➕ Thêm nhân viên",
    "employee_edit": "✏️ Chỉnh sửa hồ sơ nhân viên",
    "employment_status": "🏷️ Trạng thái làm việc",
    "employee_delete": "🗑️ Xóa nhân viên",
    "account_lock": "🔒 Khóa đăng nhập",
    "registration_lock": "🔐 Khóa quyền đăng ký",
    "auto_penalty": "⏸️ Auto Update phạt",
    "sync": "🔄 Đồng bộ dữ liệu",
    "column_config": "⚙️ Giao diện tùy chỉnh",
    "profile": "👤 Hồ sơ cá nhân",
    "birthday": "🎂 Kiểm tra sinh nhật",
    "permission_admin": "🔐 Phân quyền chức năng",
}
DEFAULT_ROLE_FEATURES = {
    "admin": set(FEATURE_DEFINITIONS),
    "quanly": {"tour", "leave", "leave_manage", "shift", "staff_list", "employee_add", "employee_edit", "employment_status", "employee_delete", "profile", "birthday"},
    "letan": {"tour", "leave", "leave_manage", "shift", "staff_list", "employee_add", "employee_edit", "employment_status", "employee_delete", "profile", "birthday"},
    "leader": {"tour", "leave", "leave_manage", "profile", "birthday"},
    "nhanvien": {"tour", "leave", "leave_manage", "profile", "birthday"},
    "locker": {"tour", "profile", "birthday"},
    "tapvu": {"birthday"},
}

# V48: Thông báo sinh nhật đầu tháng.
BIRTHDAY_NOTICE_WORKSHEET = "ThongBaoSinhNhat"
BIRTHDAY_VIEWER_ROLES = {"admin", "quanly", "letan"}
BIRTHDAY_EMPLOYEE_ROLES = {"nhanvien", "leader", "letan", "locker"}
BIRTHDAY_NOTICE_DAYS = {1, 2, 3, 4, 5}
BIRTHDAY_NOTICE_MAX_LOGINS_PER_DAY = 3
# V51: Trạng thái làm việc của nhân viên được lưu riêng, không thay đổi cấu trúc Sheet1.
EMPLOYMENT_STATUS_WORKSHEET = "TrangThaiNhanSu"
EMPLOYMENT_STATUS_HEADERS = ["STT", "Tên nhân viên", "Trạng thái", "Ngày cập nhật", "Giờ cập nhật", "Người cập nhật"]
EMPLOYMENT_STATUS_ACTIVE = "Đang làm việc"
EMPLOYMENT_STATUS_TEMP = "Tạm thời nghỉ việc"
EMPLOYMENT_STATUS_LEFT = "Đã nghỉ việc"
EMPLOYMENT_STATUS_OPTIONS = [EMPLOYMENT_STATUS_ACTIVE, EMPLOYMENT_STATUS_TEMP, EMPLOYMENT_STATUS_LEFT]
# Giữ tương thích dữ liệu trạng thái đã lưu ở các bản cũ.
EMPLOYMENT_STATUS_ALIASES = {
    normalize_login_name("Đang làm việc"): EMPLOYMENT_STATUS_ACTIVE,
    normalize_login_name("Nghỉ việc tạm thời"): EMPLOYMENT_STATUS_TEMP,
    normalize_login_name("Tạm thời nghỉ việc"): EMPLOYMENT_STATUS_TEMP,
    normalize_login_name("Đã nghỉ việc hẳn"): EMPLOYMENT_STATUS_LEFT,
    normalize_login_name("Đã nghỉ việc"): EMPLOYMENT_STATUS_LEFT,
}
STAFF_ROLE_ORDER = ["leader", "nhanvien", "quanly", "letan", "locker", "tapvu", "admin"]
EMPLOYMENT_STATUS_MANAGEABLE_ROLES = set(STAFF_ROLE_ORDER) - {"admin"}
EMPLOYEE_LEAVE_CHANGE_NOTICE_DAYS = 3
DEFAULT_LEAVE_PAGE = "📅 Đăng ký nghỉ phép"
DEFAULT_LEAVE_PAGE_SLUG = "dang-ky-thong-ke-nghi-phep"

def _set_default_page_after_login(role):
    """Lễ tân/Quản lý/Nhân viên luôn mở trang Đăng ký nghỉ phép sau đăng nhập."""
    role_norm = str(role or '').strip().lower()
    if role_norm in {"letan", "quanly", "nhanvien", "leader"}:
        st.session_state.app_page = DEFAULT_LEAVE_PAGE
        try:
            st.query_params["page"] = DEFAULT_LEAVE_PAGE_SLUG
        except Exception:
            pass

def _parse_birthday_date(value):
    """Đọc ngày sinh từ dd/mm/yyyy, dd-mm-yyyy, serial Excel; chấp nhận cả dd/mm."""
    parsed = _parse_vn_date(value)
    if parsed is not None:
        return parsed
    text = str(value or '').strip()
    if not text:
        return None
    for fmt in ('%d/%m', '%d-%m'):
        try:
            tmp = datetime.strptime(text, fmt)
            return date(2000, tmp.month, tmp.day)
        except Exception:
            pass
    return None

def get_month_birthdays(credentials_df, month=None):
    """Danh sách sinh nhật trong tháng của nhanvien/letan/locker, sắp theo ngày sinh."""
    if credentials_df is None or credentials_df.empty:
        return []
    month = int(month or get_vn_today().month)
    rows = []
    for _, r in credentials_df.iterrows():
        role = str(r.get('Phân quyền', '')).strip().lower()
        if role not in BIRTHDAY_EMPLOYEE_ROLES:
            continue
        dob = _parse_birthday_date(r.get('Ngày sinh', ''))
        if dob is None or dob.month != month:
            continue
        system_name = str(r.get('Tên nhân viên', '')).strip()
        full_name = str(r.get('Họ và tên đầy đủ', '')).strip()
        display_name = full_name or system_name
        if not display_name:
            continue
        rows.append({
            'Tên nhân viên': system_name,
            'Họ và tên': display_name,
            'Ngày sinh': f"{dob.day:02d}/{dob.month:02d}",
            'Ngày': dob.day,
            'Vai trò': role,
        })
    rows.sort(key=lambda x: (x['Ngày'], normalize_login_name(x['Họ và tên'])))
    return rows

def _get_birthday_notice_worksheet():
    client = get_gspread_client()
    if not client:
        return None
    ss = client.open_by_key(SHEET_MAT_KHAU_ID)
    ws = _get_or_create_worksheet(ss, BIRTHDAY_NOTICE_WORKSHEET, rows=1000, cols=6)
    header = _gs_call_with_backoff(ws.row_values, 1)
    wanted = ["Tài khoản", "Ngày", "Số lần đăng nhập", "Tạm tắt hôm nay"]
    if header[:4] != wanted:
        gspread_update_range(ws, 'A1:D1', [wanted])
    return ws

def register_birthday_notice_login(username):
    """Tăng bộ đếm đúng 1 lần cho một lần đăng nhập; trả về (count, muted)."""
    today_key = get_vn_today().isoformat()
    user_key = normalize_login_name(username)
    try:
        ws = _get_birthday_notice_worksheet()
        if ws is None:
            return 1, False
        vals = _gs_call_with_backoff(ws.get_all_values)
        found_row = None
        count = 0
        muted = False
        for idx, row in enumerate(vals[1:], start=2):
            if len(row) < 2:
                continue
            if normalize_login_name(row[0]) == user_key and str(row[1]).strip() == today_key:
                found_row = idx
                try:
                    count = int(float(str(row[2]).strip() or '0')) if len(row) > 2 else 0
                except Exception:
                    count = 0
                muted = str(row[3]).strip().casefold() in {'1','true','yes','x','tat','tắt'} if len(row) > 3 else False
                break
        count += 1
        if found_row:
            gspread_update_range(ws, f'C{found_row}', [[count]])
        else:
            _gs_call_with_backoff(ws.append_row, [str(username), today_key, count, ''], value_input_option='USER_ENTERED')
        return count, muted
    except Exception:
        # Nếu Google Sheets tạm lỗi, vẫn cho hiện thông báo trong phiên hiện tại.
        return 1, False

def mute_birthday_notice_today(username):
    """Tạm tắt thông báo sinh nhật cho tài khoản hiện tại đến hết ngày hôm nay."""
    today_key = get_vn_today().isoformat()
    user_key = normalize_login_name(username)
    try:
        ws = _get_birthday_notice_worksheet()
        if ws is None:
            return False, "Không kết nối được Google Sheets."
        vals = _gs_call_with_backoff(ws.get_all_values)
        found_row = None
        count = 0
        for idx, row in enumerate(vals[1:], start=2):
            if len(row) >= 2 and normalize_login_name(row[0]) == user_key and str(row[1]).strip() == today_key:
                found_row = idx
                try:
                    count = int(float(str(row[2]).strip() or '0')) if len(row) > 2 else 0
                except Exception:
                    count = 0
                break
        if found_row:
            gspread_update_range(ws, f'D{found_row}', [["1"]])
        else:
            _gs_call_with_backoff(ws.append_row, [str(username), today_key, max(1, count), '1'], value_input_option='USER_ENTERED')
        return True, "Đã tạm tắt thông báo sinh nhật đến hết hôm nay."
    except Exception as e:
        return False, f"Không thể tạm tắt thông báo: {e}"

def render_birthday_login_notice(credentials_df):
    """Hiện thông báo trong ngày 1-5, tối đa ở 3 lần đăng nhập đầu mỗi ngày."""
    role = str(st.session_state.get('current_role', '')).strip().lower()
    if role not in BIRTHDAY_VIEWER_ROLES:
        st.session_state.birthday_login_event = False
        return
    today = get_vn_today()
    if today.day not in BIRTHDAY_NOTICE_DAYS:
        st.session_state.birthday_login_event = False
        return
    birthdays = get_month_birthdays(credentials_df, today.month)
    if not birthdays:
        st.session_state.birthday_login_event = False
        return
    # Chỉ tăng số lần khi đây thực sự là một phiên vừa đăng nhập.
    if st.session_state.get('birthday_login_event', False):
        count, muted = register_birthday_notice_login(st.session_state.get('current_user', ''))
        st.session_state.birthday_notice_count_today = count
        st.session_state.birthday_notice_muted_today = muted
        st.session_state.birthday_login_event = False
    count = int(st.session_state.get('birthday_notice_count_today', 999) or 999)
    muted = bool(st.session_state.get('birthday_notice_muted_today', False))
    if muted or count > BIRTHDAY_NOTICE_MAX_LOGINS_PER_DAY:
        return

    role_labels = {'nhanvien': 'Nhân viên', 'leader': 'Leader', 'letan': 'Lễ tân', 'locker': 'Locker'}
    lines = []
    for b in birthdays:
        role_label = role_labels.get(b['Vai trò'], b['Vai trò'])
        lines.append(f"🎂 **{b['Họ và tên']}** — {b['Ngày sinh']} — {role_label}")
    st.info("🎉 **SINH NHẬT TRONG THÁNG %02d**\n\n%s" % (today.month, "\n\n".join(lines)))
    c_notice, c_mute = st.columns([4, 1])
    with c_notice:
        st.caption(f"Thông báo này xuất hiện trong ngày 01–05 và tối đa ở 3 lần đăng nhập đầu mỗi ngày. Hôm nay: lần {count}/3.")
    with c_mute:
        if st.button("🔕 Tạm tắt hôm nay", use_container_width=True, key=f"mute_birthday_{today.isoformat()}"):
            ok, msg = mute_birthday_notice_today(st.session_state.get('current_user', ''))
            if ok:
                st.session_state.birthday_notice_muted_today = True
                st.toast(msg)
                st.rerun()
            else:
                st.error(msg)

def render_manual_birthday_check(credentials_df, key_prefix="birthday_manual"):
    """Nút chủ động xem sinh nhật tháng hiện tại, dùng cho mọi vai trò."""
    today = get_vn_today()
    state_key = f"{key_prefix}_show_{today.year}_{today.month}"
    if st.button("🎂 Kiểm tra sinh nhật tháng này", use_container_width=True, key=f"{key_prefix}_button_{today.year}_{today.month}"):
        st.session_state[state_key] = not bool(st.session_state.get(state_key, False))
    if not st.session_state.get(state_key, False):
        return
    birthdays = get_month_birthdays(credentials_df, today.month)
    if not birthdays:
        st.info(f"Tháng {today.month:02d} hiện chưa có sinh nhật của Nhân viên/Lễ tân/Locker trong hồ sơ.")
        return
    role_labels = {'nhanvien': 'Nhân viên', 'leader': 'Leader', 'letan': 'Lễ tân', 'locker': 'Locker'}
    st.markdown(f"#### 🎉 Sinh nhật tháng {today.month:02d}")
    for b in birthdays:
        st.markdown(f"- 🎂 **{b['Họ và tên']}** — {b['Ngày sinh']} — {role_labels.get(b['Vai trò'], b['Vai trò'])}")


@st.cache_resource
def get_gspread_client():
    """Google API client phù hợp cả Streamlit Cloud lẫn Cloud Run.

    Ưu tiên: GOOGLE_SERVICE_ACCOUNT_JSON -> st.secrets[gcp_service_account]
    -> Application Default Credentials của Cloud Run service account.
    """
    try:
        scope = [
            "https://www.googleapis.com/auth/spreadsheets",
            "https://www.googleapis.com/auth/drive",
        ]
        env_json = str(os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "")).strip()
        if env_json:
            creds_dict = json.loads(env_json)
            creds = Credentials.from_service_account_info(creds_dict, scopes=scope)
            return gspread.authorize(creds)
        try:
            creds_dict = dict(st.secrets["gcp_service_account"])
        except Exception:
            creds_dict = {}
        if creds_dict:
            creds = Credentials.from_service_account_info(creds_dict, scopes=scope)
            return gspread.authorize(creds)
        # Cloud Run: có thể dùng chính runtime service account (ADC).
        import google.auth
        creds, _ = google.auth.default(scopes=scope)
        return gspread.authorize(creds)
    except Exception:
        return None


def get_smtp_sender_credentials():
    """Lấy Gmail sender/app-password từ Secret Manager env hoặc Streamlit secrets."""
    sender_email = str(os.getenv("SMTP_SENDER_EMAIL", "veraspabienhoa@gmail.com")).strip()
    sender_pass = str(os.getenv("SMTP_APP_PASSWORD", "")).strip()
    if not sender_pass:
        try:
            sender_pass = str(st.secrets.get("smtp_app_password", "")).strip()
        except Exception:
            pass
    return sender_email, sender_pass


def gspread_update_range(sheet, range_name, values, **kwargs):
    """Tương thích cả gspread 5.x (range trước) và 6.x (values trước)."""
    try:
        major = int(str(getattr(gspread, '__version__', '5')).split('.')[0])
    except Exception:
        major = 5
    if major >= 6:
        return sheet.update(values, range_name, **kwargs)
    return sheet.update(range_name, values, **kwargs)


def _is_google_sheets_quota_error(exc):
    """Nhận diện lỗi quota/rate-limit của Google Sheets API."""
    msg = str(exc).lower()
    return (
        ('429' in msg or 'too many requests' in msg)
        and ('quota' in msg or 'rate' in msg or 'read requests' in msg or 'write requests' in msg)
    )


def _gs_call_with_backoff(func, *args, retries=5, **kwargs):
    """
    Gọi Google Sheets API với exponential backoff khi gặp 429.
    Mục tiêu chính vẫn là GIẢM số request; retry chỉ là lớp bảo vệ khi quota đang tạm đầy.
    """
    last_exc = None
    for attempt in range(max(1, int(retries))):
        try:
            return func(*args, **kwargs)
        except Exception as exc:
            last_exc = exc
            if not _is_google_sheets_quota_error(exc) or attempt >= retries - 1:
                raise
            # 2s -> 4s -> 8s -> 16s; chỉ dùng khi thật sự gặp 429.
            time.sleep(min(2 ** (attempt + 1), 16))
    if last_exc is not None:
        raise last_exc


def _get_employment_status_worksheet():
    client = get_gspread_client()
    if not client:
        return None
    ss = client.open_by_key(SHEET_MAT_KHAU_ID)
    ws = _get_or_create_worksheet(ss, EMPLOYMENT_STATUS_WORKSHEET, rows=1000, cols=len(EMPLOYMENT_STATUS_HEADERS))
    header = _gs_call_with_backoff(ws.row_values, 1)
    if header[:len(EMPLOYMENT_STATUS_HEADERS)] != EMPLOYMENT_STATUS_HEADERS:
        gspread_update_range(ws, "A1:F1", [EMPLOYMENT_STATUS_HEADERS])
    return ws


@st.cache_data(ttl=120, show_spinner=False)
def load_employment_status_map():
    result = {}
    try:
        ws = _get_employment_status_worksheet()
        if ws is None:
            return result
        vals = _gs_call_with_backoff(ws.get_all_values)
        for row in vals[1:]:
            if len(row) < 2:
                continue
            key = normalize_login_name(row[1])
            if not key:
                continue
            status = str(row[2]).strip() if len(row) > 2 else EMPLOYMENT_STATUS_ACTIVE
            status = EMPLOYMENT_STATUS_ALIASES.get(normalize_login_name(status), EMPLOYMENT_STATUS_ACTIVE)
            result[key] = status
    except Exception:
        pass
    return result


def set_employee_employment_status(employee_name, status, updated_by):
    status = EMPLOYMENT_STATUS_ALIASES.get(normalize_login_name(status), "")
    if status not in EMPLOYMENT_STATUS_OPTIONS:
        return False, "Trạng thái không hợp lệ."
    try:
        creds = load_credentials_recent()
        target_key = normalize_login_name(employee_name)
        role = ''
        if isinstance(creds, pd.DataFrame) and not creds.empty:
            hit = creds[creds['Tên nhân viên'].apply(normalize_login_name) == target_key]
            if not hit.empty:
                role = str(hit.iloc[0].get('Phân quyền', '')).strip().lower()
        if role not in EMPLOYMENT_STATUS_MANAGEABLE_ROLES:
            return False, "Chỉ có thể cập nhật trạng thái làm việc cho tài khoản nhân sự."
        ws = _get_employment_status_worksheet()
        if ws is None:
            return False, "Không kết nối được Google Sheets."
        vals = _gs_call_with_backoff(ws.get_all_values)
        found_row = None
        for r_idx, row in enumerate(vals[1:], start=2):
            if len(row) > 1 and normalize_login_name(row[1]) == target_key:
                found_row = r_idx
                break
        now = datetime.now(VN_TZ)
        payload = [str(employee_name).strip(), status, now.strftime('%d/%m/%Y'), now.strftime('%H:%M:%S'), str(updated_by).strip()]
        if found_row:
            gspread_update_range(ws, f"B{found_row}:F{found_row}", [payload])
        else:
            stt = max(1, len(vals))
            _gs_call_with_backoff(ws.append_row, [stt] + payload, value_input_option='USER_ENTERED')
        try:
            load_employment_status_map.clear()
        except Exception:
            pass
        return True, f"Đã cập nhật {employee_name}: {status}."
    except Exception as e:
        return False, f"Lỗi cập nhật trạng thái nhân sự: {e}"



def _get_feature_permission_worksheet():
    client = get_gspread_client()
    if not client:
        return None
    ss = client.open_by_key(SHEET_MAT_KHAU_ID)
    ws = _get_or_create_worksheet(ss, FEATURE_PERMISSION_WORKSHEET, rows=1000, cols=len(FEATURE_PERMISSION_HEADERS))
    header = _gs_call_with_backoff(ws.row_values, 1)
    if header[:len(FEATURE_PERMISSION_HEADERS)] != FEATURE_PERMISSION_HEADERS:
        gspread_update_range(ws, "A1:G1", [FEATURE_PERMISSION_HEADERS])
    return ws

@st.cache_data(ttl=120, show_spinner=False)
def load_feature_permissions():
    """Trả về (role_overrides, account_overrides). Account override ưu tiên hơn role."""
    role_cfg, account_cfg = {}, {}
    try:
        ws = _get_feature_permission_worksheet()
        if ws is None:
            return role_cfg, account_cfg
        values = _gs_call_with_backoff(ws.get_all_values)
        for row in values[1:]:
            if len(row) < 4:
                continue
            scope = normalize_login_name(row[0])
            target = str(row[1]).strip()
            feature = str(row[2]).strip()
            if feature not in FEATURE_DEFINITIONS or not target:
                continue
            allowed = str(row[3]).strip().casefold() in {'1','true','yes','on','x','co','có'}
            if scope in {'role','vai tro','vai trò'}:
                role_cfg[(target.strip().lower(), feature)] = allowed
            elif scope in {'account','tai khoan','tài khoản'}:
                account_cfg[(normalize_login_name(target), feature)] = allowed
    except Exception:
        pass
    return role_cfg, account_cfg

def _clear_feature_permission_cache():
    try:
        load_feature_permissions.clear()
    except Exception:
        pass

def has_feature_access(feature, role=None, username=None):
    """Kiểm tra quyền theo thứ tự: Admin -> tài khoản -> vai trò -> mặc định hệ thống."""
    feature = str(feature or '').strip()
    if feature not in FEATURE_DEFINITIONS:
        return False
    role = str(role if role is not None else st.session_state.get('current_role', '')).strip().lower()
    username = str(username if username is not None else st.session_state.get('current_user', '')).strip()
    if role == 'admin':
        return True
    role_cfg, account_cfg = load_feature_permissions()
    account_key = (normalize_login_name(username), feature)
    if normalize_login_name(username) and account_key in account_cfg:
        return bool(account_cfg[account_key])
    role_key = (role, feature)
    if role_key in role_cfg:
        return bool(role_cfg[role_key])
    return feature in DEFAULT_ROLE_FEATURES.get(role, set())

def _rewrite_feature_permission_scope(scope, target, allowed_features, updated_by, inherit=False):
    """Ghi lại toàn bộ cấu hình của 1 vai trò/tài khoản bằng một lần rewrite sheet."""
    try:
        ws = _get_feature_permission_worksheet()
        if ws is None:
            return False, 'Không kết nối được Google Sheets.'
        values = _gs_call_with_backoff(ws.get_all_values)
        keep = [FEATURE_PERMISSION_HEADERS]
        scope_norm = normalize_login_name(scope)
        target_norm = normalize_login_name(target) if scope_norm == 'account' else str(target).strip().lower()
        for row in values[1:]:
            if len(row) < 2:
                continue
            row_scope = normalize_login_name(row[0])
            row_target = normalize_login_name(row[1]) if row_scope == 'account' else str(row[1]).strip().lower()
            if row_scope == scope_norm and row_target == target_norm:
                continue
            rr = list(row[:7]) + [''] * max(0, 7-len(row))
            keep.append(rr[:7])
        if not inherit:
            now = datetime.now(VN_TZ)
            allowed_set = set(allowed_features or [])
            for key in FEATURE_DEFINITIONS:
                keep.append([
                    scope, str(target).strip(), key, '1' if key in allowed_set else '0',
                    now.strftime('%d/%m/%Y'), now.strftime('%H:%M:%S'), str(updated_by).strip()
                ])
        _gs_call_with_backoff(ws.clear)
        end_row = max(1, len(keep))
        gspread_update_range(ws, f'A1:G{end_row}', keep)
        _clear_feature_permission_cache()
        return True, 'Đã lưu phân quyền chức năng.' if not inherit else 'Đã xóa quyền riêng; tài khoản sẽ dùng quyền theo vai trò.'
    except Exception as e:
        return False, f'Lỗi lưu phân quyền: {e}'

def save_role_feature_permissions(role, allowed_features, updated_by):
    role = str(role or '').strip().lower()
    if role == 'admin':
        return False, 'Quyền Admin luôn được giữ đầy đủ để tránh tự khóa hệ thống.'
    return _rewrite_feature_permission_scope('role', role, allowed_features, updated_by, inherit=False)

def save_account_feature_permissions(username, allowed_features, updated_by, inherit=False):
    if not str(username or '').strip():
        return False, 'Vui lòng chọn tài khoản.'
    return _rewrite_feature_permission_scope('account', str(username).strip(), allowed_features, updated_by, inherit=inherit)

def get_effective_feature_keys_for_role(role):
    role = str(role or '').strip().lower()
    return [k for k in FEATURE_DEFINITIONS if has_feature_access(k, role=role, username='')]

def _clear_payroll_config_cache():
    """Chỉ xóa cache cấu hình lương, không xóa toàn bộ cache của ứng dụng."""
    try:
        _load_payroll_config_rows_cached.clear()
    except Exception:
        pass


def _clear_dynamic_data_caches():
    """
    Xóa đúng các cache dữ liệu nghiệp vụ có thể vừa thay đổi.
    Tuyệt đối không dùng st.cache_data.clear() vì nó làm mất mọi cache và gây bão request Google Sheets.
    """
    if vpg is not None and vpg.is_enabled():
        # Invalidation dùng chung giữa mọi Cloud Run instance; lần đọc tiếp theo chỉ 1 instance refresh Sheet.
        try:
            vpg.invalidate_many("credentials", "leave_primary", "leave_secondary", "tichluy")
        except Exception:
            pass

    for fn_name in (
        'load_credentials',
        'load_credentials_recent',
        'load_backup_sheet_data',
        'load_secondary_leave_sheet_data',
        'load_loai_nghi_from_gsheet',
        'load_tichluy_tracking',
        'load_feature_permissions',
    ):
        try:
            fn = globals().get(fn_name)
            if fn is not None and hasattr(fn, 'clear'):
                fn.clear()
        except Exception:
            pass

# ==========================================================
# V84 - TRẠNG THÁI AUTO UPDATE PHẠT (DÙNG CHUNG TOÀN HỆ THỐNG)
# ==========================================================
def _get_auto_penalty_config_worksheet(client=None):
    client = client or get_gspread_client()
    if not client:
        return None
    ss = client.open_by_key(SHEET_DU_PHONG_ID)
    try:
        ws = ss.worksheet(AUTO_PENALTY_CONFIG_WORKSHEET)
    except Exception:
        ws = ss.add_worksheet(title=AUTO_PENALTY_CONFIG_WORKSHEET, rows=20, cols=8)
    return ws

def load_auto_penalty_config():
    """Đọc LIVE trạng thái RUNNING/PAUSED để nút Admin có hiệu lực trên mọi phiên/instance."""
    default = {
        "paused": False,
        "status": AUTO_PENALTY_RUNNING,
        "threshold_minutes": AUTO_PENALTY_MINUTES,
        "updated_date": "",
        "updated_time": "",
        "updated_by": "",
        "error": "",
    }
    try:
        client = get_gspread_client()
        ws = _get_auto_penalty_config_worksheet(client)
        if ws is None:
            default["error"] = "Chưa cấu hình Google Sheets."
            return default
        vals = _gs_call_with_backoff(ws.get, 'A1:F2')
        if not vals or not vals[0] or vals[0][:6] != AUTO_PENALTY_CONFIG_HEADERS:
            gspread_update_range(ws, 'A1:F1', [AUTO_PENALTY_CONFIG_HEADERS], value_input_option='USER_ENTERED')
        row = vals[1] if len(vals) > 1 else []
        if not row or str(row[0]).strip() != AUTO_PENALTY_CONFIG_KEY:
            now = datetime.now(VN_TZ)
            row = [
                AUTO_PENALTY_CONFIG_KEY, AUTO_PENALTY_RUNNING, AUTO_PENALTY_MINUTES,
                now.strftime('%d/%m/%Y'), now.strftime('%H:%M:%S'), 'Hệ thống'
            ]
            gspread_update_range(ws, 'A2:F2', [row], value_input_option='USER_ENTERED')
        row = list(row) + [""] * max(0, 6 - len(row))
        status = str(row[1] or AUTO_PENALTY_RUNNING).strip().upper()
        try:
            threshold = int(float(row[2] or AUTO_PENALTY_MINUTES))
        except Exception:
            threshold = AUTO_PENALTY_MINUTES
        # V84: ngưỡng nghiệp vụ cố định tối thiểu là 5 phút.
        threshold = max(AUTO_PENALTY_MINUTES, threshold)
        return {
            "paused": status == AUTO_PENALTY_PAUSED,
            "status": status,
            "threshold_minutes": threshold,
            "updated_date": str(row[3] or ""),
            "updated_time": str(row[4] or ""),
            "updated_by": str(row[5] or ""),
            "error": "",
        }
    except Exception as e:
        default["error"] = str(e)
        return default

def set_auto_penalty_paused(paused, updated_by):
    try:
        client = get_gspread_client()
        ws = _get_auto_penalty_config_worksheet(client)
        if ws is None:
            return False, "Chưa cấu hình quyền kết nối Google Sheets."
        now = datetime.now(VN_TZ)
        status = AUTO_PENALTY_PAUSED if bool(paused) else AUTO_PENALTY_RUNNING
        gspread_update_range(ws, 'A1:F1', [AUTO_PENALTY_CONFIG_HEADERS], value_input_option='USER_ENTERED')
        gspread_update_range(ws, 'A2:F2', [[
            AUTO_PENALTY_CONFIG_KEY, status, AUTO_PENALTY_MINUTES,
            now.strftime('%d/%m/%Y'), now.strftime('%H:%M:%S'), str(updated_by or 'Admin')
        ]], value_input_option='USER_ENTERED')
        state_vi = "TẠM DỪNG" if paused else "HOẠT ĐỘNG"
        return True, f"Auto Update phạt đã chuyển sang trạng thái {state_vi}. Ngưỡng tự động: từ {AUTO_PENALTY_MINUTES} phút."
    except Exception as e:
        return False, f"Không cập nhật được trạng thái Auto Update: {e}"

def is_auto_penalty_paused():
    return bool(load_auto_penalty_config().get("paused", False))


def get_postgres_runtime_status():
    """Dùng cho kiểm tra triển khai: không làm app lỗi nếu PostgreSQL tạm unavailable."""
    if vpg is None or not vpg.is_enabled():
        return False, "PostgreSQL chưa bật; hệ thống đang dùng chế độ Google Sheets dự phòng."
    try:
        return vpg.healthcheck()
    except Exception as exc:
        return False, f"PostgreSQL lỗi: {exc}"


# ==========================================================
# V84 - ĐỒNG BỘ EXCEL -> GOOGLE SHEET: 2 PHIÊN BẢN
#   V1: cho phép ghi đè, paste toàn bộ dữ liệu từ A2:J...
#   V2: không ghi đè, chỉ thêm dòng mới đúng LAST ROW trong A:J
# ==========================================================
def _load_excel_leave_rows_for_google_sync():
    file_id = SHEET_CHINH_ID
    temp_file = f"temp_sync_{os.getpid()}_{int(time.time())}.xlsb"
    try:
        download_file_from_google_drive(file_id, temp_file)
        xls = pd.read_excel(temp_file, sheet_name='LichNghi', engine='pyxlsb')
        df_excel = xls.iloc[:, :10].copy()
        if df_excel.shape[1] < 10:
            return pd.DataFrame(), "Sheet LichNghi của Excel chưa đủ 10 cột A:J."

        df_excel = df_excel.iloc[:, :10]
        df_excel.columns = [
            "Ngày", "Tên nhân viên", "Lý do nghỉ", "Chi tiết", "Số ngày tính",
            "Số ngày phép cộng dồn", "Phạt vi phạm", "Ngày cập nhật",
            "Giờ cập nhật", "Người cập nhật"
        ]

        def clean_date(val):
            if val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip().lower() in {"", "nan", "nat", "none"}:
                return ""
            try:
                if isinstance(val, (int, float)) and not isinstance(val, bool):
                    return pd.to_datetime(val, unit='D', origin='1899-12-30').strftime('%d/%m/%Y')
                if hasattr(val, 'strftime'):
                    return val.strftime('%d/%m/%Y')
                return pd.to_datetime(str(val).strip().split(' ')[0], dayfirst=True).strftime('%d/%m/%Y')
            except Exception:
                return str(val).strip()

        def clean_text(val):
            try:
                if pd.isna(val):
                    return ""
            except Exception:
                pass
            text = str(val).strip()
            return "" if text.casefold() in {"nan", "nat", "none"} else text

        df_excel['Ngày'] = df_excel['Ngày'].apply(clean_date)
        df_excel['Ngày cập nhật'] = df_excel['Ngày cập nhật'].apply(clean_date)
        for c in df_excel.columns:
            if c not in {'Ngày', 'Ngày cập nhật'}:
                df_excel[c] = df_excel[c].apply(clean_text)

        # Bỏ dòng hoàn toàn trống. Dấu * ở tên chỉ được bỏ khi đối chiếu, không tự sửa dữ liệu gốc.
        df_excel = df_excel[
            df_excel.apply(lambda r: any(str(v).strip() for v in r.tolist()), axis=1)
        ].reset_index(drop=True)
        return df_excel, ""
    except Exception as e:
        return pd.DataFrame(), f"Không đọc được Excel LichNghi: {e}"
    finally:
        try:
            if os.path.exists(temp_file):
                os.remove(temp_file)
        except Exception:
            pass

def _leave_sync_merge_key(row):
    if isinstance(row, pd.Series):
        getv = row.get
    else:
        getv = lambda k, d='': row.get(k, d)
    return (
        normalize_schedule_date(getv('Ngày', '')),
        normalize_employee_match_name(getv('Tên nhân viên', '')),
        normalize_leave_reason(getv('Lý do nghỉ', getv('Loại nghỉ', ''))),
    )

def _ensure_leave_sheet_header(sheet_dp):
    header = [
        "Ngày", "Tên nhân viên", "Lý do nghỉ", "Chi tiết", "Số ngày tính",
        "Số ngày phép cộng dồn", "Phạt vi phạm", "Ngày cập nhật",
        "Giờ cập nhật", "Người cập nhật"
    ]
    current = _gs_call_with_backoff(sheet_dp.get, 'A1:J1')
    current = current[0] if current else []
    if not any(str(v).strip() for v in current):
        gspread_update_range(sheet_dp, 'A1:J1', [header], value_input_option='USER_ENTERED')
    return header

def admin_sync_excel_to_gsheet_overwrite():
    """Phiên bản 1: xóa dữ liệu cũ A2:J rồi paste toàn bộ Excel bắt đầu đúng A2."""
    try:
        client = get_gspread_client()
        if not client:
            return False, "Chưa cấu hình quyền kết nối Google Sheets."
        df_excel, err = _load_excel_leave_rows_for_google_sync()
        if err:
            return False, err
        sheet_dp = client.open_by_key(SHEET_DU_PHONG_ID).get_worksheet(0)
        _ensure_leave_sheet_header(sheet_dp)
        # Chỉ xóa vùng dữ liệu A:J, không đụng header hàng 1 và không đụng các cột khác.
        _gs_call_with_backoff(sheet_dp.batch_clear, ['A2:J'])
        values = df_excel.iloc[:, :10].values.tolist() if not df_excel.empty else []
        if values:
            last_row = len(values) + 1
            gspread_update_range(sheet_dp, f'A2:J{last_row}', values, value_input_option='USER_ENTERED')
        _clear_dynamic_data_caches()
        return True, f"Phiên bản 1 hoàn tất: đã GHI ĐÈ vùng A2:J và paste {len(values)} dòng từ Excel vào Sheet1."
    except Exception as e:
        return False, f"Lỗi đồng bộ Phiên bản 1: {e}"

def admin_sync_excel_to_gsheet_append():
    """Phiên bản 2: không sửa dòng hiện có; chỉ ghi các dòng mới vào đúng last row A:J."""
    try:
        client = get_gspread_client()
        if not client:
            return False, "Chưa cấu hình quyền kết nối Google Sheets."
        df_excel, err = _load_excel_leave_rows_for_google_sync()
        if err:
            return False, err
        sheet_dp = client.open_by_key(SHEET_DU_PHONG_ID).get_worksheet(0)
        _ensure_leave_sheet_header(sheet_dp)

        live = _live_sheet_to_leave_df(sheet_dp)
        existing_keys = set()
        if isinstance(live, pd.DataFrame) and not live.empty:
            existing_keys = {_leave_sync_merge_key(r) for _, r in live.iterrows()}

        rows = []
        seen_new = set()
        for _, r in df_excel.iterrows():
            key = _leave_sync_merge_key(r)
            if key in existing_keys or key in seen_new:
                continue
            if not key[0] or not key[1] or not key[2]:
                continue
            seen_new.add(key)
            rows.append(r.iloc[:10].tolist())

        if not rows:
            return True, "Phiên bản 2: không có dòng mới; dữ liệu hiện tại trên Google Sheet được giữ nguyên."

        target_row = _next_data_row_a_to_j(sheet_dp)
        end_row = target_row + len(rows) - 1
        # Ghi RANGE chính xác A:J tại last row; tuyệt đối không overwrite các dòng hiện hữu.
        gspread_update_range(sheet_dp, f'A{target_row}:J{end_row}', rows, value_input_option='USER_ENTERED')
        _clear_dynamic_data_caches()
        return True, f"Phiên bản 2 hoàn tất: đã thêm {len(rows)} dòng mới vào đúng A{target_row}:J{end_row}; không ghi đè dữ liệu cũ."
    except Exception as e:
        return False, f"Lỗi đồng bộ Phiên bản 2: {e}"

def admin_sync_excel_to_gsheet():
    """Giữ tương thích code cũ: mặc định dùng Phiên bản 2 an toàn, không ghi đè."""
    return admin_sync_excel_to_gsheet_append()

# --- ĐỒNG BỘ GOOGLE SHEETS SANG EXCEL (TẠO FILE DOWNLOAD CHỈ THÊM MỚI) ---
def admin_sync_gsheet_to_excel(df_gsheet, df_excel_goc):
    df_gsheet['Merge_Key'] = df_gsheet.apply(lambda r: '|'.join(_leave_sync_merge_key(r)), axis=1)
    df_excel_goc['Merge_Key'] = df_excel_goc.apply(lambda r: '|'.join((normalize_schedule_date(r.iloc[0]), normalize_employee_match_name(r.iloc[1]), normalize_leave_reason(r.iloc[2]))), axis=1)
    
    new_rows = df_gsheet[~df_gsheet['Merge_Key'].isin(df_excel_goc['Merge_Key'])].copy()
    
    if new_rows.empty:
        return df_excel_goc, False
        
    new_rows = new_rows.drop(columns=['Merge_Key'], errors='ignore')
    df_excel_merged = pd.concat([df_excel_goc.drop(columns=['Merge_Key'], errors='ignore'), new_rows], ignore_index=True)
    return df_excel_merged, True

# --- HÀM TẢI MẬT KHẨU, PHÂN QUYỀN VÀ TRẠNG THÁI ĐĂNG NHẬP ---
@st.cache_resource(show_spinner=False)
def ensure_credential_control_columns():
    """Tạo các cột điều khiển nếu Sheet mật khẩu cũ chưa có."""
    try:
        client = get_gspread_client()
        if not client: return
        sheet = client.open_by_key(SHEET_MAT_KHAU_ID).get_worksheet(0)
        header = _gs_call_with_backoff(sheet.row_values, 1)
        wanted = ["Khóa đăng nhập", "Remember Token Hash", "Remember Token Expiry"]
        # Sau khi chèn J/K: R=Khóa, S=Token Hash, T=Token Expiry
        if len(header) < 20 or header[17:20] != wanted:
            gspread_update_range(sheet, 'R1:T1', [wanted])
    except Exception:
        pass

def _load_credentials_from_sheets():
    try:
        client = get_gspread_client()
        if client:
            sheet = client.open_by_key(SHEET_MAT_KHAU_ID).get_worksheet(0)
            rows = _gs_call_with_backoff(sheet.get_all_values)
            if len(rows) > 1:
                data_list = []
                for idx, row in enumerate(rows[1:], start=2):
                    stt = row[0] if len(row) > 0 else idx - 1
                    ten = row[1] if len(row) > 1 else ""
                    pwd = str(row[2]) if len(row) > 2 and str(row[2]) != "" else "123456"
                    role = row[3] if len(row) > 3 else "nhanvien"
                    fullname = str(row[4]).strip() if len(row) > 4 else ""
                    dob = str(row[5]).strip() if len(row) > 5 else ""
                    phone = str(row[6]).strip() if len(row) > 6 else ""
                    email = str(row[7]).strip() if len(row) > 7 else ""
                    address = str(row[8]).strip() if len(row) > 8 else ""
                    bank_account = str(row[9]).strip() if len(row) > 9 else ""
                    bank_name = str(row[10]).strip() if len(row) > 10 else ""
                    ps_thang = str(row[11]).strip() if len(row) > 11 else "0"
                    cp_thang = str(row[12]).strip() if len(row) > 12 else "0"
                    pn_nam = str(row[13]).strip() if len(row) > 13 else "0"
                    ca_lam_viec = str(row[14]).strip() if len(row) > 14 else ""
                    ngay_bd = str(row[15]).strip() if len(row) > 15 else ""
                    chu_ky = str(row[16]).strip() if len(row) > 16 else ""
                    login_locked = str(row[17]).strip() if len(row) > 17 else ""
                    remember_hash = str(row[18]).strip() if len(row) > 18 else ""
                    remember_expiry = str(row[19]).strip() if len(row) > 19 else ""

                    # Bỏ các dòng tiêu đề phụ bị đặt lẫn trong dữ liệu tài khoản.
                    ten_norm = normalize_login_name(ten)
                    if ten_norm in {"ten nhan vien", "ten he thong", "username", "user name"}:
                        continue
                    if str(ten).strip() != "":
                        data_list.append({
                            'STT': stt, 'Tên nhân viên': str(ten).strip(), 'Mật khẩu': pwd,
                            'Phân quyền': str(role).strip().lower() if str(role).strip() else 'nhanvien',
                            'Họ và tên đầy đủ': fullname, 'Ngày sinh': dob, 'Điện thoại': phone,
                            'Email': email, 'Địa chỉ': address, 'Số tài khoản ngân hàng': bank_account,
                            'Tên ngân hàng': bank_name, 'Phát sinh tháng': ps_thang,
                            'Có phép tháng': cp_thang, 'Phép năm': pn_nam, 'Ca làm việc': ca_lam_viec,
                            'Ngày bắt đầu ca': ngay_bd, 'Chu kỳ': chu_ky,
                            'Khóa đăng nhập': login_locked, 'Remember Token Hash': remember_hash,
                            'Remember Token Expiry': remember_expiry
                        })
                return pd.DataFrame(data_list)
    except Exception:
        pass
    return pd.DataFrame(columns=[
        'STT', 'Tên nhân viên', 'Mật khẩu', 'Phân quyền', 'Họ và tên đầy đủ', 'Ngày sinh',
        'Điện thoại', 'Email', 'Địa chỉ', 'Số tài khoản ngân hàng', 'Tên ngân hàng',
        'Phát sinh tháng', 'Có phép tháng', 'Phép năm', 'Ca làm việc', 'Ngày bắt đầu ca',
        'Chu kỳ', 'Khóa đăng nhập',
        'Remember Token Hash', 'Remember Token Expiry'
    ])

@st.cache_data(ttl=120, show_spinner=False)
def load_credentials():
    """V75: đọc qua PostgreSQL dùng chung giữa các Cloud Run instance; Google Sheets là nguồn đồng bộ dự phòng."""
    if vpg is not None and vpg.is_enabled():
        return vpg.load_dataset(
            "credentials",
            _load_credentials_from_sheets,
            ttl_seconds=int(os.getenv("VERA_PG_TTL_CREDENTIALS", "30")),
        )
    return _load_credentials_from_sheets()

@st.cache_data(ttl=10, show_spinner=False)
def load_credentials_recent():
    """Ảnh chụp Sheet1 gần thời gian thực cho TOÀN BỘ trường hồ sơ.

    TTL 10 giây giúp các trang luôn nhận dữ liệu nguồn rất mới nhưng vẫn tránh lặp
    request Google Sheets trên từng thao tác gõ/chọn của Streamlit. Mọi tác vụ quan trọng
    như gửi email vẫn dùng load_credentials_fresh() để đọc trực tiếp ngay lúc bấm.
    """
    try:
        load_credentials.clear()
    except Exception:
        pass
    return load_credentials()

def load_credentials_fresh():
    """Đọc hồ sơ mới nhất; trên Cloud Run refresh một lần rồi chia sẻ snapshot cho mọi instance."""
    try:
        load_credentials.clear()
    except Exception:
        pass
    try:
        load_credentials_recent.clear()
    except Exception:
        pass
    if vpg is not None and vpg.is_enabled():
        try:
            return vpg.load_dataset(
                "credentials", _load_credentials_from_sheets,
                ttl_seconds=int(os.getenv("VERA_PG_TTL_CREDENTIALS", "30")),
                force_refresh=True,
            )
        except Exception:
            pass
    return _load_credentials_from_sheets()

def load_credentials_fresh_for_email():
    """Giữ tương thích tên hàm cũ; thực tế làm mới TOÀN BỘ hồ sơ, không chỉ Email."""
    return load_credentials_fresh()

def latest_credential_row_from_credentials(credentials_df, username):
    """Lấy dòng hồ sơ mới nhất theo Tên Hệ thống, so khớp không dấu/không phân biệt hoa thường."""
    if not isinstance(credentials_df, pd.DataFrame) or credentials_df.empty:
        return None
    if 'Tên nhân viên' not in credentials_df.columns:
        return None
    target = normalize_login_name(username)
    matched = credentials_df[
        credentials_df['Tên nhân viên'].astype(str).apply(normalize_login_name) == target
    ]
    if matched.empty:
        return None
    return matched.iloc[-1]

def latest_email_from_credentials(credentials_df, username):
    """Lấy Email mới nhất theo Tên Hệ thống."""
    row = latest_credential_row_from_credentials(credentials_df, username)
    if row is None:
        return ""
    return str(row.get('Email', '')).strip()

PAYROLL_PROFILE_SOURCE_MAP = {
    'Họ và tên': 'Họ và tên đầy đủ',
    'Email': 'Email',
    'Số tài khoản ngân hàng': 'Số tài khoản ngân hàng',
    'Tên ngân hàng': 'Tên ngân hàng',
}

def apply_latest_profile_fields_to_payroll(payroll_df, credentials_df=None, only_current_nhanvien=False):
    """Đồng bộ mọi trường hồ sơ của bảng lương từ Sheet1 nguồn.

    Các khoản tiền nghiệp vụ vẫn giữ nguyên. Những trường hồ sơ đang dùng trong bảng
    lương/export/email (Họ tên, Email, tài khoản và tên ngân hàng) luôn lấy theo nguồn.
    Khi only_current_nhanvien=True, vai trò mới nhất cũng được dùng để chỉ giữ tài khoản
    hiện vẫn có role `nhanvien` hoặc `leader`.
    """
    if payroll_df is None or not isinstance(payroll_df, pd.DataFrame) or payroll_df.empty:
        return payroll_df.copy() if isinstance(payroll_df, pd.DataFrame) else pd.DataFrame()
    d = payroll_df.copy()
    creds = credentials_df if isinstance(credentials_df, pd.DataFrame) else load_credentials_recent()
    if creds is None or creds.empty or 'Tên nhân viên' not in creds.columns or 'Tên Hệ thống' not in d.columns:
        return d

    cred_map = {}
    for _, cr in creds.iterrows():
        key = normalize_login_name(cr.get('Tên nhân viên', ''))
        if key:
            cred_map[key] = cr

    keep_mask = []
    for idx, rr in d.iterrows():
        key = normalize_login_name(rr.get('Tên Hệ thống', ''))
        cr = cred_map.get(key)
        keep = True
        if cr is not None:
            for dest, src in PAYROLL_PROFILE_SOURCE_MAP.items():
                if dest not in d.columns:
                    d[dest] = ''
                val = str(cr.get(src, '')).strip()
                if dest == 'Số tài khoản ngân hàng':
                    val = val.replace("'", '')
                d.at[idx, dest] = val
            if only_current_nhanvien:
                keep = str(cr.get('Phân quyền', '')).strip().lower() in PAYROLL_ELIGIBLE_ROLES
        elif only_current_nhanvien:
            keep = False
        keep_mask.append(bool(keep))

    if only_current_nhanvien and len(keep_mask) == len(d):
        d = d.loc[keep_mask].copy()
        if 'TT' in d.columns:
            d = d.reset_index(drop=True)
            d['TT'] = range(1, len(d) + 1)
    return d

def set_accounts_login_lock(usernames, locked=True):
    try:
        client = get_gspread_client()
        if not client: return False, "Chưa cấu hình quyền kết nối."
        sheet = client.open_by_key(SHEET_MAT_KHAU_ID).get_worksheet(0)
        values = _gs_call_with_backoff(sheet.get_all_values)
        targets = {normalize_login_name(x) for x in usernames}
        changed = 0
        for r_idx, row in enumerate(values[1:], start=2):
            if len(row) > 1 and normalize_login_name(row[1]) in targets:
                sheet.update_cell(r_idx, 18, 'KHÓA' if locked else '')
                if locked:
                    sheet.update_cell(r_idx, 19, '')
                    sheet.update_cell(r_idx, 20, '')
                changed += 1
        _clear_dynamic_data_caches()
        return True, f"Đã {'khóa' if locked else 'mở khóa'} {changed} tài khoản."
    except Exception as e:
        return False, f"Lỗi cập nhật khóa đăng nhập: {e}"

def create_remember_token(username, days=None):
    """Lưu HASH token ở Google Sheet và duy trì cho tới khi người dùng chủ động Đăng xuất."""
    try:
        client = get_gspread_client()
        if not client: return None
        sheet = client.open_by_key(SHEET_MAT_KHAU_ID).get_worksheet(0)
        values = _gs_call_with_backoff(sheet.get_all_values)
        target = normalize_login_name(username)
        for r_idx, row in enumerate(values[1:], start=2):
            if len(row) > 1 and normalize_login_name(row[1]) == target:
                token = secrets.token_urlsafe(32)
                token_hash = hashlib.sha256(token.encode('utf-8')).hexdigest()
                sheet.update_cell(r_idx, 19, token_hash)
                # Không đặt ngày hết hạn: token chỉ bị xóa khi Đăng xuất hoặc tài khoản bị khóa.
                sheet.update_cell(r_idx, 20, '')
                _clear_dynamic_data_caches()
                return token
    except Exception:
        pass
    return None

def revoke_remember_token(username):
    try:
        client = get_gspread_client()
        if not client: return
        sheet = client.open_by_key(SHEET_MAT_KHAU_ID).get_worksheet(0)
        values = _gs_call_with_backoff(sheet.get_all_values)
        target = normalize_login_name(username)
        for r_idx, row in enumerate(values[1:], start=2):
            if len(row) > 1 and normalize_login_name(row[1]) == target:
                sheet.update_cell(r_idx, 19, '')
                sheet.update_cell(r_idx, 20, '')
                break
        _clear_dynamic_data_caches()
    except Exception:
        pass

def validate_remember_token(token, credentials_df):
    """Token hợp lệ cho tới khi bị thu hồi/khóa; không tự hết hạn theo thời gian."""
    if not token or credentials_df.empty:
        return None
    token_hash = hashlib.sha256(str(token).encode('utf-8')).hexdigest()
    for _, row in credentials_df.iterrows():
        saved_hash = str(row.get('Remember Token Hash', '')).strip()
        if not saved_hash or not hmac.compare_digest(token_hash, saved_hash):
            continue
        if is_locked_value(row.get('Khóa đăng nhập', '')):
            return None
        return row
    return None

# --- CẬP NHẬT THÔNG TIN CÁ NHÂN ---
def update_user_profile(username, new_pass, fullname, dob, phone, email, address, bank_account="", bank_name="", new_role=None):
    try:
        client = get_gspread_client()
        if not client: return False, "Chưa cấu hình quyền kết nối."
        sheet = client.open_by_key(SHEET_MAT_KHAU_ID).get_worksheet(0)
        # Tìm không phân biệt dấu / HOA thường để đồng nhất với đăng nhập.
        values = _gs_call_with_backoff(sheet.get_all_values)
        target = normalize_login_name(username)
        row_idx = None
        for i, row in enumerate(values[1:], start=2):
            if len(row) > 1 and normalize_login_name(row[1]) == target:
                row_idx = i
                break
        if row_idx:
            if new_pass: sheet.update_cell(row_idx, 3, str(new_pass))
            if new_role is not None and str(new_role).strip():
                sheet.update_cell(row_idx, 4, str(new_role).strip().lower())
            sheet.update_cell(row_idx, 5, str(fullname))
            sheet.update_cell(row_idx, 6, str(dob))
            sheet.update_cell(row_idx, 7, f"'{phone}")
            sheet.update_cell(row_idx, 8, str(email))
            sheet.update_cell(row_idx, 9, str(address))
            # Hai cột mới được chèn giữa I và J.
            sheet.update_cell(row_idx, 10, f"'{bank_account}" if str(bank_account).strip() else "")
            sheet.update_cell(row_idx, 11, str(bank_name))
            _clear_dynamic_data_caches()
            if new_role is not None and str(new_role).strip():
                try:
                    sync_tichluy_roles_and_stt(load_credentials_fresh())
                except Exception:
                    pass
            return True, "Cập nhật hồ sơ thành công!"
        return False, "Không tìm thấy tài khoản."
    except Exception as e:
        return False, f"Lỗi cập nhật: {e}"

# --- THIẾT LẬP CA LÀM VIỆC / TEMPLATE EXCEL ---
SHIFT_OPTIONS = [
    "Ca 1 (10:00 - 23:00)",
    "Ca 2 (13:00 - 00:00)",
    "Cố định Ca 1 (Không đổi)",
    "Cố định Ca 2 (Không đổi)",
]
SHIFT_CYCLE_OPTIONS = [
    "Luân phiên (14 ngày)",
    "Theo chu kỳ Tháng",
    "Cố định (Không đổi)",
]

def get_nhanvien_shift_dataframe(credentials_df):
    """Trang Thiết lập ca chỉ hiển thị đúng role `nhanvien`, sắp xếp A→Z."""
    cols = ['Tên nhân viên', 'Ca làm việc', 'Ngày bắt đầu ca', 'Chu kỳ']
    if credentials_df is None or credentials_df.empty:
        return pd.DataFrame(columns=cols)
    d = credentials_df.copy()
    for c in cols:
        if c not in d.columns:
            d[c] = ''
    if 'Phân quyền' in d.columns:
        d = d[d['Phân quyền'].astype(str).str.strip().str.lower().eq('nhanvien')].copy()
    d = d[cols].copy()
    if not d.empty:
        d['_sort'] = d['Tên nhân viên'].astype(str).apply(normalize_login_name)
        d = d.sort_values('_sort', kind='stable').drop(columns=['_sort']).reset_index(drop=True)
    return d

def build_shift_template_excel_bytes(credentials_df):
    """Tạo template Excel phân ca có dropdown ở Ca làm việc và Chu kỳ luân phiên."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    d = get_nhanvien_shift_dataframe(credentials_df)
    wb = Workbook()
    ws = wb.active
    ws.title = 'ThietLapCa'
    headers = ['Tên nhân viên', 'Ca làm việc', 'Ngày bắt đầu ca', 'Chu kỳ luân phiên']
    ws.append(headers)
    for _, r in d.iterrows():
        ws.append([
            str(r.get('Tên nhân viên','')).strip(),
            str(r.get('Ca làm việc','')).strip(),
            str(r.get('Ngày bắt đầu ca','')).strip(),
            str(r.get('Chu kỳ','')).strip(),
        ])

    fill = PatternFill('solid', fgColor='A1948C')
    thin = Side(style='thin', color='D9D9D9')
    for cell in ws[1]:
        cell.fill = fill
        cell.font = Font(bold=True, color='000000')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in ws.iter_rows(min_row=1, max_row=max(2, ws.max_row), min_col=1, max_col=4):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical='center', wrap_text=False)
    widths = [24, 30, 22, 24]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:D{max(2, ws.max_row)}'

    dm = wb.create_sheet('DanhMuc')
    dm['A1'] = 'Ca làm việc'; dm['B1'] = 'Chu kỳ luân phiên'
    for i, v in enumerate(SHIFT_OPTIONS, start=2): dm.cell(i, 1, v)
    for i, v in enumerate(SHIFT_CYCLE_OPTIONS, start=2): dm.cell(i, 2, v)
    dm.sheet_state = 'hidden'

    max_target_row = max(300, ws.max_row + 100)
    dv_shift = DataValidation(type='list', formula1=f"=DanhMuc!$A$2:$A${len(SHIFT_OPTIONS)+1}", allow_blank=True)
    dv_cycle = DataValidation(type='list', formula1=f"=DanhMuc!$B$2:$B${len(SHIFT_CYCLE_OPTIONS)+1}", allow_blank=True)
    ws.add_data_validation(dv_shift); ws.add_data_validation(dv_cycle)
    dv_shift.add(f'B2:B{max_target_row}')
    dv_cycle.add(f'D2:D{max_target_row}')

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def read_shift_template_excel(uploaded_file, credentials_df):
    """Đọc template đã nhập và chỉ nhận tên thuộc role nhanvien hiện tại."""
    try:
        raw = pd.read_excel(uploaded_file, sheet_name='ThietLapCa', dtype=str, engine='openpyxl').fillna('')
    except Exception as e:
        return pd.DataFrame(), f'Không đọc được file template: {e}'
    raw = raw.rename(columns={'Chu kỳ luân phiên': 'Chu kỳ'})
    required = ['Tên nhân viên', 'Ca làm việc', 'Ngày bắt đầu ca', 'Chu kỳ']
    missing = [c for c in required if c not in raw.columns]
    if missing:
        return pd.DataFrame(), 'Thiếu cột: ' + ', '.join(missing)
    allowed = get_nhanvien_shift_dataframe(credentials_df)
    allowed_keys = {normalize_login_name(x) for x in allowed['Tên nhân viên'].astype(str).tolist()}
    raw = raw[required].copy()
    raw = raw[raw['Tên nhân viên'].astype(str).apply(normalize_login_name).isin(allowed_keys)].copy()
    raw['Ca làm việc'] = raw['Ca làm việc'].astype(str).where(raw['Ca làm việc'].astype(str).isin(SHIFT_OPTIONS), '')
    raw['Chu kỳ'] = raw['Chu kỳ'].astype(str).where(raw['Chu kỳ'].astype(str).isin(SHIFT_CYCLE_OPTIONS), '')
    raw['_sort'] = raw['Tên nhân viên'].astype(str).apply(normalize_login_name)
    raw = raw.sort_values('_sort', kind='stable').drop(columns=['_sort']).reset_index(drop=True)
    return raw, ''

# --- GHI NHẬN HÀNG LOẠT CA LÀM VIỆC TỪ DATAFRAME ---
def batch_update_shift_schedule(edited_df):
    """Ghi đồng loạt O:P:Q một lần, không update lại toàn bộ Sheet1."""
    try:
        client = get_gspread_client()
        if not client: return False, "Chưa cấu hình quyền kết nối."
        sheet = client.open_by_key(SHEET_MAT_KHAU_ID).get_worksheet(0)
        all_vals = _gs_call_with_backoff(sheet.get_all_values)
        if not all_vals or len(all_vals) < 2:
            return False, "Sheet1 chưa có dữ liệu nhân viên."

        shift_map = {}
        for _, r in edited_df.iterrows():
            nv_name = normalize_login_name(r.get('Tên nhân viên', ''))
            if not nv_name:
                continue
            shift_map[nv_name] = {
                'ca': str(r.get('Ca làm việc', '')).replace("nan", "").strip(),
                'ngay': str(r.get('Ngày bắt đầu ca', '')).replace("nan", "").strip(),
                'chuky': str(r.get('Chu kỳ', '')).replace("nan", "").strip()
            }

        out = []
        updated = 0
        for row in all_vals[1:]:
            rr = list(row)
            while len(rr) < 17:
                rr.append("")
            key = normalize_login_name(rr[1] if len(rr) > 1 else '')
            ca, ngay, chuky = rr[14], rr[15], rr[16]
            if key in shift_map:
                ca = shift_map[key]['ca']; ngay = shift_map[key]['ngay']; chuky = shift_map[key]['chuky']
                updated += 1
            out.append([ca, ngay, chuky])

        if out:
            gspread_update_range(sheet, f"O2:Q{len(out)+1}", out, value_input_option='USER_ENTERED')
        _clear_dynamic_data_caches()
        return True, f"Đã lưu cấu hình ca cho {updated} nhân viên."
    except Exception as e:
        return False, f"Lỗi cập nhật: {e}"

# --- DANH SÁCH NHÂN SỰ: SẮP XẾP / EXPORT / IMPORT ---
STAFF_EXPORT_COLUMNS = [
    'Tên nhân viên', 'Họ và tên đầy đủ', 'Phân quyền', 'Trạng thái làm việc',
    'Điện thoại', 'Email', 'Địa chỉ', 'Số tài khoản ngân hàng',
    'Tên ngân hàng', 'Khóa đăng nhập'
]


def normalize_employment_status_value(value, default=EMPLOYMENT_STATUS_ACTIVE):
    key = normalize_login_name(value)
    if not key:
        return default
    return EMPLOYMENT_STATUS_ALIASES.get(key, default)


def build_staff_list_dataframe(credentials_df):
    """Danh sách nhân sự chuẩn: tên A→Z, role chuẩn hóa, trạng thái theo 3 mức mới."""
    if credentials_df is None or not isinstance(credentials_df, pd.DataFrame) or credentials_df.empty:
        return pd.DataFrame(columns=STAFF_EXPORT_COLUMNS)
    d = credentials_df.copy()
    employment_map = load_employment_status_map()
    if 'Tên nhân viên' not in d.columns:
        d['Tên nhân viên'] = ''
    if 'Phân quyền' not in d.columns:
        d['Phân quyền'] = 'nhanvien'
    d['Phân quyền'] = d['Phân quyền'].astype(str).str.strip().str.lower()
    d['Trạng thái làm việc'] = d['Tên nhân viên'].astype(str).apply(
        lambda name: employment_map.get(normalize_login_name(name), EMPLOYMENT_STATUS_ACTIVE)
    )
    for col in STAFF_EXPORT_COLUMNS:
        if col not in d.columns:
            d[col] = ''
    # Thứ tự nhân sự: Phân quyền theo thứ tự chuẩn; trong từng nhóm role, Tên nhân viên A→Z.
    # Các role còn lại (tapvu/admin) được đặt sau nhóm người dùng yêu cầu.
    role_rank = {role: i for i, role in enumerate(STAFF_ROLE_ORDER)}
    d['_staff_name_sort'] = d['Tên nhân viên'].astype(str).apply(normalize_login_name)
    d['_staff_role_rank'] = d['Phân quyền'].map(role_rank).fillna(len(role_rank)).astype(int)
    d = d.sort_values(['_staff_role_rank', '_staff_name_sort'], kind='stable')
    return d[STAFF_EXPORT_COLUMNS].reset_index(drop=True)


def staff_list_to_excel(df):
    """Export Danh sách nhân sự với dropdown Phân quyền/Trạng thái để có thể sửa rồi import lại."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    d = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame(columns=STAFF_EXPORT_COLUMNS)
    for col in STAFF_EXPORT_COLUMNS:
        if col not in d.columns:
            d[col] = ''
    d = d[STAFF_EXPORT_COLUMNS].copy()

    wb = Workbook()
    ws = wb.active
    ws.title = 'DanhSachNhanSu'
    ws.append(list(d.columns))
    for _, row in d.iterrows():
        ws.append([str(row.get(c, '') if not pd.isna(row.get(c, '')) else '') for c in d.columns])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    widths = {
        'Tên nhân viên': 24, 'Họ và tên đầy đủ': 28, 'Phân quyền': 14,
        'Trạng thái làm việc': 22, 'Điện thoại': 16, 'Email': 30, 'Địa chỉ': 42,
        'Số tài khoản ngân hàng': 22, 'Tên ngân hàng': 24, 'Khóa đăng nhập': 18,
    }
    for idx, col in enumerate(d.columns, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col, 18)

    max_row = max(2, ws.max_row)
    if 'Phân quyền' in d.columns:
        col_idx = d.columns.get_loc('Phân quyền') + 1
        role_list = ','.join(STAFF_ROLE_ORDER)
        dv_role = DataValidation(type='list', formula1=f'"{role_list}"', allow_blank=False)
        ws.add_data_validation(dv_role)
        dv_role.add(f'{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{max_row}')
    if 'Trạng thái làm việc' in d.columns:
        col_idx = d.columns.get_loc('Trạng thái làm việc') + 1
        status_list = ','.join(EMPLOYMENT_STATUS_OPTIONS)
        dv_status = DataValidation(type='list', formula1=f'"{status_list}"', allow_blank=False)
        ws.add_data_validation(dv_status)
        dv_status.add(f'{get_column_letter(col_idx)}2:{get_column_letter(col_idx)}{max_row}')

    note = wb.create_sheet('HuongDan')
    note['A1'] = 'HƯỚNG DẪN IMPORT DANH SÁCH NHÂN SỰ'
    note['A1'].font = Font(bold=True)
    note['A3'] = '1. Không đổi Tên nhân viên vì đây là khóa đối chiếu tài khoản.'
    note['A4'] = '2. Có thể sửa các cột hồ sơ, Phân quyền, Trạng thái làm việc và Khóa đăng nhập.'
    note['A5'] = '3. Import không ghi đè Mật khẩu, Ngày sinh, quỹ phép, ca làm việc hoặc Remember Token (trừ khi khóa đăng nhập).'
    note.column_dimensions['A'].width = 110

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def read_staff_list_import(uploaded_file):
    """Đọc file export Danh sách nhân sự; Tên nhân viên là khóa, không tạo/xóa tài khoản."""
    try:
        raw = pd.read_excel(uploaded_file, dtype=str, engine='openpyxl').fillna('')
    except Exception as e:
        return pd.DataFrame(), f"Không đọc được file Excel: {e}"
    if 'Tên nhân viên' not in raw.columns:
        return pd.DataFrame(), "File import phải có cột 'Tên nhân viên'."
    raw = raw.copy()
    raw['Tên nhân viên'] = raw['Tên nhân viên'].astype(str).str.strip()
    raw = raw[raw['Tên nhân viên'] != ''].copy()
    if raw.empty:
        return pd.DataFrame(), "File import không có nhân viên."

    keys = raw['Tên nhân viên'].apply(normalize_login_name)
    duplicated = raw.loc[keys.duplicated(keep=False), 'Tên nhân viên'].astype(str).tolist()
    if duplicated:
        return pd.DataFrame(), "Tên nhân viên bị trùng trong file import: " + ", ".join(sort_employee_names(duplicated))

    if 'Phân quyền' in raw.columns:
        roles = raw['Phân quyền'].astype(str).str.strip().str.lower()
        invalid_roles = sorted({x for x in roles if x and x not in ALL_ACCOUNT_ROLES})
        if invalid_roles:
            return pd.DataFrame(), "Phân quyền không hợp lệ: " + ", ".join(invalid_roles)
        raw['Phân quyền'] = roles

    if 'Trạng thái làm việc' in raw.columns:
        invalid_statuses = []
        normalized_statuses = []
        for val in raw['Trạng thái làm việc'].astype(str).tolist():
            if not str(val).strip():
                normalized_statuses.append('')
                continue
            key = normalize_login_name(val)
            if key not in EMPLOYMENT_STATUS_ALIASES:
                invalid_statuses.append(str(val).strip())
                normalized_statuses.append('')
            else:
                normalized_statuses.append(EMPLOYMENT_STATUS_ALIASES[key])
        if invalid_statuses:
            return pd.DataFrame(), "Trạng thái làm việc không hợp lệ: " + ", ".join(sorted(set(invalid_statuses)))
        raw['Trạng thái làm việc'] = normalized_statuses

    raw['_staff_import_key'] = raw['Tên nhân viên'].apply(normalize_login_name)
    return raw, ''


def batch_import_staff_list(import_df, updated_by, actor_role="admin"):
    """Import lại danh sách nhân sự hiện có, giữ giới hạn phân quyền của Admin/Lễ tân/Quản lý."""
    if import_df is None or not isinstance(import_df, pd.DataFrame) or import_df.empty:
        return False, "Không có dữ liệu để import."
    try:
        client = get_gspread_client()
        if not client:
            return False, "Chưa cấu hình quyền kết nối Google Sheets."
        sheet = client.open_by_key(SHEET_MAT_KHAU_ID).get_worksheet(0)
        all_vals = _gs_call_with_backoff(sheet.get_all_values)
        if not all_vals or len(all_vals) < 2:
            return False, "Sheet1 chưa có dữ liệu nhân sự."

        import_map = {
            normalize_login_name(r.get('Tên nhân viên', '')): r
            for _, r in import_df.iterrows()
            if normalize_login_name(r.get('Tên nhân viên', ''))
        }
        existing_keys = {
            normalize_login_name(row[1] if len(row) > 1 else '')
            for row in all_vals[1:]
            if normalize_login_name(row[1] if len(row) > 1 else '')
        }
        unknown = [str(r.get('Tên nhân viên', '')).strip() for _, r in import_df.iterrows()
                   if normalize_login_name(r.get('Tên nhân viên', '')) not in existing_keys]
        if unknown:
            return False, "Không tìm thấy trong hệ thống: " + ", ".join(sort_employee_names(unknown))

        actor_role = str(actor_role or '').strip().lower()
        current_status_map = load_employment_status_map()
        if actor_role != 'admin':
            # Lễ tân/Quản lý có thể import file đầy đủ, nhưng chỉ được THAY ĐỔI
            # các tài khoản thuộc nhanvien/locker/tapvu và không được nâng role.
            for row in all_vals[1:]:
                rr = list(row)
                while len(rr) < 20:
                    rr.append('')
                key = normalize_login_name(rr[1])
                item = import_map.get(key)
                if item is None:
                    continue
                current_role = str(rr[3]).strip().lower() or 'nhanvien'
                imported_role = str(item.get('Phân quyền', current_role)).strip().lower() if 'Phân quyền' in import_df.columns else current_role
                if current_role in FRONTDESK_MANAGEABLE_ROLES:
                    if imported_role and imported_role not in FRONTDESK_MANAGEABLE_ROLES:
                        return False, f"Không được đổi {rr[1]} sang phân quyền '{imported_role}'."
                    continue

                def _same_text(col, current):
                    return col not in import_df.columns or str(item.get(col, '')).replace("'", '').strip() == str(current).replace("'", '').strip()

                changed = False
                if 'Phân quyền' in import_df.columns and imported_role != current_role:
                    changed = True
                changed = changed or not _same_text('Họ và tên đầy đủ', rr[4])
                changed = changed or not _same_text('Điện thoại', rr[6])
                changed = changed or not _same_text('Email', rr[7])
                changed = changed or not _same_text('Địa chỉ', rr[8])
                changed = changed or not _same_text('Số tài khoản ngân hàng', rr[9])
                changed = changed or not _same_text('Tên ngân hàng', rr[10])
                if 'Khóa đăng nhập' in import_df.columns:
                    changed = changed or (is_locked_value(item.get('Khóa đăng nhập', '')) != is_locked_value(rr[17]))
                if 'Trạng thái làm việc' in import_df.columns:
                    imported_status = normalize_employment_status_value(item.get('Trạng thái làm việc', ''))
                    current_status = current_status_map.get(key, EMPLOYMENT_STATUS_ACTIVE)
                    changed = changed or imported_status != current_status
                if changed:
                    return False, (
                        f"Lễ tân/Quản lý không được thay đổi tài khoản {rr[1]} ({current_role}). "
                        "Chỉ được import thay đổi cho nhanvien, locker hoặc tapvu."
                    )

        out_dk = []  # D:K = role, họ tên, ngày sinh(preserve), phone, email, address, bank acc, bank
        out_rt = []  # R:T = khóa login + remember token hash/expiry (token chỉ xóa khi khóa)
        updated_count = 0
        imported_status_by_key = {}
        imported_role_by_key = {}

        for row in all_vals[1:]:
            rr = list(row)
            while len(rr) < 20:
                rr.append('')
            key = normalize_login_name(rr[1])
            item = import_map.get(key)
            role, fullname, dob, phone, email, address, bank_acc, bank_name = rr[3:11]
            lock_val, token_hash, token_expiry = rr[17], rr[18], rr[19]
            if item is not None:
                updated_count += 1
                if 'Phân quyền' in import_df.columns and str(item.get('Phân quyền', '')).strip():
                    role = str(item.get('Phân quyền', '')).strip().lower()
                imported_role_by_key[key] = role
                if 'Họ và tên đầy đủ' in import_df.columns:
                    fullname = str(item.get('Họ và tên đầy đủ', '')).strip()
                if 'Điện thoại' in import_df.columns:
                    phone = str(item.get('Điện thoại', '')).replace("'", '').strip()
                if 'Email' in import_df.columns:
                    email = str(item.get('Email', '')).strip()
                if 'Địa chỉ' in import_df.columns:
                    address = str(item.get('Địa chỉ', '')).strip()
                if 'Số tài khoản ngân hàng' in import_df.columns:
                    bank_acc = str(item.get('Số tài khoản ngân hàng', '')).replace("'", '').strip()
                if 'Tên ngân hàng' in import_df.columns:
                    bank_name = str(item.get('Tên ngân hàng', '')).strip()
                if 'Khóa đăng nhập' in import_df.columns:
                    lock_val = 'KHÓA' if is_locked_value(item.get('Khóa đăng nhập', '')) else ''
                    if lock_val:
                        token_hash = ''
                        token_expiry = ''
                if 'Trạng thái làm việc' in import_df.columns:
                    raw_status = str(item.get('Trạng thái làm việc', '')).strip()
                    if raw_status:
                        imported_status_by_key[key] = normalize_employment_status_value(raw_status)

            phone_write = f"'{phone}" if str(phone).strip() else ''
            bank_write = f"'{bank_acc}" if str(bank_acc).strip() else ''
            out_dk.append([role, fullname, dob, phone_write, email, address, bank_write, bank_name])
            out_rt.append([lock_val, token_hash, token_expiry])

        # Hai lần ghi range lớn thay cho hàng trăm update_cell -> giảm nguy cơ Sheets 429.
        last_row = len(out_dk) + 1
        if out_dk:
            gspread_update_range(sheet, f'D2:K{last_row}', out_dk, value_input_option='USER_ENTERED')
            gspread_update_range(sheet, f'R2:T{last_row}', out_rt, value_input_option='USER_ENTERED')

        # Cập nhật bảng trạng thái làm việc theo cùng danh sách nhân sự.
        if imported_status_by_key:
            status_ws = _get_employment_status_worksheet()
            existing_status = load_employment_status_map()
            now = datetime.now(VN_TZ)
            status_rows = []
            stt = 0
            for row in all_vals[1:]:
                name = str(row[1]).strip() if len(row) > 1 else ''
                key = normalize_login_name(name)
                if not key:
                    continue
                role_now = imported_role_by_key.get(key, str(row[3]).strip().lower() if len(row) > 3 else 'nhanvien')
                if role_now == 'admin':
                    continue
                stt += 1
                status = imported_status_by_key.get(key, existing_status.get(key, EMPLOYMENT_STATUS_ACTIVE))
                status_rows.append([
                    stt, name, status, now.strftime('%d/%m/%Y'), now.strftime('%H:%M:%S'), str(updated_by).strip()
                ])
            if status_ws is not None:
                try:
                    _gs_call_with_backoff(status_ws.batch_clear, ['A2:F1000'])
                except Exception:
                    pass
                if status_rows:
                    gspread_update_range(status_ws, f'A2:F{len(status_rows)+1}', status_rows, value_input_option='USER_ENTERED')

        _clear_dynamic_data_caches()
        try:
            load_employment_status_map.clear()
        except Exception:
            pass
        try:
            sync_tichluy_roles_and_stt(load_credentials_fresh())
        except Exception:
            pass
        return True, f"Đã import và cập nhật {updated_count} nhân viên. Mật khẩu, ngày sinh, quỹ phép, ca làm việc và Remember Token không bị ghi đè."
    except Exception as e:
        return False, f"Lỗi import danh sách nhân sự: {e}"


# --- TẢI DỮ LIỆU TỪ GOOGLE SHEET DỰ PHÒNG ---
def _load_backup_sheet_data_from_sheets():
    """Đọc trực tiếp A:J của sheet lịch nghỉ chính, không phụ thuộc tên header và luôn giữ source row."""
    expected = [
        "Ngày", "Tên nhân viên", "Lý do nghỉ", "Chi tiết", "Số ngày tính",
        "Số ngày phép cộng dồn", "Phạt vi phạm", "Ngày cập nhật", "Giờ cập nhật", "Người cập nhật"
    ]
    try:
        client = get_gspread_client()
        if not client:
            return pd.DataFrame(columns=expected + ['__source_sheet_id', '__source_row'])
        sheet = client.open_by_key(SHEET_DU_PHONG_ID).get_worksheet(0)
        values = _gs_call_with_backoff(sheet.get, 'A:J')
        if not values or len(values) < 2:
            # Một số phiên bản gspread / thay đổi định dạng sheet có thể trả rỗng/chỉ header cho range.
            # Fallback get_all_values bảo đảm Chi tiết danh sách không bị trắng.
            values = _gs_call_with_backoff(sheet.get_all_values)
            values = [list(r[:10]) for r in values]
        if not values or len(values) < 2:
            return pd.DataFrame(columns=expected + ['__source_sheet_id', '__source_row'])
        rows = []
        for sheet_row, row in enumerate(values[1:], start=2):
            r = list(row[:10]) + [""] * max(0, 10 - len(row))
            if not any(str(v).strip() for v in r):
                continue
            item = dict(zip(expected, r[:10]))
            item['__source_sheet_id'] = SHEET_DU_PHONG_ID
            item['__source_row'] = sheet_row
            rows.append(item)
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=expected + ['__source_sheet_id', '__source_row'])
    except Exception:
        return pd.DataFrame(columns=expected + ['__source_sheet_id', '__source_row'])

@st.cache_data(ttl=20, show_spinner=False)
def load_backup_sheet_data():
    """V75: đọc qua PostgreSQL dùng chung giữa các Cloud Run instance; Google Sheets là nguồn đồng bộ dự phòng."""
    if vpg is not None and vpg.is_enabled():
        return vpg.load_dataset(
            "leave_primary",
            _load_backup_sheet_data_from_sheets,
            ttl_seconds=int(os.getenv("VERA_PG_TTL_LEAVE_PRIMARY", "45")),
        )
    return _load_backup_sheet_data_from_sheets()

def _load_secondary_leave_sheet_data_from_sheets():
    """Đọc Sheet1 của Google Sheet thứ hai, chuẩn hóa về đúng A:J của lịch nghỉ."""
    expected = [
        "Ngày", "Tên nhân viên", "Lý do nghỉ", "Chi tiết", "Số ngày tính",
        "Số ngày phép cộng dồn", "Phạt vi phạm", "Ngày cập nhật", "Giờ cập nhật", "Người cập nhật"
    ]
    try:
        client = get_gspread_client()
        if not client:
            return pd.DataFrame(columns=expected)
        sheet = client.open_by_key(SHEET_LICH_NGHI_2_ID).get_worksheet(0)
        values = _gs_call_with_backoff(sheet.get, 'A:J')
        if not values or len(values) < 2:
            # Một số phiên bản gspread / thay đổi định dạng sheet có thể trả rỗng/chỉ header cho range.
            # Fallback get_all_values bảo đảm Chi tiết danh sách không bị trắng.
            values = _gs_call_with_backoff(sheet.get_all_values)
            values = [list(r[:10]) for r in values]
        if not values or len(values) < 2:
            return pd.DataFrame(columns=expected)

        rows = []
        for sheet_row, row in enumerate(values[1:], start=2):
            r = list(row[:10]) + [""] * max(0, 10 - len(row))
            if not any(str(v).strip() for v in r):
                continue
            item = dict(zip(expected, r[:10]))
            item['__source_sheet_id'] = SHEET_LICH_NGHI_2_ID
            item['__source_row'] = sheet_row
            rows.append(item)
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=expected + ['__source_sheet_id', '__source_row'])
    except Exception:
        return pd.DataFrame(columns=expected + ['__source_sheet_id', '__source_row'])

@st.cache_data(ttl=30, show_spinner=False)
def load_secondary_leave_sheet_data():
    """V75: đọc qua PostgreSQL dùng chung giữa các Cloud Run instance; Google Sheets là nguồn đồng bộ dự phòng."""
    if vpg is not None and vpg.is_enabled():
        return vpg.load_dataset(
            "leave_secondary",
            _load_secondary_leave_sheet_data_from_sheets,
            ttl_seconds=int(os.getenv("VERA_PG_TTL_LEAVE_SECONDARY", "90")),
        )
    return _load_secondary_leave_sheet_data_from_sheets()

@st.cache_data(ttl=120, show_spinner=False)
def load_loai_nghi_from_gsheet():
    try:
        client = get_gspread_client()
        if client:
            sheet = client.open_by_key(SHEET_DU_PHONG_ID).worksheet("LoaiNghi")
            rows = _gs_call_with_backoff(sheet.get_all_values)
            if len(rows) > 1:
                return pd.DataFrame(rows[1:], columns=rows[0])
    except Exception:
        pass
    return pd.DataFrame()

# --- GHI VÀ XÓA LỊCH ---
def _next_data_row_a_to_j(sheet):
    """Tìm dòng kế tiếp sau last row thực tế trong vùng A:J."""
    values = _gs_call_with_backoff(sheet.get, 'A:J')
    last_non_empty = 0
    for idx, row in enumerate(values, start=1):
        if any(str(v).strip() != "" for v in row[:10]):
            last_non_empty = idx
    return max(2, last_non_empty + 1)

def _live_sheet_to_leave_df(sheet):
    """Đọc trực tiếp A:J để kiểm tra trùng/thứ tự ngay trước khi ghi."""
    try:
        values = _gs_call_with_backoff(sheet.get, 'A:J')
        if not values or len(values) < 2:
            return pd.DataFrame(columns=[
                "Ngày", "Tên nhân viên", "Lý do nghỉ", "Chi tiết", "Số ngày tính",
                "Số ngày phép cộng dồn", "Phạt vi phạm", "Ngày cập nhật", "Giờ cập nhật", "Người cập nhật"
            ])
        header = [str(x).strip() for x in values[0][:10]]
        expected = [
            "Ngày", "Tên nhân viên", "Lý do nghỉ", "Chi tiết", "Số ngày tính",
            "Số ngày phép cộng dồn", "Phạt vi phạm", "Ngày cập nhật", "Giờ cập nhật", "Người cập nhật"
        ]
        if len(header) < 10 or not header[0]:
            header = expected
        rows = []
        for row in values[1:]:
            r = list(row[:10]) + [""] * max(0, 10 - len(row))
            if any(str(v).strip() for v in r):
                rows.append(r[:10])
        df = pd.DataFrame(rows, columns=header[:10]) if rows else pd.DataFrame(columns=header[:10])
        if 'Loại nghỉ' in df.columns and 'Lý do nghỉ' not in df.columns:
            df = df.rename(columns={'Loại nghỉ': 'Lý do nghỉ'})
        for c in expected:
            if c not in df.columns:
                df[c] = ""
        return df[expected].copy()
    except Exception:
        return pd.DataFrame()


def _daily_leave_group(reason, reason_type_map=None):
    """
    Phân nhóm đăng ký theo Loại nghỉ (cột C của LoaiNghi), fallback theo Lý do nghỉ.
    Trả về: co_phep / khong_phep / phat_sinh / "".
    """
    reason_clean = clean_leave_reason_display(reason)
    mapping = reason_type_map if isinstance(reason_type_map, dict) else _leave_reason_type_map(
        globals().get('df_loai_nghi', pd.DataFrame())
    )
    leave_type = clean_leave_reason_display(
        mapping.get(normalize_leave_reason(reason_clean), "")
    )

    type_key = normalize_login_name(leave_type)
    reason_key = normalize_login_name(reason_clean)

    if "khong phep" in type_key:
        return "khong_phep"
    if "phat sinh" in type_key:
        return "phat_sinh"
    if "co phep" in type_key:
        return "co_phep"

    # Fallback nếu danh mục tạm thời chưa tải được.
    if "khong phep" in reason_key:
        return "khong_phep"
    if "phat sinh" in reason_key:
        return "phat_sinh"
    if (
        "co phep" in reason_key
        or "nghi phep" in reason_key
        or "nghi dam hieu" in reason_key
        or re.search(r"(^|\s)cp($|\s)", reason_key)
    ):
        return "co_phep"
    return ""


def _validate_daily_employee_registration_rule(df_sources, ngay, employee, new_reason, new_days):
    """
    V86.11 - Quy tắc bắt buộc cho MỘT NHÂN VIÊN trong MỘT NGÀY:

    1) Số ngày tính của một dòng đăng ký chỉ được 0, 0.5 hoặc 1.
    2) Chỉ được có tối đa 1 dòng có Số ngày tính > 0.
       Vì vậy 0.5 + 0.5 trong cùng ngày cũng bị chặn.
    3) Tối đa 1 dòng thuộc từng nhóm:
       - Có phép
       - Không phép
       - Phát sinh

    Các dòng Số ngày tính = 0 vẫn phải tuân thủ quy tắc không lặp nhóm.
    """
    try:
        new_days = float(new_days or 0)
    except Exception:
        new_days = 0.0

    # Chỉ chấp nhận đúng 0 / 0.5 / 1 cho đăng ký lịch nghỉ.
    allowed_day_values = (0.0, 0.5, 1.0)
    if not any(abs(new_days - x) < 1e-9 for x in allowed_day_values):
        return False, "Số ngày tính trong 1 ngày chỉ được phép là 0, 0.5 hoặc 1."

    target_dt = pd.to_datetime(ngay, errors="coerce", dayfirst=True)
    if pd.isna(target_dt):
        return False, "Không xác định được ngày đăng ký."
    target_date = target_dt.date()

    if df_sources is None or not isinstance(df_sources, pd.DataFrame) or df_sources.empty:
        return True, ""

    required = {"Ngày", "Tên nhân viên", "Lý do nghỉ"}
    if not required.issubset(set(df_sources.columns)):
        return True, ""

    d = df_sources.copy()
    d["_date_rule"] = pd.to_datetime(d["Ngày"], errors="coerce", dayfirst=True).dt.date
    d["_emp_rule"] = d["Tên nhân viên"].astype(str).apply(normalize_login_name)
    d = d[
        (d["_date_rule"] == target_date)
        & (d["_emp_rule"] == normalize_login_name(employee))
    ].copy()

    if d.empty:
        return True, ""

    # --- Quy tắc: trong ngày chỉ được 1 dòng có Số ngày tính > 0 ---
    if new_days > 0:
        if "Số ngày tính" in d.columns:
            existing_days = pd.to_numeric(d["Số ngày tính"], errors="coerce").fillna(0.0)
        else:
            existing_days = pd.Series([0.0] * len(d), index=d.index)

        positive_rows = d[existing_days > 0]
        if not positive_rows.empty:
            existing_desc = []
            for _, r in positive_rows.iterrows():
                try:
                    day_val = float(pd.to_numeric(r.get("Số ngày tính", 0), errors="coerce") or 0)
                except Exception:
                    day_val = 0.0
                existing_desc.append(
                    f"{clean_leave_reason_display(r.get('Lý do nghỉ', ''))} ({day_val:g} ngày)"
                )
            return False, (
                "Trong cùng 1 ngày, mỗi nhân viên chỉ được có 1 dòng có Số ngày tính > 0. "
                "Không cho phép 0.5 + 0.5 = 1. "
                f"Đã có: {', '.join(existing_desc)}."
            )

    # --- Quy tắc: không được có 2 lần cùng nhóm Có phép/Không phép/Phát sinh ---
    reason_type_map = _leave_reason_type_map(globals().get("df_loai_nghi", pd.DataFrame()))
    new_group = _daily_leave_group(new_reason, reason_type_map)

    if new_group:
        existing_groups = d["Lý do nghỉ"].astype(str).apply(
            lambda x: _daily_leave_group(x, reason_type_map)
        )
        same_group_rows = d[existing_groups == new_group]
        if not same_group_rows.empty:
            labels = {
                "co_phep": "CÓ phép",
                "khong_phep": "KHÔNG phép",
                "phat_sinh": "PHÁT SINH",
            }
            old_reasons = [
                clean_leave_reason_display(x)
                for x in same_group_rows["Lý do nghỉ"].astype(str).tolist()
                if clean_leave_reason_display(x)
            ]
            return False, (
                f"Trong cùng 1 ngày, một nhân viên không được có 2 lần {labels.get(new_group, new_group)}. "
                f"Đã có: {', '.join(old_reasons)}."
            )

    return True, ""


def _leave_exists_in_sources(df_sources, ngay, nv, loai_nghi):
    """
    Chỉ xem là TRÙNG khi cùng Ngày + cùng Nhân viên + cùng Lý do nghỉ.
    Vì vậy trong cùng một ngày, một nhân viên được phép có nhiều dòng vi phạm/phạt > 0
    nếu mỗi dòng là một Lý do nghỉ khác nhau.
    """
    if df_sources is None or df_sources.empty:
        return False
    target_date = pd.to_datetime(ngay, errors='coerce', dayfirst=True)
    if pd.isna(target_date):
        return False
    target_date = target_date.date()
    d = df_sources.copy()
    d['Ngày_cmp'] = pd.to_datetime(d['Ngày'], errors='coerce', dayfirst=True).dt.date
    name_cmp = d['Tên nhân viên'].astype(str).apply(normalize_login_name)
    reason_cmp = d['Lý do nghỉ'].astype(str).apply(normalize_leave_reason)
    return bool(((d['Ngày_cmp'] == target_date) &
                 (name_cmp == normalize_login_name(nv)) &
                 (reason_cmp == normalize_leave_reason(loai_nghi))).any())


def _progressive_ordinal_and_bonus(df_sources, ngay, loai_nghi):
    """
    Tính thứ tự RIÊNG cho từng loại vi phạm trong cùng ngày:
    - Nghỉ không phép
    - Đi trễ không phép
    - Về sớm không phép (kể cả biến thể Ra sớm không phép)

    Người 1/2: +0; Người 3: +100.000; Người 4: +200.000; ...
    """
    canonical = get_progressive_penalty_reason(loai_nghi)
    if canonical is None:
        return 1, 0

    target_date = pd.to_datetime(ngay, errors='coerce', dayfirst=True)
    if pd.isna(target_date) or df_sources is None or df_sources.empty:
        ordinal = 1
    else:
        target_date = target_date.date()
        d = df_sources.copy()
        d['Ngày_cmp'] = pd.to_datetime(d['Ngày'], errors='coerce', dayfirst=True).dt.date
        canonical_series = d['Lý do nghỉ'].astype(str).apply(get_progressive_penalty_reason)
        mask = (d['Ngày_cmp'] == target_date) & canonical_series.eq(canonical)
        ordinal = int(mask.sum()) + 1

    bonus = max(0, ordinal - 2) * 100000
    return ordinal, bonus


def _unexcused_ordinal_and_bonus(df_sources, ngay):
    """Alias tương thích code cũ cho riêng Nghỉ không phép."""
    return _progressive_ordinal_and_bonus(df_sources, ngay, "Nghỉ không phép")


def save_lich_nghi_to_backup_sheet(ngay, nv, loai_nghi, chi_tiet, so_ngay, so_ngay_cong_don, phat_vi_pham, updated_by, df_main_source=None):
    """
    Chỉ ghi lịch vào Google Sheet dự phòng (SHEET_DU_PHONG_ID), Sheet1, đúng A:J ở last row.
    KHÔNG ghi lịch đăng ký mới sang file chính (SHEET_CHINH_ID).
    Trước khi ghi sẽ đọc LIVE Sheet1 để:
    - chặn trùng cùng nhân viên + ngày + loại nghỉ;
    - tính thứ tự riêng cho Nghỉ không phép / Đi trễ không phép / Về sớm không phép và tiền phạt lũy tiến.
    """
    try:
        client = get_gspread_client()
        if not client:
            return False, "Chưa cấu hình quyền kết nối Google Sheets."

        ngay_cn = get_vn_today().strftime('%d/%m/%Y')
        gio_cn = datetime.now(VN_TZ).strftime('%H:%M:%S')

        sheet_dp = client.open_by_key(SHEET_DU_PHONG_ID).get_worksheet(0)
        header = [
            "Ngày", "Tên nhân viên", "Lý do nghỉ", "Chi tiết", "Số ngày tính",
            "Số ngày phép cộng dồn", "Phạt vi phạm", "Ngày cập nhật",
            "Giờ cập nhật", "Người cập nhật"
        ]
        current_header = sheet_dp.get('A1:J1')
        current_header = current_header[0] if current_header else []
        if not any(str(v).strip() for v in current_header):
            gspread_update_range(sheet_dp, 'A1:J1', [header], value_input_option='USER_ENTERED')

        live_backup = _live_sheet_to_leave_df(sheet_dp)
        combined_live = combine_leave_sources_for_daily_stats(df_main_source, live_backup)

        # Chỉ chặn trùng CÙNG lý do. Cùng ngày + cùng nhân viên nhưng lý do khác
        # (kể cả các lý do có Phạt vi phạm > 0) vẫn được phép ghi thành các dòng riêng.
        if _leave_exists_in_sources(combined_live, ngay, nv, loai_nghi):
            return False, f"Nhân viên '{nv}' đã có đúng lý do '{clean_leave_reason_display(loai_nghi)}' trong ngày {ngay}. Lý do khác vẫn được phép ghi riêng."

        save_detail = str(chi_tiet).strip()
        save_penalty = float(phat_vi_pham) if phat_vi_pham is not None else 0.0
        ordinal_note = ""
        progressive_reason = get_progressive_penalty_reason(loai_nghi)
        if progressive_reason:
            ordinal, extra_penalty = _progressive_ordinal_and_bonus(combined_live, ngay, loai_nghi)
            ordinal_note = f"Người Thứ {ordinal} {progressive_reason.lower()}"
            save_detail = f"{ordinal_note} | {save_detail}" if save_detail else ordinal_note
            save_penalty += extra_penalty

        row_values = [
            str(ngay),
            str(nv),
            clean_leave_reason_display(loai_nghi),
            save_detail,
            float(so_ngay) if so_ngay is not None else 0.0,
            float(so_ngay_cong_don),
            save_penalty,
            str(ngay_cn),
            str(gio_cn),
            str(updated_by),
        ]

        target_row = _next_data_row_a_to_j(sheet_dp)
        gspread_update_range(sheet_dp, f"A{target_row}:J{target_row}", [row_values], value_input_option='USER_ENTERED')

        _clear_dynamic_data_caches()
        if ordinal_note:
            extra = max(0, save_penalty - float(phat_vi_pham or 0))
            return True, f"{ordinal_note}. Phạt lũy tiến cộng thêm {extra:,.0f} VNĐ; tổng phạt {save_penalty:,.0f} VNĐ."
        return True, "Đã ghi nhận lịch nghỉ thành công vào Google Sheet dự phòng!"
    except Exception as e:
        return False, f"Lỗi ghi dữ liệu: {e}"

def delete_backup_row(row_index_1_based, updated_by=None):
    """Xóa 1 dòng ở Sheet dự phòng và tự xếp lại Người Thứ X/phạt lũy tiến nếu cần."""
    try:
        client = get_gspread_client()
        sheet = client.open_by_key(SHEET_DU_PHONG_ID).get_worksheet(0)
        actor = str(updated_by or st.session_state.get("current_user", "Hệ thống"))

        # Đọc bản ghi trước khi xóa để biết nhóm nào cần xếp lại.
        row_values = sheet.get(f'A{row_index_1_based}:J{row_index_1_based}')
        deleted_row = None
        affected_groups = set()
        if row_values and row_values[0]:
            expected = [
                "Ngày", "Tên nhân viên", "Lý do nghỉ", "Chi tiết", "Số ngày tính",
                "Số ngày phép cộng dồn", "Phạt vi phạm", "Ngày cập nhật",
                "Giờ cập nhật", "Người cập nhật"
            ]
            vals = list(row_values[0][:10]) + [""] * max(0, 10 - len(row_values[0]))
            deleted_row = dict(zip(expected, vals[:10]))
            deleted_row['__source_sheet_id'] = SHEET_DU_PHONG_ID
            deleted_row['__source_row'] = int(row_index_1_based)
            group_key = _progressive_group_key(deleted_row)
            if group_key:
                affected_groups.add(group_key)

        sheet.delete_rows(row_index_1_based)

        rebalanced = 0
        if affected_groups:
            rebalanced = rebalance_progressive_penalty_groups(client, affected_groups, actor)

        _clear_dynamic_data_caches()
        if rebalanced:
            return True, f"Đã xóa lịch nghỉ và tự xếp lại thứ tự/phạt cho {rebalanced} bản ghi còn lại."
        return True, "Đã xóa lịch nghỉ thành công!"
    except Exception as e:
        return False, f"Lỗi xóa dòng: {e}"


def _find_schedule_row_index(sheet, original_row):
    """Tìm dòng Google Sheet theo Ngày + Nhân viên + Lý do (bộ ba đang được hệ thống chặn trùng)."""
    values = _gs_call_with_backoff(sheet.get_all_values)
    if len(values) < 2:
        return None
    headers = values[0]
    target_key = schedule_key(original_row)
    for idx, vals in enumerate(values[1:], start=2):
        row_dict = {headers[i]: vals[i] if i < len(vals) else '' for i in range(len(headers))}
        if schedule_key(row_dict) == target_key:
            return idx
    return None


def _parse_leave_number(value, default=0.0, money=False):
    """Chuẩn hóa số lấy từ sheet LoaiNghi, hỗ trợ dấu chấm/phẩy và ký hiệu tiền."""
    try:
        if value is None or pd.isna(value):
            return float(default)
        s = str(value).strip()
        if s.lower() in ["", "-", "nan", "none", "nat"]:
            return float(default)
        if money:
            s = (s.replace('.', '').replace(',', '').replace(' ', '')
                   .replace('đ', '').replace('Đ', '').replace('VNĐ', '').replace('VND', ''))
        else:
            s = s.replace(',', '.')
        return float(s)
    except Exception:
        return float(default)


def build_leave_reason_catalog(source_df=None):
    """
    Tạo danh mục Lý do nghỉ -> Số ngày tính / Phạt vi phạm từ sheet LoaiNghi.
    Giữ tên hiển thị sạch, không có tiền tố biểu tượng đỏ.
    """
    source = source_df if source_df is not None else globals().get('df_loai_nghi', pd.DataFrame())
    catalog = {}
    if source is None or source.empty:
        return catalog

    for _, row in source.iterrows():
        vals = row.tolist()
        name = str(vals[1]).strip() if len(vals) > 1 else ""
        if not name or name.lower() in ["nan", "none"]:
            name = str(row.get('Lý do nghỉ', row.get('Loại nghỉ', ''))).strip()
        name = clean_leave_reason_display(name)
        if not name or name.lower() in ["nan", "none", "loại nghỉ", "lý do nghỉ"]:
            continue

        days = _parse_leave_number(vals[4] if len(vals) > 4 else 0, 0.0, money=False)
        penalty = _parse_leave_number(vals[5] if len(vals) > 5 else 0, 0.0, money=True)
        catalog[normalize_leave_reason(name)] = {
            'name': name,
            'days': float(days),
            'penalty': float(penalty),
        }
    return catalog


# ==========================================================
# V84 - AUTO UPDATE VI PHẠM: BẢNG TOUR + TIMESOFT
# ==========================================================
def _auto_penalty_catalog_item(reason_name, catalog=None):
    catalog = catalog if catalog is not None else build_leave_reason_catalog(globals().get('df_loai_nghi', pd.DataFrame()))
    exact = catalog.get(normalize_leave_reason(reason_name))
    if exact:
        return exact
    wanted = normalize_login_name(reason_name)
    for item in catalog.values():
        if normalize_login_name(item.get('name', '')) == wanted:
            return item
    return None

def _canonical_system_employee_name(raw_name):
    """Đối chiếu tên nguồn ngoài với Tên nhân viên hệ thống; bỏ dấu * ở cuối khi so sánh."""
    cleaned = clean_employee_match_name(raw_name)
    if not cleaned:
        return ""
    creds = globals().get('df_credentials', pd.DataFrame())
    if isinstance(creds, pd.DataFrame) and not creds.empty and 'Tên nhân viên' in creds.columns:
        target = normalize_employee_match_name(cleaned)
        for name in creds['Tên nhân viên'].dropna().astype(str).tolist():
            if normalize_employee_match_name(name) == target:
                # Luôn ghi tên chuẩn đang lưu trong hệ thống, không ghi dấu * từ Bảng tour.
                return str(name).strip()
        # Có danh sách nhân sự nhưng không khớp -> không tự tạo bản ghi phạt sai tên.
        return ""
    # Fallback chỉ dùng khi bảng hồ sơ chưa tải được.
    return cleaned

def _tour_late_minutes(value):
    """Đọc cột Vào trễ thành số phút; hỗ trợ số phút, timedelta, time và chuỗi HH:MM[:SS]."""
    try:
        if value is None or pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, timedelta):
        return max(0.0, value.total_seconds() / 60.0)
    # pandas Timedelta
    try:
        if isinstance(value, pd.Timedelta):
            return max(0.0, value.total_seconds() / 60.0)
    except Exception:
        pass
    # datetime.time hoặc datetime/datetime-like.
    if hasattr(value, 'hour') and hasattr(value, 'minute') and not isinstance(value, numbers.Number):
        try:
            return max(0.0, float(value.hour * 60 + value.minute + getattr(value, 'second', 0) / 60.0))
        except Exception:
            pass
    text = str(value).strip()
    if not text or text.casefold() in {'nan', 'none', 'nat', '<na>'}:
        return None
    # Nếu là số thuần thì Bảng tour đang lưu trực tiếp số phút.
    try:
        return max(0.0, float(text.replace(',', '.')))
    except Exception:
        pass
    m = re.fullmatch(r'\s*(\d{1,3}):(\d{1,2})(?::(\d{1,2}))?\s*', text)
    if m:
        h, mi, sec = int(m.group(1)), int(m.group(2)), int(m.group(3) or 0)
        return max(0.0, h * 60 + mi + sec / 60.0)
    # Chuỗi dạng "5 phút", "10 min"...
    m = re.search(r'(-?\d+(?:[\.,]\d+)?)', text)
    if m:
        try:
            return max(0.0, float(m.group(1).replace(',', '.')))
        except Exception:
            pass
    return None

def _outside_late_reason_for_minutes(minutes, catalog=None):
    """Chọn đúng bậc Ra ngoài vào muộn theo số phút và ưu tiên tên đang cấu hình trong LoaiNghi."""
    try:
        m = float(minutes)
    except Exception:
        return None
    if m < AUTO_PENALTY_MINUTES:
        return None
    if m < 30:
        candidates = ["Ra ngoài vào muộn dưới 30 phút"]
    elif m < 60:
        candidates = ["Ra ngoài vào muộn dưới 60 phút"]
    elif m < 120:
        candidates = ["Ra ngoài vào muộn dưới 120 phút"]
    else:
        # Nếu danh mục có bậc >=120 thì dùng đúng bậc đó; nếu chưa có, dùng bậc cao nhất hiện tại.
        candidates = [
            "Ra ngoài vào muộn từ 120 phút trở lên",
            "Ra ngoài vào muộn trên 120 phút",
            "Ra ngoài vào muộn dưới 120 phút",
        ]
    for reason in candidates:
        item = _auto_penalty_catalog_item(reason, catalog)
        if item:
            return item.get('name', reason)
    return None

def _auto_result(source):
    return {
        "source": source, "added": 0, "skipped": 0, "errors": 0,
        "eligible": 0, "messages": [], "paused": False, "unmatched": []
    }

CA1_CLEANING_REASON = "KHÔNG dọn vệ sinh ca 1"
CA1_LATE_REASON_KEYS = {
    normalize_login_name("Đi trễ nhỏ hơn hoặc bằng 30 phút"),
    normalize_login_name("Đi trễ nhỏ hơn hoặc bằng 60 phút"),
    normalize_login_name("Đi trễ lớn 60 phút và nhỏ hơn hoặc bằng 120 phút"),
}
CA1_SUPPORT_REASON_KEYS = {
    normalize_login_name("Hỗ trợ Ca 1 đi trễ 2 tiếng"),
    normalize_login_name("Hỗ trợ Ca 1 đi trễ 3 tiếng"),
    normalize_login_name("Hỗ trợ Ca 2 đi trễ 1 tiếng"),
}


def _shift_base_number(value):
    key = normalize_login_name(value)
    if "ca 1" in key:
        return 1
    if "ca 2" in key:
        return 2
    return None


def _credential_effective_shift_for_week(cred_row, target_date=None):
    """Xác định ca đang làm của cả tuần dựa trên cấu hình Ca/Ngày bắt đầu/Chu kỳ."""
    target_date = target_date or get_vn_today()
    monday = target_date - timedelta(days=target_date.weekday())
    base = _shift_base_number(cred_row.get("Ca làm việc", ""))
    if base not in {1, 2}:
        return None

    shift_key = normalize_login_name(cred_row.get("Ca làm việc", ""))
    cycle_key = normalize_login_name(cred_row.get("Chu kỳ", ""))
    if "co dinh" in shift_key or "co dinh" in cycle_key:
        return base

    raw_start = cred_row.get("Ngày bắt đầu ca", "")
    try:
        anchor_dt = pd.to_datetime(raw_start, dayfirst=True, errors="coerce")
        anchor = anchor_dt.date() if pd.notna(anchor_dt) else None
    except Exception:
        anchor = None
    if not anchor or monday <= anchor:
        return base

    if "14 ngay" in cycle_key or "luan phien" in cycle_key:
        periods = max(0, (monday - anchor).days // 14)
    elif "thang" in cycle_key:
        periods = max(0, (monday.year - anchor.year) * 12 + monday.month - anchor.month)
    else:
        return base
    return base if periods % 2 == 0 else (2 if base == 1 else 1)


def auto_update_ca1_cleaning_from_leave_data(actor="AUTO UPDATE - CA1"):
    """Auto ghi 'KHÔNG dọn vệ sinh ca 1' cho dữ liệu hôm nay theo quy tắc V84.7.

    Điều kiện đồng thời:
    1) Role nhanvien và ca hiệu lực của tuần hiện tại là Ca 1.
    2) Hôm nay có một trong 3 loại Đi trễ 30/60/120 đã chỉ định.
    3) Hôm nay KHÔNG có một trong 3 loại Hỗ trợ Ca đi trễ đã chỉ định.
    """
    result = _auto_result("KHÔNG dọn vệ sinh ca 1")
    cfg = load_auto_penalty_config()
    if cfg.get("paused"):
        result["paused"] = True
        result["messages"].append("Auto Update đang tạm dừng bởi Admin.")
        return result

    try:
        primary = _load_backup_sheet_data_from_sheets()
        secondary = _load_secondary_leave_sheet_data_from_sheets()
        leave_df = pd.concat([primary, secondary], ignore_index=True, sort=False)
    except Exception as exc:
        result["errors"] += 1
        result["messages"].append(f"Không đọc được lịch nghỉ LIVE: {exc}")
        return result

    if leave_df.empty:
        return result

    today = get_vn_today()
    day_key = today.strftime("%d/%m/%Y")
    d = leave_df.copy()
    d["__date"] = pd.to_datetime(d.get("Ngày"), dayfirst=True, errors="coerce").dt.date
    d = d[d["__date"].eq(today)].copy()
    if d.empty:
        return result
    d["__emp"] = d.get("Tên nhân viên", "").astype(str).apply(normalize_login_name)
    d["__reason"] = d.get("Lý do nghỉ", "").astype(str).apply(normalize_login_name)

    late_by_emp = {}
    support_by_emp = {}
    for _, r in d.iterrows():
        ekey = str(r.get("__emp", "")).strip()
        rkey = str(r.get("__reason", "")).strip()
        reason_display = str(r.get("Lý do nghỉ", "") or "").strip()
        if not ekey:
            continue
        if rkey in CA1_LATE_REASON_KEYS:
            late_by_emp.setdefault(ekey, []).append(reason_display)
        if rkey in CA1_SUPPORT_REASON_KEYS:
            support_by_emp.setdefault(ekey, []).append(reason_display)

    if not late_by_emp:
        return result

    try:
        creds = load_credentials_fresh()
    except Exception:
        creds = globals().get("df_credentials", pd.DataFrame())
    if creds is None or creds.empty:
        result["errors"] += 1
        result["messages"].append("Không đọc được cấu hình ca nhân viên.")
        return result

    catalog = build_leave_reason_catalog(globals().get("df_loai_nghi", pd.DataFrame()))
    reason_item = _auto_penalty_catalog_item(CA1_CLEANING_REASON, catalog)
    if not reason_item:
        result["errors"] += 1
        result["messages"].append(f"Sheet LoaiNghi chưa có '{CA1_CLEANING_REASON}'.")
        return result

    for _, cred in creds.iterrows():
        if str(cred.get("Phân quyền", "")).strip().lower() != "nhanvien":
            continue
        employee = str(cred.get("Tên nhân viên", "") or "").strip()
        ekey = normalize_login_name(employee)
        if ekey not in late_by_emp:
            continue
        if _credential_effective_shift_for_week(cred, today) != 1:
            result["skipped"] += 1
            continue

        result["eligible"] += 1
        if support_by_emp.get(ekey):
            result["skipped"] += 1
            result["messages"].append(
                f"{employee}: bỏ qua vì có {support_by_emp[ekey][0]}."
            )
            continue

        late_reason = late_by_emp[ekey][0]
        week_start = today - timedelta(days=today.weekday())
        week_end = week_start + timedelta(days=6)
        detail = (
            f"Auto Update Ca 1 · tuần {week_start.strftime('%d/%m')}–{week_end.strftime('%d/%m/%Y')} · "
            f"phát hiện '{late_reason}' · không có Hỗ trợ Ca đi trễ"
        )
        ok, msg = save_lich_nghi_to_backup_sheet(
            day_key,
            employee,
            reason_item.get("name", CA1_CLEANING_REASON),
            detail,
            float(reason_item.get("days", 0) or 0),
            0.0,
            float(reason_item.get("penalty", 0) or 0),
            actor,
            df_main_source=globals().get("df_lich", pd.DataFrame()),
        )
        if ok:
            result["added"] += 1
        elif "không được đăng ký trùng" in str(msg).lower() or "đã có loại nghỉ" in str(msg).lower():
            result["skipped"] += 1
        else:
            result["errors"] += 1
            result["messages"].append(f"{employee}: {msg}")
    return result


def auto_update_outside_late_from_tour(df_tour, actor="AUTO UPDATE - BẢNG TOUR"):
    """
    Tự ghi vi phạm Ra ngoài vào muộn từ Bảng tour.
    Chỉ xử lý khi cột `Vào trễ` >= 5 phút. `Cẩm Nhung *` được đối chiếu như `Cẩm Nhung`.
    """
    result = _auto_result("Bảng tour")
    cfg = load_auto_penalty_config()
    if cfg.get('paused'):
        result['paused'] = True
        result['messages'].append('Auto Update đang tạm dừng bởi Admin.')
        return result
    if not isinstance(df_tour, pd.DataFrame) or df_tour.empty:
        return result

    late_col = _find_tour_col(df_tour, "Vào trễ")
    name_col = _find_tour_col(df_tour, "Tên nhân viên") or _find_tour_col(df_tour, "Tên Nhân Viên")
    out_col = _find_tour_col(df_tour, "Giờ ra")
    in_col = _find_tour_col(df_tour, "Giờ vào")
    if late_col is None or name_col is None:
        result['errors'] += 1
        result['messages'].append("Bảng tour chưa có cột 'Vào trễ' hoặc 'Tên nhân viên'.")
        return result

    catalog = build_leave_reason_catalog(globals().get('df_loai_nghi', pd.DataFrame()))
    threshold = max(AUTO_PENALTY_MINUTES, int(cfg.get('threshold_minutes', AUTO_PENALTY_MINUTES) or AUTO_PENALTY_MINUTES))
    today = get_vn_today()
    main_source = globals().get('df_lich', pd.DataFrame())

    for _, row in df_tour.iterrows():
        minutes = _tour_late_minutes(row.get(late_col, ""))
        if minutes is None or float(minutes) < threshold:
            continue
        result['eligible'] += 1
        raw_name = row.get(name_col, "")
        employee = _canonical_system_employee_name(raw_name)
        if not employee:
            result['skipped'] += 1
            result['unmatched'].append(str(raw_name or '(không có tên)'))
            continue
        reason = _outside_late_reason_for_minutes(minutes, catalog)
        if not reason:
            result['errors'] += 1
            result['messages'].append(
                f"{employee}: chưa tìm thấy loại 'Ra ngoài vào muộn' phù hợp trong sheet LoaiNghi."
            )
            continue
        item = _auto_penalty_catalog_item(reason, catalog) or {}
        base_penalty = float(item.get('penalty', 0) or 0)
        days = float(item.get('days', 0) or 0)
        details = [f"Auto Update Bảng tour · vào muộn {int(round(float(minutes)))} phút"]
        if out_col is not None and str(row.get(out_col, '')).strip():
            details.append(f"Giờ ra {str(row.get(out_col)).strip()}")
        if in_col is not None and str(row.get(in_col, '')).strip():
            details.append(f"Giờ vào {str(row.get(in_col)).strip()}")
        ok, msg = save_lich_nghi_to_backup_sheet(
            today.strftime('%d/%m/%Y'), employee, reason, " · ".join(details),
            days, 0.0, base_penalty, actor, df_main_source=main_source
        )
        if ok:
            result['added'] += 1
        elif 'không được đăng ký trùng' in str(msg).lower() or 'đã có loại nghỉ' in str(msg).lower():
            result['skipped'] += 1
        else:
            result['errors'] += 1
            result['messages'].append(f"{employee}: {msg}")
    return result

def _timesoft_row_value(row, candidates):
    for c in candidates:
        if c in row.index:
            val = row.get(c)
            if val is not None and str(val).strip().casefold() not in {"", "nan", "none", "nat"}:
                return val
    return ""

def _parse_minutes_late_from_timesoft_row(row):
    direct = _timesoft_row_value(row, [
        'TotalMinuteInGoLate', 'TotalMinuteGoLate', 'MinuteInGoLate', 'GoLateMinute', 'LateMinute'
    ])
    try:
        if str(direct).strip() != "":
            return max(0.0, float(str(direct).replace(',', '.').strip()))
    except Exception:
        pass

    # Fallback: tự so giờ check-in với giờ bắt đầu ca nếu TimeSoft không trả cột phút trễ.
    start = _timesoft_row_value(row, ['StartWorkTime', 'WorkTimeStart', 'ShiftStartTime'])
    checkin = _timesoft_row_value(row, ['MachineTimeCheckInStr', 'CheckInTimeStr', 'CheckInTime'])
    if not start or not checkin:
        return None
    def _time_minutes(v):
        text = str(v).strip()
        m = re.search(r'(\d{1,2}):(\d{2})(?::(\d{2}))?', text)
        if not m:
            return None
        h, mi = int(m.group(1)), int(m.group(2))
        return h * 60 + mi
    sm = _time_minutes(start)
    cm = _time_minutes(checkin)
    if sm is None or cm is None:
        return None
    diff = cm - sm
    if diff < -12 * 60:
        diff += 24 * 60
    return max(0.0, float(diff))

TIMESOFT_SUPPORT_LATE_ALLOWANCES = {
    normalize_login_name("Hỗ trợ Ca 1 đi trễ 2 tiếng"): 120,
    normalize_login_name("Hỗ trợ Ca 1 sau 0:0H đi trễ 3 tiếng"): 180,
    normalize_login_name("Hỗ trợ Ca 2 sau 0:0H đi trễ 1 tiếng"): 60,
}


def _timesoft_support_for_day(employee, target_date):
    """Đọc Hỗ trợ cùng Ngày + Nhân viên từ dữ liệu lịch nghỉ đã hợp nhất."""
    sources = []
    for name in ("df_lich", "df_leave_secondary", "df_backup"):
        value = globals().get(name)
        if isinstance(value, pd.DataFrame) and not value.empty:
            sources.append(value)
    if not sources:
        return [], 0, ""

    data = pd.concat(sources, ignore_index=True, sort=False)
    if "Ngày" not in data.columns or "Tên nhân viên" not in data.columns or "Lý do nghỉ" not in data.columns:
        return [], 0, ""

    td = pd.to_datetime(target_date, errors="coerce", dayfirst=True)
    if pd.isna(td):
        return [], 0, ""
    target = td.date()

    dates = pd.to_datetime(data["Ngày"], errors="coerce", dayfirst=True).dt.date
    emp_keys = data["Tên nhân viên"].astype(str).apply(normalize_login_name)
    target_emp = normalize_login_name(employee)
    reasons = data.loc[(dates == target) & (emp_keys == target_emp), "Lý do nghỉ"].astype(str).tolist()
    supports = [r.strip() for r in reasons if "ho tro" in normalize_login_name(r)]

    if not supports:
        return [], 0, ""

    matched = []
    unknown = []
    for reason in supports:
        allowance = TIMESOFT_SUPPORT_LATE_ALLOWANCES.get(normalize_login_name(reason))
        if allowance is None:
            unknown.append(reason)
        else:
            matched.append((allowance, reason))

    # Hỗ trợ khác: giữ hành vi cũ, bỏ qua Auto phạt.
    if unknown:
        return supports, None, unknown[0]
    allowance, reason = max(matched, key=lambda x: x[0])
    return supports, int(allowance), reason


def _timesoft_late_vs_shift(row):
    """Ưu tiên check-in - giờ bắt đầu ca, fallback về parser hiện có."""
    start = _timesoft_row_value(row, ['StartWorkTime', 'WorkTimeStart', 'ShiftStartTime'])
    checkin = _timesoft_row_value(row, ['MachineTimeCheckInStr', 'CheckInTimeStr', 'CheckInTime'])

    def _tm(v):
        m = re.search(r'(\d{1,2}):(\d{2})(?::(\d{2}))?', str(v or '').strip())
        if not m:
            return None
        return int(m.group(1)) * 60 + int(m.group(2)) + int(m.group(3) or 0) / 60.0

    sm, cm = _tm(start), _tm(checkin)
    if sm is not None and cm is not None:
        diff = cm - sm
        if diff < -12 * 60:
            diff += 24 * 60
        return max(0.0, float(diff))
    return _parse_minutes_late_from_timesoft_row(row)


def auto_update_checkin_late_from_timesoft(checkin_df, actor="AUTO UPDATE - TIMESOFT"):
    """Tự ghi Đi trễ không phép khi check-in muộn hơn ca quy định từ 5 phút trở lên."""
    result = _auto_result("TimeSoft")
    cfg = load_auto_penalty_config()
    if cfg.get('paused'):
        result['paused'] = True
        result['messages'].append('Auto Update đang tạm dừng bởi Admin.')
        return result
    if not isinstance(checkin_df, pd.DataFrame) or checkin_df.empty:
        return result

    catalog = build_leave_reason_catalog(globals().get('df_loai_nghi', pd.DataFrame()))
    reason_item = _auto_penalty_catalog_item('Đi trễ không phép', catalog)
    if not reason_item:
        result['errors'] += 1
        result['messages'].append("Sheet LoaiNghi chưa có 'Đi trễ không phép', nên hệ thống không tự ghi phạt.")
        return result
    reason = reason_item.get('name', 'Đi trễ không phép')
    base_penalty = float(reason_item.get('penalty', 0) or 0)
    days = float(reason_item.get('days', 0) or 0)
    threshold = max(AUTO_PENALTY_MINUTES, int(cfg.get('threshold_minutes', AUTO_PENALTY_MINUTES) or AUTO_PENALTY_MINUTES))
    main_source = globals().get('df_lich', pd.DataFrame())

    for _, row in checkin_df.iterrows():
        minutes = _timesoft_late_vs_shift(row)
        if minutes is None or float(minutes) < threshold:
            continue
        result['eligible'] += 1
        raw_name = _timesoft_row_value(row, [
            'employeeInfo.Name', 'EmployeeName', 'employeeName', 'Name', 'FullName'
        ])
        employee = _canonical_system_employee_name(raw_name)
        if not employee:
            result['skipped'] += 1
            result['unmatched'].append(str(raw_name or '(không có tên)'))
            continue
        raw_date = _timesoft_row_value(row, ['WorkDateStr', 'WorkDate', 'CreateDateStr', 'CreateDate'])
        parsed_date = _parse_vn_date(raw_date)
        if not parsed_date:
            try:
                parsed_date = pd.to_datetime(raw_date, dayfirst=True, errors='coerce').date()
            except Exception:
                parsed_date = None
        if not parsed_date or pd.isna(parsed_date):
            result['skipped'] += 1
            result['messages'].append(f"{employee}: không xác định được ngày chấm công.")
            continue

        support_reasons, allowance, support_reason = _timesoft_support_for_day(employee, parsed_date)
        if support_reasons:
            if allowance is None:
                result['skipped'] += 1
                result['messages'].append(f"{employee}: bỏ qua Auto phạt vì có Hỗ trợ '{support_reason}'.")
                continue
            if float(minutes) <= float(allowance):
                result['skipped'] += 1
                result['messages'].append(
                    f"{employee}: Hỗ trợ cho phép trễ {allowance} phút; thực tế {int(round(float(minutes)))} phút → không phạt."
                )
                continue

        shift_start = _timesoft_row_value(row, ['StartWorkTime', 'WorkTimeStart', 'ShiftStartTime'])
        checkin_time = _timesoft_row_value(row, ['MachineTimeCheckInStr', 'CheckInTimeStr', 'CheckInTime'])
        detail = f"Auto Update TimeSoft · check-in muộn {int(round(float(minutes)))} phút"
        if support_reasons and allowance is not None and float(minutes) > float(allowance):
            detail += f" · Hỗ trợ cho phép {allowance} phút nhưng đã vượt {int(round(float(minutes) - float(allowance)))} phút"
        if shift_start:
            detail += f" · Ca bắt đầu {shift_start}"
        if checkin_time:
            detail += f" · Check-in {checkin_time}"
        ok, msg = save_lich_nghi_to_backup_sheet(
            parsed_date.strftime('%d/%m/%Y'), employee, reason, detail,
            days, 0.0, base_penalty, actor, df_main_source=main_source
        )
        if ok:
            result['added'] += 1
        elif 'không được đăng ký trùng' in str(msg).lower() or 'đã có loại nghỉ' in str(msg).lower():
            result['skipped'] += 1
        else:
            result['errors'] += 1
            result['messages'].append(f"{employee}: {msg}")
    return result

def run_auto_penalty_now(tour_df=None, checkin_df=None, actor="AUTO UPDATE"):
    """Chạy cả 2 nguồn; dùng cho trang Admin và các lần đồng bộ thủ công."""
    if tour_df is None:
        try:
            tour_df, _ = load_bang_tour_input()
        except Exception:
            tour_df = pd.DataFrame()
    r_tour = auto_update_outside_late_from_tour(tour_df, actor=f"{actor} - BẢNG TOUR")
    r_ts = auto_update_checkin_late_from_timesoft(checkin_df, actor=f"{actor} - TIMESOFT") if isinstance(checkin_df, pd.DataFrame) else _auto_result("TimeSoft")
    r_ca1 = auto_update_ca1_cleaning_from_leave_data(actor=f"{actor} - CA1")
    return {"tour": r_tour, "timesoft": r_ts, "ca1": r_ca1}


def _vn_weekday_label(value):
    """Trả về Thứ 2..Thứ 7 / Chủ nhật từ giá trị ngày; không làm thay đổi dữ liệu nguồn."""
    d = _parse_vn_date(value)
    if d is None:
        return ""
    labels = {
        0: "Thứ 2",
        1: "Thứ 3",
        2: "Thứ 4",
        3: "Thứ 5",
        4: "Thứ 6",
        5: "Thứ 7",
        6: "Chủ nhật",
    }
    return labels.get(d.weekday(), "")


def add_weekday_column(df):
    """Thêm cột hiển thị `Thứ ngày` ngay bên phải cột `Ngày`."""
    if not isinstance(df, pd.DataFrame):
        return df
    d = df.copy()
    if "Ngày" not in d.columns:
        return d
    if "Thứ ngày" in d.columns:
        d = d.drop(columns=["Thứ ngày"])
    pos = d.columns.get_loc("Ngày") + 1
    d.insert(pos, "Thứ ngày", d["Ngày"].apply(_vn_weekday_label))
    return d


def _leave_reason_type_map(source_df=None):
    """
    Mapping chuẩn từ sheet LoaiNghi của file 1Kz0...:
    - Cột B = Lý do nghỉ
    - Cột C = Loại nghỉ

    Trả về dict theo khóa normalize_leave_reason(Lý do nghỉ) -> Loại nghỉ.
    """
    source = source_df if isinstance(source_df, pd.DataFrame) else globals().get('df_loai_nghi', pd.DataFrame())
    result = {}
    if source is None or not isinstance(source, pd.DataFrame) or source.empty:
        return result

    for _, row in source.iterrows():
        vals = row.tolist()

        # Ưu tiên tuyệt đối vị trí vật lý B/C như người dùng quy định.
        reason = str(vals[1]).strip() if len(vals) > 1 else ""
        leave_type = str(vals[2]).strip() if len(vals) > 2 else ""

        # Fallback theo tên header nếu file thay đổi cách load nhưng vẫn giữ ý nghĩa cột.
        if not reason or reason.casefold() in {"nan", "none"}:
            reason = str(row.get('Lý do nghỉ', '') or '').strip()
        if not leave_type or leave_type.casefold() in {"nan", "none"}:
            leave_type = str(row.get('Loại nghỉ', '') or '').strip()

        reason = clean_leave_reason_display(reason)
        leave_type = clean_leave_reason_display(leave_type)

        if not reason or reason.casefold() in {"nan", "none", "lý do nghỉ"}:
            continue
        if leave_type.casefold() in {"nan", "none", "loại nghỉ"}:
            leave_type = ""

        key = normalize_leave_reason(reason)
        if key and key not in result:
            result[key] = leave_type

    return result


def _leave_type_series_from_reason(df):
    """Trả về Loại nghỉ chuẩn theo mapping LoaiNghi: cột B (Lý do) -> cột C (Loại nghỉ)."""
    if not isinstance(df, pd.DataFrame) or df.empty:
        return pd.Series(dtype=str)
    d = df.copy()
    if 'Lý do nghỉ' not in d.columns:
        return pd.Series([''] * len(d), index=d.index, dtype=str)
    mapping = _leave_reason_type_map(globals().get('df_loai_nghi', pd.DataFrame()))
    return d['Lý do nghỉ'].astype(str).apply(
        lambda v: clean_leave_reason_display(mapping.get(normalize_leave_reason(v), ''))
    )


def add_source_leave_type_column(df):
    """
    V86.1: đồng nhất dữ liệu hiển thị lịch nghỉ:
    - `Lý do nghỉ` = giá trị Lý do đã lưu, đối chiếu danh mục cột B của sheet LoaiNghi.
    - `Loại nghỉ` = giá trị cột C tương ứng trong sheet LoaiNghi.
    - `Thứ ngày` đặt ngay sau `Ngày`.

    Không còn sao chép Lý do nghỉ sang Loại nghỉ.
    """
    if not isinstance(df, pd.DataFrame):
        return df

    d = add_weekday_column(df.copy())
    if 'Lý do nghỉ' not in d.columns:
        return d

    # Làm sạch Lý do nghỉ để khớp đúng với cột B của danh mục.
    d['Lý do nghỉ'] = d['Lý do nghỉ'].apply(clean_leave_reason_display)

    reason_type_map = _leave_reason_type_map(globals().get('df_loai_nghi', pd.DataFrame()))
    type_values = d['Lý do nghỉ'].apply(
        lambda v: reason_type_map.get(normalize_leave_reason(v), "")
    )

    if 'Loại nghỉ' in d.columns:
        d = d.drop(columns=['Loại nghỉ'])

    pos = d.columns.get_loc('Lý do nghỉ') + 1
    d.insert(pos, 'Loại nghỉ', type_values)
    return d


def get_leave_reason_options(source_df=None, extra_values=None):
    """Danh sách dropdown Lý do nghỉ luôn chứa ĐÚNG chuỗi hiện có của dữ liệu lịch sử.

    Streamlit SelectboxColumn không chỉ so khớp theo ý nghĩa mà yêu cầu chuỗi trong ô phải
    khớp CHÍNH XÁC với một option. Vì danh mục LoaiNghi có thể dùng kiểu chữ khác nhau
    (ví dụ ``Nghỉ KHÔNG phép``) trong khi dữ liệu lịch sử là ``Nghỉ không phép``, nếu chỉ
    khử trùng theo dạng chuẩn hóa thì ô sẽ bị hiển thị trắng dù dữ liệu vẫn còn nguyên.

    V69 giữ cả biến thể danh mục và biến thể chính xác đang tồn tại trong bảng. Nhờ vậy mọi
    loại có chữ KHÔNG phép (Nghỉ/Đi trễ/Về sớm, kể cả CUỐI TUẦN) luôn hiện đúng trong editor.
    """
    catalog = build_leave_reason_catalog(source_df)
    options = []
    seen_exact = set()

    def _append_exact(value):
        clean = clean_leave_reason_display(value)
        if not clean or clean.casefold() in ['nan', 'none', 'nat', 'loại nghỉ', 'lý do nghỉ']:
            return
        # Chỉ bỏ trùng khi chuỗi sau làm sạch thực sự giống HỆT nhau. Không dùng
        # normalize_leave_reason ở đây vì khác kiểu HOA/thường cũng làm SelectboxColumn
        # coi là hai giá trị khác nhau.
        if clean not in seen_exact:
            options.append(clean)
            seen_exact.add(clean)

    for item in [v['name'] for v in catalog.values()]:
        _append_exact(item)
    if extra_values is not None:
        for val in extra_values:
            _append_exact(val)
    return options



def _schedule_compare_value(column_name, value):
    """Chuẩn hóa giá trị để phát hiện dòng lịch nghỉ thật sự đã thay đổi."""
    if column_name == 'Ngày':
        return normalize_schedule_date(value)
    if column_name == 'Tên nhân viên':
        return normalize_login_name(value)
    if column_name == 'Lý do nghỉ':
        return normalize_leave_reason(clean_leave_reason_display(value))
    if column_name == 'Chi tiết':
        return str(value or '').strip()
    return str(value if value is not None else '').strip()


def get_changed_schedule_positions(original_df, edited_df, editable_columns=None):
    """Trả về các vị trí dòng đã đổi, không phụ thuộc checkbox Chọn."""
    if original_df is None or edited_df is None:
        return []
    editable_columns = editable_columns or ['Ngày', 'Tên nhân viên', 'Lý do nghỉ', 'Chi tiết']
    changed = []
    limit = min(len(original_df), len(edited_df))
    for pos in range(limit):
        old = original_df.iloc[pos]
        new = edited_df.iloc[pos]
        if any(
            c in new.index and _schedule_compare_value(c, old.get(c, '')) != _schedule_compare_value(c, new.get(c, ''))
            for c in editable_columns
        ):
            changed.append(pos)
    return changed


LEADER_POLICY_LEAVE_REASONS = {
    normalize_login_name("Leader nghỉ phép theo chính sách"),
    normalize_login_name("Leader về sớm về sớm theo chính sách"),
    normalize_login_name("Leader đi trễ sớm theo chính sách"),
}
BEREAVEMENT_LEAVE_REASON = "Nghỉ đám hiếu"


def is_leader_policy_leave_reason(value):
    return normalize_login_name(clean_leave_reason_display(value)) in LEADER_POLICY_LEAVE_REASONS


def is_bereavement_leave_reason(value):
    return normalize_login_name(clean_leave_reason_display(value)) == normalize_login_name(BEREAVEMENT_LEAVE_REASON)


def is_special_day_rule_exempt(role, reason):
    role = str(role or "").strip().lower()
    if role == "leader" and is_leader_policy_leave_reason(reason):
        return True
    if role in {"admin", "letan", "quanly"} and is_bereavement_leave_reason(reason):
        return True
    if is_video_leave_reason(reason):
        return True
    return False


def _filter_active_employees_for_leave_stats(df):
    """Chỉ nhân sự Đang làm việc được tính thống kê/hạn mức nghỉ."""
    if not isinstance(df, pd.DataFrame) or df.empty or "Tên nhân viên" not in df.columns:
        return df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    status_map = load_employment_status_map()
    if not status_map:
        return df.copy()
    d = df.copy()
    keys = d["Tên nhân viên"].astype(str).apply(normalize_login_name)
    statuses = keys.map(lambda k: status_map.get(k, EMPLOYMENT_STATUS_ACTIVE))
    return d[statuses.eq(EMPLOYMENT_STATUS_ACTIVE)].copy()


VIDEO_LEAVE_REASON = "Nghỉ phép quay video"


def is_video_leave_reason(value):
    """Loại nghỉ đặc biệt: không chịu hạn mức ngày/tháng/cuối tuần và không chiếm suất nghỉ của người khác."""
    return normalize_login_name(clean_leave_reason_display(value)) == normalize_login_name(VIDEO_LEAVE_REASON)


def is_employee_khong_phep_leave_reason(value):
    """Nhận diện Loại nghỉ KHÔNG phép cho quyền Sửa/Hủy của Nhân viên/Leader."""
    key = normalize_login_name(clean_leave_reason_display(value))
    return bool(key and "khong phep" in key)


def _leave_rows_counting_toward_quota(df):
    """Loại Nghỉ phép quay video khỏi mọi phép đếm hạn mức nhân sự."""
    if not isinstance(df, pd.DataFrame) or df.empty or "Lý do nghỉ" not in df.columns:
        return df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    return df[~df["Lý do nghỉ"].astype(str).apply(is_video_leave_reason)].copy()


def is_employee_co_phep_leave_reason(value):
    """Nhận diện nhóm Loại nghỉ CÓ phép theo cùng quy tắc đang dùng ở phần thống kê."""
    key = normalize_login_name(clean_leave_reason_display(value))
    if not key or "khong phep" in key or "nghi phat sinh" in key:
        return False
    excluded_keywords = [
        "di tre", "khong don ve sinh", "loi vi pham", "qua tour", "xuong phong",
        "ra som", "vao muon", "di tua", "ngung nhan", "ho tro ca"
    ]
    return not any(kw in key for kw in excluded_keywords)


def validate_employee_leave_change_permission(original_row, edited_row=None, current_user=None, today=None, action="sửa"):
    """
    Nhân viên/Leader:
    - chỉ thao tác lịch của chính mình;
    - CÓ phép: Sửa/Hủy trước ít nhất 3 ngày;
    - KHÔNG phép: được Sửa/Hủy, không áp dụng mốc 3 ngày nhưng chỉ với ngày tương lai (> hôm nay);
    - Nghỉ phép quay video: miễn mọi giới hạn thời gian/loại nghỉ, nhưng vẫn không được đổi sang người khác.
    """
    today = today or get_vn_today()
    actor = normalize_login_name(current_user or st.session_state.get("current_user", ""))
    owner = normalize_login_name(original_row.get("Tên nhân viên", ""))
    role_now = str(st.session_state.get("current_role", "") or "").strip().lower()
    if not actor or owner != actor:
        return False, "Nhân viên chỉ được thao tác lịch nghỉ của chính mình."

    old_reason = original_row.get("Lý do nghỉ", original_row.get("Loại nghỉ", ""))
    old_is_video = is_video_leave_reason(old_reason)
    old_is_khong_phep = is_employee_khong_phep_leave_reason(old_reason)
    old_is_co_phep = is_employee_co_phep_leave_reason(old_reason)
    old_is_leader_policy = role_now == "leader" and is_leader_policy_leave_reason(old_reason)

    old_dt = pd.to_datetime(original_row.get("Ngày"), errors="coerce", dayfirst=True)
    old_date = old_dt.date() if pd.notna(old_dt) else None

    # Nghỉ phép quay video và 3 lý do chính sách Leader: miễn giới hạn 3 ngày/ngày trong tuần.
    if old_is_video or old_is_leader_policy:
        if edited_row is not None:
            new_owner = normalize_login_name(edited_row.get("Tên nhân viên", original_row.get("Tên nhân viên", "")))
            if new_owner != actor:
                return False, "Nhân viên không được đổi lịch sang tên người khác."
            if old_is_leader_policy:
                new_reason = edited_row.get("Lý do nghỉ", edited_row.get("Loại nghỉ", old_reason))
                if not (is_leader_policy_leave_reason(new_reason) or is_video_leave_reason(new_reason)):
                    return False, "Lịch chính sách Leader chỉ được đổi sang lý do chính sách Leader khác hoặc Nghỉ phép quay video."
        return True, ""

    if old_is_khong_phep:
        # Không phép không chịu mốc 3 ngày, nhưng hôm nay và quá khứ đều bị khóa.
        if old_date is None or old_date <= today:
            return False, f"Nhân viên chỉ được {action} lịch Nghỉ KHÔNG phép ở ngày tương lai; không được sửa/hủy ngày hiện tại hoặc quá khứ."
    elif old_is_co_phep:
        min_date = today + timedelta(days=EMPLOYEE_LEAVE_CHANGE_NOTICE_DAYS)
        if old_date is None or old_date < min_date:
            return False, (
                f"Nhân viên chỉ được {action} lịch Nghỉ CÓ phép trước ít nhất {EMPLOYEE_LEAVE_CHANGE_NOTICE_DAYS} ngày "
                f"(ngày nghỉ từ {min_date.strftime('%d/%m/%Y')} trở đi)."
            )
    else:
        return False, f"Nhân viên chỉ được {action} lịch Nghỉ CÓ phép, Nghỉ KHÔNG phép hoặc '{VIDEO_LEAVE_REASON}' của chính mình."

    if edited_row is not None:
        new_owner = normalize_login_name(edited_row.get("Tên nhân viên", original_row.get("Tên nhân viên", "")))
        if new_owner != actor:
            return False, "Nhân viên không được đổi lịch sang tên người khác."

        new_reason = edited_row.get("Lý do nghỉ", edited_row.get("Loại nghỉ", old_reason))
        new_is_video = is_video_leave_reason(new_reason)
        new_is_khong_phep = is_employee_khong_phep_leave_reason(new_reason)
        new_is_co_phep = is_employee_co_phep_leave_reason(new_reason)
        new_dt = pd.to_datetime(edited_row.get("Ngày"), errors="coerce", dayfirst=True)
        new_date = new_dt.date() if pd.notna(new_dt) else None

        if new_is_video:
            return True, ""

        if old_is_khong_phep:
            if not new_is_khong_phep:
                return False, "Lịch Nghỉ KHÔNG phép chỉ được sửa sang một Lý do nghỉ KHÔNG phép khác hoặc Nghỉ phép quay video."
            if new_date is None or new_date <= today:
                return False, "Ngày sau khi sửa của Nghỉ KHÔNG phép phải là ngày tương lai; không được chọn hôm nay hoặc quá khứ."
        else:
            if not new_is_co_phep:
                return False, "Lịch Nghỉ CÓ phép chỉ được thay đổi giữa các Loại nghỉ CÓ phép hoặc Nghỉ phép quay video."
            min_date = today + timedelta(days=EMPLOYEE_LEAVE_CHANGE_NOTICE_DAYS)
            if new_date is None or new_date < min_date:
                return False, (
                    f"Ngày nghỉ sau khi sửa phải cách hiện tại ít nhất {EMPLOYEE_LEAVE_CHANGE_NOTICE_DAYS} ngày "
                    f"(từ {min_date.strftime('%d/%m/%Y')} trở đi)."
                )
    return True, ""


def validate_schedule_delete_permission(original_row, role, current_user=None, today=None):
    """Kiểm tra quyền hủy/xóa lịch theo vai trò. Admin luôn toàn quyền."""
    today = today or get_vn_today()
    role = str(role or "").strip().lower()
    if role == "admin":
        return True, ""
    if role in EMPLOYEE_LIKE_ROLES:
        return validate_employee_leave_change_permission(
            original_row, None, current_user=current_user, today=today, action="hủy"
        )
    if role in {"letan", "quanly"}:
        old_reason = original_row.get("Lý do nghỉ", original_row.get("Loại nghỉ", ""))
        if is_video_leave_reason(old_reason) or is_bereavement_leave_reason(old_reason):
            return True, ""
        dt = pd.to_datetime(original_row.get("Ngày"), errors="coerce", dayfirst=True)
        if pd.notna(dt) and dt.date() < today:
            return False, "Lễ tân/Quản lý không được xóa lịch trong quá khứ."
    return True, ""


def validate_schedule_edit_permission(original_row, edited_row, role, today=None, current_user=None):
    """Giữ đúng giới hạn sửa lịch của từng vai trò trước khi batch-save."""
    today = today or get_vn_today()
    old_dt = pd.to_datetime(original_row.get('Ngày'), errors='coerce', dayfirst=True)
    new_dt = pd.to_datetime(edited_row.get('Ngày'), errors='coerce', dayfirst=True)
    old_date = old_dt.date() if pd.notna(old_dt) else today
    new_date = new_dt.date() if pd.notna(new_dt) else None
    if new_date is None:
        return False, "Ngày nghỉ không hợp lệ."

    role = str(role or '').strip().lower()
    if role == "admin":
        return True, ""
    if role in EMPLOYEE_LIKE_ROLES:
        permitted, message = validate_employee_leave_change_permission(
            original_row, edited_row, current_user=current_user, today=today, action="thay đổi"
        )
        if not permitted:
            return False, message
        new_reason = edited_row.get("Lý do nghỉ", edited_row.get("Loại nghỉ", original_row.get("Lý do nghỉ", "")))
        role_now = str(role or "").strip().lower()
        if not (is_video_leave_reason(new_reason) or (role_now == "leader" and is_leader_policy_leave_reason(new_reason))):
            _, emp_max_date = employee_registration_window(today)
            if new_date > emp_max_date:
                return False, f"Nhân viên chỉ được sửa lịch đến hết {emp_max_date.strftime('%d/%m/%Y')}."
    elif role in {'letan', 'quanly'}:
        old_reason = original_row.get("Lý do nghỉ", original_row.get("Loại nghỉ", ""))
        new_reason = edited_row.get("Lý do nghỉ", edited_row.get("Loại nghỉ", old_reason))
        special_frontdesk = (
            is_video_leave_reason(old_reason) or is_video_leave_reason(new_reason)
            or is_bereavement_leave_reason(old_reason) or is_bereavement_leave_reason(new_reason)
        )
        if not special_frontdesk:
            if old_date < today or new_date < today:
                return False, "Lễ tân/Quản lý không được sửa lịch trong quá khứ."

    if not str(edited_row.get('Tên nhân viên', '')).strip():
        return False, "Tên nhân viên không được để trống."
    if not str(edited_row.get('Lý do nghỉ', '')).strip():
        return False, "Lý do nghỉ không được để trống."
    return True, ""

def _exclude_original_from_leave_df(df_sources, original_row):
    """Loại đúng bản ghi đang sửa ra khỏi tập dữ liệu dùng để tính lại."""
    if df_sources is None or df_sources.empty:
        return pd.DataFrame(columns=df_sources.columns if hasattr(df_sources, 'columns') else [])
    d = df_sources.copy()

    source_id = str(original_row.get('__source_sheet_id', '')).strip()
    source_row = original_row.get('__source_row', '')
    if source_id and source_row not in ['', None] and '__source_sheet_id' in d.columns and '__source_row' in d.columns:
        try:
            target_row = int(float(source_row))
            row_num = pd.to_numeric(d['__source_row'], errors='coerce')
            exact_mask = (d['__source_sheet_id'].astype(str).str.strip() == source_id) & (row_num == target_row)
            if exact_mask.any():
                return d.loc[~exact_mask].copy()
        except Exception:
            pass

    original_key = schedule_key(original_row)
    keep_mask = d.apply(lambda r: schedule_key(r) != original_key, axis=1)
    return d.loc[keep_mask].copy()


def _strip_generated_progressive_prefix(detail):
    """Bỏ tiền tố 'Người Thứ ...' do hệ thống từng tự thêm để tránh lặp khi sửa."""
    import re
    s = str(detail or '').strip()
    pattern = (
        r'^Người\s+Thứ\s+\d+\s+'
        r'(?:nghỉ\s+không\s+phép|đi\s+trễ\s+không\s+phép|về\s+sớm\s+không\s+phép|ra\s+sớm\s+không\s+phép)'
        r'\s*(?:\|\s*)?'
    )
    return re.sub(pattern, '', s, flags=re.IGNORECASE).strip()


def _get_existing_progressive_ordinal(original_row, all_leave_data=None):
    """
    Lấy đúng thứ tự Người Thứ X của bản ghi hiện hữu để GIỮ NGUYÊN khi sửa
    mà vẫn cùng ngày + cùng nhóm vi phạm lũy tiến.

    Ưu tiên:
    1) Đọc trực tiếp "Người Thứ X" đã lưu trong cột Chi tiết.
    2) Nếu dữ liệu cũ chưa có tiền tố này, suy ra vị trí từ chính bản ghi hiện hữu
       trong dữ liệu 2 nguồn (không coi thao tác sửa là một lượt vi phạm mới).
    """
    import re

    # 1. Bản ghi mới của hệ thống luôn có tiền tố này -> đây là nguồn chính xác nhất.
    detail = str(original_row.get('Chi tiết', '') or '')
    m = re.search(r'Người\s+Thứ\s+(\d+)', detail, flags=re.IGNORECASE)
    if m:
        try:
            return max(1, int(m.group(1)))
        except Exception:
            pass

    # 2. Tương thích dữ liệu cũ chưa ghi "Người Thứ X".
    canonical = get_progressive_penalty_reason(original_row.get('Lý do nghỉ', ''))
    ngay = normalize_schedule_date(original_row.get('Ngày', ''))
    if not canonical or not ngay or all_leave_data is None or getattr(all_leave_data, 'empty', True):
        return None

    d = all_leave_data.copy()
    if 'Ngày' not in d.columns or 'Lý do nghỉ' not in d.columns:
        return None

    d['_date_keep_ord'] = d['Ngày'].apply(normalize_schedule_date)
    d['_reason_keep_ord'] = d['Lý do nghỉ'].astype(str).apply(get_progressive_penalty_reason)
    same = d[(d['_date_keep_ord'] == ngay) & (d['_reason_keep_ord'] == canonical)].copy()
    if same.empty:
        return None

    # Cố gắng tìm đúng bản ghi theo source sheet + source row.
    src_id = str(original_row.get('__source_sheet_id', '') or '').strip()
    src_row = original_row.get('__source_row', None)
    if src_id and src_row not in [None, ''] and '__source_sheet_id' in same.columns and '__source_row' in same.columns:
        try:
            target_row = int(float(src_row))
            same['_src_row_num'] = pd.to_numeric(same['__source_row'], errors='coerce')
            same['_src_sheet_text'] = same['__source_sheet_id'].astype(str).str.strip()
            # Thứ tự lịch sử theo số dòng trong nguồn; nếu có nhiều nguồn thì giữ thứ tự ổn định
            # theo thứ tự hiện hữu trong DataFrame hợp nhất.
            same = same.reset_index(drop=False).rename(columns={'index': '_original_index'})
            match = same[(same['_src_sheet_text'] == src_id) & (same['_src_row_num'] == target_row)]
            if not match.empty:
                matched_original_index = match.iloc[0]['_original_index']
                # Dùng thứ tự xuất hiện trong tập cùng ngày/cùng loại.
                positions = list(same['_original_index'])
                return positions.index(matched_original_index) + 1
        except Exception:
            pass

    # Fallback theo khóa lịch nếu source metadata không còn.
    target_key = schedule_key(original_row)
    same = same.reset_index(drop=False).rename(columns={'index': '_original_index'})
    for idx, r in same.iterrows():
        if schedule_key(r) == target_key:
            return int(idx) + 1
    return None


def recalculate_schedule_fields(original_row, edited_row, updated_by, all_leave_data=None, source_df=None):
    """
    Tự động tính lại các cột phụ thuộc khi sửa lịch:
    - Số ngày tính: theo LoaiNghi
    - Số ngày phép cộng dồn: tổng tháng của nhân viên, loại bản ghi cũ rồi cộng giá trị mới
    - Phạt vi phạm: theo LoaiNghi + phạt lũy tiến nếu thuộc 3 nhóm vi phạm
    - Ngày/Giờ/Người cập nhật: theo thời điểm và tài khoản đang thao tác
    """
    catalog = build_leave_reason_catalog(source_df)
    result = edited_row.copy()

    ngay = normalize_schedule_date(result.get('Ngày', original_row.get('Ngày', '')))
    nv = str(result.get('Tên nhân viên', original_row.get('Tên nhân viên', ''))).strip()
    reason = clean_leave_reason_display(result.get('Lý do nghỉ', original_row.get('Lý do nghỉ', '')))
    key = normalize_leave_reason(reason)
    defaults = catalog.get(key)

    if defaults:
        reason = defaults['name']
        so_ngay = float(defaults['days'])
        base_penalty = float(defaults['penalty'])
    else:
        # Với dữ liệu lịch sử không còn trong LoaiNghi, giữ giá trị cũ để tránh làm mất dữ liệu.
        so_ngay = _parse_leave_number(original_row.get('Số ngày tính', result.get('Số ngày tính', 0)), 0.0)
        base_penalty = _parse_leave_number(original_row.get('Phạt vi phạm', result.get('Phạt vi phạm', 0)), 0.0, money=True)

    others = _exclude_original_from_leave_df(all_leave_data, original_row)

    # Tính số ngày phép cộng dồn trong cùng tháng/năm của đúng nhân viên.
    dt = pd.to_datetime(ngay, errors='coerce', dayfirst=True)
    accumulated = float(so_ngay)
    if pd.notna(dt) and others is not None and not others.empty:
        d = others.copy()
        d['_dt_calc'] = pd.to_datetime(d['Ngày'], errors='coerce', dayfirst=True)
        d['_days_calc'] = pd.to_numeric(d['Số ngày tính'], errors='coerce').fillna(0)
        same_emp = d['Tên nhân viên'].astype(str).apply(normalize_login_name).eq(normalize_login_name(nv))
        same_month = d['_dt_calc'].dt.month.eq(dt.month) & d['_dt_calc'].dt.year.eq(dt.year)
        accumulated = float(d.loc[same_emp & same_month, '_days_calc'].sum()) + float(so_ngay)

    # Phạt lũy tiến cho Nghỉ không phép / Đi trễ không phép / Về sớm không phép.
    final_penalty = float(base_penalty)
    detail = _strip_generated_progressive_prefix(result.get('Chi tiết', original_row.get('Chi tiết', '')))
    progressive_reason = get_progressive_penalty_reason(reason)
    if progressive_reason:
        # Nếu chỉ sửa nội dung nhưng vẫn cùng ngày + cùng nhóm vi phạm, giữ đúng
        # thứ tự Người Thứ đã ghi trước đó. Nếu đổi ngày/đổi loại thì tính lại thứ tự.
        ordinal = None
        original_reason = clean_leave_reason_display(original_row.get('Lý do nghỉ', ''))
        original_canonical = get_progressive_penalty_reason(original_reason)
        original_date = normalize_schedule_date(original_row.get('Ngày', ''))

        # QUY TẮC QUAN TRỌNG KHI SỬA:
        # Nếu vẫn cùng NGÀY + cùng NHÓM VI PHẠM thì đây vẫn là cùng một người/lượt cũ.
        # Tuyệt đối không đẩy Người Thứ 1 thành Người Thứ 2/3 chỉ vì bấm Sửa/Lưu lại.
        if original_canonical == progressive_reason and original_date == ngay:
            ordinal = _get_existing_progressive_ordinal(original_row, all_leave_data)

        # Chỉ cấp thứ tự mới khi thực sự đổi ngày hoặc đổi sang nhóm vi phạm khác,
        # hoặc dữ liệu lịch sử quá cũ không thể xác định thứ tự cũ.
        if ordinal is None:
            ordinal, _ = _progressive_ordinal_and_bonus(others, ngay, reason)
        extra_penalty = max(0, int(ordinal) - 2) * 100000
        final_penalty += float(extra_penalty)
        ordinal_note = f"Người Thứ {ordinal} {progressive_reason.lower()}"
        detail = f"{ordinal_note} | {detail}" if detail else ordinal_note

    now_vn = datetime.now(VN_TZ)
    result['Ngày'] = ngay
    result['Tên nhân viên'] = nv
    result['Lý do nghỉ'] = reason
    result['Chi tiết'] = detail
    result['Số ngày tính'] = float(so_ngay)
    result['Số ngày phép cộng dồn'] = float(accumulated)
    result['Phạt vi phạm'] = float(final_penalty)
    result['Ngày cập nhật'] = now_vn.strftime('%d/%m/%Y')
    result['Giờ cập nhật'] = now_vn.strftime('%H:%M:%S')
    result['Người cập nhật'] = str(updated_by)
    return result


def _load_live_two_leave_sheets(client):
    """Đọc trực tiếp hai Google Sheet lịch nghỉ để tính/sửa bằng dữ liệu mới nhất."""
    primary = client.open_by_key(SHEET_DU_PHONG_ID).get_worksheet(0)
    secondary = client.open_by_key(SHEET_LICH_NGHI_2_ID).get_worksheet(0)

    df_primary = _live_sheet_to_leave_df(primary)
    if not df_primary.empty:
        df_primary['__source_sheet_id'] = SHEET_DU_PHONG_ID
        # Gắn row sheet theo thứ tự A:J đã đọc; dùng key vẫn là lớp dự phòng chính nếu có dòng trống.
        df_primary['__source_row'] = range(2, len(df_primary) + 2)

    df_secondary = _live_sheet_to_leave_df(secondary)
    if not df_secondary.empty:
        df_secondary['__source_sheet_id'] = SHEET_LICH_NGHI_2_ID
        df_secondary['__source_row'] = range(2, len(df_secondary) + 2)

    return combine_leave_sources_for_daily_stats(df_secondary, df_primary)


def _read_leave_sheet_with_source(sheet, source_id):
    """Đọc A:J và giữ chính xác số dòng vật lý của Google Sheet để có thể cập nhật lại đúng dòng."""
    expected = [
        "Ngày", "Tên nhân viên", "Lý do nghỉ", "Chi tiết", "Số ngày tính",
        "Số ngày phép cộng dồn", "Phạt vi phạm", "Ngày cập nhật",
        "Giờ cập nhật", "Người cập nhật"
    ]
    try:
        values = _gs_call_with_backoff(sheet.get, 'A:J')
        rows = []
        if not values or len(values) < 2:
            return pd.DataFrame(columns=expected + ['__source_sheet_id', '__source_row'])
        for sheet_row, values_row in enumerate(values[1:], start=2):
            vals = list(values_row[:10]) + [""] * max(0, 10 - len(values_row))
            if not any(str(v).strip() for v in vals[:10]):
                continue
            item = dict(zip(expected, vals[:10]))
            item['__source_sheet_id'] = str(source_id)
            item['__source_row'] = int(sheet_row)
            rows.append(item)
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=expected + ['__source_sheet_id', '__source_row'])
    except Exception:
        return pd.DataFrame(columns=expected + ['__source_sheet_id', '__source_row'])


def _extract_progressive_ordinal(detail):
    """Lấy số X từ tiền tố 'Người Thứ X ...'."""
    try:
        m = re.search(r'Người\s+Thứ\s+(\d+)', str(detail or ''), flags=re.IGNORECASE)
        return max(1, int(m.group(1))) if m else None
    except Exception:
        return None


def _progressive_group_key(row):
    """Khóa nhóm phạt lũy tiến = (ngày, loại chuẩn)."""
    canonical = get_progressive_penalty_reason(row.get('Lý do nghỉ', ''))
    ngay = normalize_schedule_date(row.get('Ngày', ''))
    if not canonical or not ngay:
        return None
    return (str(ngay), str(canonical))


def _existing_base_penalty(row, catalog):
    """Lấy mức phạt gốc, tách phần lũy tiến khỏi tổng tiền hiện có khi cần."""
    reason = clean_leave_reason_display(row.get('Lý do nghỉ', ''))
    key = normalize_leave_reason(reason)
    if key in catalog:
        return float(catalog[key].get('penalty', 0) or 0)

    canonical = get_progressive_penalty_reason(reason)
    if canonical:
        for item in catalog.values():
            if get_progressive_penalty_reason(item.get('name', '')) == canonical:
                return float(item.get('penalty', 0) or 0)

    current_total = _parse_leave_number(row.get('Phạt vi phạm', 0), 0.0, money=True)
    old_ordinal = _extract_progressive_ordinal(row.get('Chi tiết', ''))
    old_extra = max(0, int(old_ordinal or 1) - 2) * 100000
    return max(0.0, float(current_total) - float(old_extra))


def rebalance_progressive_penalty_groups(client, affected_groups, updated_by):
    """
    Xếp lại toàn bộ Người Thứ X và mức phạt lũy tiến của các nhóm bị ảnh hưởng.

    Ví dụ sau khi Người Thứ 1 bị xóa/đổi sang Có phép:
      cũ 2 -> mới 1
      cũ 3 -> mới 2 (bỏ +100.000)
      cũ 4 -> mới 3 (chỉ còn +100.000)
    và ghi ngược vào đúng Google Sheet/dòng vật lý.

    Nếu cùng một lịch xuất hiện ở cả hai nguồn, lịch đó chỉ chiếm 1 vị trí thứ tự,
    nhưng mọi bản sao vật lý của nó đều được cập nhật để hai nguồn nhất quán.
    """
    clean_groups = set()
    for item in affected_groups or []:
        try:
            ngay, canonical = item
            if ngay and canonical:
                clean_groups.add((str(ngay), str(canonical)))
        except Exception:
            continue
    if not clean_groups:
        return 0

    # Đọc dữ liệu LIVE, giữ row vật lý của cả hai nguồn.
    sheet_map = {
        SHEET_DU_PHONG_ID: client.open_by_key(SHEET_DU_PHONG_ID).get_worksheet(0),
        SHEET_LICH_NGHI_2_ID: client.open_by_key(SHEET_LICH_NGHI_2_ID).get_worksheet(0),
    }
    frames = [
        _read_leave_sheet_with_source(sheet_map[SHEET_DU_PHONG_ID], SHEET_DU_PHONG_ID),
        _read_leave_sheet_with_source(sheet_map[SHEET_LICH_NGHI_2_ID], SHEET_LICH_NGHI_2_ID),
    ]
    frames = [f for f in frames if f is not None and not f.empty]
    if not frames:
        return 0
    raw_all = pd.concat(frames, ignore_index=True)
    raw_all['_reb_date'] = raw_all['Ngày'].apply(normalize_schedule_date)
    raw_all['_reb_reason'] = raw_all['Lý do nghỉ'].astype(str).apply(get_progressive_penalty_reason)

    catalog = build_leave_reason_catalog(globals().get('df_loai_nghi', pd.DataFrame()))
    now_vn = datetime.now(VN_TZ)
    actor = str(updated_by or 'Hệ thống')
    update_date = now_vn.strftime('%d/%m/%Y')
    update_time = now_vn.strftime('%H:%M:%S')
    updated_physical_rows = 0

    for ngay, canonical in sorted(clean_groups):
        group = raw_all[(raw_all['_reb_date'] == ngay) & (raw_all['_reb_reason'] == canonical)].copy()
        if group.empty:
            continue

        # Một lịch logic = Ngày + Nhân viên + Lý do. Gom mọi bản sao vật lý của lịch đó.
        logical = {}
        for _, r in group.iterrows():
            key = schedule_key(r)
            logical.setdefault(key, []).append(r.copy())

        ordered = []
        for logical_key, physical_rows in logical.items():
            # Ưu tiên thứ tự Người Thứ X đã lưu; đây là thứ tự lịch sử đáng tin cậy nhất.
            ordinals = [
                _extract_progressive_ordinal(r.get('Chi tiết', ''))
                for r in physical_rows
            ]
            ordinals = [x for x in ordinals if x is not None]
            old_ordinal = min(ordinals) if ordinals else None

            # Representative ưu tiên Sheet dự phòng, rồi theo row vật lý.
            representative = sorted(
                physical_rows,
                key=lambda r: (
                    0 if str(r.get('__source_sheet_id', '')) == SHEET_DU_PHONG_ID else 1,
                    int(float(r.get('__source_row', 10**9) or 10**9)),
                )
            )[0]
            fallback_row = int(float(representative.get('__source_row', 10**9) or 10**9))
            ordered.append((old_ordinal, fallback_row, logical_key, representative, physical_rows))

        # Những bản ghi có Người Thứ cũ được giữ đúng trật tự cũ; dữ liệu rất cũ không có tiền tố xếp sau theo row.
        ordered.sort(key=lambda x: (x[0] is None, x[0] if x[0] is not None else 10**9, x[1], normalize_login_name(x[3].get('Tên nhân viên', ''))))

        for new_ordinal, (_, _, logical_key, representative, physical_rows) in enumerate(ordered, start=1):
            base_penalty = _existing_base_penalty(representative, catalog)
            extra_penalty = max(0, new_ordinal - 2) * 100000
            new_penalty = float(base_penalty) + float(extra_penalty)
            prefix = f"Người Thứ {new_ordinal} {canonical.lower()}"

            for physical in physical_rows:
                source_id = str(physical.get('__source_sheet_id', '')).strip()
                try:
                    row_idx = int(float(physical.get('__source_row')))
                except Exception:
                    continue
                target = sheet_map.get(source_id)
                if target is None:
                    continue

                user_note = _strip_generated_progressive_prefix(physical.get('Chi tiết', ''))
                new_detail = f"{prefix} | {user_note}" if user_note else prefix

                # Chỉ thay phần cần thiết; giữ nguyên Số ngày tính và Số ngày phép cộng dồn hiện có.
                e_val = physical.get('Số ngày tính', '')
                f_val = physical.get('Số ngày phép cộng dồn', '')
                values_d_to_j = [[
                    new_detail,
                    e_val,
                    f_val,
                    new_penalty,
                    update_date,
                    update_time,
                    actor,
                ]]
                gspread_update_range(target, f'D{row_idx}:J{row_idx}', values_d_to_j, value_input_option='USER_ENTERED')
                updated_physical_rows += 1

    if updated_physical_rows:
        _clear_dynamic_data_caches()
    return updated_physical_rows


def update_schedule_record(original_row, edited_row, updated_by):
    """
    Sửa đúng dòng ở Google Sheet nguồn của bản ghi đang hiển thị.
    Sau khi sửa, tự động xếp lại Người Thứ X/phạt lũy tiến của nhóm cũ và nhóm mới.
    """
    try:
        client = get_gspread_client()
        if not client:
            return False, "Chưa cấu hình quyền kết nối Google Sheets."

        # Nhớ nhóm cũ trước khi thay đổi để sau đó có thể co lại thứ tự 2->1, 3->2...
        affected_groups = set()
        old_group = _progressive_group_key(original_row)
        if old_group:
            affected_groups.add(old_group)

        # Đọc LIVE cả hai nguồn để tránh dùng cache khi tính lại hoặc kiểm tra trùng.
        live_all = _load_live_two_leave_sheets(client)
        recalculated = recalculate_schedule_fields(
            original_row,
            edited_row,
            updated_by,
            all_leave_data=live_all,
            source_df=globals().get('df_loai_nghi', pd.DataFrame()),
        )

        ngay = normalize_schedule_date(recalculated.get('Ngày', ''))
        nv = str(recalculated.get('Tên nhân viên', '')).strip()
        lydo = clean_leave_reason_display(recalculated.get('Lý do nghỉ', ''))
        if not nv or not lydo:
            return False, "Tên nhân viên và Lý do nghỉ không được để trống."

        # Chỉ chặn nếu sửa thành CÙNG Ngày + Nhân viên + Lý do đã tồn tại.
        # Nếu lý do khác nhau thì được phép có nhiều dòng trong cùng ngày, kể cả Phạt vi phạm > 0.
        others = _exclude_original_from_leave_df(live_all, original_row)
        if _leave_exists_in_sources(others, ngay, nv, lydo):
            return False, f"'{nv}' đã có đúng lý do '{lydo}' trong ngày {ngay}. Có thể ghi thêm nếu là lý do khác."

        new_values = [
            ngay,
            nv,
            lydo,
            str(recalculated.get('Chi tiết', '')).strip(),
            float(recalculated.get('Số ngày tính', 0) or 0),
            float(recalculated.get('Số ngày phép cộng dồn', 0) or 0),
            float(recalculated.get('Phạt vi phạm', 0) or 0),
            str(recalculated.get('Ngày cập nhật', '')),
            str(recalculated.get('Giờ cập nhật', '')),
            str(recalculated.get('Người cập nhật', updated_by)),
        ]

        source_id = str(original_row.get('__source_sheet_id', '')).strip() or SHEET_DU_PHONG_ID
        target = client.open_by_key(source_id).get_worksheet(0)
        row_idx = _find_schedule_row_index(target, original_row)
        if not row_idx:
            return False, "Không tìm thấy dòng tương ứng trong Google Sheet nguồn."
        gspread_update_range(target, f'A{row_idx}:J{row_idx}', [new_values], raw=False)

        # Nhóm mới cũng phải được chuẩn hóa. Nếu đổi Không phép -> Có phép thì new_group=None,
        # nhưng old_group vẫn được xếp lại để Người 2 trở thành Người 1, v.v.
        new_group = _progressive_group_key(recalculated)
        if new_group:
            affected_groups.add(new_group)

        rebalanced = rebalance_progressive_penalty_groups(client, affected_groups, updated_by)

        _clear_dynamic_data_caches()
        if rebalanced:
            return True, f"Đã cập nhật lịch nghỉ và tự xếp lại thứ tự/phạt lũy tiến cho {rebalanced} bản ghi trong nhóm bị ảnh hưởng."
        return True, "Đã cập nhật lịch nghỉ thành công."
    except Exception as e:
        return False, f"Lỗi cập nhật lịch nghỉ: {e}"

def delete_schedule_records(original_rows, updated_by=None):
    """Xóa nhiều lịch đúng nguồn rồi tự xếp lại thứ tự/phạt của mọi nhóm vi phạm bị ảnh hưởng."""
    try:
        client = get_gspread_client()
        if not client:
            return False, "Chưa cấu hình quyền kết nối Google Sheets."
        actor = str(updated_by or st.session_state.get("current_user", "Hệ thống"))

        affected_groups = set()
        grouped = {}
        for row in original_rows:
            group_key = _progressive_group_key(row)
            if group_key:
                affected_groups.add(group_key)
            source_id = str(row.get('__source_sheet_id', '')).strip() or SHEET_DU_PHONG_ID
            grouped.setdefault(source_id, []).append(row)

        deleted = 0
        for source_id, rows in grouped.items():
            target = client.open_by_key(source_id).get_worksheet(0)
            indices = []
            for row in rows:
                idx = _find_schedule_row_index(target, row)
                if idx:
                    indices.append(idx)
            for idx in sorted(set(indices), reverse=True):
                target.delete_rows(idx)
                deleted += 1

        rebalanced = rebalance_progressive_penalty_groups(client, affected_groups, actor) if affected_groups else 0

        _clear_dynamic_data_caches()
        if rebalanced:
            return True, f"Đã xóa {deleted} dòng và tự xếp lại thứ tự/phạt cho {rebalanced} bản ghi còn lại."
        return True, f"Đã xóa {deleted} dòng lịch nghỉ từ đúng Google Sheet nguồn."
    except Exception as e:
        return False, f"Lỗi xóa lịch nghỉ: {e}"


# --- HÀM TẢI FILE TỪ DRIVE ---
def download_file_from_google_drive(id, destination):
    """Tải file nhị phân từ Google Drive, hỗ trợ trang confirm của file lớn."""
    session = requests.Session()
    errors = []

    # Endpoint usercontent thường ổn định hơn với file Excel nhị phân công khai.
    urls = [
        f"https://drive.usercontent.google.com/download?id={id}&export=download&confirm=t",
        f"https://drive.google.com/uc?export=download&id={id}&confirm=t",
    ]

    for url in urls:
        try:
            response = session.get(url, stream=True, timeout=60, allow_redirects=True)
            response.raise_for_status()

            # Một số file lớn vẫn trả trang confirm; thử lấy token từ HTML/cookie.
            ctype = str(response.headers.get('Content-Type', '')).lower()
            if 'text/html' in ctype:
                html = response.text
                token = next((v for k, v in response.cookies.items() if k.startswith('download_warning')), None)
                if not token:
                    m = re.search(r'confirm=([0-9A-Za-z_-]+)', html)
                    token = m.group(1) if m else None
                if token:
                    response = session.get(
                        "https://drive.google.com/uc",
                        params={'export': 'download', 'id': id, 'confirm': token},
                        stream=True, timeout=60, allow_redirects=True
                    )
                    response.raise_for_status()

            with open(destination, "wb") as f:
                for chunk in response.iter_content(1024 * 1024):
                    if chunk:
                        f.write(chunk)

            if os.path.exists(destination) and os.path.getsize(destination) > 0:
                return destination
        except Exception as e:
            errors.append(str(e))
            try:
                if os.path.exists(destination):
                    os.remove(destination)
            except Exception:
                pass

    raise RuntimeError("Không tải được file Google Drive: " + " | ".join(errors[-2:]))


def _excel_col_letter(idx):
    n = idx + 1
    out = ""
    while n:
        n, rem = divmod(n - 1, 26)
        out = chr(65 + rem) + out
    return out


@st.cache_data(ttl=15, show_spinner=False)
def load_bang_tour_input():
    """Đọc sheet Input từ file Bảng tour dạng .XLSM trên Google Drive."""
    temp_file = f"temp_bangtour_{os.getpid()}_{int(time.time())}.xlsm"
    try:
        download_file_from_google_drive(BANG_TOUR_FILE_ID, temp_file)

        # XLSM là gói ZIP. Nếu không phải ZIP thì gần như chắc chắn Google Drive
        # đã trả HTML (đăng nhập/xác nhận quyền truy cập), không phải file Excel.
        if not zipfile.is_zipfile(temp_file):
            preview = ""
            try:
                preview = open(temp_file, 'r', encoding='utf-8', errors='ignore').read(180)
            except Exception:
                pass
            hint = " Google Drive đang trả trang HTML thay vì file XLSM." if '<html' in preview.lower() else ""
            return pd.DataFrame(), (
                "File tải về không phải XLSM hợp lệ." + hint +
                " Hãy đặt quyền file thành 'Bất kỳ ai có đường liên kết' hoặc cấp quyền cho service account."
            )

        # File đã xác nhận là .xlsm nên dùng openpyxl trực tiếp; không dùng pyxlsb.
        raw = pd.read_excel(temp_file, sheet_name="Input", header=None, engine="openpyxl")
        if raw.empty:
            return pd.DataFrame(), "Sheet Input đang trống."

        # Hệ thống Tour Vera dùng dòng 20 làm header và dữ liệu từ dòng 21.
        header_idx = 19 if len(raw) > 19 else 0
        # VBA người dùng gửi có rule đến cột X -> giữ tối đa A:X.
        max_cols = min(24, raw.shape[1])
        raw = raw.iloc[:, :max_cols]
        header_vals = raw.iloc[header_idx].tolist()

        headers = []
        seen = {}
        for i, v in enumerate(header_vals):
            txt = "" if pd.isna(v) else str(v).strip()
            if not txt or txt.lower() == "nan":
                txt = _excel_col_letter(i)
            if txt in seen:
                seen[txt] += 1
                txt = f"{txt} ({_excel_col_letter(i)})"
            else:
                seen[txt] = 1
            headers.append(txt)

        df = raw.iloc[header_idx + 1:].copy()
        df.columns = headers
        df = df.dropna(how="all").reset_index(drop=True)
        df.attrs["excel_header_row"] = header_idx + 1
        # Lưu vị trí cột gốc Excel để việc đổi thứ tự hiển thị không làm sai rule màu.
        df.attrs["excel_col_index"] = {headers[i]: i for i in range(len(headers))}
        return df, ""
    except ValueError as e:
        return pd.DataFrame(), f"Không đọc được sheet Input trong file XLSM: {e}"
    except Exception as e:
        return pd.DataFrame(), f"Lỗi tải/đọc Bảng tour XLSM: {e}"
    finally:
        try:
            if os.path.exists(temp_file):
                os.remove(temp_file)
        except Exception:
            pass


def _tour_text(v):
    if pd.isna(v):
        return ""
    return str(v).strip()


def _tour_num(v):
    try:
        if pd.isna(v) or str(v).strip() == "":
            return None
        return float(str(v).replace(",", "").strip())
    except Exception:
        return None


def _find_tour_col(df, wanted):
    """Tìm tên cột theo kiểu không dấu/không phân biệt hoa thường."""
    wanted_norm = remove_vietnamese_accents(str(wanted)).casefold().strip()
    exact = []
    contains = []
    for c in df.columns:
        norm = remove_vietnamese_accents(str(c)).casefold().strip()
        if norm == wanted_norm:
            exact.append(c)
        elif wanted_norm in norm:
            contains.append(c)
    return exact[0] if exact else (contains[0] if contains else None)


def reorder_bang_tour_columns(df):
    """
    Chỉ đổi thứ tự HIỂN THỊ:
    Tên Nhân Viên -> Trạng Thái -> Thời gian còn lại -> các cột còn lại.
    Vị trí cột Excel gốc vẫn được giữ trong df.attrs để tô màu đúng.
    """
    if df.empty:
        return df

    name_col = _find_tour_col(df, "Tên nhân viên")
    status_col = _find_tour_col(df, "Trạng thái")
    remain_col = _find_tour_col(df, "Thời gian còn lại")

    cols = list(df.columns)
    moved = [c for c in [status_col, remain_col] if c is not None]
    base = [c for c in cols if c not in moved]

    if name_col and name_col in base:
        pos = base.index(name_col) + 1
        for c in reversed(moved):
            base.insert(pos, c)
    else:
        # Nếu workbook đổi tên cột thì vẫn ưu tiên đưa hai cột này ra đầu.
        base = moved + base

    out = df.loc[:, base].copy()
    out.attrs.update(df.attrs)
    return out


def _tour_norm_token(v):
    """Chuẩn hóa text Tour: không dấu, không phân biệt hoa/thường, bỏ _ và - thừa."""
    txt = remove_vietnamese_accents(_tour_text(v)).casefold()
    txt = txt.replace("_", " ").replace("-", " ")
    return " ".join(txt.split())


def prepare_bang_tour_display(df):
    """
    Chuẩn bị dữ liệu HIỂN THỊ cho Bảng tour sau khi đọc từ file:
    - Tên Nhân Viên -> Trạng Thái -> Thời gian còn lại -> các cột khác.
    - DANG CHO -> Đang chờ; DANG THUC HIEN -> Đang thực hiện.
    - Mọi giá trị số hiển thị dạng SỐ NGUYÊN, không có phần thập phân.
    - None / NaN / NaT / <NA> được đổi thành ô trống thật sự.
    - Thời gian còn lại <= -15 được làm trống trên giao diện (không ghi ngược file nguồn).
    - Cột Thời gian: nếu là số âm nhỏ hơn -180 (độ âm vượt 180 phút) thì làm trống trên giao diện.
    """
    out = reorder_bang_tour_columns(df).copy()

    # Ghi nhớ các dòng có thời gian <= -15 trước khi làm trống để không nhầm thành "Đang rảnh".
    expired_indices = set()
    remain_col = _find_tour_col(out, "Thời gian còn lại")
    if remain_col is not None:
        raw_remain = out[remain_col].apply(_tour_num)
        expired_indices = set(raw_remain[raw_remain.apply(lambda x: x is not None and x <= -15)].index.tolist())

        def fmt_remaining(v):
            n = _tour_num(v)
            if n is None or n <= -15:
                return ""
            return str(int(round(n)))

        out[remain_col] = out[remain_col].apply(fmt_remaining)

    # V86.2: cột "Thời gian" có thể chứa số âm bất thường rất lớn
    # (ví dụ -66605391). Nếu độ âm vượt 180 phút thì không hiển thị.
    # Chỉ thay đổi giao diện, không ghi ngược về TourVera.xlsm.
    time_col = _find_tour_col(out, "Thời gian")
    if time_col is not None and time_col != remain_col:
        def fmt_tour_time(v):
            n = _tour_num(v)
            if n is not None:
                if n < -180:
                    return ""
                # Giữ kiểu hiển thị số nguyên giống các cột số khác.
                return str(int(round(n)))
            if pd.isna(v):
                return ""
            s = str(v).strip()
            return "" if s.casefold() in {"nan", "none", "nat", "<na>"} else s

        out[time_col] = out[time_col].apply(fmt_tour_time)

    status_col = _find_tour_col(out, "Trạng thái")
    if status_col is not None:
        def fmt_status(v):
            token = _tour_norm_token(v)
            if token == "dang cho":
                return "Đang chờ"
            if token == "dang thuc hien":
                return "Đang thực hiện"
            if pd.isna(v):
                return ""
            s = str(v).strip()
            return "" if s.casefold() in {"none", "nan", "nat", "<na>"} else s
        out[status_col] = out[status_col].apply(fmt_status)

    def clean_and_integer_tour_value(v):
        # Ẩn hoàn toàn giá trị thiếu thật sự.
        try:
            if pd.isna(v):
                return ""
        except Exception:
            pass

        if isinstance(v, str):
            s = v.strip()
            if s.casefold() in {"none", "nan", "nat", "<na>"}:
                return ""
            # Nếu chuỗi chỉ là một số thì hiển thị thành số nguyên.
            if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", s):
                try:
                    return str(int(round(float(s))))
                except Exception:
                    return s
            return s

        # Giá trị số từ openpyxl/pandas -> số nguyên hiển thị.
        if isinstance(v, bool):
            return v
        try:
            if isinstance(v, numbers.Number):
                return str(int(round(float(v))))
        except Exception:
            pass
        return v

    for c in out.columns:
        out[c] = out[c].apply(clean_and_integer_tour_value)

    out.attrs.update(df.attrs)
    out.attrs["_tour_expired_indices"] = expired_indices
    return out

def calculate_bang_tour_stats(df):
    """Tính bảng thống kê số lượng cho Bảng tour từ dữ liệu gốc vừa tải."""
    if df.empty:
        return pd.DataFrame(columns=["Chỉ số", "Số lượng"])

    status_col = _find_tour_col(df, "Trạng thái")
    remain_col = _find_tour_col(df, "Thời gian còn lại")
    work_col = _find_tour_col(df, "Đi làm")
    shift_col = _find_tour_col(df, "Vào ca")
    break_col = _find_tour_col(df, "Break")

    # Nếu tên cột workbook hơi khác, thử các biến thể thường gặp.
    if break_col is None:
        break_col = _find_tour_col(df, "Breaktime")
    if shift_col is None:
        shift_col = _find_tour_col(df, "Ca")

    blank = pd.Series([""] * len(df), index=df.index, dtype=object)
    status_s = df[status_col].apply(_tour_norm_token) if status_col else blank.copy()
    work_s = df[work_col].apply(_tour_norm_token) if work_col else blank.copy()
    shift_s = df[shift_col].apply(_tour_norm_token) if shift_col else blank.copy()
    break_s = df[break_col].apply(_tour_norm_token) if break_col else blank.copy()
    remain_num = df[remain_col].apply(_tour_num) if remain_col else pd.Series([None] * len(df), index=df.index)
    remain_num = pd.to_numeric(remain_num, errors='coerce')
    status_num = pd.to_numeric(df[status_col], errors='coerce') if status_col else pd.Series([float('nan')] * len(df), index=df.index)

    dang_thuc_hien_mask = status_s.eq("dang thuc hien")
    dang_cho_mask = status_s.eq("dang cho")

    # "Sắp xong": hỗ trợ đúng cả hai trường hợp dữ liệu:
    # - Nếu cột Trạng thái có giá trị số: đếm giá trị <= 30 theo yêu cầu.
    # - Với cấu trúc hiện tại Trạng thái là chữ DANG THUC HIEN: dùng Thời gian còn lại <= 30.
    #   Loại <= -15 vì các giá trị này được làm trống khỏi bảng hiển thị.
    sap_xong_mask = (
        (status_num.notna() & (status_num <= 30))
        | (dang_thuc_hien_mask & remain_num.notna() & (remain_num <= 30) & (remain_num > -15))
    )

    active_shift_mask = shift_s.isin(["ca 1", "ca 2"])
    di_lam_mask = work_s.eq("di lam")
    nghi_phep_mask = work_s.eq("nghi phep")
    idle_time_mask = remain_num.isna() | remain_num.eq(0)
    dang_ranh_mask = idle_time_mask & active_shift_mask & di_lam_mask

    ca1_mask = shift_s.eq("ca 1")
    ca2_mask = shift_s.eq("ca 2")
    break_mask = break_s.eq("break")

    dang_thuc_hien = int(dang_thuc_hien_mask.sum())
    dang_cho = int(dang_cho_mask.sum())
    sap_xong = int(sap_xong_mask.sum())
    dang_ranh = int(dang_ranh_mask.sum())
    co_the_len_tour = sap_xong + dang_ranh

    # V85.1: giao diện Thống kê Bảng tour chỉ hiển thị duy nhất
    # "Có thể lên tour". Vẫn giữ nguyên công thức = Sắp xong + Đang rảnh.
    rows = [
        ("Có thể lên tour", co_the_len_tour),
    ]
    return pd.DataFrame(rows, columns=["Chỉ số", "Số lượng"])


def style_bang_tour(df):
    """
    Tô nguyên dòng Bảng tour theo quy tắc vận hành Vera.
    - Nghỉ phép: nền trắng, chữ mờ.
    - Đi làm: nền trắng, chữ đen.
    - Thời gian >= 15: xanh lá.
    - 0 <= thời gian < 15: vàng.
    - -15 < thời gian < 0: đỏ.
    - Đang rảnh: xanh nhạt + chữ đậm.
    - Break: cam (ưu tiên cao nhất).
    - Header: nền rgb(161,148,140) / #A1948C, chữ đen đậm.
    """
    remain_col = _find_tour_col(df, "Thời gian còn lại")
    work_col = _find_tour_col(df, "Đi làm")
    shift_col = _find_tour_col(df, "Vào ca")
    break_col = _find_tour_col(df, "Break")
    if break_col is None:
        break_col = _find_tour_col(df, "Breaktime")
    if shift_col is None:
        shift_col = _find_tour_col(df, "Ca")

    expired_indices = set(df.attrs.get("_tour_expired_indices", set()))

    def row_style(row):
        work_norm = _tour_norm_token(row.get(work_col, "")) if work_col else ""
        shift_norm = _tour_norm_token(row.get(shift_col, "")) if shift_col else ""
        break_norm = _tour_norm_token(row.get(break_col, "")) if break_col else ""
        remain_num = _tour_num(row.get(remain_col, "")) if remain_col else None

        bg = "#FFFFFF"
        fg = "#000000"
        weight = "400"

        if work_norm == "nghi phep":
            bg, fg, weight = "#FFFFFF", "#A6A6A6", "400"
        elif work_norm == "di lam":
            bg, fg, weight = "#FFFFFF", "#000000", "400"

        # Màu theo thời gian, không ghi đè dòng Nghỉ phép.
        if remain_num is not None and work_norm != "nghi phep":
            if remain_num >= 15:
                bg, fg, weight = "#92D050", "#000000", "600"
            elif 0 <= remain_num < 15:
                bg, fg, weight = "#FFD966", "#000000", "600"
            elif -15 < remain_num < 0:
                bg, fg, weight = "#FF6666", "#000000", "600"

        # Đang rảnh: thời gian trống/0 + Ca 1/Ca 2 + Đi làm; loại trừ dòng <= -15 đã bị làm trống.
        is_idle = (
            row.name not in expired_indices
            and work_norm == "di lam"
            and shift_norm in {"ca 1", "ca 2"}
            and (remain_num is None or remain_num == 0)
        )
        if is_idle:
            bg, fg, weight = "#D9EAD3", "#000000", "700"

        # Break ưu tiên cuối cùng.
        if break_norm == "break":
            bg, fg, weight = "#F4B183", "#000000", "700"

        css = (
            f"background-color:{bg};"
            f"color:{fg};"
            f"font-weight:{weight};"
            "white-space:nowrap;"
        )
        return [css] * len(row)

    styler = df.style.apply(row_style, axis=1).format(na_rep="")
    styler = styler.set_table_styles([
        {
            "selector": "th",
            "props": [
                ("background-color", "#A1948C"),
                ("color", "#000000"),
                ("font-weight", "700"),
                ("text-align", "center"),
                ("white-space", "normal"),
                ("overflow-wrap", "anywhere"),
                ("word-break", "break-word"),
                ("line-height", "1.15"),
            ],
        },
        {"selector": "td", "props": [("white-space", "nowrap")]},
    ])

    status_col = _find_tour_col(df, "Trạng thái")
    if status_col is not None:
        styler = styler.set_properties(
            subset=[status_col],
            **{"white-space": "nowrap", "min-width": "135px", "width": "135px"}
        )
    return styler

def combine_leave_sources_for_daily_stats(*sources):
    """
    Hợp nhất một hoặc nhiều nguồn lịch nghỉ. Loại trùng theo:
    Ngày + Tên nhân viên + Lý do nghỉ. Nguồn truyền vào SAU sẽ được ưu tiên
    khi cùng một bản ghi xuất hiện ở nhiều nguồn.
    """
    expected = [
        'Ngày', 'Tên nhân viên', 'Lý do nghỉ', 'Chi tiết', 'Số ngày tính',
        'Số ngày phép cộng dồn', 'Phạt vi phạm', 'Ngày cập nhật',
        'Giờ cập nhật', 'Người cập nhật'
    ]
    meta_cols = ['__source_sheet_id', '__source_row']
    prepared = []
    for source in sources:
        if source is None or source.empty:
            continue
        d = source.copy()
        if 'Loại nghỉ' in d.columns and 'Lý do nghỉ' not in d.columns:
            d = d.rename(columns={'Loại nghỉ': 'Lý do nghỉ'})
        for col in expected:
            if col not in d.columns:
                d[col] = ""
        for col in meta_cols:
            if col not in d.columns:
                d[col] = ""
        d = d[expected + meta_cols].copy()
        # Đọc từng ô để không làm mất các dòng có ngày dạng serial Excel/Google Sheets
        # hoặc dữ liệu ngày bị trộn dd/mm/yyyy, dd-mm-yyyy, yyyy-mm-dd.
        d['Ngày'] = d['Ngày'].apply(_parse_vn_date)
        d = d[d['Ngày'].notna()].copy()
        for _money_col in ['Số ngày tính', 'Số ngày phép cộng dồn']:
            d[_money_col] = d[_money_col].apply(lambda _v: _parse_leave_number(_v, 0.0, money=False))
        d['Phạt vi phạm'] = d['Phạt vi phạm'].apply(lambda _v: _parse_leave_number(_v, 0.0, money=True))
        prepared.append(d)

    if not prepared:
        return pd.DataFrame(columns=expected + meta_cols)

    combined = pd.concat(prepared, ignore_index=True)
    combined['_key'] = (
        combined['Ngày'].astype(str) + '|' +
        combined['Tên nhân viên'].apply(normalize_name).str.casefold() + '|' +
        combined['Lý do nghỉ'].astype(str).str.strip().str.casefold()
    )
    combined = combined.drop_duplicates(subset=['_key'], keep='last').drop(columns=['_key'])
    return combined.reset_index(drop=True)


@st.cache_data(ttl=60)
def load_lich_nghi(url):
    try:
        file_id = url.split('/d/')[1].split('/')[0]
        temp_file = "temp_lichnghi.xlsb"
        download_file_from_google_drive(file_id, temp_file)
        
        xls = pd.read_excel(temp_file, sheet_name=['LichNghi', 'DanhSachNV', 'LoaiNghi'], engine='pyxlsb')
        df_lich = xls['LichNghi'].iloc[:, :10]
        df_lich.columns = ['Ngày', 'Tên nhân viên', 'Lý do nghỉ', 'Chi tiết', 'Số ngày tính', 'Số ngày phép cộng dồn', 'Phạt vi phạm', 'Ngày cập nhật', 'Giờ cập nhật', 'Người cập nhật']
        
        if os.path.exists(temp_file): os.remove(temp_file)
            
        def safe_date_parse(val):
            try:
                if pd.isna(val): return pd.NaT
                if hasattr(val, 'date'): return val.date() 
                if isinstance(val, (int, float)): return pd.to_datetime(val, unit='D', origin='1899-12-30').date()
                s = str(val).strip().split(' ')[0]
                return pd.to_datetime(s, dayfirst=True).date()
            except: return pd.NaT
                
        df_lich['Ngày'] = df_lich['Ngày'].apply(safe_date_parse)
        df_lich = df_lich.dropna(subset=['Ngày'])
        df_lich['Số ngày tính'] = pd.to_numeric(df_lich['Số ngày tính'].astype(str).str.replace(',', '').str.replace('-', '').str.strip(), errors='coerce').fillna(0)
        df_lich['Phạt vi phạm'] = pd.to_numeric(df_lich['Phạt vi phạm'].astype(str).str.replace(',', '').str.replace('-', '').str.strip(), errors='coerce').fillna(0)
        
        # FIX FORMAT NGÀY GIỜ CẬP NHẬT TỪ EXCEL SERIAL
        def format_excel_date(val):
            if pd.isna(val) or str(val).strip() == "": return ""
            try:
                if isinstance(val, (int, float)):
                    return pd.to_datetime(val, unit='D', origin='1899-12-30').strftime('%d/%m/%Y')
                if hasattr(val, 'strftime'): return val.strftime('%d/%m/%Y')
                return str(val).split(' ')[0]
            except: return str(val)

        def format_excel_time(val):
            if pd.isna(val) or str(val).strip() == "": return ""
            try:
                if isinstance(val, (int, float)):
                    s = int(round(val * 86400))
                    return f"{s // 3600:02d}:{(s % 3600) // 60:02d}:{s % 60:02d}"
                if hasattr(val, 'strftime'): return val.strftime('%H:%M:%S')
                return str(val)
            except: return str(val)

        if 'Ngày cập nhật' in df_lich.columns:
            df_lich['Ngày cập nhật'] = df_lich['Ngày cập nhật'].apply(format_excel_date)
        if 'Giờ cập nhật' in df_lich.columns:
            df_lich['Giờ cập nhật'] = df_lich['Giờ cập nhật'].apply(format_excel_time)
            
        df_nv_excel = xls['DanhSachNV'].dropna(subset=['Tên nhân viên'])
        return df_lich, df_nv_excel, xls['LoaiNghi']
    except Exception as e:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()

@st.cache_data(show_spinner=False)
def to_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='DuLieuLichNghi')
    return output.getvalue()


# ==========================================================
# BẢNG LƯƠNG - quyền truy cập được quản lý theo vai trò/tài khoản
# ==========================================================
PAYROLL_COLUMNS = [
    "TT", "Tên Hệ thống", "Họ và tên", "Tiền Lương", "Tiền Hỗ Trợ Hoàn Lại",
    "Tích lũy", "Chi Phí Sinh Hoạt", "Tiền phạt trong tháng", "Vi phạm kỳ trước", "Tiền ứng lương",
    "Tiền hỗ trợ Locker", "Số tiền thực nhận", "Email",
    "Số tài khoản ngân hàng", "Tên ngân hàng", "Số dòng Tip"
]
PAYROLL_ADJUSTMENT_COLUMNS = [
    "Tiền Hỗ Trợ Hoàn Lại", "Tích lũy", "Chi Phí Sinh Hoạt",
    "Tiền ứng lương", "Tiền hỗ trợ Locker"
]

# Tiêu đề hiển thị chuẩn cho toàn bộ bảng lương (web + Excel).
PAYROLL_DISPLAY_LABELS = {
    "TT": "TT",
    "Tên Hệ thống": "Tên Hệ thống",
    "Tiền Lương": "Tiền Lương",
    "Tiền Hỗ Trợ Hoàn Lại": "Hỗ Trợ Hoàn Lại",
    "Tích lũy": "Tích lũy",
    "Chi Phí Sinh Hoạt": "Phí Sinh Hoạt",
    "Tiền phạt trong tháng": "Vi phạm",
    "Vi phạm kỳ trước": "Vi phạm kỳ trước",
    "Tiền ứng lương": "Tiền ứng",
    "Tiền hỗ trợ Locker": "Tiền hỗ trợ Locker",
    "Số tiền thực nhận": "Thực nhận",
    "Số tài khoản ngân hàng": "Tài khoản ngân hàng",
    "Tên ngân hàng": "Tên ngân hàng",
    "Email": "Email",
}
PAYROLL_HISTORY_HEADERS = [
    "Mã bản lưu", "Từ ngày", "Đến ngày", "Ngày lưu", "Giờ lưu", "Người lưu", "Nguồn dữ liệu",
    "TT", "Tên Hệ thống", "Họ và tên", "Tiền Lương", "Tiền Hỗ Trợ Hoàn Lại",
    "Hỗ trợ dạy nghề", "Học phí", "Tích lũy", "Chi Phí Sinh Hoạt",
    "Tiền phạt trong tháng", "Tiền ứng lương", "Tiền hỗ trợ Locker", "Số tiền thực nhận",
    "Email", "Số tài khoản ngân hàng", "Tên ngân hàng", "Số dòng Tip", "Vi phạm kỳ trước"
]


def _get_or_create_worksheet(spreadsheet, title, rows=1000, cols=30):
    try:
        return spreadsheet.worksheet(title)
    except Exception:
        return spreadsheet.add_worksheet(title=title, rows=rows, cols=cols)


def _get_usage_guide_worksheet():
    try:
        client = get_gspread_client()
        if not client:
            return None, "Chưa cấu hình quyền kết nối Google Sheets."
        ss = client.open_by_key(SHEET_MAT_KHAU_ID)
        ws = _get_or_create_worksheet(ss, USAGE_GUIDE_WORKSHEET, rows=200, cols=10)
        if int(getattr(ws, "col_count", 0) or 0) < 10:
            _gs_call_with_backoff(ws.resize, cols=10)
        return ws, ""
    except Exception as e:
        return None, f"Không mở được kho Hướng dẫn sử dụng: {e}"


@st.cache_data(ttl=300, show_spinner=False)
def load_usage_guide_document():
    ws, err = _get_usage_guide_worksheet()
    if err or ws is None:
        return None, None, err
    try:
        meta_rows = _gs_call_with_backoff(ws.get, 'A1:J2')
        if not meta_rows or len(meta_rows) < 2:
            return None, None, ""
        headers = list(meta_rows[0]) + [""] * max(0, 10 - len(meta_rows[0]))
        vals = list(meta_rows[1]) + [""] * max(0, 10 - len(meta_rows[1]))
        meta = {str(headers[i] or USAGE_GUIDE_META_HEADERS[i]): vals[i] for i in range(10)}
        filename = str(meta.get("Tên file", "")).strip()
        chunk_count = int(float(str(meta.get("Số chunk", "0") or "0")))
        if not filename or chunk_count <= 0:
            return None, None, ""
        rows = _gs_call_with_backoff(ws.get, f'A4:B{3 + chunk_count}')
        chunks = [str(r[1]).strip() for r in (rows or []) if len(r) >= 2 and str(r[1]).strip()]
        if len(chunks) != chunk_count:
            return meta, None, f"Tài liệu bị thiếu dữ liệu ({len(chunks)}/{chunk_count} phần)."
        raw = base64.b64decode(''.join(chunks).encode('ascii'))
        expected_sha = str(meta.get("SHA256", "")).strip().lower()
        if expected_sha and hashlib.sha256(raw).hexdigest() != expected_sha:
            return meta, None, "Không thể mở tài liệu vì kiểm tra toàn vẹn SHA256 không khớp."
        return meta, raw, ""
    except Exception as e:
        return None, None, f"Lỗi đọc Hướng dẫn sử dụng: {e}"


def _clear_usage_guide_cache():
    try:
        load_usage_guide_document.clear()
    except Exception:
        pass


def save_usage_guide_document(uploaded_file, title, version, note, actor):
    if uploaded_file is None:
        return False, "Chưa chọn file tài liệu."
    raw = uploaded_file.getvalue()
    if not raw:
        return False, "File tài liệu đang trống."
    if len(raw) > USAGE_GUIDE_MAX_BYTES:
        return False, f"File vượt giới hạn {USAGE_GUIDE_MAX_BYTES // (1024*1024)} MB. Hãy giảm dung lượng PDF/ảnh trước khi tải lên."

    filename = str(getattr(uploaded_file, 'name', '') or 'huong-dan').strip()
    mime = str(getattr(uploaded_file, 'type', '') or '').strip().lower()
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    if ext not in {'pdf', 'png', 'jpg', 'jpeg', 'webp'}:
        return False, "Chỉ hỗ trợ PDF, PNG, JPG/JPEG hoặc WEBP."
    if not mime:
        mime = 'application/pdf' if ext == 'pdf' else f'image/{"jpeg" if ext in {"jpg","jpeg"} else ext}'

    b64 = base64.b64encode(raw).decode('ascii')
    chunks = [b64[i:i + USAGE_GUIDE_CHUNK_SIZE] for i in range(0, len(b64), USAGE_GUIDE_CHUNK_SIZE)]
    now = datetime.now(VN_TZ).strftime('%d/%m/%Y %H:%M:%S')
    meta = [
        str(title or filename).strip(), str(version or '').strip(), filename, mime,
        str(len(raw)), hashlib.sha256(raw).hexdigest(), now, str(actor or '').strip(),
        str(note or '').strip(), str(len(chunks))
    ]

    ws, err = _get_usage_guide_worksheet()
    if err or ws is None:
        return False, err or "Không mở được kho Hướng dẫn sử dụng."
    try:
        _gs_call_with_backoff(ws.resize, rows=max(30, 4 + len(chunks)), cols=10)
        _gs_call_with_backoff(ws.clear)
        gspread_update_range(ws, 'A1:J2', [USAGE_GUIDE_META_HEADERS, meta], raw=False)
        gspread_update_range(ws, 'A3:B3', [["STT chunk", "Dữ liệu base64"]], raw=False)
        for start in range(0, len(chunks), 80):
            part = chunks[start:start + 80]
            row_start = 4 + start
            row_end = row_start + len(part) - 1
            values = [[start + i + 1, chunk] for i, chunk in enumerate(part)]
            gspread_update_range(ws, f'A{row_start}:B{row_end}', values, raw=False)
        _clear_usage_guide_cache()
        return True, f"Đã lưu '{filename}' ({len(raw)/1024/1024:.2f} MB) làm Hướng dẫn sử dụng."
    except Exception as e:
        return False, f"Lỗi lưu Hướng dẫn sử dụng: {e}"


def update_usage_guide_metadata(title, version, note, actor):
    meta, raw, err = load_usage_guide_document()
    if err:
        return False, err
    if not meta or raw is None:
        return False, "Chưa có tài liệu để sửa thông tin."
    ws, err = _get_usage_guide_worksheet()
    if err or ws is None:
        return False, err or "Không mở được kho Hướng dẫn sử dụng."
    try:
        vals = [
            str(title or meta.get('Tên tài liệu', '')).strip(), str(version or '').strip(),
            str(meta.get('Tên file', '')).strip(), str(meta.get('MIME', '')).strip(),
            str(meta.get('Dung lượng', '')).strip(), str(meta.get('SHA256', '')).strip(),
            datetime.now(VN_TZ).strftime('%d/%m/%Y %H:%M:%S'), str(actor or '').strip(),
            str(note or '').strip(), str(meta.get('Số chunk', '')).strip(),
        ]
        gspread_update_range(ws, 'A1:J2', [USAGE_GUIDE_META_HEADERS, vals], raw=False)
        _clear_usage_guide_cache()
        return True, "Đã cập nhật thông tin Hướng dẫn sử dụng."
    except Exception as e:
        return False, f"Lỗi sửa thông tin tài liệu: {e}"


def delete_usage_guide_document(actor):
    ws, err = _get_usage_guide_worksheet()
    if err or ws is None:
        return False, err or "Không mở được kho Hướng dẫn sử dụng."
    try:
        _gs_call_with_backoff(ws.clear)
        gspread_update_range(ws, 'A1:J1', [USAGE_GUIDE_META_HEADERS], raw=False)
        _clear_usage_guide_cache()
        return True, f"Đã xóa Hướng dẫn sử dụng bởi {actor}."
    except Exception as e:
        return False, f"Lỗi xóa Hướng dẫn sử dụng: {e}"


def render_protected_usage_guide(meta, raw):
    if not meta or raw is None:
        st.info("Hiện chưa có Hướng dẫn sử dụng được Admin tải lên.")
        return

    filename = str(meta.get('Tên file', '')).strip()
    mime = str(meta.get('MIME', '')).strip().lower()
    title = str(meta.get('Tên tài liệu', '')).strip() or filename
    version = str(meta.get('Phiên bản', '')).strip()
    note = str(meta.get('Ghi chú', '')).strip()
    user_mark = f"VERA SPA • {st.session_state.get('current_user','')} • {datetime.now(VN_TZ).strftime('%d/%m/%Y %H:%M')}"
    payload = base64.b64encode(raw).decode('ascii')

    st.markdown(f"### 📘 {title}")
    meta_bits = []
    if version:
        meta_bits.append(f"Phiên bản: **{version}**")
    if meta.get('Cập nhật lúc'):
        meta_bits.append(f"Cập nhật: **{meta.get('Cập nhật lúc')}**")
    if meta_bits:
        st.caption(" · ".join(meta_bits))
    if note:
        st.info(note)

    safe_payload = json.dumps(payload)
    safe_mime = json.dumps(mime)
    safe_mark = json.dumps(user_mark)
    is_pdf = mime == 'application/pdf' or filename.lower().endswith('.pdf')
    pdf_flag = 'true' if is_pdf else 'false'

    components.html(f"""<!doctype html>
<html><head><meta charset='utf-8'>
<style>
html,body{{margin:0;padding:0;background:#eef1f5;font-family:Arial,sans-serif;user-select:none;-webkit-user-select:none;-webkit-touch-callout:none;}}
#shell{{height:850px;overflow:auto;padding:12px;box-sizing:border-box;position:relative;}}
.page{{position:relative;margin:0 auto 16px auto;background:white;box-shadow:0 2px 10px rgba(0,0,0,.15);width:min(100%,1000px);}}
.page canvas,.page img{{display:block;width:100%;height:auto;pointer-events:none;-webkit-user-drag:none;}}
.watermark{{position:absolute;inset:0;pointer-events:none;overflow:hidden;z-index:5;}}
.watermark span{{position:absolute;left:8%;right:8%;top:45%;transform:rotate(-28deg);font-size:clamp(18px,3vw,34px);font-weight:800;color:rgba(80,80,80,.16);text-align:center;white-space:nowrap;}}
.notice{{padding:12px;border:1px solid #ddd;border-radius:8px;background:white;color:#555;}}
@media print{{html,body,#shell{{display:none!important;}}}}
</style></head>
<body oncontextmenu='return false;'>
<div id='shell'><div id='status' class='notice'>Đang tải tài liệu…</div></div>
<script>
const B64={safe_payload}; const MIME={safe_mime}; const MARK={safe_mark}; const IS_PDF={pdf_flag};
const shell=document.getElementById('shell'); const status=document.getElementById('status');
function lock(e){{e.preventDefault();e.stopPropagation();return false;}}
['copy','cut','dragstart','contextmenu'].forEach(ev=>document.addEventListener(ev,lock,true));
document.addEventListener('keydown',e=>{{const k=(e.key||'').toLowerCase();if((e.ctrlKey||e.metaKey)&&['c','s','p','u','a'].includes(k))lock(e);if(e.key==='PrintScreen')lock(e);}},true);
function wm(parent){{const w=document.createElement('div');w.className='watermark';const s=document.createElement('span');s.textContent=MARK;w.appendChild(s);parent.appendChild(w);}}
function bytesFromB64(s){{const bin=atob(s);const out=new Uint8Array(bin.length);for(let i=0;i<bin.length;i++)out[i]=bin.charCodeAt(i);return out;}}
function showError(msg){{status.textContent=msg;status.style.color='#b00020';}}
if(!IS_PDF){{status.remove();const page=document.createElement('div');page.className='page';const img=document.createElement('img');img.alt='Hướng dẫn sử dụng';img.draggable=false;img.src=`data:${{MIME}};base64,${{B64}}`;page.appendChild(img);wm(page);shell.appendChild(page);}}
else{{const sc=document.createElement('script');sc.src='https://cdnjs.cloudflare.com/ajax/libs/pdf.js/3.11.174/pdf.min.js';sc.onload=async()=>{{try{{pdfjsLib.GlobalWorkerOptions.workerSrc='https://cdnjs.cloudflare.com/ajax/libs/pdf.js/3.11.174/pdf.worker.min.js';const pdf=await pdfjsLib.getDocument({{data:bytesFromB64(B64)}}).promise;status.remove();for(let n=1;n<=pdf.numPages;n++){{const p=await pdf.getPage(n);const viewport=p.getViewport({{scale:1.55}});const wrap=document.createElement('div');wrap.className='page';const canvas=document.createElement('canvas');canvas.width=viewport.width;canvas.height=viewport.height;wrap.appendChild(canvas);wm(wrap);shell.appendChild(wrap);await p.render({{canvasContext:canvas.getContext('2d'),viewport}}).promise;}}}}catch(err){{showError('Không thể hiển thị PDF. Vui lòng báo Admin kiểm tra lại file.');}}}};sc.onerror=()=>showError('Không tải được thư viện hiển thị PDF. Hãy kiểm tra kết nối Internet.');document.head.appendChild(sc);}}
</script></body></html>""", height=880, scrolling=False)


# ==========================================================
# V85.2 - THỨ TỰ MENU CHỨC NĂNG RIÊNG CHO ADMIN
# Lưu vào Google Sheet để giữ nguyên sau reload/deploy.
# ==========================================================
@st.cache_resource(show_spinner=False)
def _ensure_admin_menu_config_storage():
    try:
        client = get_gspread_client()
        if not client:
            return None, "Chưa cấu hình quyền kết nối Google Sheets."
        ss = client.open_by_key(SHEET_MAT_KHAU_ID)
        ws = _get_or_create_worksheet(ss, ADMIN_MENU_CONFIG_WORKSHEET, rows=30, cols=6)
        if int(getattr(ws, "col_count", 0) or 0) < 4:
            _gs_call_with_backoff(ws.resize, cols=6)
        header = _gs_call_with_backoff(ws.row_values, 1)
        if header[:4] != ADMIN_MENU_CONFIG_HEADERS:
            gspread_update_range(ws, "A1:D1", [ADMIN_MENU_CONFIG_HEADERS])
        return ws, ""
    except Exception as e:
        return None, f"Lỗi khởi tạo cấu hình MENU admin: {e}"


@st.cache_data(ttl=300, show_spinner=False)
def load_admin_menu_order():
    ws, err = _ensure_admin_menu_config_storage()
    if err or ws is None:
        return [], err
    try:
        values = _gs_call_with_backoff(ws.get_all_values)
        for row in values[1:]:
            if not row or str(row[0]).strip() != ADMIN_MENU_CONFIG_KEY:
                continue
            try:
                raw = json.loads(row[1]) if len(row) > 1 and str(row[1]).strip() else []
            except Exception:
                raw = []
            if isinstance(raw, list):
                return [str(x) for x in raw if str(x).strip()], ""
        return [], ""
    except Exception as e:
        return [], f"Lỗi đọc thứ tự MENU admin: {e}"


def _clear_admin_menu_order_cache():
    try:
        load_admin_menu_order.clear()
    except Exception:
        pass


def save_admin_menu_order(order, username):
    ws, err = _ensure_admin_menu_config_storage()
    if err or ws is None:
        return False, err or "Không mở được sheet cấu hình MENU admin."
    try:
        clean_order = []
        seen = set()
        for item in order or []:
            item = str(item).strip()
            if item and item not in seen:
                clean_order.append(item)
                seen.add(item)

        values = _gs_call_with_backoff(ws.get_all_values)
        row_idx = None
        for idx, row in enumerate(values[1:], start=2):
            if row and str(row[0]).strip() == ADMIN_MENU_CONFIG_KEY:
                row_idx = idx
                break

        now = datetime.now(VN_TZ).strftime("%d/%m/%Y %H:%M:%S")
        row_value = [
            ADMIN_MENU_CONFIG_KEY,
            json.dumps(clean_order, ensure_ascii=False),
            now,
            str(username),
        ]
        if row_idx:
            gspread_update_range(ws, f"A{row_idx}:D{row_idx}", [row_value])
        else:
            _gs_call_with_backoff(ws.append_row, row_value, value_input_option="USER_ENTERED")
        _clear_admin_menu_order_cache()
        return True, "Đã lưu thứ tự MENU CHỨC NĂNG cho tài khoản Admin."
    except Exception as e:
        return False, f"Lỗi lưu thứ tự MENU admin: {e}"


def admin_menu_order_for_pages(default_pages):
    """Áp thứ tự đã lưu cho Admin; các role khác giữ nguyên thứ tự mặc định."""
    base = [str(x) for x in (default_pages or [])]
    if str(st.session_state.get("current_role", "")).strip().lower() != "admin":
        return base
    saved, _ = load_admin_menu_order()
    valid = [x for x in saved if x in base]
    missing = [x for x in base if x not in valid]
    return valid + missing


def _move_menu_item(order, item, target_index):
    """Di chuyển item đến vị trí target_index (0-based), các nút khác tự dồn lại."""
    arr = [x for x in order if x != item]
    target_index = max(0, min(int(target_index), len(arr)))
    arr.insert(target_index, item)
    return arr

# ==========================================================
# V71 - CẤU HÌNH GIAO DIỆN TOÀN HỆ THỐNG (ADMIN)
# Tiêu đề lớn / con / nhỏ / label-nội dung / bảng.
# Lưu vào Google Sheet CauHinhGiaoDien để toàn bộ tài khoản dùng cùng mặc định.
# ==========================================================
UI_THEME_HEADERS = ["ThemeKey", "Cấu hình JSON", "Cập nhật lúc", "Người cập nhật"]
UI_THEME_KEY = "global"
UI_THEME_GROUP_ORDER = ["main_title", "sub_title", "small_title", "label_content", "table"]
UI_THEME_GROUP_LABELS = {
    "main_title": "Tiêu đề lớn",
    "sub_title": "Tiêu đề con",
    "small_title": "Tiêu đề nhỏ",
    "label_content": "Label / Nội dung",
    "table": "Bảng",
}
UI_THEME_FONT_OPTIONS = [
    "Roboto", "Arial", "Tahoma", "Verdana", "Times New Roman", "Georgia",
    "Courier New", "Cinzel Decorative"
]
UI_THEME_EFFECT_OPTIONS = ["Không", "Bóng nhẹ", "Bóng nổi", "Hover nâng", "Gradient nhẹ"]
UI_THEME_DEFAULT = {
    "main_title": {"desktop_size": 28, "mobile_size": 28, "font_family": "Roboto", "text_color": "#222222", "bg_color": "#D9D9D9", "effect": "Không"},
    "sub_title": {"desktop_size": 22, "mobile_size": 22, "font_family": "Roboto", "text_color": "#222222", "bg_color": "#D9D9D9", "effect": "Không"},
    "small_title": {"desktop_size": 18, "mobile_size": 18, "font_family": "Roboto", "text_color": "#222222", "bg_color": "#D9D9D9", "effect": "Không"},
    "label_content": {"desktop_size": 16, "mobile_size": 16, "font_family": "Roboto", "text_color": "#333333", "bg_color": "#D9D9D9", "effect": "Không"},
    "table": {"desktop_size": 13, "mobile_size": 13, "font_family": "Roboto", "text_color": "#333333", "bg_color": "#D9D9D9", "effect": "Không"},
}


def _valid_theme_hex(value, fallback):
    value = str(value or "").strip()
    if re.fullmatch(r"#[0-9A-Fa-f]{6}", value):
        return value.upper()
    return str(fallback).upper()


def _normalized_theme_config(raw=None):
    raw = raw if isinstance(raw, dict) else {}
    result = {}
    for key in UI_THEME_GROUP_ORDER:
        base = dict(UI_THEME_DEFAULT[key])
        incoming = raw.get(key, {}) if isinstance(raw.get(key, {}), dict) else {}
        try:
            desktop_size = max(8, min(48, int(float(incoming.get("desktop_size", base["desktop_size"])))))
        except Exception:
            desktop_size = int(base["desktop_size"])
        try:
            mobile_size = max(8, min(48, int(float(incoming.get("mobile_size", base["mobile_size"])))))
        except Exception:
            mobile_size = int(base["mobile_size"])
        font_family = str(incoming.get("font_family", base["font_family"]))
        if font_family not in UI_THEME_FONT_OPTIONS:
            font_family = base["font_family"]
        effect = str(incoming.get("effect", base["effect"]))
        if effect not in UI_THEME_EFFECT_OPTIONS:
            effect = base["effect"]
        result[key] = {
            "desktop_size": desktop_size,
            "mobile_size": mobile_size,
            "font_family": font_family,
            "text_color": _valid_theme_hex(incoming.get("text_color"), base["text_color"]),
            "bg_color": _valid_theme_hex(incoming.get("bg_color"), base["bg_color"]),
            "effect": effect,
        }
    return result


@st.cache_resource(show_spinner=False)
def _ensure_ui_theme_storage():
    try:
        client = get_gspread_client()
        if not client:
            return None, "Chưa cấu hình quyền kết nối Google Sheets."
        ss = client.open_by_key(SHEET_MAT_KHAU_ID)
        ws = _get_or_create_worksheet(ss, UI_THEME_WORKSHEET, rows=20, cols=6)
        if int(getattr(ws, "col_count", 0) or 0) < 4:
            _gs_call_with_backoff(ws.resize, cols=6)
        header = _gs_call_with_backoff(ws.row_values, 1)
        if header[:4] != UI_THEME_HEADERS:
            gspread_update_range(ws, "A1:D1", [UI_THEME_HEADERS])
        return ws, ""
    except Exception as e:
        return None, f"Lỗi khởi tạo cấu hình giao diện: {e}"


@st.cache_data(ttl=300, show_spinner=False)
def load_ui_theme_config():
    ws, err = _ensure_ui_theme_storage()
    if err or ws is None:
        return _normalized_theme_config(), err
    try:
        values = _gs_call_with_backoff(ws.get_all_values)
        for row in values[1:]:
            if not row or str(row[0]).strip() != UI_THEME_KEY:
                continue
            try:
                raw = json.loads(row[1]) if len(row) > 1 and str(row[1]).strip() else {}
            except Exception:
                raw = {}
            return _normalized_theme_config(raw), ""
        return _normalized_theme_config(), ""
    except Exception as e:
        return _normalized_theme_config(), f"Lỗi đọc cấu hình giao diện: {e}"


def _clear_ui_theme_cache():
    try:
        load_ui_theme_config.clear()
    except Exception:
        pass


def save_ui_theme_config(config, username):
    ws, err = _ensure_ui_theme_storage()
    if err or ws is None:
        return False, err or "Không mở được sheet cấu hình giao diện."
    try:
        cfg = _normalized_theme_config(config)
        values = _gs_call_with_backoff(ws.get_all_values)
        row_idx = None
        for idx, row in enumerate(values[1:], start=2):
            if row and str(row[0]).strip() == UI_THEME_KEY:
                row_idx = idx
                break
        now = datetime.now(VN_TZ).strftime("%d/%m/%Y %H:%M:%S")
        row_value = [UI_THEME_KEY, json.dumps(cfg, ensure_ascii=False), now, str(username)]
        if row_idx:
            gspread_update_range(ws, f"A{row_idx}:D{row_idx}", [row_value])
        else:
            _gs_call_with_backoff(ws.append_row, row_value, value_input_option="USER_ENTERED")
        _clear_ui_theme_cache()
        return True, "Đã lưu giao diện làm mặc định cho toàn hệ thống."
    except Exception as e:
        return False, f"Lỗi lưu cấu hình giao diện: {e}"


def _theme_effect_css(effect, bg_color):
    effect = str(effect)
    if effect == "Bóng nhẹ":
        return "box-shadow:0 2px 6px rgba(0,0,0,.12)!important;"
    if effect == "Bóng nổi":
        return "box-shadow:0 5px 14px rgba(0,0,0,.18)!important;transform:translateY(-1px);"
    if effect == "Gradient nhẹ":
        return f"background:linear-gradient(135deg,{bg_color} 0%,rgba(255,255,255,.88) 100%)!important;"
    return ""


def _theme_hover_css(effect):
    if str(effect) == "Hover nâng":
        return "transform:translateY(-2px)!important;box-shadow:0 4px 12px rgba(0,0,0,.14)!important;"
    return ""


def render_global_ui_theme_css(config=None):
    """Áp dụng theme đã lưu cho cả desktop + mobile bằng CSS override cuối cùng."""
    cfg = _normalized_theme_config(config)
    selectors = {
        # Thêm selector có data-testid để thắng các rule màu/font mặc định của Streamlit.
        "main_title": "[data-testid='stMarkdownContainer'] h1,h1,.custom-main-title",
        "sub_title": "[data-testid='stMarkdownContainer'] h2,[data-testid='stMarkdownContainer'] h3,h2,h3",
        "small_title": "[data-testid='stMarkdownContainer'] h4,[data-testid='stMarkdownContainer'] h5,[data-testid='stMarkdownContainer'] h6,h4,h5,h6,[data-testid='stExpander'] details summary p",
        "label_content": "p,.stText,[data-testid='stMarkdownContainer'],[data-testid='stWidgetLabel'],[data-testid='stWidgetLabel'] p,[data-testid='stWidgetLabel'] span,[data-testid='stWidgetLabel'] div,input,textarea,button,[data-baseweb='select']",
        "table": "table,table th,table td,[data-testid='stDataFrame'],[data-testid='stDataEditor'],[data-testid='stTable']",
    }
    desktop_rules = []
    mobile_rules = []
    hover_rules = []
    for key in UI_THEME_GROUP_ORDER:
        item = cfg[key]
        selector = selectors[key]
        font = str(item["font_family"]).replace("'", "")
        color = item["text_color"]
        bg = item["bg_color"]
        effect_css = _theme_effect_css(item["effect"], bg)
        # Background applies to headings/labels/table headers, not table data cells or free paragraphs.
        if key == "label_content":
            bg_rule = ""
            desktop_rules.append(
                f"{selector}{{font-family:'{font}',sans-serif!important;font-size:{item['desktop_size']}px!important;color:{color}!important;}}"
            )
            desktop_rules.append(
                f"[data-testid='stWidgetLabel'],label[data-testid='stWidgetLabel'],div[data-testid='stWidgetLabel'],"
                f"[data-testid='stSelectbox']>label,[data-testid='stMultiSelect']>label,[data-testid='stDateInput']>label,"
                f"[data-testid='stTextInput']>label,[data-testid='stTextArea']>label,[data-testid='stNumberInput']>label,"
                f"[data-testid='stFileUploader']>label,[data-testid='stRadio']>label,[data-testid='stCheckbox']>label,fieldset>legend"
                f"{{background:{bg}!important;color:{color}!important;{effect_css}}}"
            )
            desktop_rules.append(
                f"[data-testid='stWidgetLabel'] p,[data-testid='stWidgetLabel'] span,[data-testid='stWidgetLabel'] div,fieldset>legend"
                f"{{color:{color}!important;font-family:'{font}',sans-serif!important;font-size:{item['desktop_size']}px!important;}}"
            )
        elif key == "table":
            desktop_rules.append(
                f"{selector}{{font-family:'{font}',sans-serif!important;font-size:{item['desktop_size']}px!important;color:{color}!important;}}"
            )
            desktop_rules.append(
                f"table th,[data-testid='stDataFrame'] [role='columnheader'],[data-testid='stDataEditor'] [role='columnheader']"
                f"{{background:{bg}!important;color:{color}!important;{effect_css}}}"
            )
        else:
            desktop_rules.append(
                f"{selector}{{font-family:'{font}',sans-serif!important;font-size:{item['desktop_size']}px!important;color:{color}!important;background:{bg}!important;{effect_css}transition:transform .15s ease,box-shadow .15s ease!important;}}"
            )
        if item["effect"] == "Hover nâng":
            hover_rules.append(f"{selector}:hover{{{_theme_hover_css(item['effect'])}}}")
        mobile_rules.append(f"{selector}{{font-size:{item['mobile_size']}px!important;}}")

    # Mobile: giảm padding/gap chứ không ép giảm cỡ chữ; Admin có cột riêng để tự đặt nếu muốn.
    css = "\n".join(desktop_rules + hover_rules)
    mobile_css = "\n".join(mobile_rules)
    st.markdown(f"""
<style id="vera-global-theme-v71">
{css}
@media (max-width:768px) {{
{mobile_css}
  h1,h2,h3,h4,h5,h6,.custom-main-title {{padding:.32rem .48rem!important;margin-top:.18rem!important;margin-bottom:.3rem!important;overflow-wrap:anywhere!important;}}
  [data-testid='stWidgetLabel'] {{padding:.22rem .4rem!important;overflow-wrap:anywhere!important;}}
  .block-container {{padding-left:.4rem!important;padding-right:.4rem!important;}}
  table th,table td {{padding-left:4px!important;padding-right:4px!important;}}
}}
</style>
""", unsafe_allow_html=True)


def _theme_editor_df(config):
    cfg = _normalized_theme_config(config)
    rows = []
    for key in UI_THEME_GROUP_ORDER:
        item = cfg[key]
        rows.append({
            "Nhóm": UI_THEME_GROUP_LABELS[key],
            "Font chữ": item["font_family"],
            "Cỡ chữ Web": item["desktop_size"],
            "Cỡ chữ Mobile": item["mobile_size"],
            "Màu chữ": item["text_color"],
            "Màu nền": item["bg_color"],
            "Hiệu ứng": item["effect"],
        })
    return pd.DataFrame(rows)


def _theme_from_editor_df(df):
    result = {}
    label_to_key = {v: k for k, v in UI_THEME_GROUP_LABELS.items()}
    if not isinstance(df, pd.DataFrame):
        return _normalized_theme_config()
    for _, row in df.iterrows():
        key = label_to_key.get(str(row.get("Nhóm", "")).strip())
        if not key:
            continue
        result[key] = {
            "font_family": row.get("Font chữ", "Roboto"),
            "desktop_size": row.get("Cỡ chữ Web", UI_THEME_DEFAULT[key]["desktop_size"]),
            "mobile_size": row.get("Cỡ chữ Mobile", UI_THEME_DEFAULT[key]["mobile_size"]),
            "text_color": row.get("Màu chữ", UI_THEME_DEFAULT[key]["text_color"]),
            "bg_color": row.get("Màu nền", UI_THEME_DEFAULT[key]["bg_color"]),
            "effect": row.get("Hiệu ứng", "Không"),
        }
    return _normalized_theme_config(result)


def render_admin_theme_config_panel():
    """Bảng tùy chỉnh font/cỡ/màu/hiệu ứng, chỉ Admin mới nhìn thấy và lưu được."""
    if st.session_state.get("current_role") != "admin":
        return
    current, err = load_ui_theme_config()
    with st.expander("🎨 Cấu hình Font · Màu · Cỡ chữ · Hiệu ứng", expanded=True):
        st.caption(
            "Mặc định: Tiêu đề lớn 28px · Tiêu đề con 22px · Tiêu đề nhỏ 18px · "
            "Label/Nội dung 16px · Bảng 13px. Có cột riêng cho Web và Mobile. "
            "Màu nhập theo mã HEX, ví dụ #D9D9D9."
        )
        if err:
            st.warning(err)
        editor_key = "admin_global_theme_editor_v71"
        edited = st.data_editor(
            _theme_editor_df(current),
            key=editor_key,
            hide_index=True,
            num_rows="fixed",
            width="stretch",
            height="content",
            disabled=["Nhóm"],
            row_height=42,
            column_config={
                "Nhóm": st.column_config.TextColumn("Nhóm", disabled=True, width=150),
                "Font chữ": st.column_config.SelectboxColumn("Font chữ", options=UI_THEME_FONT_OPTIONS, width=145),
                "Cỡ chữ Web": st.column_config.NumberColumn("Cỡ chữ Web", min_value=8, max_value=48, step=1, format="%d", width=105),
                "Cỡ chữ Mobile": st.column_config.NumberColumn("Cỡ chữ Mobile", min_value=8, max_value=48, step=1, format="%d", width=120),
                "Màu chữ": st.column_config.TextColumn("Màu chữ", width=105, help="HEX: #RRGGBB"),
                "Màu nền": st.column_config.TextColumn("Màu nền", width=105, help="HEX: #RRGGBB"),
                "Hiệu ứng": st.column_config.SelectboxColumn("Hiệu ứng", options=UI_THEME_EFFECT_OPTIONS, width=125),
            },
        )
        preview_cfg = _theme_from_editor_df(edited)
        # Preview dùng HTML đơn giản, không ghi dữ liệu cho tới khi bấm Lưu.
        p1, p2, p3 = preview_cfg["main_title"], preview_cfg["sub_title"], preview_cfg["small_title"]
        st.markdown(
            f"<div style=\"font-family:'{p1['font_family']}';font-size:{p1['desktop_size']}px;color:{p1['text_color']};background:{p1['bg_color']};padding:6px 10px;border-radius:6px;margin:3px 0;\">Tiêu đề lớn – Xem trước</div>"
            f"<div style=\"font-family:'{p2['font_family']}';font-size:{p2['desktop_size']}px;color:{p2['text_color']};background:{p2['bg_color']};padding:5px 9px;border-radius:6px;margin:3px 0;\">Tiêu đề con – Xem trước</div>"
            f"<div style=\"font-family:'{p3['font_family']}';font-size:{p3['desktop_size']}px;color:{p3['text_color']};background:{p3['bg_color']};padding:4px 8px;border-radius:6px;margin:3px 0;\">Tiêu đề nhỏ – Xem trước</div>",
            unsafe_allow_html=True,
        )
        c1, c2 = st.columns(2)
        with c1:
            if st.button("💾 Lưu giao diện làm mặc định", use_container_width=True, key="save_global_theme_v71"):
                ok, msg = save_ui_theme_config(preview_cfg, st.session_state.current_user)
                (st.success if ok else st.error)(msg)
                if ok:
                    st.rerun()
        with c2:
            if st.button("♻️ Khôi phục 28 · 22 · 18 · 16 · 13", use_container_width=True, key="reset_global_theme_v71"):
                ok, msg = save_ui_theme_config(UI_THEME_DEFAULT, st.session_state.current_user)
                (st.success if ok else st.error)(msg)
                if ok:
                    st.rerun()


# ==========================================================
# CẤU HÌNH HIỂN THỊ CỘT TOÀN HỆ THỐNG
# V37: thứ tự + độ rộng + độ cao dòng + font + kiểu chữ + căn lề + wrap text
# ==========================================================
TABLE_LAYOUT_LABELS = {
    "tour_main": "Bảng tour",
    "staff_list": "Danh sách nhân sự",
    "payroll_current": "Bảng lương",
    "payroll_history": "Bảng lương đã lưu / chỉnh sửa",
    "leave_detail": "Chi tiết danh sách nghỉ",
    "leave_manage": "Quản lý lịch nghỉ",
}
TABLE_LAYOUT_STATIC_COLUMNS = {
    "staff_list": [
        "Tên nhân viên", "Họ và tên đầy đủ", "Phân quyền", "Trạng thái làm việc", "Điện thoại", "Email",
        "Địa chỉ", "Số tài khoản ngân hàng", "Tên ngân hàng", "Khóa đăng nhập"
    ],
    "payroll_current": [
        "TT", "Tên Hệ thống", "Tiền Lương", "Tiền Hỗ Trợ Hoàn Lại", "Tích lũy",
        "Chi Phí Sinh Hoạt", "Tiền phạt trong tháng", "Vi phạm kỳ trước", "Tiền ứng lương",
        "Tiền hỗ trợ Locker", "Số tiền thực nhận", "Số tài khoản ngân hàng", "Tên ngân hàng", "Email"
    ],
    "payroll_history": [
        "TT", "Tên Hệ thống", "Tiền Lương", "Tiền Hỗ Trợ Hoàn Lại", "Tích lũy",
        "Chi Phí Sinh Hoạt", "Tiền phạt trong tháng", "Vi phạm kỳ trước", "Tiền ứng lương",
        "Tiền hỗ trợ Locker", "Số tiền thực nhận", "Số tài khoản ngân hàng", "Tên ngân hàng", "Email"
    ],
    "leave_detail": [
        "Chọn", "Ngày", "Thứ ngày", "Tên nhân viên", "Lý do nghỉ", "Loại nghỉ", "Chi tiết", "Số ngày tính",
        "Số ngày phép cộng dồn", "Phạt vi phạm", "Ngày cập nhật", "Giờ cập nhật", "Người cập nhật"
    ],
    "leave_manage": [
        "Ngày", "Thứ ngày", "Tên nhân viên", "Lý do nghỉ", "Loại nghỉ", "Chi tiết", "Số ngày tính",
        "Số ngày phép cộng dồn", "Phạt vi phạm", "Ngày cập nhật", "Giờ cập nhật", "Người cập nhật"
    ],
}

TABLE_LAYOUT_FONT_OPTIONS = [
    "Roboto", "Arial", "Tahoma", "Verdana", "Times New Roman", "Georgia", "Courier New"
]
TABLE_LAYOUT_FONT_STYLE_OPTIONS = ["Thường", "Đậm", "Nghiêng", "Đậm + Nghiêng"]
TABLE_LAYOUT_ALIGN_OPTIONS = ["left", "center", "right"]
TABLE_LAYOUT_DEFAULT_ROW_HEIGHT = 36
TABLE_LAYOUT_DEFAULT_FONT_SIZE = 13


def _default_column_width(column_name):
    name = str(column_name)
    if name in {"TT", "Chọn"}: return 65
    if name in {"Ngày", "Giờ cập nhật", "Ngày cập nhật"}: return 115
    if name == "Thứ ngày": return 95
    if name in {"Tên nhân viên", "Tên Hệ thống", "Phân quyền"}: return 150
    if "Email" in name: return 210
    if "Địa chỉ" in name or "Chi tiết" in name: return 240
    if "ngân hàng" in name.casefold(): return 190
    if any(x in name for x in ["Tiền", "Phạt", "Tích lũy", "Phí", "Số ngày"]): return 125
    return 140


def _default_column_alignment(column_name):
    name = str(column_name)
    if name in {"TT", "Chọn"}: return "center"
    if any(x in name for x in ["Tiền", "Phạt", "Tích lũy", "Phí", "Số ngày", "Thời gian còn lại"]):
        return "right"
    return "left"


def _default_column_visual_style(column_name):
    return {
        "font_family": "Roboto",
        "font_size": TABLE_LAYOUT_DEFAULT_FONT_SIZE,
        "font_style": "Thường",
        "align": _default_column_alignment(column_name),
        "wrap": True,
    }


@st.cache_resource(show_spinner=False)
def _ensure_ui_layout_storage():
    try:
        client = get_gspread_client()
        if not client:
            return None, "Chưa cấu hình quyền kết nối Google Sheets."
        ss = client.open_by_key(SHEET_MAT_KHAU_ID)
        ws = _get_or_create_worksheet(ss, UI_LAYOUT_WORKSHEET, rows=100, cols=8)
        if int(getattr(ws, "col_count", 0) or 0) < 7:
            _gs_call_with_backoff(ws.resize, cols=8)
        header = _gs_call_with_backoff(ws.row_values, 1)
        # Giữ 6 cột cũ để tương thích, thêm JSON kiểu hiển thị ở cột G.
        wanted = [
            "TableKey", "Tên bảng", "Thứ tự cột JSON", "Độ rộng cột JSON",
            "Cập nhật lúc", "Người cập nhật", "Kiểu hiển thị JSON"
        ]
        if header[:len(wanted)] != wanted:
            gspread_update_range(ws, "A1:G1", [wanted])
        return ws, ""
    except Exception as e:
        return None, f"Lỗi khởi tạo giao diện tùy chỉnh: {e}"


@st.cache_data(ttl=300, show_spinner=False)
def load_table_layouts():
    ws, err = _ensure_ui_layout_storage()
    if err or ws is None:
        return {}, err
    try:
        values = _gs_call_with_backoff(ws.get_all_values)
        result = {}
        for row_idx, row in enumerate(values[1:], start=2):
            if not row or not str(row[0]).strip():
                continue
            key = str(row[0]).strip()
            try:
                order = json.loads(row[2]) if len(row) > 2 and str(row[2]).strip() else []
            except Exception:
                order = []
            try:
                widths = json.loads(row[3]) if len(row) > 3 and str(row[3]).strip() else {}
            except Exception:
                widths = {}
            try:
                visual = json.loads(row[6]) if len(row) > 6 and str(row[6]).strip() else {}
            except Exception:
                visual = {}
            result[key] = {
                "row": row_idx,
                "order": order if isinstance(order, list) else [],
                "widths": widths if isinstance(widths, dict) else {},
                "visual": visual if isinstance(visual, dict) else {},
                "updated_at": str(row[4]).strip() if len(row) > 4 else "",
                "updated_by": str(row[5]).strip() if len(row) > 5 else "",
            }
        return result, ""
    except Exception as e:
        return {}, f"Lỗi đọc giao diện tùy chỉnh: {e}"


def _clear_table_layout_cache():
    try:
        load_table_layouts.clear()
    except Exception:
        pass


def get_table_layout(table_key, available_columns):
    available = [str(c) for c in available_columns]
    layouts, _ = load_table_layouts()
    cfg = layouts.get(str(table_key), {})
    saved_order = [str(c) for c in cfg.get("order", []) if str(c) in available]
    order = saved_order + [c for c in available if c not in saved_order]
    # V86.1: cột Loại nghỉ lấy từ cột C của sheet LoaiNghi; Lý do nghỉ đối chiếu theo cột B.
    # Nếu cấu hình cột cũ chưa biết cột mới, vẫn ép vị trí mặc định ngay sau Lý do nghỉ.
    if str(table_key) in {"leave_detail", "leave_manage"} and "Thứ ngày" in order and "Ngày" in order:
        order = [c for c in order if c != "Thứ ngày"]
        order.insert(order.index("Ngày") + 1, "Thứ ngày")
    if str(table_key) in {"leave_detail", "leave_manage"} and "Loại nghỉ" in order and "Lý do nghỉ" in order:
        order = [c for c in order if c != "Loại nghỉ"]
        order.insert(order.index("Lý do nghỉ") + 1, "Loại nghỉ")
    saved_widths = cfg.get("widths", {}) if isinstance(cfg.get("widths", {}), dict) else {}
    widths = {}
    for c in available:
        try:
            widths[c] = max(50, min(800, int(float(saved_widths.get(c, _default_column_width(c))))))
        except Exception:
            widths[c] = _default_column_width(c)
    return order, widths


def get_table_visual_settings(table_key, available_columns):
    """Đọc cấu hình kiểu hiển thị của bảng, tự bù mặc định cho cột mới."""
    available = [str(c) for c in available_columns]
    layouts, _ = load_table_layouts()
    cfg = layouts.get(str(table_key), {})
    raw = cfg.get("visual", {}) if isinstance(cfg.get("visual", {}), dict) else {}
    try:
        row_height = int(float(raw.get("row_height", TABLE_LAYOUT_DEFAULT_ROW_HEIGHT)))
    except Exception:
        row_height = TABLE_LAYOUT_DEFAULT_ROW_HEIGHT
    row_height = max(24, min(120, row_height))
    raw_cols = raw.get("columns", {}) if isinstance(raw.get("columns", {}), dict) else {}
    col_styles = {}
    for c in available:
        d = _default_column_visual_style(c)
        saved = raw_cols.get(c, {}) if isinstance(raw_cols.get(c, {}), dict) else {}
        font_family = str(saved.get("font_family", d["font_family"]))
        if font_family not in TABLE_LAYOUT_FONT_OPTIONS:
            font_family = d["font_family"]
        try:
            font_size = max(8, min(30, int(float(saved.get("font_size", d["font_size"])))))
        except Exception:
            font_size = d["font_size"]
        font_style = str(saved.get("font_style", d["font_style"]))
        if font_style not in TABLE_LAYOUT_FONT_STYLE_OPTIONS:
            font_style = d["font_style"]
        align = str(saved.get("align", d["align"])).lower()
        if align not in TABLE_LAYOUT_ALIGN_OPTIONS:
            align = d["align"]
        wrap = saved.get("wrap", d["wrap"])
        if isinstance(wrap, str):
            wrap = wrap.strip().lower() in {"1", "true", "yes", "y", "có", "co"}
        else:
            wrap = bool(wrap)
        col_styles[c] = {
            "font_family": font_family,
            "font_size": font_size,
            "font_style": font_style,
            "align": align,
            "wrap": wrap,
        }
    return row_height, col_styles


def layout_row_height(table_key, fallback=TABLE_LAYOUT_DEFAULT_ROW_HEIGHT):
    try:
        row_height, _ = get_table_visual_settings(table_key, [])
        return int(row_height)
    except Exception:
        return int(fallback)


def _font_style_css(style_name):
    style_name = str(style_name)
    if style_name == "Đậm":
        return "700", "normal"
    if style_name == "Nghiêng":
        return "400", "italic"
    if style_name == "Đậm + Nghiêng":
        return "700", "italic"
    return "400", "normal"



def neutralize_khong_phep_rows_styler(data_or_styler, reason_col="Lý do nghỉ"):
    """Legacy helper. V86.3 không còn dùng helper này cho Chi tiết/Quản lý lịch nghỉ.

    Không tô nền, không đổi màu chữ, không bold/italic/underline. Hàm này được áp dụng
    sau cấu hình hiển thị cột để chắc chắn mọi conditional formatting cũ không còn ảnh hưởng.
    """
    if isinstance(data_or_styler, pd.DataFrame):
        styler = data_or_styler.style
        data = data_or_styler
    else:
        styler = data_or_styler
        try:
            data = styler.data
        except Exception:
            return data_or_styler

    if not isinstance(data, pd.DataFrame) or data.empty or reason_col not in data.columns:
        return styler

    def _neutral_row(row):
        reason_key = remove_vietnamese_accents(str(row.get(reason_col, ""))).casefold()
        if "khong phep" not in reason_key:
            return [""] * len(row)
        # Ép màu chữ tối dễ đọc thay vì ``inherit``. Một số theme/Styler cũ có thể
        # để màu kế thừa thành trắng hoặc xám quá sáng, khiến người dùng tưởng dữ liệu mất.
        neutral_css = (
            "background-color: transparent !important;"
            "color: #262730 !important;"
            "font-weight: normal !important;"
            "font-style: normal !important;"
            "text-decoration: none !important;"
            "opacity: 1 !important;"
        )
        return [neutral_css] * len(row)

    return styler.apply(_neutral_row, axis=1)

def apply_table_visual_styler(data_or_styler, table_key, columns=None):
    """
    Áp dụng font/cỡ chữ/kiểu chữ/căn lề/wrap cho bảng đọc (st.dataframe).
    Giữ nguyên Styler sẵn có để không làm mất màu điều kiện của Tour/Lịch nghỉ.
    """
    if isinstance(data_or_styler, pd.DataFrame):
        columns = list(data_or_styler.columns) if columns is None else list(columns)
        styler = data_or_styler.style
    else:
        styler = data_or_styler
        if columns is None:
            try:
                columns = list(styler.data.columns)
            except Exception:
                columns = []
        else:
            columns = list(columns)
    _, col_styles = get_table_visual_settings(table_key, columns)
    header_rules = []
    for idx, c in enumerate(columns):
        if c not in col_styles:
            continue
        cfg = col_styles[c]
        fw, fs = _font_style_css(cfg.get("font_style"))
        wrap = bool(cfg.get("wrap", True))
        props = {
            "font-family": f"'{cfg.get('font_family', 'Roboto')}', sans-serif",
            "font-size": f"{int(cfg.get('font_size', TABLE_LAYOUT_DEFAULT_FONT_SIZE))}px",
            "font-weight": fw,
            "font-style": fs,
            "text-align": cfg.get("align", "left"),
            "white-space": "normal" if wrap else "nowrap",
            "overflow-wrap": "anywhere" if wrap else "normal",
            "word-break": "break-word" if wrap else "normal",
        }
        try:
            styler = styler.set_properties(subset=[c], **props)
        except Exception:
            pass
        # Dòng tiêu đề cũng dùng font/căn lề theo cột và luôn cho phép wrap nếu cột hẹp.
        header_rules.append({
            "selector": f"th.col_heading.level0.col{idx}",
            "props": [
                ("font-family", f"'{cfg.get('font_family', 'Roboto')}', sans-serif"),
                ("font-size", f"{int(cfg.get('font_size', TABLE_LAYOUT_DEFAULT_FONT_SIZE))}px"),
                ("font-weight", fw), ("font-style", fs),
                ("text-align", cfg.get("align", "left")),
                ("white-space", "normal"), ("overflow-wrap", "anywhere"),
                ("word-break", "break-word"), ("line-height", "1.15"),
            ]
        })
    try:
        if header_rules:
            styler = styler.set_table_styles(header_rules, overwrite=False)
    except Exception:
        pass
    return styler


def table_layout_html_css(table_key, columns, selector):
    """Sinh CSS theo từng cột cho các bảng HTML (hiện dùng ở Bảng lương tổng hợp)."""
    row_height, col_styles = get_table_visual_settings(table_key, columns)
    css = [f"{selector} tbody tr{{height:{int(row_height)}px;}}"]
    for idx, c in enumerate(columns, start=1):
        cfg = col_styles.get(c, _default_column_visual_style(c))
        fw, fs = _font_style_css(cfg.get("font_style"))
        wrap = bool(cfg.get("wrap", True))
        white = "normal" if wrap else "nowrap"
        overflow = "anywhere" if wrap else "normal"
        word_break = "break-word" if wrap else "normal"
        font = str(cfg.get("font_family", "Roboto")).replace("'", "")
        size = int(cfg.get("font_size", TABLE_LAYOUT_DEFAULT_FONT_SIZE))
        align = cfg.get("align", "left")
        css.append(
            f"{selector} th:nth-child({idx}),{selector} td:nth-child({idx}){{"
            f"font-family:'{font}',sans-serif!important;font-size:{size}px!important;"
            f"font-weight:{fw}!important;font-style:{fs}!important;text-align:{align}!important;"
            f"white-space:{white}!important;overflow-wrap:{overflow}!important;word-break:{word_break}!important;}}"
        )
        # Tiêu đề luôn wrap để không bị mất chữ khi cột nhỏ.
        css.append(
            f"{selector} th:nth-child({idx}){{white-space:nowrap!important;overflow-wrap:normal!important;word-break:normal!important;}}"
        )
    return "\n".join(css)


def apply_table_layout_df(df, table_key):
    if not isinstance(df, pd.DataFrame):
        return df, {}
    order, widths = get_table_layout(table_key, list(df.columns))
    return df[order].copy(), widths


def table_layout_column_config(table_key, columns, label_map=None):
    _, widths = get_table_layout(table_key, columns)
    label_map = label_map or {}
    cfg = {}
    for c in columns:
        try:
            cfg[c] = st.column_config.Column(label_map.get(c, c), width=int(widths.get(c, _default_column_width(c))))
        except Exception:
            cfg[c] = st.column_config.TextColumn(label_map.get(c, c), width="medium")
    return cfg


def layout_width(table_key, column_name, fallback=None):
    _, widths = get_table_layout(table_key, [column_name])
    value = int(widths.get(column_name, _default_column_width(column_name)))
    return value if value else (fallback or "medium")


def save_table_layout_config(table_key, order, widths, username, visual=None):
    ws, err = _ensure_ui_layout_storage()
    if err or ws is None:
        return False, err or "Không mở được sheet giao diện tùy chỉnh."
    try:
        layouts, _ = load_table_layouts()
        cfg = layouts.get(table_key, {})
        row_idx = cfg.get("row")
        now = datetime.now(VN_TZ).strftime("%d/%m/%Y %H:%M:%S")
        if visual is None:
            visual = cfg.get("visual", {}) if isinstance(cfg.get("visual", {}), dict) else {}
        values = [[
            table_key, TABLE_LAYOUT_LABELS.get(table_key, table_key),
            json.dumps(list(order), ensure_ascii=False),
            json.dumps({str(k): int(v) for k, v in widths.items()}, ensure_ascii=False),
            now, str(username), json.dumps(visual, ensure_ascii=False)
        ]]
        if row_idx:
            gspread_update_range(ws, f"A{row_idx}:G{row_idx}", values)
        else:
            _gs_call_with_backoff(ws.append_row, values[0], value_input_option="USER_ENTERED")
        _clear_table_layout_cache()
        return True, "Đã lưu cấu hình hiển thị và áp dụng cho toàn hệ thống."
    except Exception as e:
        return False, f"Lỗi lưu giao diện tùy chỉnh: {e}"


def render_admin_quick_layout_default(table_key, columns, key_suffix=""):
    """
    Nút lưu bố cục ngay tại từng bảng, chỉ Admin.
    Streamlit không trả lại kích thước cột khi người dùng kéo trực tiếp bằng chuột, vì vậy
    khối này cung cấp chỉnh nhanh Vị trí/Độ rộng ngay cạnh bảng rồi lưu làm mặc định toàn hệ thống.
    """
    if st.session_state.get('current_role') != 'admin':
        return
    cols = [str(c) for c in (columns or []) if str(c)]
    if not cols:
        return
    safe_suffix = re.sub(r'[^a-zA-Z0-9_]+', '_', str(key_suffix or table_key))
    with st.expander("⭐ Lưu bố cục bảng này làm mặc định", expanded=False):
        components.html(r"""
        <script>
        (function(){
          try {
            const doc = window.parent.document;
            const wanted = '⭐ Lưu bố cục bảng này làm mặc định';
            const nodes = doc.querySelectorAll('[data-testid="stExpander"] summary, details summary');
            nodes.forEach((node) => {
              if ((node.innerText || '').trim().includes(wanted)) {
                node.style.fontSize = '13px';
                node.querySelectorAll('p,span').forEach((x) => { x.style.fontSize = '13px'; });
              }
            });
          } catch (e) {}
        })();
        </script>
        """, height=0, width=0)
        st.caption("Chỉnh nhanh thứ tự/độ rộng rồi lưu. Font, căn lề, Wrap Text và độ cao dòng được giữ theo cấu hình hiện tại.")
        order, widths = get_table_layout(table_key, cols)
        quick_df = pd.DataFrame([
            {"Tên cột": c, "Vị trí": i, "Độ rộng (px)": int(widths.get(c, _default_column_width(c)))}
            for i, c in enumerate(order, start=1)
        ])
        quick_edit = st.data_editor(
            quick_df, hide_index=True, num_rows='fixed', width='stretch', height='content',
            disabled=['Tên cột'], key=f"quick_layout_{table_key}_{safe_suffix}",
            column_config={
                'Tên cột': st.column_config.TextColumn('Tên cột', disabled=True, width=220),
                'Vị trí': st.column_config.NumberColumn('Vị trí', min_value=1, max_value=max(1, len(quick_df)), step=1, format='%d', width=85),
                'Độ rộng (px)': st.column_config.NumberColumn('Độ rộng (px)', min_value=50, max_value=800, step=10, format='%d', width=120),
            }
        )
        if st.button("💾 Lưu làm mặc định", use_container_width=True, key=f"quick_layout_save_{table_key}_{safe_suffix}"):
            rows = quick_edit.to_dict('records')
            rows = sorted(rows, key=lambda r: (int(float(r.get('Vị trí', 9999) or 9999)), normalize_login_name(r.get('Tên cột',''))))
            new_order = [str(r.get('Tên cột','')).strip() for r in rows if str(r.get('Tên cột','')).strip()]
            new_widths = {
                str(r.get('Tên cột','')).strip(): max(50, min(800, int(float(r.get('Độ rộng (px)', 140) or 140))))
                for r in rows if str(r.get('Tên cột','')).strip()
            }
            row_height, styles = get_table_visual_settings(table_key, new_order)
            visual = {"row_height": row_height, "columns": styles}
            ok, msg = save_table_layout_config(table_key, new_order, new_widths, st.session_state.current_user, visual=visual)
            (st.success if ok else st.error)(msg)
            if ok:
                st.rerun()


def get_table_columns_for_settings(table_key):
    if table_key == "tour_main":
        try:
            dft, _ = load_bang_tour_input()
            if isinstance(dft, pd.DataFrame) and not dft.empty:
                return list(prepare_bang_tour_display(dft).columns)
        except Exception:
            pass
        return ["Tên Nhân Viên", "Trạng Thái", "Thời gian còn lại", "Đi làm", "Vào ca", "Break"]
    return list(TABLE_LAYOUT_STATIC_COLUMNS.get(table_key, []))


def _layout_editor_rows_from_saved(table_key, available_cols):
    current_order, current_widths = get_table_layout(table_key, available_cols)
    _, current_styles = get_table_visual_settings(table_key, current_order)
    rows = []
    for pos, col in enumerate(current_order, start=1):
        vis = current_styles.get(col, _default_column_visual_style(col))
        rows.append({
            "Tên cột": col,
            "Vị trí": pos,
            "Độ rộng (px)": int(current_widths.get(col, _default_column_width(col))),
            "Font chữ": vis.get("font_family", "Roboto"),
            "Cỡ chữ": int(vis.get("font_size", TABLE_LAYOUT_DEFAULT_FONT_SIZE)),
            "Kiểu chữ": vis.get("font_style", "Thường"),
            "Căn lề": vis.get("align", _default_column_alignment(col)),
            "Wrap text": bool(vis.get("wrap", True)),
        })
    return rows


def _layout_editor_on_change(table_key, editor_key, rows_state_key, version_state_key):
    """
    V37 - Đổi vị trí theo logic chèn:
    ví dụ cột ở vị trí 6 chuyển thành 3 => vị trí 3,4,5 cũ tự thành 4,5,6.
    Không bao giờ để trùng Vị trí.
    """
    rows = [dict(r) for r in st.session_state.get(rows_state_key, [])]
    if not rows:
        return
    widget_state = st.session_state.get(editor_key, {})
    edits = widget_state.get("edited_rows", {}) if isinstance(widget_state, dict) else {}
    if not isinstance(edits, dict) or not edits:
        return

    # Map row index -> tên cột từ trạng thái trước khi di chuyển.
    original_names = {i: str(r.get("Tên cột", "")) for i, r in enumerate(rows)}

    # Áp dụng các chỉnh sửa không phải vị trí trước.
    for idx_raw, changes in edits.items():
        try:
            idx = int(idx_raw)
        except Exception:
            continue
        if idx < 0 or idx >= len(rows) or not isinstance(changes, dict):
            continue
        target_name = original_names.get(idx, "")
        target_row = next((r for r in rows if str(r.get("Tên cột", "")) == target_name), None)
        if target_row is None:
            continue
        for field, value in changes.items():
            if field == "Vị trí":
                continue
            target_row[field] = value

    # Xử lý vị trí theo cơ chế remove + insert, rồi đánh lại 1..N.
    for idx_raw, changes in edits.items():
        if not isinstance(changes, dict) or "Vị trí" not in changes:
            continue
        try:
            idx = int(idx_raw)
            target_pos = int(float(changes.get("Vị trí")))
        except Exception:
            continue
        target_name = original_names.get(idx, "")
        if not target_name:
            continue
        ordered = sorted(rows, key=lambda r: int(float(r.get("Vị trí", 9999))))
        current_idx = next((i for i, r in enumerate(ordered) if str(r.get("Tên cột", "")) == target_name), None)
        if current_idx is None:
            continue
        target_pos = max(1, min(len(ordered), target_pos))
        item = ordered.pop(current_idx)
        ordered.insert(target_pos - 1, item)
        for pos, r in enumerate(ordered, start=1):
            r["Vị trí"] = pos
        rows = ordered

    # Chuẩn hóa kiểu dữ liệu.
    for pos, r in enumerate(sorted(rows, key=lambda x: int(float(x.get("Vị trí", 9999)))), start=1):
        r["Vị trí"] = pos
        try: r["Độ rộng (px)"] = max(50, min(800, int(float(r.get("Độ rộng (px)", 140)))))
        except Exception: r["Độ rộng (px)"] = 140
        try: r["Cỡ chữ"] = max(8, min(30, int(float(r.get("Cỡ chữ", TABLE_LAYOUT_DEFAULT_FONT_SIZE)))))
        except Exception: r["Cỡ chữ"] = TABLE_LAYOUT_DEFAULT_FONT_SIZE
        if str(r.get("Font chữ", "Roboto")) not in TABLE_LAYOUT_FONT_OPTIONS: r["Font chữ"] = "Roboto"
        if str(r.get("Kiểu chữ", "Thường")) not in TABLE_LAYOUT_FONT_STYLE_OPTIONS: r["Kiểu chữ"] = "Thường"
        if str(r.get("Căn lề", "left")) not in TABLE_LAYOUT_ALIGN_OPTIONS: r["Căn lề"] = "left"
        r["Wrap text"] = bool(r.get("Wrap text", True))

    st.session_state[rows_state_key] = sorted(rows, key=lambda x: int(x["Vị trí"]))
    st.session_state[version_state_key] = int(st.session_state.get(version_state_key, 0) or 0) + 1


@st.cache_resource(show_spinner=False)
def _ensure_payroll_storage():
    """
    Tạo/lấy sheet lưu lương + cấu hình CHỈ MỘT LẦN cho mỗi tiến trình Streamlit.
    V27 gọi hàm này lặp lại trong cùng một rerun, làm phát sinh nhiều request metadata/read.
    """
    client = get_gspread_client()
    if not client:
        return None, None, "Chưa cấu hình quyền kết nối Google Sheets."
    try:
        ss = client.open_by_key(SHEET_MAT_KHAU_ID)
        ws_pay = _get_or_create_worksheet(ss, PAYROLL_STORAGE_WORKSHEET, rows=3000, cols=30)
        ws_cfg = _get_or_create_worksheet(ss, PAYROLL_CONFIG_WORKSHEET, rows=30, cols=5)

        pay_header = _gs_call_with_backoff(ws_pay.row_values, 1)
        if not pay_header or pay_header[:len(PAYROLL_HISTORY_HEADERS)] != PAYROLL_HISTORY_HEADERS:
            gspread_update_range(ws_pay, "A1:Y1", [PAYROLL_HISTORY_HEADERS])

        cfg_vals = _gs_call_with_backoff(ws_cfg.get, 'A:B')
        if not cfg_vals:
            gspread_update_range(ws_cfg, "A1:B6", [
                ["Key", "Value"],
                ["letan_payroll_access", "0"],
                ["default_living_expense", "150000"],
                ["default_locker_support", "80000"],
                ["employee_payroll_overrides_json", "{}"],
                ["leader_responsibility_allowance", "0"],
            ])
        else:
            # Bổ sung key thiếu chỉ ở lần khởi tạo tài nguyên, không kiểm tra lại ở mỗi rerun.
            existing_keys = {str(r[0]).strip() for r in cfg_vals[1:] if r}
            additions = []
            if "default_living_expense" not in existing_keys:
                additions.append(["default_living_expense", "150000"])
            if "default_locker_support" not in existing_keys:
                additions.append(["default_locker_support", "80000"])
            if "employee_payroll_overrides_json" not in existing_keys:
                additions.append(["employee_payroll_overrides_json", "{}"])
            if "letan_payroll_access" not in existing_keys:
                additions.append(["letan_payroll_access", "0"])
            if "leader_responsibility_allowance" not in existing_keys:
                additions.append(["leader_responsibility_allowance", "0"])
            if additions:
                next_row = max(2, len(cfg_vals) + 1)
                gspread_update_range(ws_cfg, f"A{next_row}:B{next_row + len(additions) - 1}", additions)
        return ws_pay, ws_cfg, ""
    except Exception as e:
        return None, None, f"Lỗi khởi tạo vùng lưu bảng lương: {e}"


@st.cache_data(ttl=300, show_spinner=False)
def _load_payroll_config_rows_cached():
    """Một lần đọc A:B dùng chung cho quyền Lễ tân, mức mặc định và mức riêng NV."""
    _, ws_cfg, err = _ensure_payroll_storage()
    if err or ws_cfg is None:
        return [], err or "Không mở được sheet cấu hình."
    try:
        vals = _gs_call_with_backoff(ws_cfg.get, 'A:B')
        return vals or [], ""
    except Exception as e:
        return [], f"Lỗi đọc cấu hình lương: {e}"


def _payroll_config_dict():
    vals, err = _load_payroll_config_rows_cached()
    cfg = {}
    if vals:
        for row in vals[1:]:
            if row:
                cfg[str(row[0]).strip()] = row[1] if len(row) > 1 else ''
    return cfg, vals, err


def _payroll_config_key_rows(vals):
    rows = {}
    for idx, row in enumerate((vals or [])[1:], start=2):
        if row:
            rows[str(row[0]).strip()] = idx
    return rows


def get_payroll_letan_enabled():
    try:
        cfg, _, _ = _payroll_config_dict()
        value = str(cfg.get('letan_payroll_access', '0')).strip().lower()
        return value in {"1", "true", "yes", "on", "mở", "mo"}
    except Exception:
        return False


def set_payroll_letan_enabled(enabled):
    try:
        _, ws_cfg, err = _ensure_payroll_storage()
        if err or ws_cfg is None:
            return False, err or "Không mở được sheet cấu hình."
        _, vals, read_err = _payroll_config_dict()
        if read_err:
            return False, read_err
        key_rows = _payroll_config_key_rows(vals)
        target_row = key_rows.get('letan_payroll_access', max(2, len(vals) + 1))
        gspread_update_range(ws_cfg, f"A{target_row}:B{target_row}", [["letan_payroll_access", "1" if enabled else "0"]])
        _clear_payroll_config_cache()
        return True, "Đã mở quyền xem Bảng lương cho Lễ tân." if enabled else "Đã đóng quyền xem Bảng lương của Lễ tân."
    except Exception as e:
        return False, f"Lỗi cập nhật quyền Lễ tân: {e}"


def get_payroll_default_amounts():
    """Đọc hai mức mặc định từ snapshot cấu hình đã cache."""
    living, locker = 150000.0, 80000.0
    try:
        cfg, _, _ = _payroll_config_dict()
        living = _money_to_float(cfg.get('default_living_expense', living)) or living
        locker = _money_to_float(cfg.get('default_locker_support', locker)) or locker
    except Exception:
        pass
    return float(living), float(locker)


def set_payroll_default_amounts(living_expense, locker_support):
    try:
        _, ws_cfg, err = _ensure_payroll_storage()
        if err or ws_cfg is None:
            return False, err or "Không mở được sheet cấu hình."
        _, vals, read_err = _payroll_config_dict()
        if read_err:
            return False, read_err
        key_rows = _payroll_config_key_rows(vals)
        next_row = max(2, len(vals) + 1)
        updates = []
        for key, value in [
            ('default_living_expense', int(round(_money_to_float(living_expense)))),
            ('default_locker_support', int(round(_money_to_float(locker_support)))),
        ]:
            row_idx = key_rows.get(key)
            if row_idx is None:
                row_idx = next_row
                next_row += 1
            updates.append((row_idx, key, str(value)))
        # Hai write nhỏ, nhưng KHÔNG phát sinh thêm read nào.
        for row_idx, key, value in updates:
            gspread_update_range(ws_cfg, f"A{row_idx}:B{row_idx}", [[key, value]])
        _clear_payroll_config_cache()
        return True, "Đã lưu mức Chi phí sinh hoạt và Hỗ trợ Locker mặc định."
    except Exception as e:
        return False, f"Lỗi lưu mức mặc định: {e}"


def get_leader_responsibility_allowance():
    try:
        cfg, _, _ = _payroll_config_dict()
        return float(_money_to_float(cfg.get('leader_responsibility_allowance', 0)))
    except Exception:
        return 0.0

def set_leader_responsibility_allowance(amount):
    try:
        _, ws_cfg, err = _ensure_payroll_storage()
        if err or ws_cfg is None:
            return False, err or "Không mở được sheet cấu hình."
        _, vals, read_err = _payroll_config_dict()
        if read_err:
            return False, read_err
        key_rows = _payroll_config_key_rows(vals)
        target_row = key_rows.get('leader_responsibility_allowance', max(2, len(vals) + 1))
        value = int(round(_money_to_float(amount)))
        gspread_update_range(ws_cfg, f"A{target_row}:B{target_row}", [["leader_responsibility_allowance", str(value)]])
        _clear_payroll_config_cache()
        return True, f"Đã lưu tiền trách nhiệm Leader: {value:,.0f} đ".replace(',', '.')
    except Exception as e:
        return False, f"Lỗi lưu tiền trách nhiệm Leader: {e}"

def get_payroll_employee_overrides():
    """Đọc mức riêng từ cùng snapshot cấu hình cache, không gọi Sheets API thêm lần nữa."""
    try:
        cfg, _, _ = _payroll_config_dict()
        raw = cfg.get('employee_payroll_overrides_json', '{}') or '{}'
        data = json.loads(str(raw))
        if not isinstance(data, dict):
            return {}
        cleaned = {}
        for key, value in data.items():
            if not isinstance(value, dict):
                continue
            norm_key = normalize_login_name(key)
            if not norm_key:
                continue
            cleaned[norm_key] = {
                "name": str(value.get("name", key)).strip(),
                "living": float(_money_to_float(value.get("living", 0))),
                "locker": float(_money_to_float(value.get("locker", 0))),
            }
        return cleaned
    except Exception:
        return {}


def _write_payroll_employee_overrides(overrides):
    try:
        _, ws_cfg, err = _ensure_payroll_storage()
        if err or ws_cfg is None:
            return False, err or "Không mở được sheet cấu hình."
        _, vals, read_err = _payroll_config_dict()
        if read_err:
            return False, read_err
        key_rows = _payroll_config_key_rows(vals)
        target_row = key_rows.get('employee_payroll_overrides_json', max(2, len(vals) + 1))
        payload = json.dumps(overrides or {}, ensure_ascii=False, separators=(",", ":"))
        gspread_update_range(ws_cfg, f"A{target_row}:B{target_row}", [["employee_payroll_overrides_json", payload]])
        _clear_payroll_config_cache()
        return True, "Đã lưu mức riêng theo nhân viên."
    except Exception as e:
        return False, f"Lỗi lưu mức riêng theo nhân viên: {e}"


def set_payroll_employee_overrides(employee_names, living_expense, locker_support):
    names = [str(x).strip() for x in (employee_names or []) if str(x).strip()]
    if not names:
        return False, "Vui lòng chọn ít nhất 1 nhân viên."
    overrides = get_payroll_employee_overrides()
    living = int(round(_money_to_float(living_expense)))
    locker = int(round(_money_to_float(locker_support)))
    for name in names:
        key = normalize_login_name(name)
        overrides[key] = {"name": name, "living": living, "locker": locker}
    ok, msg = _write_payroll_employee_overrides(overrides)
    if ok:
        return True, f"Đã áp dụng mức riêng cho {len(names)} nhân viên."
    return ok, msg


def clear_payroll_employee_overrides(employee_names):
    names = [str(x).strip() for x in (employee_names or []) if str(x).strip()]
    if not names:
        return False, "Vui lòng chọn ít nhất 1 nhân viên."
    overrides = get_payroll_employee_overrides()
    removed = 0
    for name in names:
        key = normalize_login_name(name)
        if key in overrides:
            overrides.pop(key, None)
            removed += 1
    ok, msg = _write_payroll_employee_overrides(overrides)
    if ok:
        return True, f"Đã xóa mức riêng của {removed} nhân viên; các nhân viên này sẽ dùng mức mặc định chung."
    return ok, msg


def _apply_payroll_override_to_current_session(employee_names, living_expense, locker_support):
    """Cập nhật ngay bảng lương đang mở nếu đã tính trước đó."""
    cur = st.session_state.get('payroll_current_df')
    if not isinstance(cur, pd.DataFrame) or cur.empty or 'Tên Hệ thống' not in cur.columns:
        return
    selected = {normalize_login_name(x) for x in (employee_names or [])}
    d = cur.copy()
    mask = d['Tên Hệ thống'].apply(normalize_login_name).isin(selected)
    if mask.any():
        d.loc[mask, 'Chi Phí Sinh Hoạt'] = float(_money_to_float(living_expense))
        d.loc[mask, 'Tiền hỗ trợ Locker'] = float(_money_to_float(locker_support))
        st.session_state.payroll_current_df = recalculate_payroll_net(d)


def _money_to_float(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 0.0
    if isinstance(value, numbers.Number):
        try: return float(value)
        except Exception: return 0.0
    text = str(value).strip()
    if not text or text.lower() in {"none", "nan", "nat", "-"}:
        return 0.0
    neg = text.startswith('-')
    digits = re.sub(r"[^0-9]", "", text)
    if not digits:
        return 0.0
    val = float(digits)
    return -val if neg else val


def _filter_real_payroll_rows(df):
    """Loại các dòng tiêu đề/placeholder bị đọc nhầm thành nhân viên thật."""
    if df is None or not isinstance(df, pd.DataFrame) or df.empty:
        return df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
    d = df.copy()
    if 'Tên Hệ thống' in d.columns:
        bad_names = {"ten nhan vien", "ten he thong", "username", "user name"}
        mask_bad = d['Tên Hệ thống'].astype(str).apply(normalize_login_name).isin(bad_names)
        d = d[~mask_bad].copy()
    if 'TT' in d.columns:
        d = d.reset_index(drop=True)
        d['TT'] = range(1, len(d) + 1)
    return d



# ==========================================================
# TÍCH LŨY NHÂN VIÊN
# ==========================================================
TICHLUY_HEADER_ALIASES = {
    "STT": ["STT", "TT", "Thứ tự", "Số thứ tự"],
    "Tên nhân viên": ["Tên nhân viên", "Tên Hệ thống", "Nhân viên", "Username"],
    "Ngày bắt đầu làm": ["Ngày bắt đầu làm", "Ngày bắt đầu đi làm", "Ngày vào làm", "Ngày bắt đầu"],
    "Mục tiêu tích lũy": ["Mục tiêu tích lũy", "Mục tiêu", "Tổng cần tích lũy"],
    "Đã tích lũy": ["Đã tích lũy", "Tích lũy", "Số tiền tích lũy", "Đã đóng"],
    "Còn lại": ["Còn lại", "Còn phải tích lũy", "Số tiền còn lại"],
    "Kỳ gần nhất": ["Kỳ gần nhất", "Kỳ đóng gần nhất"],
    "Số tiền kỳ gần nhất": ["Số tiền kỳ gần nhất", "Tiền kỳ gần nhất"],
    "Chi tiết các kỳ": ["Chi tiết các kỳ", "Lịch sử kỳ", "Lịch sử tích lũy"],
}


def _tichluy_header_positions(header):
    """Map tên cột chuẩn -> index 0-based, chấp nhận các tên cột người dùng đã tạo trước đó."""
    positions = {}
    normalized = [normalize_login_name(x) for x in (header or [])]
    for canonical in TICHLUY_HEADERS:
        aliases = TICHLUY_HEADER_ALIASES.get(canonical, [canonical])
        alias_keys = {normalize_login_name(x) for x in aliases}
        for i, key in enumerate(normalized):
            if key in alias_keys:
                positions[canonical] = i
                break
    return positions


@st.cache_resource(show_spinner=False)
def _ensure_tichluy_sheet():
    """
    Lấy/tạo sheet TichLuy. Nếu người dùng đã tạo cột trước đó thì GIỮ NGUYÊN,
    chỉ bổ sung các cột hệ thống còn thiếu ở bên phải, tránh ghi đè dữ liệu hiện hữu.
    """
    client = get_gspread_client()
    if not client:
        return None, "Chưa cấu hình quyền Google Sheets."
    try:
        ss = client.open_by_key(SHEET_MAT_KHAU_ID)
        ws = _get_or_create_worksheet(ss, TICHLUY_WORKSHEET, rows=1000, cols=20)
        header = _gs_call_with_backoff(ws.row_values, 1)
        if not header or not any(str(x).strip() for x in header):
            gspread_update_range(ws, "A1:I1", [TICHLUY_HEADERS])
        else:
            positions = _tichluy_header_positions(header)
            # STT luôn ở cột A. Nếu sheet cũ chưa có STT, chèn một cột mới ở đầu
            # để KHÔNG ghi đè / làm mất cột Tên nhân viên hiện có.
            if "STT" not in positions or positions.get("STT") != 0:
                try:
                    _gs_call_with_backoff(ss.batch_update, {
                        "requests": [{
                            "insertDimension": {
                                "range": {
                                    "sheetId": ws.id, "dimension": "COLUMNS",
                                    "startIndex": 0, "endIndex": 1
                                },
                                "inheritFromBefore": False
                            }
                        }]
                    })
                    gspread_update_range(ws, "A1:A1", [["STT"]])
                    header = ["STT"] + list(header)
                except Exception:
                    # Fallback tương thích gspread cũ.
                    _gs_call_with_backoff(ws.insert_cols, [["STT"]], col=1)
                    header = ["STT"] + list(header)
                positions = _tichluy_header_positions(header)

            missing = [h for h in TICHLUY_HEADERS if h not in positions]
            if missing:
                start_col = len(header) + 1
                start_a1 = gspread.utils.rowcol_to_a1(1, start_col)
                end_a1 = gspread.utils.rowcol_to_a1(1, start_col + len(missing) - 1)
                gspread_update_range(ws, f"{start_a1}:{end_a1}", [missing])
        return ws, ""
    except Exception as e:
        return None, f"Không mở được sheet TichLuy: {e}"


def _load_tichluy_tracking_from_sheets():
    """Đọc TichLuy một lần/cache; tự nhận diện vị trí cột để tương thích sheet đã có."""
    try:
        ws, err = _ensure_tichluy_sheet()
        if err or ws is None:
            return pd.DataFrame(columns=TICHLUY_HEADERS)
        values = _gs_call_with_backoff(ws.get_all_values)
        if not values:
            return pd.DataFrame(columns=TICHLUY_HEADERS)
        header = values[0]
        positions = _tichluy_header_positions(header)
        rows = []
        # Dùng cache tài khoản hiện có để ẩn hoàn toàn quanly/letan/locker/tapvu/admin khỏi
        # nghiệp vụ TichLuy, kể cả khi sheet cũ vẫn còn dòng từ phiên bản trước.
        role_map = _credential_role_map(load_credentials_recent())
        for sheet_row, row in enumerate(values[1:], start=2):
            if not any(str(v).strip() for v in row):
                continue
            item = {}
            for canonical in TICHLUY_HEADERS:
                pos = positions.get(canonical)
                item[canonical] = row[pos] if pos is not None and pos < len(row) else ''
            emp_key = normalize_login_name(item.get('Tên nhân viên', ''))
            emp_role = role_map.get(emp_key, 'nhanvien')
            if emp_role in TICHLUY_EXCLUDED_ROLES:
                continue
            item['__sheet_row'] = sheet_row
            rows.append(item)
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=TICHLUY_HEADERS + ['__sheet_row'])
    except Exception:
        return pd.DataFrame(columns=TICHLUY_HEADERS + ['__sheet_row'])

@st.cache_data(ttl=30, show_spinner=False)
def load_tichluy_tracking():
    """V75: đọc qua PostgreSQL dùng chung giữa các Cloud Run instance; Google Sheets là nguồn đồng bộ dự phòng."""
    if vpg is not None and vpg.is_enabled():
        return vpg.load_dataset(
            "tichluy",
            _load_tichluy_tracking_from_sheets,
            ttl_seconds=int(os.getenv("VERA_PG_TTL_TICHLUY", "90")),
        )
    return _load_tichluy_tracking_from_sheets()


def _parse_vn_date(value):
    """Đọc ngày Việt Nam và cả Excel serial date (ví dụ 46088) từ Google Sheets."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    # Google Sheets có thể trả ngày dưới dạng số serial của Excel/Sheets.
    # Epoch tương thích Excel là 30/12/1899 (serial 1 = 31/12/1899).
    try:
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            serial = float(value)
            if 20000 <= serial <= 100000:
                return (date(1899, 12, 30) + timedelta(days=int(serial)))
    except Exception:
        pass

    text = str(value or '').strip()
    if not text or text.casefold() in {'nan','none','nat'}:
        return None

    # Trường hợp serial date được trả về dưới dạng chuỗi, ví dụ "46088" hoặc "46088.0".
    try:
        serial = float(text.replace(',', '.'))
        if 20000 <= serial <= 100000:
            return (date(1899, 12, 30) + timedelta(days=int(serial)))
    except Exception:
        pass

    for fmt in ('%d/%m/%Y','%d-%m-%Y','%Y-%m-%d','%d/%m/%y'):
        try:
            return datetime.strptime(text, fmt).date()
        except Exception:
            pass
    try:
        dt = pd.to_datetime(text, dayfirst=True, errors='coerce')
        return None if pd.isna(dt) else dt.date()
    except Exception:
        return None



def prepare_leave_editor_types(df):
    """Chuẩn hóa dtype trước khi truyền vào st.data_editor.

    Google Sheets trả phần lớn dữ liệu dưới dạng chuỗi. Streamlit các bản mới kiểm tra
    chặt kiểu dữ liệu giữa DataFrame và column_config; vì vậy DateColumn/NumberColumn
    có thể ném StreamlitAPIException nếu cột nguồn vẫn là object/string.

    Hàm này chỉ chuẩn hóa bản sao dùng cho giao diện editor, KHÔNG thay đổi dữ liệu nguồn.
    """
    if not isinstance(df, pd.DataFrame):
        return df
    d = df.copy()

    if 'Chọn' in d.columns:
        d['Chọn'] = d['Chọn'].fillna(False).astype(bool)

    if 'Ngày' in d.columns:
        def _to_editor_date(v):
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return None
            if isinstance(v, datetime):
                return v.date()
            if isinstance(v, date):
                return v
            parsed = pd.to_datetime(str(v).strip(), errors='coerce', dayfirst=True)
            return None if pd.isna(parsed) else parsed.date()
        d['Ngày'] = d['Ngày'].apply(_to_editor_date)

    for col in ['Số ngày tính', 'Số ngày phép cộng dồn', 'Phạt vi phạm']:
        if col in d.columns:
            is_money = col == 'Phạt vi phạm'
            def _to_editor_number(v, _money=is_money):
                if v is None:
                    return None
                s = str(v).strip()
                if not s or s.casefold() in {'nan', 'none', 'nat', '-'}:
                    return None
                try:
                    return float(_parse_leave_number(v, 0.0, money=_money))
                except Exception:
                    try:
                        return float(v)
                    except Exception:
                        return None
            d[col] = pd.to_numeric(d[col].apply(_to_editor_number), errors='coerce')

    # Các cột cấu hình TextColumn/SelectboxColumn phải luôn là chuỗi thuần.
    for col in [
        'Tên nhân viên', 'Lý do nghỉ', 'Loại nghỉ', 'Chi tiết',
        'Ngày cập nhật', 'Giờ cập nhật', 'Người cập nhật'
    ]:
        if col in d.columns:
            d[col] = d[col].apply(
                lambda v: '' if v is None or (isinstance(v, float) and pd.isna(v)) else str(v)
            )

    return d

def _tichluy_period_key(start_date, end_date):
    return f"{start_date.isoformat()}|{end_date.isoformat()}"


def _parse_tichluy_history(value):
    try:
        data = json.loads(str(value or '{}'))
        if not isinstance(data, dict):
            return {}
        return {str(k): float(_money_to_float(v)) for k, v in data.items()}
    except Exception:
        return {}


def _is_tichluy_completed(target_value, accumulated_value, remaining_value=''):
    """
    Xác định hồ sơ đã hoàn thành đóng Tích lũy.
    Dữ liệu Admin nhập tay ở D/E/F được xem là hoàn thành khi E >= D (>0).
    Khi đã hoàn thành, hệ thống tuyệt đối không tự sửa D/E/F/G/H/I của dòng đó.
    """
    try:
        target = float(_money_to_float(target_value))
        accumulated = float(_money_to_float(accumulated_value))
        return target > 0 and accumulated >= target
    except Exception:
        return False


def _first_pay_period_for_start(start_work_date):
    """Kỳ đầu theo ngày vào làm, dùng đúng chuẩn 01-15 / 16-cuối tháng."""
    return _official_payroll_period(
        start_work_date.year, start_work_date.month, 1 if start_work_date.day <= 15 else 2
    )


def _parse_tichluy_period_text(value):
    """Đọc cột Kỳ gần nhất dạng dd/mm/yyyy - dd/mm/yyyy hoặc yyyy-mm-dd|yyyy-mm-dd."""
    text = str(value or '').strip()
    if not text:
        return None, None
    if '|' in text:
        parts = [x.strip() for x in text.split('|', 1)]
    elif ' - ' in text:
        parts = [x.strip() for x in text.split(' - ', 1)]
    else:
        return None, None
    if len(parts) != 2:
        return None, None
    return _parse_vn_date(parts[0]), _parse_vn_date(parts[1])


def _select_preferred_tichluy_rows(tracking):
    """
    Sheet TichLuy của dữ liệu cũ có thể trùng Tên nhân viên.
    Chỉ chọn 1 dòng tốt nhất cho mỗi người:
    1) ưu tiên dòng có Ngày bắt đầu làm hợp lệ;
    2) ưu tiên dòng có Chi tiết các kỳ / Kỳ gần nhất;
    3) ưu tiên dòng có dữ liệu số Mục tiêu/Đã tích lũy/Còn lại;
    4) nếu vẫn bằng nhau, lấy dòng nằm dưới cùng (STT/sheet row mới hơn).
    """
    if tracking is None or tracking.empty:
        return tracking
    best = {}
    for _, r in tracking.iterrows():
        key = normalize_login_name(r.get('Tên nhân viên', ''))
        if not key:
            continue
        completed_ok = 1 if _is_tichluy_completed(
            r.get('Mục tiêu tích lũy', ''), r.get('Đã tích lũy', ''), r.get('Còn lại', '')
        ) else 0
        start_ok = 1 if _parse_vn_date(r.get('Ngày bắt đầu làm', '')) else 0
        hist = _parse_tichluy_history(r.get('Chi tiết các kỳ', ''))
        period_start, period_end = _parse_tichluy_period_text(r.get('Kỳ gần nhất', ''))
        history_ok = 1 if hist or (period_start and period_end) else 0
        numeric_ok = 0
        for c in ('Mục tiêu tích lũy', 'Đã tích lũy', 'Còn lại'):
            raw = str(r.get(c, '') or '').strip()
            if raw not in {'', '-', 'None', 'nan'}:
                numeric_ok += 1
        try:
            sheet_row = int(r.get('__sheet_row', 0) or 0)
        except Exception:
            sheet_row = 0
        # Dòng đã hoàn thành do Admin nhập tay luôn được ưu tiên nếu tên bị trùng.
        score = (completed_ok, start_ok, history_ok, numeric_ok, sheet_row)
        if key not in best or score > best[key][0]:
            best[key] = (score, r)
    if not best:
        return tracking.iloc[0:0].copy()
    return pd.DataFrame([item[1] for item in best.values()]).reset_index(drop=True)


def get_tichluy_charge_map(start_date, end_date, employee_names=None, for_existing_snapshot=False):
    """
    Số Tích lũy tự động của kỳ:
    - mục tiêu mặc định 5.000.000;
    - mỗi kỳ 500.000, kỳ cuối chỉ thu phần còn thiếu;
    - kỳ đầu tiên kể từ ngày bắt đầu làm: nếu số ngày từ ngày vào làm đến cuối kỳ < 10 thì không thu;
    - khi một kỳ đã được ghi nhận trong TichLuy, bảng lương MỚI không thu lại;
      còn bản lương lịch sử đang sửa giữ đúng số của kỳ đã ghi nhận.
    """
    tracking = _select_preferred_tichluy_rows(load_tichluy_tracking())
    wanted = {normalize_login_name(x) for x in (employee_names or []) if str(x).strip()} if employee_names else None
    result, info = {}, {}
    if tracking is None or tracking.empty:
        return result, info
    period_key = _tichluy_period_key(start_date, end_date)
    for _, r in tracking.iterrows():
        name = str(r.get('Tên nhân viên','')).strip()
        key = normalize_login_name(name)
        if not key or (wanted is not None and key not in wanted):
            continue
        start_work = _parse_vn_date(r.get('Ngày bắt đầu làm',''))
        target = float(_money_to_float(r.get('Mục tiêu tích lũy', TICHLUY_TARGET_DEFAULT)) or TICHLUY_TARGET_DEFAULT)
        accumulated = float(_money_to_float(r.get('Đã tích lũy',0)))
        completed_manual = _is_tichluy_completed(
            r.get('Mục tiêu tích lũy', target), r.get('Đã tích lũy', accumulated), r.get('Còn lại', '')
        )
        # Dòng đã hoàn thành do Admin nhập tay được khóa nghiệp vụ: luôn còn lại = 0,
        # không phụ thuộc F đang để trống, dấu '-' hay số cũ.
        if completed_manual:
            remaining = 0.0
        else:
            # Cột F = Còn lại là nguồn ưu tiên nếu có số hợp lệ; nếu trống / '-' thì suy ra D - E.
            remaining_raw = str(r.get('Còn lại', '') or '').strip()
            if remaining_raw and remaining_raw not in {'-', '–', '—'}:
                remaining = max(0.0, float(_money_to_float(remaining_raw)))
                # Nếu E đang trống nhưng F có dữ liệu thì suy ngược Đã tích lũy để giữ D/E/F nhất quán.
                if not str(r.get('Đã tích lũy', '') or '').strip():
                    accumulated = max(0.0, target - remaining)
            else:
                remaining = max(0.0, target - accumulated)
        hist = _parse_tichluy_history(r.get('Chi tiết các kỳ',''))
        history_total, history_keys = _matching_tichluy_history(hist, start_date, end_date)
        # Một kỳ chính thức chỉ thu tối đa mức kỳ. Nếu bản cũ từng tạo nhiều key 16→ngày hiện tại
        # thì chỉ coi là đã đóng tối đa một kỳ, không thu thêm lần nữa.
        existing_amount = min(float(TICHLUY_PERIOD_DEFAULT), max(0.0, history_total))

        # Tương thích dữ liệu TichLuy cũ: có thể chưa có JSON "Chi tiết các kỳ"
        # nhưng đã có "Kỳ gần nhất" + "Số tiền kỳ gần nhất".
        if existing_amount <= 0:
            last_start, last_end = _parse_tichluy_period_text(r.get('Kỳ gần nhất', ''))
            if last_start and last_end:
                last_match = (last_start == start_date and last_end == end_date)
                if not last_match and _is_official_payroll_period(start_date, end_date):
                    cls, cle = _canonicalize_payroll_period(last_start, last_end)
                    last_match = (cls == start_date and cle == end_date)
                if last_match:
                    existing_amount = min(
                        float(TICHLUY_PERIOD_DEFAULT),
                        max(0.0, float(_money_to_float(r.get('Số tiền kỳ gần nhất', 0))))
                    )

        charge = 0.0
        reason = ''
        if existing_amount > 0:
            charge = existing_amount if for_existing_snapshot else 0.0
            reason = 'Kỳ này đã được ghi nhận trước đó.'
        elif remaining <= 0:
            reason = 'Đã đủ mục tiêu tích lũy.'
        elif start_work is None:
            # Dữ liệu TichLuy cũ có nhiều nhân viên đã tồn tại trước khi hệ thống bắt đầu
            # lưu Ngày bắt đầu làm. Nếu D/E/F cho thấy vẫn còn phải tích lũy thì vẫn thu
            # theo kỳ; chỉ các nhân viên mới (được thêm từ hệ thống) mới cần áp dụng chính
            # xác quy tắc kỳ đầu <10 ngày vì các dòng mới luôn có ngày bắt đầu làm.
            charge = min(float(TICHLUY_PERIOD_DEFAULT), remaining)
            reason = 'Hồ sơ cũ chưa có Ngày bắt đầu làm; thu theo số Còn lại trong TichLuy.'
        elif end_date < start_work:
            reason = 'Kỳ lương trước ngày bắt đầu làm.'
        else:
            first_start, first_end = _first_pay_period_for_start(start_work)
            is_first_period = not (end_date < first_start or start_date > first_end)
            first_days = (first_end - start_work).days + 1
            if is_first_period and first_days < 10:
                reason = f'Kỳ đầu chỉ có {first_days} ngày kể từ ngày bắt đầu làm (<10 ngày), tạm không thu.'
            else:
                charge = min(float(TICHLUY_PERIOD_DEFAULT), remaining)
                reason = 'Thu tích lũy theo kỳ.'
        result[key] = float(charge)
        info[key] = {
            'name': name, 'start_date': start_work, 'target': target, 'accumulated': accumulated,
            'remaining': remaining, 'charge': float(charge), 'reason': reason,
            'sheet_row': int(r.get('__sheet_row', 0) or 0),
        }
    return result, info



def _credential_role_map(credentials_df=None):
    """Tên đăng nhập chuẩn hóa -> vai trò, dùng chung để lọc danh sách nghiệp vụ."""
    credentials_df = load_credentials_recent() if credentials_df is None else credentials_df
    result = {}
    if credentials_df is None or credentials_df.empty:
        return result
    for _, r in credentials_df.iterrows():
        key = normalize_login_name(r.get('Tên nhân viên', ''))
        if key:
            result[key] = str(r.get('Phân quyền', 'nhanvien')).strip().lower()
    return result


def get_leave_eligible_employee_names(credentials_df, excel_df=None):
    """Danh sách đăng ký nghỉ: ẩn letan, locker, tapvu kể cả khi tên còn nằm trong file Excel cũ."""
    role_map = _credential_role_map(credentials_df)
    excluded_names = {k for k, role in role_map.items() if role in LEAVE_EXCLUDED_ROLES}
    names = []
    if credentials_df is not None and not credentials_df.empty and 'Tên nhân viên' in credentials_df.columns:
        for _, r in credentials_df.iterrows():
            name = str(r.get('Tên nhân viên', '')).strip()
            role = str(r.get('Phân quyền', 'nhanvien')).strip().lower()
            if name and role not in LEAVE_EXCLUDED_ROLES:
                names.append(name)
    if excel_df is not None and not excel_df.empty and 'Tên nhân viên' in excel_df.columns:
        for name in excel_df['Tên nhân viên'].dropna().astype(str).str.strip().tolist():
            if name and normalize_login_name(name) not in excluded_names:
                names.append(name)
    # Giữ tên hiển thị gốc nhưng loại trùng theo chuẩn hóa.
    by_key = {}
    for name in names:
        by_key.setdefault(normalize_login_name(name), name)
    return sorted(by_key.values(), key=lambda x: normalize_login_name(x))


def renumber_credential_sheet_stt(sheet=None):
    """Đánh lại STT cột A của Sheet1 theo các tài khoản có tên ở cột B, chỉ 1 read + 1 write."""
    try:
        if sheet is None:
            client = get_gspread_client()
            if not client:
                return False, 'Chưa cấu hình Google Sheets.'
            sheet = client.open_by_key(SHEET_MAT_KHAU_ID).get_worksheet(0)
        names = _gs_call_with_backoff(sheet.col_values, 2)
        if len(names) <= 1:
            return True, 'Sheet1 chưa có nhân viên để đánh STT.'
        seq = 0
        values = []
        for name in names[1:]:
            if str(name).strip():
                seq += 1
                values.append([seq])
            else:
                values.append([''])
        if values:
            gspread_update_range(sheet, f"A2:A{len(values)+1}", values, value_input_option='USER_ENTERED')
        return True, f'Đã sắp xếp lại STT Sheet1: {seq} tài khoản.'
    except Exception as e:
        return False, f'Lỗi đánh STT Sheet1: {e}'


def sync_tichluy_roles_and_stt(credentials_df=None):
    """
    Đồng bộ TichLuy với danh sách nhân viên hiện tại nhưng KHÔNG sửa dữ liệu tích lũy đã nhập tay:
    - chỉ giữ các vai trò được phép tham gia TichLuy;
    - tự thêm các nhân viên hiện tại còn thiếu vào cuối sheet;
    - KHÔNG ghi đè D/E/F của bất kỳ dòng đã có;
    - đặc biệt dòng đã hoàn thành (E >= D > 0) tuyệt đối không bị hệ thống sửa D:I;
    - cột A được đánh lại từ dòng 2: 1,2,3... theo đúng thứ tự dòng hiện tại.
    """
    try:
        credentials_df = load_credentials_recent() if credentials_df is None else credentials_df
        role_map = _credential_role_map(credentials_df)
        ws, err = _ensure_tichluy_sheet()
        if err or ws is None:
            return False, err or 'Không mở được TichLuy.'

        values = _gs_call_with_backoff(ws.get_all_values)
        if not values:
            values = [list(TICHLUY_HEADERS)]
        header = values[0]
        pos = _tichluy_header_positions(header)
        name_pos = pos.get('Tên nhân viên')
        if name_pos is None:
            return False, 'TichLuy chưa có cột Tên nhân viên.'

        # Danh sách nhân viên hiện tại cần có trong TichLuy, giữ đúng thứ tự Sheet1.
        eligible = []
        eligible_keys = set()
        if isinstance(credentials_df, pd.DataFrame) and not credentials_df.empty:
            for _, cr in credentials_df.iterrows():
                name = str(cr.get('Tên nhân viên', '')).strip()
                role = str(cr.get('Phân quyền', 'nhanvien')).strip().lower()
                key = normalize_login_name(name)
                if not key or role in TICHLUY_EXCLUDED_ROLES:
                    continue
                if key in {'ten nhan vien', 'ten he thong', 'username', 'user name'} or key in eligible_keys:
                    continue
                eligible.append((name, key))
                eligible_keys.add(key)

        # Chỉ xóa dòng thuộc tài khoản hiện tại có role bị loại hoặc tên không còn tồn tại.
        # Không chỉnh D:I ở những dòng được giữ lại.
        delete_sheet_rows = []
        existing_keys = set()
        kept_names = []
        for sheet_row, row in enumerate(values[1:], start=2):
            name = row[name_pos] if name_pos < len(row) else ''
            if not str(name).strip():
                continue
            key = normalize_login_name(name)
            role = role_map.get(key)
            if key not in eligible_keys or role in TICHLUY_EXCLUDED_ROLES:
                delete_sheet_rows.append(sheet_row)
            else:
                existing_keys.add(key)
                kept_names.append(str(name).strip())

        if delete_sheet_rows:
            client = get_gspread_client()
            ss = client.open_by_key(SHEET_MAT_KHAU_ID)
            requests = []
            for r in sorted(delete_sheet_rows, reverse=True):
                requests.append({
                    'deleteDimension': {
                        'range': {
                            'sheetId': ws.id, 'dimension': 'ROWS',
                            'startIndex': r - 1, 'endIndex': r
                        }
                    }
                })
            _gs_call_with_backoff(ss.batch_update, {'requests': requests})

        # Tự thêm những nhân viên hiện tại còn thiếu. Với nhân sự cũ chưa có ngày vào làm,
        # để C trống để Admin bổ sung; D/E/F có mặc định và có thể nhập tay nếu đã hoàn thành.
        missing = [(name, key) for name, key in eligible if key not in existing_keys]
        if missing:
            rows_to_append = []
            for name, key in missing:
                row = [''] * len(header)
                defaults = {
                    'STT': '',
                    'Tên nhân viên': name,
                    'Ngày bắt đầu làm': '',
                    'Mục tiêu tích lũy': TICHLUY_TARGET_DEFAULT,
                    'Đã tích lũy': 0,
                    'Còn lại': TICHLUY_TARGET_DEFAULT,
                    'Kỳ gần nhất': '',
                    'Số tiền kỳ gần nhất': 0,
                    'Chi tiết các kỳ': '{}',
                }
                for canonical, value in defaults.items():
                    idx = pos.get(canonical)
                    if idx is not None and idx < len(row):
                        row[idx] = value
                rows_to_append.append(row)
            # Một request thay vì append_row từng nhân viên để giảm quota.
            start_row = max(2, len(values) - len(delete_sheet_rows) + 1)
            end_row = start_row + len(rows_to_append) - 1
            end_col = gspread.utils.rowcol_to_a1(1, len(header)).rstrip('1')
            gspread_update_range(ws, f'A{start_row}:{end_col}{end_row}', rows_to_append, value_input_option='USER_ENTERED')

        # Sau xóa/thêm, đọc đúng cột Tên nhân viên và đánh STT A2 = 1,2,3... theo từng dòng thực tế.
        # Nếu có dòng trống xen giữa thì A của dòng đó để trống; các dòng có tên vẫn chạy số liên tục.
        # Đây là cột duy nhất hệ thống luôn được phép thay đổi trên các dòng đã hoàn thành.
        names_after = _gs_call_with_backoff(ws.col_values, name_pos + 1)
        stt_values = []
        seq = 0
        for name in names_after[1:] if len(names_after) > 1 else []:
            if str(name).strip():
                seq += 1
                stt_values.append([seq])
            else:
                stt_values.append([''])
        if stt_values:
            gspread_update_range(ws, f'A2:A{len(stt_values) + 1}', stt_values, value_input_option='USER_ENTERED')

        try:
            load_tichluy_tracking.clear()
            if vpg is not None and vpg.is_enabled():
                try:
                    vpg.invalidate_dataset("tichluy")
                except Exception:
                    pass
        except Exception:
            pass
        return True, (
            f'Đã đồng bộ TichLuy: {seq} dòng nhân viên, thêm {len(missing)} người còn thiếu; '
            'D/E/F của các dòng hiện có được giữ nguyên.'
        )
    except Exception as e:
        return False, f'Lỗi đồng bộ STT/TichLuy: {e}'

def ensure_employee_in_tichluy(employee_name, start_work_date=None):
    """Thêm nhân viên vào TichLuy nếu chưa có; ngày bắt đầu = ngày tạo tài khoản."""
    try:
        name = str(employee_name or '').strip()
        if not name:
            return False, 'Thiếu tên nhân viên.'
        start_work_date = start_work_date or get_vn_today()
        ws, err = _ensure_tichluy_sheet()
        if err or ws is None:
            return False, err or 'Không mở được TichLuy.'
        values = _gs_call_with_backoff(ws.get_all_values)
        header = values[0] if values else list(TICHLUY_HEADERS)
        pos = _tichluy_header_positions(header)
        key = normalize_login_name(name)
        name_pos = pos.get('Tên nhân viên', 0)
        for row in values[1:] if values else []:
            if name_pos < len(row) and normalize_login_name(row[name_pos]) == key:
                return True, 'Nhân viên đã có trong TichLuy.'
        row = [''] * len(header)
        defaults = {
            'STT': '',
            'Tên nhân viên': name,
            'Ngày bắt đầu làm': start_work_date.strftime('%d/%m/%Y'),
            'Mục tiêu tích lũy': TICHLUY_TARGET_DEFAULT,
            'Đã tích lũy': 0,
            'Còn lại': TICHLUY_TARGET_DEFAULT,
            'Kỳ gần nhất': '', 'Số tiền kỳ gần nhất': 0, 'Chi tiết các kỳ': '{}'
        }
        for canonical, value in defaults.items():
            if canonical in pos:
                row[pos[canonical]] = value
        ws.append_row(row, value_input_option='USER_ENTERED')
        try: load_tichluy_tracking.clear()
        except Exception: pass
        return True, 'Đã thêm vào TichLuy.'
    except Exception as e:
        return False, f'Lỗi thêm TichLuy: {e}'


def ensure_employee_in_leave_employee_list(employee_name, start_work_date=None):
    """
    Đồng bộ nhân viên mới sang file lịch nghỉ 1Kz0... vào sheet DanhSachNV.
    Không chèn dòng giả vào Sheet1 A:J vì Sheet1 là dữ liệu lịch nghỉ nghiệp vụ.
    """
    try:
        name = str(employee_name or '').strip()
        if not name:
            return False, 'Thiếu tên nhân viên.'
        start_work_date = start_work_date or get_vn_today()
        client = get_gspread_client()
        if not client:
            return False, 'Chưa cấu hình Google Sheets.'
        ss = client.open_by_key(SHEET_DU_PHONG_ID)
        ws = None
        for title in ('DanhSachNV', 'Danh sách NV', 'NhanVien', 'Nhân viên'):
            try:
                ws = ss.worksheet(title)
                break
            except Exception:
                pass
        if ws is None:
            ws = ss.add_worksheet(title='DanhSachNV', rows=1000, cols=5)
        header = _gs_call_with_backoff(ws.row_values, 1)
        if not header or normalize_login_name(header[0] if header else '') not in {'ten nhan vien','ten he thong'}:
            gspread_update_range(ws, 'A1:B1', [['Tên nhân viên','Ngày bắt đầu làm']])
        values = _gs_call_with_backoff(ws.get, 'A:B')
        key = normalize_login_name(name)
        for row in values[1:] if values else []:
            if row and normalize_login_name(row[0]) == key:
                return True, 'Nhân viên đã có trong DanhSachNV của file lịch nghỉ.'
        ws.append_row([name, start_work_date.strftime('%d/%m/%Y')], value_input_option='USER_ENTERED')
        return True, 'Đã thêm vào DanhSachNV của file lịch nghỉ.'
    except Exception as e:
        return False, f'Lỗi đồng bộ danh sách nhân viên lịch nghỉ: {e}'


def record_tichluy_contributions(payroll_df, start_date, end_date):
    """Ghi/ghi đè số Tích lũy của kỳ vào TichLuy theo khóa kỳ, tránh cộng trùng khi lưu lại."""
    try:
        if payroll_df is None or payroll_df.empty:
            return True, 'Không có dữ liệu Tích lũy cần cập nhật.'
        ws, err = _ensure_tichluy_sheet()
        if err or ws is None:
            return False, err or 'Không mở được TichLuy.'
        values = _gs_call_with_backoff(ws.get_all_values)
        if not values:
            return False, 'Sheet TichLuy chưa có tiêu đề.'
        header = values[0]
        pos = _tichluy_header_positions(header)
        rows_by_key = {}
        role_map = _credential_role_map(load_credentials_recent())
        name_pos = pos.get('Tên nhân viên', 0)
        for r_idx, row in enumerate(values[1:], start=2):
            if name_pos < len(row) and str(row[name_pos]).strip():
                full_row = list(row) + [''] * max(0, len(header)-len(row))
                key = normalize_login_name(full_row[name_pos])
                def _rv(canonical, default=''):
                    idx = pos.get(canonical)
                    return full_row[idx] if idx is not None and idx < len(full_row) else default
                completed_score = 1 if _is_tichluy_completed(
                    _rv('Mục tiêu tích lũy'), _rv('Đã tích lũy'), _rv('Còn lại')
                ) else 0
                score = (completed_score, r_idx)
                if key not in rows_by_key or score > rows_by_key[key][0]:
                    rows_by_key[key] = (score, r_idx, full_row[:len(header)])
        period_key = _tichluy_period_key(start_date, end_date)
        updates = []
        created = 0
        for _, pr in payroll_df.iterrows():
            name = str(pr.get('Tên Hệ thống','')).strip()
            key = normalize_login_name(name)
            if not key:
                continue
            role = role_map.get(key)
            if role is None or role in TICHLUY_EXCLUDED_ROLES:
                continue
            amount = max(0.0, float(_money_to_float(pr.get('Tích lũy',0))))
            if key not in rows_by_key:
                # Dữ liệu cũ thiếu hồ sơ TichLuy: tạo dòng để Admin bổ sung Ngày bắt đầu làm.
                new_row = [''] * len(header)
                defaults = {
                    'Tên nhân viên': name, 'Ngày bắt đầu làm': '', 'Mục tiêu tích lũy': TICHLUY_TARGET_DEFAULT,
                    'Đã tích lũy': 0, 'Còn lại': TICHLUY_TARGET_DEFAULT, 'Kỳ gần nhất': '',
                    'Số tiền kỳ gần nhất': 0, 'Chi tiết các kỳ': '{}'
                }
                for canonical, value in defaults.items():
                    if canonical in pos: new_row[pos[canonical]] = value
                ws.append_row(new_row, value_input_option='USER_ENTERED')
                created += 1
                continue
            _score, r_idx, row = rows_by_key[key]
            def g(canonical, default=''):
                i = pos.get(canonical)
                return row[i] if i is not None and i < len(row) else default
            target = float(_money_to_float(g('Mục tiêu tích lũy')) or TICHLUY_TARGET_DEFAULT)
            current_total = float(_money_to_float(g('Đã tích lũy')))
            # Admin có thể nhập tay D/E/F cho người đã hoàn thành. Nếu E >= D > 0,
            # tuyệt đối bỏ qua dòng này: không sửa D/E/F và cũng không sửa G/H/I.
            if _is_tichluy_completed(g('Mục tiêu tích lũy'), g('Đã tích lũy'), g('Còn lại')):
                continue
            hist = _parse_tichluy_history(g('Chi tiết các kỳ'))
            old_history_total, matched_history_keys = _matching_tichluy_history(hist, start_date, end_date)
            old_display_amount = min(float(TICHLUY_PERIOD_DEFAULT), max(0.0, old_history_total))
            amount_to_record = old_display_amount if old_display_amount > 0 and amount <= 0 else amount

            # Nếu dữ liệu cũ đã ghi cùng kỳ dưới nhiều key rút gọn (vd. 16→17, 16→20),
            # gom về một key chính thức và loại phần cộng trùng khỏi tổng tích lũy.
            amount_to_subtract = old_history_total
            new_total = max(0.0, min(target, current_total - amount_to_subtract + amount_to_record))
            for legacy_key in matched_history_keys:
                hist.pop(legacy_key, None)
            hist[period_key] = float(amount_to_record)
            remaining = max(0.0, target - new_total)
            values_to_set = {
                'Mục tiêu tích lũy': target,
                'Đã tích lũy': new_total,
                'Còn lại': remaining,
                'Kỳ gần nhất': f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}",
                'Số tiền kỳ gần nhất': amount_to_record,
                'Chi tiết các kỳ': json.dumps(hist, ensure_ascii=False, separators=(',',':')),
            }
            for canonical, value in values_to_set.items():
                if canonical in pos:
                    row[pos[canonical]] = value
            updates.append((r_idx, row[:len(header)]))
        # Ghi full row để giữ nguyên mọi cột do người dùng tự thêm trong TichLuy.
        end_col_a1 = gspread.utils.rowcol_to_a1(1, len(header)).rstrip('1')
        for r_idx, row in updates:
            gspread_update_range(ws, f'A{r_idx}:{end_col_a1}{r_idx}', [row])
        try: load_tichluy_tracking.clear()
        except Exception: pass
        return True, f'Đã cập nhật Tích lũy cho {len(updates)} nhân viên' + (f'; tạo bổ sung {created} hồ sơ thiếu.' if created else '.')
    except Exception as e:
        return False, f'Lỗi cập nhật TichLuy: {e}'


def get_employee_violation_details(employee_name, start_date, end_date, leave_df=None):
    """Chi tiết các dòng có Phạt vi phạm > 0 của một nhân viên trong đúng kỳ lương."""
    cols = ['Ngày', 'Lý do nghỉ', 'Chi tiết', 'Phạt vi phạm']
    try:
        d = leave_df.copy() if isinstance(leave_df, pd.DataFrame) else load_backup_sheet_data()
        if d is None or d.empty or 'Tên nhân viên' not in d.columns:
            return pd.DataFrame(columns=cols)
        d = d.copy()
        if 'Lý do nghỉ' not in d.columns and 'Loại nghỉ' in d.columns:
            d = d.rename(columns={'Loại nghỉ':'Lý do nghỉ'})
        for c in cols:
            if c not in d.columns: d[c] = ''
        # Dùng cùng bộ parse ngày với logic tổng Vi phạm để email/Excel chi tiết
        # luôn khớp 100% với số tiền trên bảng lương.
        d['__date'] = d['Ngày'].apply(_parse_vn_date)
        d['__key'] = d['Tên nhân viên'].apply(normalize_login_name)
        d['__penalty'] = d['Phạt vi phạm'].apply(_money_to_float)
        key = normalize_login_name(employee_name)
        d = d[(d['__key'] == key) & (d['__date'] >= start_date) & (d['__date'] <= end_date) & (d['__penalty'] > 0)].copy()
        if d.empty:
            return pd.DataFrame(columns=cols)
        d['Ngày'] = d['__date'].apply(lambda x: x.strftime('%d/%m/%Y') if x else '')
        d['Phạt vi phạm'] = d['__penalty']
        return d[cols].reset_index(drop=True)
    except Exception:
        return pd.DataFrame(columns=cols)


def _standardize_payroll_source(raw_df):
    """Chuẩn hóa đúng theo quy tắc người dùng: B=Thời gian, F=Loại, G=Tiền, I=Nhân viên."""
    if raw_df is None or raw_df.empty:
        return pd.DataFrame(columns=["Thời gian", "Sản phẩm/ Dịch vụ/ PT", "Tổng tiền", "NV tư vấn"])
    raw = raw_df.copy()
    # Tìm dòng tiêu đề; nếu không tìm được vẫn dùng vị trí cột B/F/G/I.
    header_idx = None
    for i in range(min(20, len(raw))):
        vals = [str(x).strip().casefold() for x in raw.iloc[i].tolist()]
        joined = " | ".join(vals)
        if "thời gian" in joined and ("sản phẩm" in joined or "dịch vụ" in joined) and "tổng tiền" in joined:
            header_idx = i
            break
    if header_idx is not None:
        data = raw.iloc[header_idx + 1:].copy()
    else:
        data = raw.copy()
    while data.shape[1] < 9:
        data[data.shape[1]] = ""
    out = pd.DataFrame({
        "Thời gian": data.iloc[:, 1],
        "Sản phẩm/ Dịch vụ/ PT": data.iloc[:, 5],
        "Tổng tiền": data.iloc[:, 6],
        "NV tư vấn": data.iloc[:, 8],
    })
    out = out.replace({None: ""})
    out["Thời gian_DT"] = pd.to_datetime(out["Thời gian"], dayfirst=True, errors="coerce")
    out["Tổng tiền"] = out["Tổng tiền"].apply(_money_to_float)
    out["NV tư vấn"] = out["NV tư vấn"].astype(str).str.strip()
    out["Sản phẩm/ Dịch vụ/ PT"] = out["Sản phẩm/ Dịch vụ/ PT"].astype(str).str.strip()
    return out.dropna(subset=["Thời gian_DT"])



def _timesoft_payroll_norm_col(value):
    """Tên cột chuẩn hóa để dò schema TimeSoft kể cả khi API đổi nhẹ tên field."""
    return re.sub(r"[^a-z0-9]+", "", remove_vietnamese_accents(str(value or "")).casefold())


def _timesoft_expand_payroll_candidate_frames(invoice_df):
    """
    TimeSoft có thể trả chi tiết dịch vụ trực tiếp ở Data hoặc lồng trong list chi tiết.
    Trả về nhiều DataFrame ứng viên để bộ dò tự chọn cấu trúc chứa dòng Tip tốt nhất.
    """
    if not isinstance(invoice_df, pd.DataFrame) or invoice_df.empty:
        return []

    frames = [("Data", invoice_df.copy())]

    # pd.json_normalize vẫn giữ các list-of-dict dưới dạng object; thử bung từng list đó.
    for col in invoice_df.columns:
        try:
            sample = None
            for v in invoice_df[col].tolist()[:100]:
                if isinstance(v, list) and v:
                    sample = v
                    break
                if isinstance(v, str) and v.strip().startswith("["):
                    try:
                        parsed = json.loads(v)
                        if isinstance(parsed, list) and parsed:
                            sample = parsed
                            break
                    except Exception:
                        pass
            if not isinstance(sample, list) or not sample or not isinstance(sample[0], dict):
                continue

            parent_rows = []
            child_rows = []
            for _, row in invoice_df.iterrows():
                raw_child = row.get(col)
                if isinstance(raw_child, str) and raw_child.strip().startswith("["):
                    try:
                        raw_child = json.loads(raw_child)
                    except Exception:
                        raw_child = []
                if not isinstance(raw_child, list):
                    raw_child = []
                for child in raw_child:
                    if not isinstance(child, dict):
                        continue
                    parent_rows.append(row.drop(labels=[col]).to_dict())
                    child_rows.append(child)

            if child_rows:
                parent_df = pd.DataFrame(parent_rows).reset_index(drop=True)
                child_df = pd.json_normalize(child_rows, sep=".").add_prefix("detail.").reset_index(drop=True)
                frames.append((f"Data.{col}", pd.concat([parent_df, child_df], axis=1)))
        except Exception:
            continue

    return frames


def _timesoft_pick_payroll_column(df, role):
    """
    Dò cột TimeSoft theo tên field + chất lượng dữ liệu.
    role: time | item | money | employee
    """
    if not isinstance(df, pd.DataFrame) or df.empty:
        return None, -1

    exact_aliases = {
        "time": {
            "createtime", "createtimestr", "createdate", "createdatestr", "invoicetime",
            "invoicedate", "datecreate", "datecreatestr", "time", "datetime", "ngay", "thoigian"
        },
        "item": {
            "productservicename", "serviceproductname", "servicename", "productname",
            "itemname", "nameproduct", "nameservice", "productservice", "service",
            "product", "sanphamdichvupt", "sanphamdichvu", "dichvu", "sanpham"
        },
        "money": {
            "totalmoney", "amount", "money", "totalamount", "actualmoney", "price",
            "totalprice", "paymentamount", "revenue", "value", "tongtien", "thanhtien"
        },
        "employee": {
            "employeename", "staffname", "consultantname", "advisorname", "salename",
            "employeeinfoname", "employee.name", "creatorname", "createdbyname",
            "nvtu van", "nvtuvan", "nhanvientuvan", "nhanvien"
        },
    }

    role_tokens = {
        "time": ["time", "date", "ngay", "thoigian", "create"],
        "item": ["service", "product", "item", "dichvu", "sanpham", "productservice", "name"],
        "money": ["money", "amount", "price", "revenue", "payment", "total", "tien", "thanhtien"],
        "employee": ["employee", "staff", "consult", "advisor", "sale", "nhanvien", "tuvan", "creator"],
    }

    avoid_tokens = {
        "time": ["end", "update", "modify", "checkout"],
        "item": ["employee", "customer", "branch", "room"],
        "money": ["discount", "debt", "tax", "vat", "fee"],
        "employee": ["customer", "client", "branch", "room"],
    }

    best_col, best_score = None, -10**9

    for col in df.columns:
        n = _timesoft_payroll_norm_col(col)
        s = df[col].dropna()
        if s.empty:
            continue
        sample = s.head(250)

        score = 0.0
        if n in exact_aliases.get(role, set()):
            score += 80
        score += 10 * sum(1 for tok in role_tokens.get(role, []) if tok in n)
        score -= 12 * sum(1 for tok in avoid_tokens.get(role, []) if tok in n)

        if role == "time":
            try:
                parsed = pd.to_datetime(sample.astype(str), dayfirst=True, errors="coerce")
                score += float(parsed.notna().mean()) * 45
            except Exception:
                pass

        elif role == "item":
            vals = sample.astype(str).str.strip()
            if len(vals):
                tip_ratio = vals.str.casefold().str.startswith("tip").mean()
                score += float(tip_ratio) * 160
                score += float(vals.ne("").mean()) * 10

        elif role == "money":
            vals = sample.apply(_money_to_float)
            if len(vals):
                numeric_ratio = pd.Series(vals).apply(lambda x: isinstance(x, numbers.Number)).mean()
                nonzero_ratio = pd.Series(vals).abs().gt(0).mean()
                score += float(numeric_ratio) * 15 + float(nonzero_ratio) * 25

        elif role == "employee":
            vals = sample.astype(str).str.strip()
            if len(vals):
                nonempty = vals.ne("").mean()
                # Tên nhân viên thường có chữ và không quá dài.
                human_like = vals.apply(
                    lambda x: bool(re.search(r"[A-Za-zÀ-ỹ]", x)) and len(x) <= 100
                ).mean()
                score += float(nonempty) * 12 + float(human_like) * 18

        if score > best_score:
            best_col, best_score = col, score

    return best_col, best_score


PAYROLL_TIMESOFT_EXPORT_COLUMNS = [
    "STT", "Thời gian", "Mã hóa đơn", "Mã KH2", "Tên KH",
    "Sản phẩm/ Dịch vụ/ PT", "Tổng tiền", "Giảm giá",
    "NV tư vấn", "Ghi chú", "Nhân viên tư vấn",
]


def _timesoft_payroll_known_employee_keys():
    """Tập tên nhân viên hệ thống để chọn đúng cột NV tư vấn trong dữ liệu TimeSoft."""
    try:
        creds = load_credentials_recent()
        if isinstance(creds, pd.DataFrame) and not creds.empty and "Tên nhân viên" in creds.columns:
            return {
                normalize_employee_match_name(v)
                for v in creds["Tên nhân viên"].astype(str).tolist()
                if normalize_employee_match_name(v)
            }
    except Exception:
        pass
    return set()


def _timesoft_pick_employee_column_for_tip(frame, item_col):
    """
    Chọn cột NV tư vấn theo chính các dòng Tip.

    V85.6: trước đây bộ dò có thể chọn nhầm một cột tên người khác trong JSON TimeSoft
    (ví dụ tên khách/creator), làm toàn bộ salary_map không khớp tài khoản và Tiền Lương = 0.
    Nay ưu tiên cột có tên giống NV tư vấn/Nhân viên tư vấn và có giá trị khớp hồ sơ nhân viên.
    """
    if not isinstance(frame, pd.DataFrame) or frame.empty or item_col not in frame.columns:
        return None, -1

    item_vals = frame[item_col].astype(str).str.strip()
    tip_mask = item_vals.str.casefold().str.startswith("tip")
    known_keys = _timesoft_payroll_known_employee_keys()

    exact_aliases = {
        "nvtuvan", "nhanvientuvan", "consultantname", "advisorname", "salename",
        "employeename", "employeeinfoname", "employee.name", "staffname",
        "detail.nvtuvan", "detail.nhanvientuvan", "detail.employeename",
        "detail.employeeinfoname", "detail.employee.name",
    }
    best_col, best_score = None, -10**9
    for col in frame.columns:
        if col == item_col:
            continue
        n = _timesoft_payroll_norm_col(col)
        vals = frame.loc[tip_mask, col] if bool(tip_mask.any()) else frame[col]
        vals = vals.dropna().astype(str).str.strip()
        if vals.empty:
            continue

        score = 0.0
        if n in exact_aliases:
            score += 250
        if any(tok in n for tok in ["tuvan", "consult", "advisor", "sale"]):
            score += 90
        if any(tok in n for tok in ["employee", "staff", "nhanvien"]):
            score += 45
        if any(tok in n for tok in ["customer", "client", "khach", "creator", "branch", "room"]):
            score -= 180

        human_ratio = vals.apply(lambda x: bool(re.search(r"[A-Za-zÀ-ỹ]", x)) and len(x) <= 100).mean()
        score += float(human_ratio) * 30

        if known_keys:
            normalized = vals.apply(normalize_employee_match_name)
            match_ratio = normalized.isin(known_keys).mean()
            match_count = int(normalized.isin(known_keys).sum())
            # Tín hiệu mạnh nhất: tên trong dòng Tip phải khớp danh sách tài khoản Vera.
            score += float(match_ratio) * 700 + min(match_count, 50) * 8

        if score > best_score:
            best_col, best_score = col, score

    # Fallback về bộ dò chung nếu không tìm được cột có dữ liệu.
    if best_col is None:
        return _timesoft_pick_payroll_column(frame, "employee")
    return best_col, best_score


def _timesoft_pick_optional_export_column(frame, aliases, tokens=(), avoid_tokens=()):
    """Dò cột phụ để dựng đúng 11 cột như file Excel xuất từ ReportSummaryInvoice."""
    if not isinstance(frame, pd.DataFrame) or frame.empty:
        return None
    alias_norm = {_timesoft_payroll_norm_col(x) for x in aliases}
    best_col, best_score = None, -10**9
    for col in frame.columns:
        n = _timesoft_payroll_norm_col(col)
        score = 0
        if n in alias_norm:
            score += 100
        score += 15 * sum(1 for t in tokens if t in n)
        score -= 25 * sum(1 for t in avoid_tokens if t in n)
        if score > best_score and score > 0:
            best_col, best_score = col, score
    return best_col


def _timesoft_build_excel_report_frame(frame, t_col, i_col, m_col, e_col):
    """
    Dựng dữ liệu TimeSoft về ĐÚNG format file Excel mẫu của người dùng:
    A:STT, B:Thời gian, C:Mã hóa đơn, D:Mã KH2, E:Tên KH,
    F:Sản phẩm/Dịch vụ/PT, G:Tổng tiền, H:Giảm giá,
    I:NV tư vấn, J:Ghi chú, K:Nhân viên tư vấn.

    build_payroll_table sau đó vẫn dùng đúng quy tắc lịch sử B/F/G/I.
    """
    report = pd.DataFrame(index=frame.index, columns=PAYROLL_TIMESOFT_EXPORT_COLUMNS)
    report[:] = ""

    report["Thời gian"] = frame[t_col]
    report["Sản phẩm/ Dịch vụ/ PT"] = frame[i_col]
    report["Tổng tiền"] = frame[m_col]
    report["NV tư vấn"] = frame[e_col]
    report["Nhân viên tư vấn"] = frame[e_col]

    optional_specs = {
        "STT": (["STT", "No", "Index", "RowNumber"], ["stt", "index", "row", "no"], []),
        "Mã hóa đơn": (["Mã hóa đơn", "InvoiceCode", "InvoiceNo", "CodeInvoice", "BillCode"], ["invoice", "bill", "code"], ["customer"]),
        "Mã KH2": (["Mã KH2", "CustomerCode", "Customer.Code", "ClientCode"], ["customer", "client", "code"], ["invoice"]),
        "Tên KH": (["Tên KH", "CustomerName", "Customer.Name", "ClientName"], ["customer", "client", "name"], ["employee"]),
        "Giảm giá": (["Giảm giá", "Discount", "DiscountMoney", "TotalDiscount"], ["discount", "giamgia"], []),
        "Ghi chú": (["Ghi chú", "Note", "Remark", "Description"], ["note", "remark", "ghichu"], []),
    }
    for dest, (aliases, tokens, avoid) in optional_specs.items():
        c = _timesoft_pick_optional_export_column(frame, aliases, tokens=tokens, avoid_tokens=avoid)
        if c is not None:
            report[dest] = frame[c]

    return report[PAYROLL_TIMESOFT_EXPORT_COLUMNS].copy()


def _standardize_payroll_source_from_timesoft(invoice_df):
    """
    V85.6 - Chuyển dữ liệu ReportSummaryInvoice về cùng format với file Excel mẫu.

    File Excel mẫu dùng chính xác:
    B = Thời gian, F = Sản phẩm/ Dịch vụ/ PT, G = Tổng tiền, I = NV tư vấn.
    Vì vậy TimeSoft được dựng lại thành 11 cột A:K trước, rồi đi qua cùng
    `_standardize_payroll_source()` như nguồn Upload Excel/Google Sheet.
    """
    base_cols = ["Thời gian", "Sản phẩm/ Dịch vụ/ PT", "Tổng tiền", "NV tư vấn", "Thời gian_DT"]
    if not isinstance(invoice_df, pd.DataFrame) or invoice_df.empty:
        return pd.DataFrame(columns=base_cols), "TimeSoft không có dữ liệu doanh thu trong kỳ đã chọn."

    best = None

    for frame_name, frame in _timesoft_expand_payroll_candidate_frames(invoice_df):
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            continue

        t_col, t_score = _timesoft_pick_payroll_column(frame, "time")
        i_col, i_score = _timesoft_pick_payroll_column(frame, "item")
        m_col, m_score = _timesoft_pick_payroll_column(frame, "money")

        # Chỉ sau khi đã biết cột dịch vụ mới chọn NV dựa trên các dòng Tip.
        e_col, e_score = _timesoft_pick_employee_column_for_tip(frame, i_col) if i_col is not None else (None, -1)

        # Fallback B/F/G/I CHỈ dùng khi frame thật sự có cấu trúc giống file Excel mẫu.
        if frame.shape[1] >= 9:
            cols = list(frame.columns)
            if t_col is None or t_score < 20:
                t_col = cols[1]
            if i_col is None or i_score < 20:
                i_col = cols[5]
            if m_col is None or m_score < 20:
                m_col = cols[6]
            if e_col is None or e_score < 20:
                e_col = cols[8]

        if any(c is None for c in [t_col, i_col, m_col, e_col]):
            continue
        if len({str(t_col), str(i_col), str(m_col), str(e_col)}) < 4:
            continue

        # Dựng đúng format Excel A:K rồi dùng CHUNG parser B/F/G/I.
        excel_like = _timesoft_build_excel_report_frame(frame, t_col, i_col, m_col, e_col)
        out = _standardize_payroll_source(excel_like)
        if out.empty:
            continue

        tip_mask = out["Sản phẩm/ Dịch vụ/ PT"].astype(str).str.strip().str.casefold().str.startswith("tip")
        tip_count = int(tip_mask.sum())
        tip_valid_emp = int(out.loc[tip_mask, "NV tư vấn"].astype(str).str.strip().ne("").sum()) if tip_count else 0
        tip_nonzero_money = int(out.loc[tip_mask, "Tổng tiền"].abs().gt(0).sum()) if tip_count else 0

        known_keys = _timesoft_payroll_known_employee_keys()
        tip_match_count = 0
        if tip_count and known_keys:
            tip_match_count = int(
                out.loc[tip_mask, "NV tư vấn"].astype(str)
                .apply(normalize_employee_match_name).isin(known_keys).sum()
            )

        # Frame có Tip + tên NV khớp hệ thống được ưu tiên tuyệt đối.
        quality = (
            tip_count * 2000
            + tip_valid_emp * 200
            + tip_nonzero_money * 200
            + tip_match_count * 1000
            + max(0, t_score) + max(0, i_score) + max(0, m_score) + max(0, e_score)
        )

        candidate = {
            "quality": quality,
            "frame_name": frame_name,
            "out": out,
            "excel_like": excel_like,
            "mapping": {
                "B · Thời gian": str(t_col),
                "F · Sản phẩm/ Dịch vụ/ PT": str(i_col),
                "G · Tổng tiền": str(m_col),
                "I · NV tư vấn": str(e_col),
            },
            "tip_count": tip_count,
            "tip_match_count": tip_match_count,
        }
        if best is None or candidate["quality"] > best["quality"]:
            best = candidate

    if best is None:
        cols_preview = ", ".join(map(str, list(invoice_df.columns)[:25]))
        return pd.DataFrame(columns=base_cols), (
            "Không tự nhận diện được dữ liệu TimeSoft theo format Excel báo cáo doanh thu hóa đơn. "
            f"Các cột TimeSoft nhận được: {cols_preview}"
        )

    # Không cho phép âm thầm tính bảng lương toàn 0 nếu mapping không có dòng Tip.
    if int(best.get("tip_count", 0) or 0) <= 0:
        return pd.DataFrame(columns=base_cols), (
            "TimeSoft đã trả dữ liệu nhưng chưa tìm thấy dòng 'Tip' theo format file Excel mẫu "
            "(B=Thời gian, F=Sản phẩm/Dịch vụ/PT, G=Tổng tiền, I=NV tư vấn). "
            "Hệ thống dừng tính để tránh tạo bảng lương Tiền Lương = 0 sai dữ liệu."
        )

    st.session_state["payroll_timesoft_mapping_v855"] = {
        "frame": best["frame_name"],
        "mapping": best["mapping"],
        "tip_count": best["tip_count"],
        "tip_match_count": best["tip_match_count"],
        "format": "Excel Báo cáo doanh thu hóa đơn A:K · B/F/G/I",
    }
    # Giữ một bản A:K trong session để Admin/debug/export nếu cần, không ảnh hưởng tính lương.
    st.session_state["payroll_timesoft_excel_like_v856"] = best["excel_like"]
    return best["out"], ""

def load_payroll_source_from_timesoft(start_date, end_date):
    """Lấy trực tiếp TimeSoft theo đúng kỳ lương và chuẩn hóa thành nguồn tính lương."""
    if not timesoft_is_configured():
        return pd.DataFrame(), "Chưa cấu hình tài khoản TimeSoft trong Secrets.", {}

    ok, msg, result = timesoft_direct_sync(start_date, end_date, force_login=False)
    if not ok:
        return pd.DataFrame(), msg, result or {}

    inv_df = (result or {}).get("summary_invoice_df")
    source_df, source_err = _standardize_payroll_source_from_timesoft(inv_df)
    if source_err:
        return source_df, source_err, result or {}

    return source_df, "", result or {}


@st.cache_data(ttl=60, show_spinner=False)
def load_payroll_source_from_google_sheet():
    try:
        client = get_gspread_client()
        if not client:
            return pd.DataFrame(), "Chưa cấu hình quyền Google Sheets."
        ws = client.open_by_key(PAYROLL_SOURCE_SHEET_ID).worksheet(PAYROLL_SOURCE_WORKSHEET)
        values = _gs_call_with_backoff(ws.get_all_values)
        if not values:
            return pd.DataFrame(), "Sheet dữ liệu lương đang trống."
        raw = pd.DataFrame(values)
        return _standardize_payroll_source(raw), ""
    except Exception as e:
        return pd.DataFrame(), f"Không đọc được nguồn dữ liệu lương mặc định: {e}"


def load_payroll_source_from_uploaded_excel(uploaded_file):
    try:
        if uploaded_file is None:
            return pd.DataFrame(), "Chưa chọn file dữ liệu lương."
        uploaded_file.seek(0)
        raw = pd.read_excel(uploaded_file, sheet_name=PAYROLL_SOURCE_WORKSHEET, header=None, engine="openpyxl")
        return _standardize_payroll_source(raw), ""
    except Exception as e:
        return pd.DataFrame(), f"Không đọc được sheet '{PAYROLL_SOURCE_WORKSHEET}': {e}"


def _official_payroll_period(year, month, period_no):
    """
    Trả về kỳ lương chính thức của VERA SPA.
    Kỳ 1 luôn từ ngày 01 đến hết ngày 15.
    Kỳ 2 luôn từ ngày 16 đến hết ngày cuối cùng của tháng.
    Không phụ thuộc ngày hiện tại và không cắt kỳ tại ngày hôm nay.
    """
    year, month, period_no = int(year), int(month), int(period_no)
    if period_no == 1:
        return date(year, month, 1), date(year, month, 15)
    if period_no == 2:
        last_day = calendar.monthrange(year, month)[1]
        return date(year, month, 16), date(year, month, last_day)
    raise ValueError('period_no chỉ nhận 1 hoặc 2')


def _canonicalize_payroll_period(start_date, end_date):
    """
    Chuẩn hóa một khoảng ngày thuộc cùng tháng về đúng kỳ lương chính thức.
    Hàm này dùng cho nghiệp vụ Tích lũy/lưu kỳ để tránh trường hợp kỳ 2 bị ghi 16→ngày hiện tại.
    Nếu khoảng ngày không cùng tháng hoặc cắt qua cả hai kỳ thì giữ nguyên để bảo toàn dữ liệu lịch sử cũ.
    """
    if not start_date or not end_date:
        return start_date, end_date
    if start_date.year != end_date.year or start_date.month != end_date.month:
        return start_date, end_date
    if end_date.day <= 15:
        return _official_payroll_period(start_date.year, start_date.month, 1)
    if start_date.day >= 16:
        return _official_payroll_period(start_date.year, start_date.month, 2)
    return start_date, end_date


def _is_official_payroll_period(start_date, end_date):
    if not start_date or not end_date:
        return False
    if start_date.year != end_date.year or start_date.month != end_date.month:
        return False
    period_no = 1 if start_date.day == 1 else (2 if start_date.day == 16 else 0)
    if not period_no:
        return False
    return (start_date, end_date) == _official_payroll_period(start_date.year, start_date.month, period_no)


def _matching_tichluy_history(hist, start_date, end_date):
    """
    Trả về (tổng tiền đã ghi, các key khớp) cho cùng một kỳ lương chính thức.
    Hỗ trợ dữ liệu cũ từng ghi Kỳ 2 dạng 16→ngày hiện tại hoặc Kỳ 1 dạng 01→ngày hiện tại.
    Chỉ gom legacy key khi khoảng đang tính là kỳ chính thức 01-15 / 16-cuối tháng.
    """
    if not isinstance(hist, dict):
        return 0.0, []
    exact_key = _tichluy_period_key(start_date, end_date)
    if not _is_official_payroll_period(start_date, end_date):
        amount = max(0.0, float(_money_to_float(hist.get(exact_key, 0))))
        return amount, ([exact_key] if exact_key in hist else [])

    matched = []
    total = 0.0
    for key, value in hist.items():
        ks, ke = _parse_tichluy_period_text(key)
        if not ks or not ke:
            continue
        cs, ce = _canonicalize_payroll_period(ks, ke)
        if cs == start_date and ce == end_date:
            matched.append(str(key))
            total += max(0.0, float(_money_to_float(value)))
    return total, matched


def resolve_payroll_period(preset, today=None, custom_range=None):
    """
    Kỳ lương cố định toàn hệ thống:
      • Kỳ 1: 01 → 15
      • Kỳ 2: 16 → ngày cuối tháng
    Không dùng ngày hiện tại làm ngày kết thúc kỳ.
    """
    today = today or get_vn_today()
    first_this = date(today.year, today.month, 1)
    prev_last = first_this - timedelta(days=1)
    if preset == "Kỳ 1 - Tháng này":
        return *_official_payroll_period(today.year, today.month, 1), ""
    if preset == "Kỳ 2 - Tháng này":
        return *_official_payroll_period(today.year, today.month, 2), ""
    if preset == "Kỳ 1 - Tháng trước":
        return *_official_payroll_period(prev_last.year, prev_last.month, 1), ""
    if preset == "Kỳ 2 - Tháng trước":
        return *_official_payroll_period(prev_last.year, prev_last.month, 2), ""
    return None, None, "Không xác định được kỳ lương."


def _period_penalty_by_employee(start_date, end_date, leave_primary=None, leave_secondary=None):
    """
    Tiền phạt CHỈ lấy từ Google Sheet 1Kz0... (SHEET_DU_PHONG_ID), theo kỳ đang chọn.

    Quan trọng: parse ngày từng ô bằng _parse_vn_date để hỗ trợ đồng thời
    dd-mm-yyyy, dd/mm/yyyy, yyyy-mm-dd và Excel/Google Sheets serial date.
    Cách cũ dùng pd.to_datetime cho cả Series có thể bỏ sót một số dòng khi
    dữ liệu ngày trong cùng Sheet không đồng nhất định dạng.
    """
    try:
        d = leave_primary.copy() if isinstance(leave_primary, pd.DataFrame) else load_backup_sheet_data()
        if d is None or d.empty:
            return {}
        if 'Ngày' not in d.columns or 'Tên nhân viên' not in d.columns or 'Phạt vi phạm' not in d.columns:
            return {}

        d = d.copy()
        d['Ngày_DT'] = d['Ngày'].apply(_parse_vn_date)
        d['__penalty'] = d['Phạt vi phạm'].apply(_money_to_float)
        d['__key'] = d['Tên nhân viên'].apply(normalize_login_name)

        # Chỉ giữ các dòng ngày hợp lệ, nằm trọn trong kỳ lương (inclusive 2 đầu).
        d = d[
            d['Ngày_DT'].notna()
            & (d['Ngày_DT'] >= start_date)
            & (d['Ngày_DT'] <= end_date)
        ].copy()
        if d.empty:
            return {}

        # Cộng trực tiếp TẤT CẢ dòng Phạt vi phạm của cùng nhân viên trong kỳ.
        # Không dedupe theo ngày/lý do vì một nhân viên có thể có nhiều vi phạm cùng ngày.
        return d.groupby('__key', dropna=False)['__penalty'].sum().to_dict()
    except Exception:
        return {}


def get_period_penalty_audit(employee_name, start_date, end_date, leave_primary=None):
    """Trả chi tiết các dòng phạt đã được cộng cho một nhân viên trong kỳ, không phát sinh thêm API read nếu đã truyền DataFrame."""
    try:
        d = leave_primary.copy() if isinstance(leave_primary, pd.DataFrame) else load_backup_sheet_data()
        if d is None or d.empty:
            return pd.DataFrame(columns=['Ngày', 'Tên nhân viên', 'Lý do nghỉ', 'Chi tiết', 'Phạt vi phạm'])
        for c in ['Ngày','Tên nhân viên','Lý do nghỉ','Chi tiết','Phạt vi phạm']:
            if c not in d.columns:
                d[c] = ''
        d['__date'] = d['Ngày'].apply(_parse_vn_date)
        d['__key'] = d['Tên nhân viên'].apply(normalize_login_name)
        d['__penalty'] = d['Phạt vi phạm'].apply(_money_to_float)
        key = normalize_login_name(employee_name)
        d = d[(d['__key'] == key) & d['__date'].notna() & (d['__date'] >= start_date) & (d['__date'] <= end_date) & (d['__penalty'] != 0)].copy()
        d['Ngày'] = d['__date'].apply(lambda x: x.strftime('%d/%m/%Y') if x else '')
        d['Phạt vi phạm'] = d['__penalty']
        return d[['Ngày','Tên nhân viên','Lý do nghỉ','Chi tiết','Phạt vi phạm']].reset_index(drop=True)
    except Exception:
        return pd.DataFrame(columns=['Ngày', 'Tên nhân viên', 'Lý do nghỉ', 'Chi tiết', 'Phạt vi phạm'])



def _next_official_payroll_period(start_date, end_date):
    """Trả về kỳ lương chính thức kế tiếp (01-15 hoặc 16-cuối tháng)."""
    cs, ce = _canonicalize_payroll_period(start_date, end_date)
    if cs.day == 1:
        return _official_payroll_period(cs.year, cs.month, 2)
    if cs.month == 12:
        return _official_payroll_period(cs.year + 1, 1, 1)
    return _official_payroll_period(cs.year, cs.month + 1, 1)


def _violation_debt_source_key(kind, start_date, end_date, employee_name):
    return f"{kind}|{start_date.isoformat()}|{end_date.isoformat()}|{normalize_login_name(employee_name)}"


@st.cache_resource(show_spinner=False)
def _ensure_violation_debt_storage():
    """Tạo/lấy sheet NoViPham dùng làm sổ nghĩa vụ Vi phạm chuyển kỳ."""
    client = get_gspread_client()
    if not client:
        return None, "Chưa cấu hình quyền kết nối Google Sheets."
    try:
        ss = client.open_by_key(SHEET_MAT_KHAU_ID)
        ws = _get_or_create_worksheet(ss, VIOLATION_DEBT_WORKSHEET, rows=3000, cols=20)
        header = _gs_call_with_backoff(ws.row_values, 1)
        if not header or header[:len(VIOLATION_DEBT_HEADERS)] != VIOLATION_DEBT_HEADERS:
            gspread_update_range(ws, "A1:N1", [VIOLATION_DEBT_HEADERS])
        return ws, ""
    except Exception as e:
        return None, f"Lỗi khởi tạo sheet {VIOLATION_DEBT_WORKSHEET}: {e}"


def _load_violation_debt_ledger_from_sheets():
    """Đọc sổ nợ Vi phạm một lần, có cache để tránh quota Sheets."""
    ws, err = _ensure_violation_debt_storage()
    if err or ws is None:
        return pd.DataFrame(columns=VIOLATION_DEBT_HEADERS + ['__sheet_row'])
    try:
        vals = _gs_call_with_backoff(ws.get_all_values)
        if not vals:
            return pd.DataFrame(columns=VIOLATION_DEBT_HEADERS + ['__sheet_row'])
        header = [str(x).strip() for x in vals[0]]
        pos = {name: (header.index(name) if name in header else None) for name in VIOLATION_DEBT_HEADERS}
        rows = []
        for sheet_row, row in enumerate(vals[1:], start=2):
            if not any(str(v).strip() for v in row):
                continue
            item = {name: (row[p] if p is not None and p < len(row) else '') for name, p in pos.items()}
            item['__sheet_row'] = sheet_row
            rows.append(item)
        return pd.DataFrame(rows) if rows else pd.DataFrame(columns=VIOLATION_DEBT_HEADERS + ['__sheet_row'])
    except Exception:
        return pd.DataFrame(columns=VIOLATION_DEBT_HEADERS + ['__sheet_row'])

@st.cache_data(ttl=30, show_spinner=False)
def load_violation_debt_ledger():
    """V75: đọc qua PostgreSQL dùng chung giữa các Cloud Run instance; Google Sheets là nguồn đồng bộ dự phòng."""
    if vpg is not None and vpg.is_enabled():
        return vpg.load_dataset(
            "violation_debt",
            _load_violation_debt_ledger_from_sheets,
            ttl_seconds=int(os.getenv("VERA_PG_TTL_VIOLATION_DEBT", "60")),
        )
    return _load_violation_debt_ledger_from_sheets()


def _clear_violation_debt_cache():
    try:
        load_violation_debt_ledger.clear()
        if vpg is not None and vpg.is_enabled():
            try:
                vpg.invalidate_dataset("violation_debt")
            except Exception:
                pass
    except Exception:
        pass


def _is_open_violation_debt_status(value):
    txt = normalize_login_name(value)
    return txt in {'', normalize_login_name(VIOLATION_DEBT_OPEN_STATUS), 'chua hoan thanh'}


def get_violation_debt_state(start_date, end_date, employee_names=None):
    """
    Trả về:
      due_map: nghĩa vụ từ kỳ trước đã tới hạn khấu trừ ở kỳ đang tính.
      deferred_current_map: tiền Vi phạm của chính kỳ đang tính mà Admin đã chủ động hoãn.
      active_df: các nghĩa vụ còn mở liên quan.

    Lưu ý: khoản "Tạm hoãn vi phạm" của một kỳ vẫn phải được trừ khỏi Vi phạm gốc
    của kỳ đó kể cả sau này nghĩa vụ đã được hoàn thành ở kỳ kế tiếp. Nhờ vậy khi
    mở lại/tính lại lịch sử, số của kỳ cũ không bị thay đổi ngược trở lại.
    """
    d = load_violation_debt_ledger().copy()
    if d is None or d.empty:
        return {}, {}, pd.DataFrame(columns=VIOLATION_DEBT_HEADERS)

    allowed = None
    if employee_names is not None:
        allowed = {normalize_login_name(x) for x in employee_names if normalize_login_name(x)}

    due_map = {}
    deferred_current_map = {}
    active_rows = []
    for _, r in d.iterrows():
        emp = str(r.get('Tên nhân viên', '')).strip()
        key = normalize_login_name(emp)
        if not key or (allowed is not None and key not in allowed):
            continue
        amount = max(0.0, float(_money_to_float(r.get('Số tiền', 0))))
        if amount <= 0:
            continue
        src_start = _parse_vn_date(r.get('Kỳ phát sinh từ', ''))
        src_end = _parse_vn_date(r.get('Kỳ phát sinh đến', ''))
        due_from = _parse_vn_date(r.get('Bắt đầu trừ từ', ''))
        debt_type = str(r.get('Loại', '')).strip()

        # Lịch sử tạm hoãn là thuộc tính của kỳ phát sinh, không phụ thuộc trạng thái
        # nghĩa vụ ở thời điểm hiện tại.
        if normalize_login_name(debt_type) == normalize_login_name('Tạm hoãn vi phạm') and src_start and src_end:
            cs, ce = _canonicalize_payroll_period(src_start, src_end)
            if cs == start_date and ce == end_date:
                deferred_current_map[key] = deferred_current_map.get(key, 0.0) + amount

        # Chỉ nghĩa vụ còn mở mới được cộng vào kỳ hiện tại.
        if not _is_open_violation_debt_status(r.get('Trạng thái', '')):
            continue
        if due_from and due_from <= start_date:
            due_map[key] = due_map.get(key, 0.0) + amount
        active_rows.append(r.to_dict())

    active_df = pd.DataFrame(active_rows) if active_rows else pd.DataFrame(columns=VIOLATION_DEBT_HEADERS)
    return due_map, deferred_current_map, active_df


def get_open_negative_payroll_debts():
    """
    Danh sách các khoản nợ do Thực nhận âm còn Chưa hoàn thành.

    Chỉ lấy Loại = "Âm thực nhận" trong sheet NoViPham; không trộn với
    khoản Admin chủ động "Tạm hoãn vi phạm".

    Trả về:
      summary_df: tổng nợ theo nhân viên.
      detail_df:  chi tiết từng kỳ phát sinh còn mở.
    """
    d = load_violation_debt_ledger().copy()
    if d is None or d.empty:
        empty_summary = pd.DataFrame(columns=["Tên nhân viên", "Tổng còn nợ", "Số kỳ còn nợ", "Kỳ nợ gần nhất", "Bắt đầu trừ từ"])
        empty_detail = pd.DataFrame(columns=["Tên nhân viên", "Số tiền", "Kỳ phát sinh từ", "Kỳ phát sinh đến", "Bắt đầu trừ từ", "Nội dung", "Trạng thái"])
        return empty_summary, empty_detail

    rows = []
    for _, r in d.iterrows():
        debt_type = normalize_login_name(r.get('Loại', ''))
        if debt_type != normalize_login_name('Âm thực nhận'):
            continue
        if not _is_open_violation_debt_status(r.get('Trạng thái', '')):
            continue
        emp = str(r.get('Tên nhân viên', '')).strip()
        amount = max(0.0, float(_money_to_float(r.get('Số tiền', 0))))
        if not emp or amount <= 0:
            continue
        src_start = _parse_vn_date(r.get('Kỳ phát sinh từ', ''))
        src_end = _parse_vn_date(r.get('Kỳ phát sinh đến', ''))
        due_from = _parse_vn_date(r.get('Bắt đầu trừ từ', ''))
        rows.append({
            "Tên nhân viên": emp,
            "Số tiền": int(round(amount)),
            "Kỳ phát sinh từ": src_start.strftime('%d/%m/%Y') if src_start else str(r.get('Kỳ phát sinh từ', '')).strip(),
            "Kỳ phát sinh đến": src_end.strftime('%d/%m/%Y') if src_end else str(r.get('Kỳ phát sinh đến', '')).strip(),
            "Bắt đầu trừ từ": due_from.strftime('%d/%m/%Y') if due_from else str(r.get('Bắt đầu trừ từ', '')).strip(),
            "Nội dung": str(r.get('Nội dung', '')).strip() or VIOLATION_DEBT_CONTENT,
            "Trạng thái": str(r.get('Trạng thái', '')).strip() or VIOLATION_DEBT_OPEN_STATUS,
            "__src_start": src_start,
            "__src_end": src_end,
            "__due_from": due_from,
        })

    if not rows:
        empty_summary = pd.DataFrame(columns=["Tên nhân viên", "Tổng còn nợ", "Số kỳ còn nợ", "Kỳ nợ gần nhất", "Bắt đầu trừ từ"])
        empty_detail = pd.DataFrame(columns=["Tên nhân viên", "Số tiền", "Kỳ phát sinh từ", "Kỳ phát sinh đến", "Bắt đầu trừ từ", "Nội dung", "Trạng thái"])
        return empty_summary, empty_detail

    detail = pd.DataFrame(rows)
    # Kỳ cũ trước, kỳ mới sau để Admin dễ theo dõi diễn biến.
    detail = detail.sort_values(
        by=["__src_start", "Tên nhân viên"],
        key=lambda s: s.map(lambda x: x.toordinal() if hasattr(x, 'toordinal') else 99999999) if s.name == "__src_start" else s,
        na_position='last'
    ).reset_index(drop=True)

    summary_rows = []
    for emp, grp in detail.groupby("Tên nhân viên", sort=True):
        total = int(round(grp["Số tiền"].apply(_money_to_float).sum()))
        latest = grp.copy()
        latest["__sort"] = latest["__src_start"].apply(lambda x: x.toordinal() if hasattr(x, 'toordinal') else -1)
        latest_row = latest.sort_values("__sort", ascending=False).iloc[0]
        due_dates = [x for x in grp["__due_from"].tolist() if hasattr(x, 'strftime')]
        earliest_due = min(due_dates) if due_dates else None
        latest_period = f"{latest_row['Kỳ phát sinh từ']} - {latest_row['Kỳ phát sinh đến']}"
        summary_rows.append({
            "Tên nhân viên": emp,
            "Tổng còn nợ": total,
            "Số kỳ còn nợ": int(len(grp)),
            "Kỳ nợ gần nhất": latest_period,
            "Bắt đầu trừ từ": earliest_due.strftime('%d/%m/%Y') if earliest_due else str(latest_row.get('Bắt đầu trừ từ', '')).strip(),
        })

    summary = pd.DataFrame(summary_rows).sort_values(["Tổng còn nợ", "Tên nhân viên"], ascending=[False, True]).reset_index(drop=True)
    detail = detail.drop(columns=["__src_start", "__src_end", "__due_from"], errors='ignore')
    return summary, detail



def get_open_admin_deferred_violation_debts():
    """
    Danh sách các nghĩa vụ Vi phạm do Admin chủ động tạm hoãn còn Chưa hoàn thành.

    Chỉ lấy Loại = "Tạm hoãn vi phạm" trong sheet NoViPham.
    Trả về:
      summary_df: tổng nghĩa vụ theo nhân viên.
      detail_df:  chi tiết từng kỳ Admin đã chủ động tạm hoãn.
    """
    d = load_violation_debt_ledger().copy()
    summary_cols = ["Tên nhân viên", "Tổng tạm hoãn", "Số kỳ tạm hoãn", "Kỳ tạm hoãn gần nhất", "Bắt đầu trừ từ"]
    detail_cols = ["Tên nhân viên", "Số tiền", "Kỳ phát sinh từ", "Kỳ phát sinh đến", "Bắt đầu trừ từ", "Nội dung", "Loại", "Trạng thái"]
    if d is None or d.empty:
        return pd.DataFrame(columns=summary_cols), pd.DataFrame(columns=detail_cols)

    rows = []
    target_type = normalize_login_name('Tạm hoãn vi phạm')
    for _, r in d.iterrows():
        debt_type = normalize_login_name(r.get('Loại', ''))
        if debt_type != target_type:
            continue
        if not _is_open_violation_debt_status(r.get('Trạng thái', '')):
            continue
        emp = str(r.get('Tên nhân viên', '')).strip()
        amount = max(0.0, float(_money_to_float(r.get('Số tiền', 0))))
        if not emp or amount <= 0:
            continue
        src_start = _parse_vn_date(r.get('Kỳ phát sinh từ', ''))
        src_end = _parse_vn_date(r.get('Kỳ phát sinh đến', ''))
        due_from = _parse_vn_date(r.get('Bắt đầu trừ từ', ''))
        rows.append({
            "Tên nhân viên": emp,
            "Số tiền": int(round(amount)),
            "Kỳ phát sinh từ": src_start.strftime('%d/%m/%Y') if src_start else str(r.get('Kỳ phát sinh từ', '')).strip(),
            "Kỳ phát sinh đến": src_end.strftime('%d/%m/%Y') if src_end else str(r.get('Kỳ phát sinh đến', '')).strip(),
            "Bắt đầu trừ từ": due_from.strftime('%d/%m/%Y') if due_from else str(r.get('Bắt đầu trừ từ', '')).strip(),
            "Nội dung": str(r.get('Nội dung', '')).strip() or VIOLATION_DEBT_CONTENT,
            "Loại": str(r.get('Loại', '')).strip() or 'Tạm hoãn vi phạm',
            "Trạng thái": str(r.get('Trạng thái', '')).strip() or VIOLATION_DEBT_OPEN_STATUS,
            "__src_start": src_start,
            "__due_from": due_from,
        })

    if not rows:
        return pd.DataFrame(columns=summary_cols), pd.DataFrame(columns=detail_cols)

    detail = pd.DataFrame(rows)
    detail['__sort_src'] = detail['__src_start'].apply(lambda x: x.toordinal() if hasattr(x, 'toordinal') else 99999999)
    detail = detail.sort_values(['__sort_src', 'Tên nhân viên'], ascending=[True, True]).reset_index(drop=True)

    summary_rows = []
    for emp, grp in detail.groupby('Tên nhân viên', sort=True):
        total = int(round(grp['Số tiền'].apply(_money_to_float).sum()))
        latest = grp.sort_values('__sort_src', ascending=False).iloc[0]
        due_dates = [x for x in grp['__due_from'].tolist() if hasattr(x, 'strftime')]
        earliest_due = min(due_dates) if due_dates else None
        latest_period = f"{latest['Kỳ phát sinh từ']} - {latest['Kỳ phát sinh đến']}"
        summary_rows.append({
            'Tên nhân viên': emp,
            'Tổng tạm hoãn': total,
            'Số kỳ tạm hoãn': int(len(grp)),
            'Kỳ tạm hoãn gần nhất': latest_period,
            'Bắt đầu trừ từ': earliest_due.strftime('%d/%m/%Y') if earliest_due else str(latest.get('Bắt đầu trừ từ', '')).strip(),
        })

    summary = pd.DataFrame(summary_rows).sort_values(['Tổng tạm hoãn', 'Tên nhân viên'], ascending=[False, True]).reset_index(drop=True)
    detail = detail.drop(columns=['__src_start', '__due_from', '__sort_src'], errors='ignore')
    return summary, detail


def defer_violation_to_next_period(employee_name, amount, start_date, end_date, updated_by):
    """Admin chuyển một phần/toàn bộ Vi phạm của kỳ hiện tại sang các kỳ kế tiếp."""
    employee_name = str(employee_name).strip()
    amount = max(0.0, float(_money_to_float(amount)))
    if not employee_name or amount <= 0:
        return False, "Vui lòng chọn nhân viên và nhập số tiền lớn hơn 0."

    ws, err = _ensure_violation_debt_storage()
    if err or ws is None:
        return False, err or "Không mở được sổ nghĩa vụ Vi phạm."
    try:
        vals = _gs_call_with_backoff(ws.get_all_values)
        header = [str(x).strip() for x in (vals[0] if vals else VIOLATION_DEBT_HEADERS)]
        if header[:len(VIOLATION_DEBT_HEADERS)] != VIOLATION_DEBT_HEADERS:
            gspread_update_range(ws, "A1:N1", [VIOLATION_DEBT_HEADERS])
            vals = [VIOLATION_DEBT_HEADERS] + (vals[1:] if vals else [])
            header = VIOLATION_DEBT_HEADERS[:]
        source_key = _violation_debt_source_key('DEFER', start_date, end_date, employee_name)
        next_start, next_end = _next_official_payroll_period(start_date, end_date)
        now = datetime.now(VN_TZ)
        source_col = VIOLATION_DEBT_HEADERS.index('Mã nguồn')
        amount_col = VIOLATION_DEBT_HEADERS.index('Số tiền')
        existing_row = None
        existing_amount = 0.0
        existing_status = ''
        status_col = VIOLATION_DEBT_HEADERS.index('Trạng thái')
        for row_idx, row in enumerate(vals[1:], start=2):
            key_value = row[source_col] if source_col < len(row) else ''
            if str(key_value).strip() == source_key:
                existing_row = row_idx
                existing_amount = float(_money_to_float(row[amount_col] if amount_col < len(row) else 0))
                existing_status = row[status_col] if status_col < len(row) else ''
                break
        if existing_row and not _is_open_violation_debt_status(existing_status):
            return False, "Khoản Vi phạm đã tạm hoãn của kỳ này đã được khấu trừ ở kỳ sau, không thể mở lại từ kỳ cũ."
        total_amount = existing_amount + amount
        _defer_note_text = f"Trừ kỳ lương kế tiếp: {total_amount:,.0f} đ".replace(',', '.')
        row_values = [
            (existing_row - 1 if existing_row else max(1, len(vals))),
            employee_name,
            int(round(total_amount)),
            _defer_note_text,
            "Tạm hoãn vi phạm",
            start_date.strftime('%d/%m/%Y'),
            end_date.strftime('%d/%m/%Y'),
            next_start.strftime('%d/%m/%Y'),
            VIOLATION_DEBT_OPEN_STATUS,
            source_key,
            now.strftime('%d/%m/%Y'),
            now.strftime('%H:%M:%S'),
            str(updated_by),
            "",
        ]
        if existing_row:
            gspread_update_range(ws, f"A{existing_row}:N{existing_row}", [row_values])
        else:
            ws.append_row(row_values, value_input_option='USER_ENTERED')
        _clear_violation_debt_cache()
        return True, (
            f"Đã chuyển thêm {amount:,.0f}đ Vi phạm của {employee_name} sang kỳ kế tiếp. "
            f"Tổng nghĩa vụ đang hoãn từ kỳ này: {total_amount:,.0f}đ."
        ).replace(',', '.')
    except Exception as e:
        return False, f"Lỗi lưu nghĩa vụ Vi phạm: {e}"


def _renumber_violation_debt_stt(ws):
    """Đánh lại STT cột A từ dòng 2 sau khi xóa nghĩa vụ."""
    try:
        vals = _gs_call_with_backoff(ws.get_all_values)
        if not vals or len(vals) <= 1:
            return
        # Chỉ đánh STT cho các dòng có dữ liệu thật ở B:N.
        numbers = []
        n = 0
        for row in vals[1:]:
            has_data = any(str(v).strip() for v in row[1:14]) if len(row) > 1 else False
            if has_data:
                n += 1
                numbers.append([n])
            else:
                numbers.append([''])
        if numbers:
            gspread_update_range(ws, f"A2:A{len(numbers)+1}", numbers)
    except Exception:
        pass


def update_violation_debt_obligation(sheet_row, amount, content, due_from, updated_by):
    """
    Admin sửa một nghĩa vụ Vi phạm đang mở.
    Giữ nguyên nhân viên / loại / kỳ phát sinh / mã nguồn để không phá lịch sử,
    chỉ cho chỉnh số tiền, nội dung và ngày bắt đầu trừ.
    """
    try:
        sheet_row = int(sheet_row)
    except Exception:
        return False, "Không xác định được dòng nghĩa vụ cần sửa."
    if sheet_row < 2:
        return False, "Dòng nghĩa vụ không hợp lệ."
    amount = max(0.0, float(_money_to_float(amount)))
    if amount <= 0:
        return False, "Số tiền nghĩa vụ phải lớn hơn 0. Nếu không còn nghĩa vụ, hãy dùng nút Xóa."
    if not hasattr(due_from, 'strftime'):
        due_from = _parse_vn_date(due_from)
    if due_from is None:
        return False, "Ngày bắt đầu trừ không hợp lệ."

    ws, err = _ensure_violation_debt_storage()
    if err or ws is None:
        return False, err or "Không mở được sổ nghĩa vụ Vi phạm."
    try:
        row = _gs_call_with_backoff(ws.row_values, sheet_row)
        if not row or not any(str(v).strip() for v in row):
            return False, "Nghĩa vụ này không còn tồn tại. Hãy tải lại danh sách."
        while len(row) < len(VIOLATION_DEBT_HEADERS):
            row.append('')
        pos = {name: VIOLATION_DEBT_HEADERS.index(name) for name in VIOLATION_DEBT_HEADERS}
        if not _is_open_violation_debt_status(row[pos['Trạng thái']]):
            return False, "Chỉ có thể sửa nghĩa vụ đang ở trạng thái Chưa hoàn thành."

        now = datetime.now(VN_TZ)
        row[pos['Số tiền']] = int(round(amount))
        row[pos['Nội dung']] = str(content).strip() or VIOLATION_DEBT_CONTENT
        row[pos['Bắt đầu trừ từ']] = due_from.strftime('%d/%m/%Y')
        row[pos['Ngày cập nhật']] = now.strftime('%d/%m/%Y')
        row[pos['Giờ cập nhật']] = now.strftime('%H:%M:%S')
        row[pos['Người cập nhật']] = str(updated_by)
        gspread_update_range(ws, f"A{sheet_row}:N{sheet_row}", [row[:len(VIOLATION_DEBT_HEADERS)]])
        _clear_violation_debt_cache()
        return True, "Đã cập nhật Nghĩa vụ Vi phạm."
    except Exception as e:
        return False, f"Lỗi cập nhật Nghĩa vụ Vi phạm: {e}"


def delete_violation_debt_obligation(sheet_row, updated_by=''):
    """Admin xóa hẳn một nghĩa vụ Vi phạm đang mở và đánh lại STT."""
    try:
        sheet_row = int(sheet_row)
    except Exception:
        return False, "Không xác định được dòng nghĩa vụ cần xóa."
    if sheet_row < 2:
        return False, "Dòng nghĩa vụ không hợp lệ."
    ws, err = _ensure_violation_debt_storage()
    if err or ws is None:
        return False, err or "Không mở được sổ nghĩa vụ Vi phạm."
    try:
        row = _gs_call_with_backoff(ws.row_values, sheet_row)
        if not row or not any(str(v).strip() for v in row):
            return False, "Nghĩa vụ này không còn tồn tại."
        while len(row) < len(VIOLATION_DEBT_HEADERS):
            row.append('')
        status = row[VIOLATION_DEBT_HEADERS.index('Trạng thái')]
        if not _is_open_violation_debt_status(status):
            return False, "Chỉ xóa trực tiếp nghĩa vụ đang Chưa hoàn thành."
        _gs_call_with_backoff(ws.delete_rows, sheet_row)
        _renumber_violation_debt_stt(ws)
        _clear_violation_debt_cache()
        return True, "Đã xóa Nghĩa vụ Vi phạm."
    except Exception as e:
        return False, f"Lỗi xóa Nghĩa vụ Vi phạm: {e}"


def refresh_current_payroll_violation_debt(payroll_df, start_date, end_date):
    """Tính lại riêng cột Vi phạm kỳ trước theo ledger mới nhất rồi tính lại Thực nhận."""
    if payroll_df is None or payroll_df.empty:
        return payroll_df
    d = payroll_df.copy()
    names = d.get('Tên Hệ thống', pd.Series(dtype=str)).astype(str).tolist()
    due_map, _, _ = get_violation_debt_state(start_date, end_date, names)
    if 'Vi phạm kỳ trước' not in d.columns:
        d['Vi phạm kỳ trước'] = 0.0
    for idx, r in d.iterrows():
        key = normalize_login_name(r.get('Tên Hệ thống', ''))
        d.at[idx, 'Vi phạm kỳ trước'] = float(_money_to_float(due_map.get(key, 0)))
    return recalculate_payroll_net(d)


def get_all_open_violation_debts_for_admin():
    """Trả toàn bộ nghĩa vụ Vi phạm đang mở, kể cả nhân viên không nằm trong kỳ lương hiện tại."""
    d = load_violation_debt_ledger().copy()
    if d is None or d.empty:
        return pd.DataFrame(columns=VIOLATION_DEBT_HEADERS + ['__sheet_row'])
    keep = []
    for _, r in d.iterrows():
        amount = max(0.0, float(_money_to_float(r.get('Số tiền', 0))))
        if amount <= 0 or not _is_open_violation_debt_status(r.get('Trạng thái', '')):
            continue
        keep.append(r.to_dict())
    return pd.DataFrame(keep) if keep else pd.DataFrame(columns=VIOLATION_DEBT_HEADERS + ['__sheet_row'])


def _upsert_negative_violation_debt(ws, vals, employee_name, amount, start_date, end_date, updated_by):
    """Tạo/cập nhật khoản âm Thực nhận của một kỳ, không tạo trùng khi lưu lại cùng kỳ."""
    source_key = _violation_debt_source_key('NEG', start_date, end_date, employee_name)
    source_col = VIOLATION_DEBT_HEADERS.index('Mã nguồn')
    next_start, _ = _next_official_payroll_period(start_date, end_date)
    now = datetime.now(VN_TZ)
    existing_row = None
    existing_status = ''
    for row_idx, row in enumerate(vals[1:], start=2):
        key_value = row[source_col] if source_col < len(row) else ''
        if str(key_value).strip() == source_key:
            existing_row = row_idx
            status_col = VIOLATION_DEBT_HEADERS.index('Trạng thái')
            existing_status = row[status_col] if status_col < len(row) else ''
            break

    # Nếu khoản của kỳ này đã được khấu trừ xong ở kỳ sau thì không tự mở lại do chỉnh lịch sử.
    if existing_row and not _is_open_violation_debt_status(existing_status):
        return None

    amount = max(0.0, float(_money_to_float(amount)))
    # Không tạo dòng rỗng cho nhân viên có Thực nhận >= 0. Chỉ cập nhật về hoàn thành
    # nếu trước đó đã tồn tại một khoản âm của chính kỳ này.
    if amount <= 0 and not existing_row:
        return None
    status = VIOLATION_DEBT_OPEN_STATUS if amount > 0 else VIOLATION_DEBT_DONE_STATUS
    row_values = [
        (existing_row - 1 if existing_row else max(1, len(vals))),
        str(employee_name).strip(),
        int(round(amount)),
        VIOLATION_DEBT_CONTENT,
        "Âm thực nhận",
        start_date.strftime('%d/%m/%Y'),
        end_date.strftime('%d/%m/%Y'),
        next_start.strftime('%d/%m/%Y'),
        status,
        source_key,
        now.strftime('%d/%m/%Y'),
        now.strftime('%H:%M:%S'),
        str(updated_by),
        "" if amount > 0 else f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}",
    ]
    if existing_row:
        gspread_update_range(ws, f"A{existing_row}:N{existing_row}", [row_values])
        return ('updated', existing_row, row_values)
    ws.append_row(row_values, value_input_option='USER_ENTERED')
    vals.append([str(x) for x in row_values])
    return ('appended', None, row_values)


def commit_violation_debts_after_payroll(payroll_df, start_date, end_date, saved_by):
    """
    Khi lưu bảng lương:
      1) Các nghĩa vụ cũ đã tới hạn và đã được đưa vào cột "Vi phạm kỳ trước" được đánh dấu hoàn thành.
      2) Nếu Thực nhận âm và Admin KHÔNG chủ động tạm hoãn người đó, phần âm được tự lưu thành nghĩa vụ mới.
      3) Nếu Admin đã chủ động tạm hoãn Vi phạm của người đó trong kỳ, không tự tạo thêm nợ âm để tránh trùng.
    """
    if payroll_df is None or payroll_df.empty:
        return True, ""
    ws, err = _ensure_violation_debt_storage()
    if err or ws is None:
        return False, err or "Không mở được sổ nghĩa vụ Vi phạm."
    try:
        vals = _gs_call_with_backoff(ws.get_all_values)
        if not vals:
            vals = [VIOLATION_DEBT_HEADERS]
        header = [str(x).strip() for x in vals[0]]
        if header[:len(VIOLATION_DEBT_HEADERS)] != VIOLATION_DEBT_HEADERS:
            gspread_update_range(ws, "A1:N1", [VIOLATION_DEBT_HEADERS])
            header = VIOLATION_DEBT_HEADERS[:]
            vals[0] = header
        pos = {name: header.index(name) for name in VIOLATION_DEBT_HEADERS}
        payroll_names = {
            normalize_login_name(x): str(x).strip()
            for x in payroll_df.get('Tên Hệ thống', pd.Series(dtype=str)).tolist()
            if normalize_login_name(x)
        }
        now = datetime.now(VN_TZ)
        settled_count = 0
        # Nghĩa vụ cũ tới hạn đã được cộng vào Vi phạm của kỳ này -> đóng khoản cũ.
        for row_idx, row in enumerate(vals[1:], start=2):
            emp = row[pos['Tên nhân viên']] if pos['Tên nhân viên'] < len(row) else ''
            emp_key = normalize_login_name(emp)
            if emp_key not in payroll_names:
                continue
            status = row[pos['Trạng thái']] if pos['Trạng thái'] < len(row) else ''
            if not _is_open_violation_debt_status(status):
                continue
            due_from = _parse_vn_date(row[pos['Bắt đầu trừ từ']] if pos['Bắt đầu trừ từ'] < len(row) else '')
            if not due_from or due_from > start_date:
                continue
            # Ghi trạng thái hoàn thành + kỳ đã khấu trừ; các cột khác giữ nguyên.
            gspread_update_range(ws, f"I{row_idx}:N{row_idx}", [[
                VIOLATION_DEBT_DONE_STATUS,
                row[pos['Mã nguồn']] if pos['Mã nguồn'] < len(row) else '',
                now.strftime('%d/%m/%Y'),
                now.strftime('%H:%M:%S'),
                str(saved_by),
                f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}",
            ]])
            # Đồng bộ snapshot trong bộ nhớ để bước upsert phía sau không đọc sai trạng thái.
            while len(row) < len(VIOLATION_DEBT_HEADERS):
                row.append('')
            row[pos['Trạng thái']] = VIOLATION_DEBT_DONE_STATUS
            settled_count += 1

        # V50: xác định nhân viên Admin đã chủ động tạm hoãn Vi phạm của CHÍNH kỳ này.
        # Những người này không được tự tạo thêm một khoản "Âm thực nhận" khác, tránh ghi nợ kép.
        manual_adjusted_keys = set()
        for row in vals[1:]:
            emp = row[pos['Tên nhân viên']] if pos['Tên nhân viên'] < len(row) else ''
            emp_key = normalize_login_name(emp)
            debt_type = row[pos['Loại']] if pos['Loại'] < len(row) else ''
            if normalize_login_name(debt_type) != normalize_login_name('Tạm hoãn vi phạm'):
                continue
            src_start = _parse_vn_date(row[pos['Kỳ phát sinh từ']] if pos['Kỳ phát sinh từ'] < len(row) else '')
            src_end = _parse_vn_date(row[pos['Kỳ phát sinh đến']] if pos['Kỳ phát sinh đến'] < len(row) else '')
            if src_start and src_end:
                cs, ce = _canonicalize_payroll_period(src_start, src_end)
                if cs == start_date and ce == end_date and emp_key:
                    manual_adjusted_keys.add(emp_key)

        negative_count = 0
        negative_total = 0.0
        auto_negative_names = []
        for _, r in payroll_df.iterrows():
            emp = str(r.get('Tên Hệ thống', '')).strip()
            emp_key = normalize_login_name(emp)
            if not emp_key:
                continue
            net = float(_money_to_float(r.get('Số tiền thực nhận', 0)))

            if emp_key in manual_adjusted_keys:
                # Nếu trước đó đã từng lưu tự động phần âm của cùng kỳ rồi sau đó Admin mới
                # chủ động tạm hoãn, đóng khoản tự động cũ để không tồn tại song song hai nghĩa vụ.
                _upsert_negative_violation_debt(ws, vals, emp, 0.0, start_date, end_date, saved_by)
                continue

            debt_amount = abs(net) if net < 0 else 0.0
            result = _upsert_negative_violation_debt(
                ws, vals, emp, debt_amount, start_date, end_date, saved_by
            )
            if debt_amount > 0 and result is not None:
                negative_count += 1
                negative_total += debt_amount
                auto_negative_names.append(emp)

        _clear_violation_debt_cache()
        msg_parts = []
        if settled_count:
            msg_parts.append(f"đã khấu trừ {settled_count} nghĩa vụ Vi phạm cũ")
        if negative_count:
            msg_parts.append(
                (f"đã tự lưu {negative_total:,.0f}đ phần Thực nhận âm của {negative_count} nhân viên "
                 f"không được Admin điều chỉnh sang kỳ kế tiếp: {', '.join(auto_negative_names)}").replace(',', '.')
            )
        return True, '; '.join(msg_parts)
    except Exception as e:
        return False, f"Lỗi cập nhật nghĩa vụ Vi phạm chuyển kỳ: {e}"


def is_payroll_period_2(start_date, end_date):
    """Kỳ 2 cố định: ngày 16 đến ngày cuối cùng của cùng một tháng."""
    try:
        s = pd.to_datetime(start_date).date()
        e = pd.to_datetime(end_date).date()
        last_day = calendar.monthrange(s.year, s.month)[1]
        return s.year == e.year and s.month == e.month and s.day == 16 and e.day == last_day
    except Exception:
        return False


def build_payroll_table(source_df, credentials_df, start_date, end_date, leave_primary=None, leave_secondary=None, default_living_expense=150000, default_locker_support=80000, leader_responsibility_allowance=0):
    """Tổng hợp lương: chỉ cộng G khi F bắt đầu bằng 'Tip', nhóm theo tên nhân viên ở cột I."""
    if source_df is None or source_df.empty:
        return pd.DataFrame(columns=PAYROLL_COLUMNS), []
    src = source_df.copy()
    src['Ngày'] = src['Thời gian_DT'].dt.date
    src = src[(src['Ngày'] >= start_date) & (src['Ngày'] <= end_date)]
    tip_mask = src['Sản phẩm/ Dịch vụ/ PT'].astype(str).str.strip().str.casefold().str.startswith('tip')
    tip = src[tip_mask].copy()
    tip['__key'] = tip['NV tư vấn'].apply(normalize_login_name)
    salary_map = tip.groupby('__key')['Tổng tiền'].sum().to_dict() if not tip.empty else {}
    tip_count_map = tip.groupby('__key').size().to_dict() if not tip.empty else {}

    creds = credentials_df.copy() if credentials_df is not None else pd.DataFrame()
    if creds.empty:
        return pd.DataFrame(columns=PAYROLL_COLUMNS), sorted(set(tip['NV tư vấn'].tolist())) if not tip.empty else []
    creds = creds[creds['Tên nhân viên'].astype(str).str.strip() != ''].copy()
    # Loại dòng tiêu đề phụ nếu sheet tài khoản có header lặp trong vùng dữ liệu.
    creds = creds[~creds['Tên nhân viên'].astype(str).apply(normalize_login_name).isin({
        'ten nhan vien', 'ten he thong', 'username', 'user name'
    })].copy()

    # Tính lương áp dụng cho Nhân viên và Leader. Leader có quyền nghiệp vụ giống Nhân viên.
    # Letan / quanly / locker / tapvu / admin không tạo dòng trong bảng lương mới.
    # Giữ tập khóa của toàn bộ tài khoản để Tip thuộc bộ phận bị loại không bị báo nhầm là
    # "không khớp tài khoản hệ thống".
    all_credential_keys = set(creds['Tên nhân viên'].apply(normalize_login_name).tolist())
    if 'Phân quyền' in creds.columns:
        roles = creds['Phân quyền'].astype(str).str.strip().str.lower()
        creds = creds[roles.isin(PAYROLL_ELIGIBLE_ROLES)].copy()
    else:
        # Sheet cũ chưa có cột phân quyền được xem là nhân viên để bảo toàn tương thích.
        creds = creds.copy()
    creds['__key'] = creds['Tên nhân viên'].apply(normalize_login_name)
    penalty_map = _period_penalty_by_employee(start_date, end_date, leave_primary, leave_secondary)
    due_violation_debt_map, deferred_current_violation_map, _ = get_violation_debt_state(
        start_date, end_date, creds['Tên nhân viên'].astype(str).tolist()
    )
    employee_overrides = get_payroll_employee_overrides()
    # Tích lũy tự động lấy từ sheet TichLuy. Nếu kỳ này đã được ghi nhận trước đó,
    # bảng lương vẫn phải HIỂN THỊ đúng số tiền của kỳ đó. Việc chống cộng trùng được
    # xử lý ở bước commit/save TichLuy, không được biến số hiển thị thành 0.
    tichluy_map, _tichluy_info = get_tichluy_charge_map(
        start_date, end_date, creds['Tên nhân viên'].astype(str).tolist(), for_existing_snapshot=True
    )

    # V72: Tiền trách nhiệm Leader chỉ được tự động cộng ở Kỳ 2 (16 -> cuối tháng).
    leader_allowance_active = is_payroll_period_2(start_date, end_date)

    rows = []
    for idx, (_, c) in enumerate(creds.iterrows(), start=1):
        k = c['__key']
        emp_override = employee_overrides.get(k, {})
        emp_living = emp_override.get("living", default_living_expense)
        emp_locker = emp_override.get("locker", default_locker_support)
        _emp_role = str(c.get('Phân quyền', 'nhanvien')).strip().lower()
        rows.append({
            "TT": idx,
            "Tên Hệ thống": str(c.get('Tên nhân viên', '')).strip(),
            "Họ và tên": str(c.get('Họ và tên đầy đủ', '')).strip(),
            "Tiền Lương": float(salary_map.get(k, 0)),
            # V72: Leader chỉ được cộng tiền trách nhiệm vào Hỗ Trợ Hoàn Lại ở Kỳ 2 hằng tháng.
            "Tiền Hỗ Trợ Hoàn Lại": (
                float(_money_to_float(leader_responsibility_allowance))
                if _emp_role == 'leader' and leader_allowance_active else 0.0
            ),
            "Tích lũy": float(tichluy_map.get(k, 0)),
            "Chi Phí Sinh Hoạt": float(_money_to_float(emp_living)),
            # V50: tách riêng Vi phạm phát sinh trong kỳ và nghĩa vụ Vi phạm từ kỳ trước.
            "Tiền phạt trong tháng": float(
                max(0.0, _money_to_float(penalty_map.get(k, 0)) - _money_to_float(deferred_current_violation_map.get(k, 0)))
            ),
            "Vi phạm kỳ trước": float(_money_to_float(due_violation_debt_map.get(k, 0))),
            "Tiền ứng lương": 0.0,
            "Tiền hỗ trợ Locker": float(_money_to_float(emp_locker)),
            "Số tiền thực nhận": 0.0,
            "Email": str(c.get('Email', '')).strip(),
            "Số tài khoản ngân hàng": str(c.get('Số tài khoản ngân hàng', '')).strip().replace("'", ""),
            "Tên ngân hàng": str(c.get('Tên ngân hàng', '')).strip(),
            "Số dòng Tip": int(tip_count_map.get(k, 0)),
        })
    result = pd.DataFrame(rows, columns=PAYROLL_COLUMNS)
    result = recalculate_payroll_net(result)
    unmatched = sorted({
        str(v).strip() for v in tip.loc[~tip['__key'].isin(all_credential_keys), 'NV tư vấn'].tolist()
        if str(v).strip()
    })
    return result, unmatched


def refresh_saved_payroll_from_system(payroll_df, start_date, end_date, credentials_df=None, leave_primary=None):
    """
    Cập nhật một bản lương đã lưu bằng dữ liệu hệ thống mới nhất, nhưng giữ nguyên
    các khoản nhập tay và Tiền Lương đã lưu.

    Tự cập nhật:
    - Vi phạm trong kỳ từ Google Sheet lịch nghỉ chính 1Kz0...
    - Vi phạm kỳ trước từ sổ NoViPham đã tới hạn khấu trừ
    - Tích lũy theo sheet TichLuy và quy tắc kỳ lương
    - Phí Sinh Hoạt / Hỗ trợ Locker theo mức mặc định hoặc mức riêng hiện hành
    - Tài khoản ngân hàng / Tên ngân hàng / Email từ hồ sơ nhân viên
    - Thực nhận sau khi các khoản trên thay đổi

    Không tự đổi Tiền Lương vì dữ liệu doanh thu nguồn có thể là file Excel upload
    và không được lưu như một nguồn dữ liệu vĩnh viễn trong hệ thống.
    """
    if payroll_df is None or not isinstance(payroll_df, pd.DataFrame) or payroll_df.empty:
        return pd.DataFrame(columns=PAYROLL_COLUMNS), {"updated": 0, "missing": []}

    d = payroll_df.copy()
    creds = credentials_df.copy() if isinstance(credentials_df, pd.DataFrame) else load_credentials_recent()
    leave_df = leave_primary.copy() if isinstance(leave_primary, pd.DataFrame) else load_backup_sheet_data()

    # Khi mở lại/tính lại bảng lương, bổ sung các nhân viên còn thiếu vào TichLuy trước.
    # D/E/F của các dòng đã có (đặc biệt người đã hoàn thành) không bị thay đổi.
    sync_tichluy_roles_and_stt(creds)

    # Dùng cùng snapshot cấu hình để tránh phát sinh nhiều request Google Sheets.
    default_living, default_locker = get_payroll_default_amounts()
    leader_allowance = get_leader_responsibility_allowance()
    leader_allowance_active = is_payroll_period_2(start_date, end_date)
    overrides = get_payroll_employee_overrides()
    penalty_map = _period_penalty_by_employee(start_date, end_date, leave_df, None)
    due_violation_debt_map, deferred_current_violation_map, _ = get_violation_debt_state(
        start_date, end_date, d.get('Tên Hệ thống', pd.Series(dtype=str)).astype(str).tolist()
    )
    tichluy_map, tichluy_info = get_tichluy_charge_map(
        start_date, end_date, d.get('Tên Hệ thống', pd.Series(dtype=str)).astype(str).tolist(),
        for_existing_snapshot=True
    )

    cred_map = {}
    if isinstance(creds, pd.DataFrame) and not creds.empty and 'Tên nhân viên' in creds.columns:
        for _, cr in creds.iterrows():
            key = normalize_login_name(cr.get('Tên nhân viên', ''))
            if key:
                cred_map[key] = cr

    missing = []
    updated = 0
    for idx, row in d.iterrows():
        emp_name = str(row.get('Tên Hệ thống', '')).strip()
        key = normalize_login_name(emp_name)
        if not key:
            continue

        # Tiền phạt luôn lấy lại theo đúng kỳ của bản lương đang mở.
        new_penalty = float(
            max(0.0, _money_to_float(penalty_map.get(key, 0)) - _money_to_float(deferred_current_violation_map.get(key, 0)))
        )
        if 'Tiền phạt trong tháng' in d.columns:
            d.at[idx, 'Tiền phạt trong tháng'] = new_penalty
        if 'Vi phạm kỳ trước' not in d.columns:
            d['Vi phạm kỳ trước'] = 0.0
        d.at[idx, 'Vi phạm kỳ trước'] = float(_money_to_float(due_violation_debt_map.get(key, 0)))
        # Tích lũy của bản lịch sử lấy đúng số kỳ đã ghi nhận; nếu chưa ghi thì tính theo quy tắc hiện tại.
        if 'Tích lũy' in d.columns:
            d.at[idx, 'Tích lũy'] = float(_money_to_float(tichluy_map.get(key, d.at[idx, 'Tích lũy'])))

        # Mức khấu trừ/hỗ trợ dùng mức riêng nếu có, nếu không dùng mức chung.
        ov = overrides.get(key, {}) if isinstance(overrides, dict) else {}
        living = ov.get('living', default_living)
        locker = ov.get('locker', default_locker)
        if 'Chi Phí Sinh Hoạt' in d.columns:
            d.at[idx, 'Chi Phí Sinh Hoạt'] = float(_money_to_float(living))
        if 'Tiền hỗ trợ Locker' in d.columns:
            d.at[idx, 'Tiền hỗ trợ Locker'] = float(_money_to_float(locker))

        # Đồng bộ thông tin hồ sơ mới nhất.
        cr = cred_map.get(key)
        if cr is None:
            missing.append(emp_name)
        else:
            # V72: chỉ tự động bảo đảm Tiền trách nhiệm Leader trong Kỳ 2.
            # Kỳ 1 tuyệt đối không tự cộng khoản trách nhiệm vào Hỗ Trợ Hoàn Lại.
            if (leader_allowance_active and str(cr.get('Phân quyền', '')).strip().lower() == 'leader'
                    and 'Tiền Hỗ Trợ Hoàn Lại' in d.columns):
                d.at[idx, 'Tiền Hỗ Trợ Hoàn Lại'] = max(
                    float(_money_to_float(d.at[idx, 'Tiền Hỗ Trợ Hoàn Lại'])), float(_money_to_float(leader_allowance))
                )
            if 'Số tài khoản ngân hàng' in d.columns:
                d.at[idx, 'Số tài khoản ngân hàng'] = str(cr.get('Số tài khoản ngân hàng', '')).strip().replace("'", "")
            if 'Tên ngân hàng' in d.columns:
                d.at[idx, 'Tên ngân hàng'] = str(cr.get('Tên ngân hàng', '')).strip()
            if 'Email' in d.columns:
                d.at[idx, 'Email'] = str(cr.get('Email', '')).strip()
            if 'Họ và tên' in d.columns:
                d.at[idx, 'Họ và tên'] = str(cr.get('Họ và tên đầy đủ', '')).strip()
        updated += 1

    d = recalculate_payroll_net(d)
    d = _filter_real_payroll_rows(d)
    d = apply_latest_profile_fields_to_payroll(d, creds, only_current_nhanvien=False)
    return d, {
        "updated": updated,
        "missing": sorted(set(missing)),
        "tichluy_updated": sum(1 for v in tichluy_map.values() if float(_money_to_float(v)) > 0),
        "tichluy_info": tichluy_info,
    }


def recalculate_payroll_net(df):
    d = df.copy()
    money_cols = [
        "Tiền Lương", "Tiền Hỗ Trợ Hoàn Lại", "Tích lũy",
        "Chi Phí Sinh Hoạt", "Tiền phạt trong tháng", "Vi phạm kỳ trước", "Tiền ứng lương", "Tiền hỗ trợ Locker"
    ]
    for col in money_cols:
        if col not in d.columns:
            d[col] = 0
        d[col] = pd.to_numeric(d[col], errors='coerce').fillna(0)
    # V51: Tiền Lương = 0 thì không ghi Phí Sinh Hoạt và Tiền hỗ trợ Locker.
    # Đặt ở hàm tính thực nhận để áp dụng đồng nhất cho bảng mới, bảng lịch sử và bảng chỉnh sửa.
    zero_salary_mask = d["Tiền Lương"].abs().le(1e-9)
    d.loc[zero_salary_mask, "Chi Phí Sinh Hoạt"] = 0.0
    d.loc[zero_salary_mask, "Tiền hỗ trợ Locker"] = 0.0
    net = (
        d["Tiền Lương"] + d["Tiền Hỗ Trợ Hoàn Lại"]
        - d["Tích lũy"] - d["Chi Phí Sinh Hoạt"]
        - d["Tiền phạt trong tháng"] - d["Vi phạm kỳ trước"]
        - d["Tiền ứng lương"] - d["Tiền hỗ trợ Locker"]
    )
    # V47: KHÔNG chặn ở 0. Giá trị âm phải được hiển thị đúng để biết phần nghĩa vụ
    # chưa thanh toán và chuyển sang kỳ lương kế tiếp.
    d["Số tiền thực nhận"] = net
    return d


def save_payroll_snapshot(payroll_df, start_date, end_date, source_label, saved_by):
    try:
        ws_pay, _, err = _ensure_payroll_storage()
        if err or ws_pay is None:
            return False, err or "Không mở được vùng lưu Bảng lương.", ""
        now = datetime.now(VN_TZ)
        batch_id = f"BL-{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}-{now.strftime('%Y%m%d%H%M%S')}"
        rows = []
        for _, r in payroll_df.iterrows():
            row = [
                batch_id, start_date.strftime('%d/%m/%Y'), end_date.strftime('%d/%m/%Y'),
                now.strftime('%d/%m/%Y'), now.strftime('%H:%M:%S'), str(saved_by), str(source_label),
                int(_money_to_float(r.get('TT', 0))), str(r.get('Tên Hệ thống', '')), str(r.get('Họ và tên', '')),
                float(_money_to_float(r.get('Tiền Lương', 0))), float(_money_to_float(r.get('Tiền Hỗ Trợ Hoàn Lại', 0))),
                float(_money_to_float(r.get('Hỗ trợ dạy nghề', 0))), float(_money_to_float(r.get('Học phí', 0))),
                float(_money_to_float(r.get('Tích lũy', 0))), float(_money_to_float(r.get('Chi Phí Sinh Hoạt', 0))),
                float(_money_to_float(r.get('Tiền phạt trong tháng', 0))), float(_money_to_float(r.get('Tiền ứng lương', 0))),
                float(_money_to_float(r.get('Tiền hỗ trợ Locker', 0))), float(_money_to_float(r.get('Số tiền thực nhận', 0))),
                str(r.get('Email', '')), "'" + str(r.get('Số tài khoản ngân hàng', '')).replace("'", ""),
                str(r.get('Tên ngân hàng', '')), int(_money_to_float(r.get('Số dòng Tip', 0))),
                float(_money_to_float(r.get('Vi phạm kỳ trước', 0)))
            ]
            rows.append(row)
        if rows:
            ws_pay.append_rows(rows, value_input_option='USER_ENTERED')
        tl_ok, tl_msg = record_tichluy_contributions(payroll_df, start_date, end_date)
        debt_ok, debt_msg = commit_violation_debts_after_payroll(payroll_df, start_date, end_date, saved_by)
        try:
            load_payroll_history.clear()
            if vpg is not None and vpg.is_enabled():
                try:
                    vpg.invalidate_dataset("payroll_history")
                except Exception:
                    pass
        except Exception:
            pass
        msg = f"Đã lưu bảng lương {len(rows)} nhân viên vào hệ thống."
        if not tl_ok:
            msg += f" ⚠️ {tl_msg}"
        if debt_msg:
            msg += f" · {debt_msg}"
        if not debt_ok:
            msg += f" ⚠️ {debt_msg}"
        return True, msg, batch_id
    except Exception as e:
        return False, f"Lỗi lưu bảng lương: {e}", ""


def overwrite_payroll_snapshot(batch_id, payroll_df, start_date, end_date, source_label, saved_by):
    """Ghi đè một bản lương đã lưu, giữ nguyên Mã bản lưu và cập nhật dấu thời gian/người sửa."""
    try:
        ws_pay, _, err = _ensure_payroll_storage()
        if err or ws_pay is None:
            return False, err or "Không mở được vùng lưu Bảng lương."

        batch_id = str(batch_id).strip()
        if not batch_id:
            return False, "Thiếu Mã bản lưu cần cập nhật."

        values = _gs_call_with_backoff(ws_pay.get_all_values)
        matched_rows = []
        for row_idx, row in enumerate(values[1:], start=2):
            if row and str(row[0]).strip() == batch_id:
                matched_rows.append(row_idx)

        if not matched_rows:
            return False, f"Không tìm thấy bản lương {batch_id} để ghi đè."

        # Xóa bản cũ từ dưới lên để không làm lệch chỉ số dòng.
        for row_idx in sorted(matched_rows, reverse=True):
            ws_pay.delete_rows(row_idx)

        now = datetime.now(VN_TZ)
        rows = []
        payroll_df = _filter_real_payroll_rows(recalculate_payroll_net(payroll_df))
        for _, r in payroll_df.iterrows():
            rows.append([
                batch_id, start_date.strftime('%d/%m/%Y'), end_date.strftime('%d/%m/%Y'),
                now.strftime('%d/%m/%Y'), now.strftime('%H:%M:%S'), str(saved_by), str(source_label),
                int(_money_to_float(r.get('TT', 0))), str(r.get('Tên Hệ thống', '')), str(r.get('Họ và tên', '')),
                float(_money_to_float(r.get('Tiền Lương', 0))), float(_money_to_float(r.get('Tiền Hỗ Trợ Hoàn Lại', 0))),
                0.0, 0.0,
                float(_money_to_float(r.get('Tích lũy', 0))), float(_money_to_float(r.get('Chi Phí Sinh Hoạt', 0))),
                float(_money_to_float(r.get('Tiền phạt trong tháng', 0))), float(_money_to_float(r.get('Tiền ứng lương', 0))),
                float(_money_to_float(r.get('Tiền hỗ trợ Locker', 0))), float(_money_to_float(r.get('Số tiền thực nhận', 0))),
                str(r.get('Email', '')), "'" + str(r.get('Số tài khoản ngân hàng', '')).replace("'", ""),
                str(r.get('Tên ngân hàng', '')), int(_money_to_float(r.get('Số dòng Tip', 0))),
                float(_money_to_float(r.get('Vi phạm kỳ trước', 0)))
            ])

        if rows:
            ws_pay.append_rows(rows, value_input_option='USER_ENTERED')
        tl_ok, tl_msg = record_tichluy_contributions(payroll_df, start_date, end_date)
        debt_ok, debt_msg = commit_violation_debts_after_payroll(payroll_df, start_date, end_date, saved_by)
        try:
            load_payroll_history.clear()
            if vpg is not None and vpg.is_enabled():
                try:
                    vpg.invalidate_dataset("payroll_history")
                except Exception:
                    pass
        except Exception:
            pass
        msg = f"Đã ghi đè cập nhật bản lương {batch_id} cho {len(rows)} nhân viên."
        if not tl_ok:
            msg += f" ⚠️ {tl_msg}"
        if debt_msg:
            msg += f" · {debt_msg}"
        if not debt_ok:
            msg += f" ⚠️ {debt_msg}"
        return True, msg
    except Exception as e:
        return False, f"Lỗi ghi đè bảng lương: {e}"



def delete_payroll_snapshots(batch_ids):
    """
    Xóa một hoặc nhiều BẢN LƯƠNG khỏi vùng Lịch sử bảng lương đã lưu.
    Chỉ xóa các dòng trong sheet lưu lịch sử bảng lương; không đụng tới TichLuy,
    dữ liệu lịch nghỉ/vi phạm hay hồ sơ nhân viên.
    """
    try:
        wanted = {
            str(x).strip() for x in (batch_ids or [])
            if str(x).strip()
        }
        if not wanted:
            return False, "Chưa chọn bản lương cần xóa.", []

        ws_pay, _, err = _ensure_payroll_storage()
        if err or ws_pay is None:
            return False, err or "Không mở được vùng lưu Bảng lương.", []

        values = _gs_call_with_backoff(ws_pay.get_all_values)
        if len(values) < 2:
            return False, "Lịch sử bảng lương hiện đang trống.", []

        matched_rows = []
        found_batches = set()
        # Dòng 1 là header; dữ liệu bắt đầu từ dòng 2.
        for row_idx, row in enumerate(values[1:], start=2):
            batch_id = str(row[0]).strip() if row else ""
            if batch_id in wanted:
                matched_rows.append(row_idx)
                found_batches.add(batch_id)

        if not matched_rows:
            return False, "Không tìm thấy các bản lương đã chọn trong hệ thống.", []

        # Gom các dòng liên tiếp thành block để giảm số request Google Sheets.
        blocks = []
        start = prev = matched_rows[0]
        for row_idx in matched_rows[1:]:
            if row_idx == prev + 1:
                prev = row_idx
            else:
                blocks.append((start, prev))
                start = prev = row_idx
        blocks.append((start, prev))

        # Xóa từ dưới lên để chỉ số dòng phía trên không bị thay đổi.
        for start_row, end_row in reversed(blocks):
            if start_row == end_row:
                _gs_call_with_backoff(ws_pay.delete_rows, start_row)
            else:
                _gs_call_with_backoff(ws_pay.delete_rows, start_row, end_row)

        try:
            load_payroll_history.clear()
            if vpg is not None and vpg.is_enabled():
                try:
                    vpg.invalidate_dataset("payroll_history")
                except Exception:
                    pass
        except Exception:
            pass

        deleted = [x for x in batch_ids if str(x).strip() in found_batches]
        missing = sorted(wanted - found_batches)
        msg = (
            f"Đã xóa {len(found_batches)} bản lương khỏi Lịch sử bảng lương "
            f"({len(matched_rows)} dòng dữ liệu)."
        )
        if missing:
            msg += " Không tìm thấy: " + ", ".join(missing)
        return True, msg, deleted
    except Exception as e:
        return False, f"Lỗi xóa lịch sử bảng lương: {e}", []


def _load_payroll_history_from_sheets():
    try:
        ws_pay, _, err = _ensure_payroll_storage()
        if err or ws_pay is None:
            return pd.DataFrame(columns=PAYROLL_HISTORY_HEADERS)
        values = _gs_call_with_backoff(ws_pay.get_all_values)
        if len(values) < 2:
            return pd.DataFrame(columns=PAYROLL_HISTORY_HEADERS)
        header = values[0][:len(PAYROLL_HISTORY_HEADERS)]
        rows = []
        for r in values[1:]:
            rr = list(r[:len(PAYROLL_HISTORY_HEADERS)]) + [''] * max(0, len(PAYROLL_HISTORY_HEADERS) - len(r))
            if any(str(v).strip() for v in rr): rows.append(rr[:len(PAYROLL_HISTORY_HEADERS)])
        return pd.DataFrame(rows, columns=header if len(header)==len(PAYROLL_HISTORY_HEADERS) else PAYROLL_HISTORY_HEADERS)
    except Exception:
        return pd.DataFrame(columns=PAYROLL_HISTORY_HEADERS)

@st.cache_data(ttl=30, show_spinner=False)
def load_payroll_history():
    """V75: đọc qua PostgreSQL dùng chung giữa các Cloud Run instance; Google Sheets là nguồn đồng bộ dự phòng."""
    if vpg is not None and vpg.is_enabled():
        return vpg.load_dataset(
            "payroll_history",
            _load_payroll_history_from_sheets,
            ttl_seconds=int(os.getenv("VERA_PG_TTL_PAYROLL_HISTORY", "90")),
        )
    return _load_payroll_history_from_sheets()


def payroll_history_to_table(history_df):
    cols = [c for c in PAYROLL_COLUMNS if c in history_df.columns]
    d = history_df[cols].copy()
    for col in [c for c in PAYROLL_COLUMNS if c.startswith('Tiền') or c in {'Tích lũy','Chi Phí Sinh Hoạt','Vi phạm kỳ trước','Số tiền thực nhận'}]:
        if col in d.columns: d[col] = pd.to_numeric(d[col], errors='coerce').fillna(0)
    if 'TT' in d.columns: d['TT'] = pd.to_numeric(d['TT'], errors='coerce').fillna(0).astype(int)
    if 'Số dòng Tip' in d.columns: d['Số dòng Tip'] = pd.to_numeric(d['Số dòng Tip'], errors='coerce').fillna(0).astype(int)
    return d


def build_payroll_excel_bytes(payroll_df, start_date, end_date):
    """Xuất toàn bộ bảng lương: A4 ngang, fit 1 trang chiều rộng, có đầy đủ tài khoản/ngân hàng."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    d = recalculate_payroll_net(payroll_df).copy()

    # V60: trước khi export, đồng bộ TOÀN BỘ trường hồ sơ đang dùng trong file
    # (Họ và Tên, tài khoản ngân hàng, tên ngân hàng, Email nội bộ) từ Sheet1 nguồn.
    # Export vẫn không hiển thị Email theo yêu cầu.
    try:
        _export_creds = load_credentials_recent()
        d = apply_latest_profile_fields_to_payroll(d, _export_creds, only_current_nhanvien=False)
    except Exception:
        pass

    # V45/V50: File Excel export KHÔNG xuất Email. Giữ Họ và Tên nhân viên ở cột M
    # như cấu trúc đã chốt trước đó; cột Vi phạm kỳ trước được bổ sung sau cùng (cột N).
    export_cols = [
        "TT", "Tên Hệ thống", "Tiền Lương", "Tiền Hỗ Trợ Hoàn Lại",
        "Tích lũy", "Chi Phí Sinh Hoạt", "Tiền phạt trong tháng", "Tiền ứng lương",
        "Tiền hỗ trợ Locker", "Số tiền thực nhận", "Số tài khoản ngân hàng", "Tên ngân hàng", "Họ và tên",
        "Vi phạm kỳ trước"
    ]
    for c in export_cols:
        if c not in d.columns:
            d[c] = "" if c in {"Tên Hệ thống","Số tài khoản ngân hàng","Tên ngân hàng","Họ và tên"} else 0

    wb = Workbook()
    ws = wb.active
    ws.title = "Bảng lương"
    last_col = len(export_cols)
    last_letter = get_column_letter(last_col)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws['A1'] = "BẢNG LƯƠNG NHÂN VIÊN"
    ws['A1'].font = Font(name='Arial', size=18, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill('solid', fgColor='F3E4EC')
    ws.row_dimensions[1].height = 30
    ws['A2'] = "KỲ LƯƠNG"
    ws['A2'].font = Font(name='Arial', size=11, bold=True)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=last_col)
    ws['B2'] = f"Từ ngày {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}"
    ws['B2'].font = Font(name='Arial', size=11, bold=True)

    # Dùng bộ tiêu đề chuẩn, riêng cột Họ và tên đổi tên hiển thị theo yêu cầu export.
    header_labels = dict(PAYROLL_DISPLAY_LABELS)
    header_labels["Họ và tên"] = "Họ và Tên nhân viên"
    for c, h in enumerate(export_cols, start=1):
        display_header = header_labels.get(h, h)
        cell = ws.cell(row=3, column=c, value=display_header)
        cell.font = Font(name='Arial', size=9, bold=True, color='000000')
        cell.fill = PatternFill('solid', fgColor='A1948C')
        # Riêng Tên ngân hàng không wrap text theo yêu cầu.
        cell.alignment = Alignment(
            horizontal='center', vertical='center',
            wrap_text=False if h == 'Tên ngân hàng' else True
        )
    ws.row_dimensions[3].height = 52
    thin = Side(style='thin', color='A6A6A6')

    start_row = 4
    money_cols = {"Tiền Lương", "Tiền Hỗ Trợ Hoàn Lại", "Tích lũy", "Chi Phí Sinh Hoạt", "Tiền phạt trong tháng", "Vi phạm kỳ trước", "Tiền ứng lương", "Tiền hỗ trợ Locker", "Số tiền thực nhận"}
    non_positive_fill = PatternFill('solid', fgColor='FFF2CC')
    for i, (_, r) in enumerate(d.iterrows(), start=start_row):
        for j, col in enumerate(export_cols, start=1):
            val = r.get(col, '')
            if col in money_cols:
                val = float(_money_to_float(val))
            elif col == 'TT':
                val = int(_money_to_float(val))
            elif col == 'Số tài khoản ngân hàng':
                val = str(val).replace("'", "")
            ws.cell(row=i, column=j, value=val)
        # Tài khoản ngân hàng buộc kiểu Text để giữ số 0 đầu.
        bank_col = export_cols.index('Số tài khoản ngân hàng') + 1
        ws.cell(row=i, column=bank_col).number_format = '@'

        # V45: tô vàng TOÀN BỘ dòng khi Thực nhận <= 0 để Admin dễ kiểm tra.
        if _money_to_float(r.get('Số tiền thực nhận', 0)) <= 0:
            for j in range(1, last_col + 1):
                ws.cell(row=i, column=j).fill = non_positive_fill

    total_row = start_row + len(d)
    ws.cell(total_row, 2, "TỔNG")
    for j, col in enumerate(export_cols, start=1):
        if col in money_cols:
            values = d[col].apply(_money_to_float)
            # V44: riêng cột J / Thực nhận, dòng TỔNG chỉ cộng những nhân viên có
            # Thực nhận > 0. Giá trị âm hoặc 0 không được cộng vào tổng chi trả.
            total_value = values[values > 0].sum() if col == "Số tiền thực nhận" else values.sum()
            ws.cell(total_row, j, float(total_value))
    for c in range(1, last_col + 1):
        ws.cell(total_row, c).font = Font(name='Arial', size=10, bold=True)
        ws.cell(total_row, c).fill = PatternFill('solid', fgColor='E2E3E5')

    for row in ws.iter_rows(min_row=3, max_row=total_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if cell.row >= 4:
                cell.font = Font(name='Arial', size=9, bold=(cell.row == total_row))
                # Không wrap cột Tên ngân hàng; các cột còn lại giữ wrap để gọn trên A4 ngang.
                col_name = export_cols[cell.column - 1]
                cell.alignment = Alignment(
                    vertical='center',
                    wrap_text=False if col_name == 'Tên ngân hàng' else True
                )
    for j, col in enumerate(export_cols, start=1):
        if col in money_cols:
            for row in range(4, total_row + 1):
                ws.cell(row, j).number_format = '#,##0'
                ws.cell(row, j).alignment = Alignment(horizontal='right', vertical='center')

    # Auto-fit có giới hạn để vẫn vừa A4 ngang.
    for j, col in enumerate(export_cols, start=1):
        max_len = len(col)
        for row in range(4, min(total_row, 60) + 1):
            max_len = max(max_len, len(str(ws.cell(row, j).value or '')))
        if col == 'TT':
            # Cột TT cũ tối thiểu rộng 6; giảm 80% còn khoảng 20% chiều rộng.
            width = 1.2
        elif col in money_cols:
            width = min(max(max_len + 2, 12), 17)
        elif col == 'Tên ngân hàng':
            # Không wrap nên cho phép cột rộng hơn để tên ngân hàng nằm trên một dòng.
            width = min(max(max_len + 2, 22), 32)
        elif col == 'Họ và tên':
            width = min(max(max_len + 2, 18), 28)
        else:
            width = min(max(max_len + 2, 6), 19)
        ws.column_dimensions[get_column_letter(j)].width = width
    for r in range(4, total_row + 1):
        ws.row_dimensions[r].height = 20

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f"A3:{last_letter}{max(3,total_row-1)}"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.18, right=0.18, top=0.3, bottom=0.3, header=0.12, footer=0.12)
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = '1:3'
    ws.print_area = f"A1:{last_letter}{total_row}"
    ws.sheet_view.zoomScale = 65
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()



def build_employee_payroll_excel_bytes(employee_row, start_date, end_date, violation_details=None):
    """Tạo phiếu lương cá nhân theo đúng bố cục nội dung email gửi nhân viên."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.worksheet.page import PageMargins

    emp = str(employee_row.get('Tên Hệ thống', '')).strip()
    # Bản V21 đã bỏ cột Họ và Tên khỏi bảng lương; nếu dữ liệu cũ còn cột này thì vẫn ưu tiên dùng.
    full = str(employee_row.get('Họ và tên', '')).strip()
    display_name = full if full and full.lower() not in {'nan', 'none'} else emp

    items = [
        ('Tiền Lương', 'Tiền Lương'),
        ('Tiền Hỗ Trợ Hoàn Lại', 'Hỗ trợ hoặc hoàn lại'),
        ('Tích lũy', 'Tích lũy'),
        ('Chi Phí Sinh Hoạt', 'Phí sinh hoạt'),
        ('Tiền phạt trong tháng', 'Vi phạm trong kỳ'),
        ('Vi phạm kỳ trước', 'Vi phạm kỳ trước'),
        ('Tiền ứng lương', 'Tiền ứng lương'),
        ('Tiền hỗ trợ Locker', 'Tiền hỗ trợ Locker'),
        ('Số tiền thực nhận', 'Số tiền thực nhận'),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Bảng lương'
    ws.sheet_view.showGridLines = False

    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='A1948C')

    # Lời chào / thông tin kỳ lương giống nội dung email.
    ws.merge_cells('A1:B1')
    ws['A1'] = f'Chào {display_name},'
    ws['A1'].font = Font(name='Arial', size=14, bold=False, color='000000')
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 24

    ws.merge_cells('A3:B3')
    ws['A3'] = f"VERA SPA gửi bảng lương kỳ từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}."
    ws['A3'].font = Font(name='Arial', size=12, color='000000')
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[3].height = 30

    # Bảng Khoản mục / Số tiền.
    ws['A5'] = 'Khoản mục'
    ws['B5'] = 'Số tiền'
    for c in ('A5', 'B5'):
        ws[c].font = Font(name='Arial', size=11, bold=True, color='000000')
        ws[c].fill = header_fill
        ws[c].alignment = Alignment(horizontal='center', vertical='center')
        ws[c].border = border
    ws.row_dimensions[5].height = 25

    row = 6
    for label, source_col in items:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=float(_money_to_float(employee_row.get(source_col, 0))))
        ws.cell(row=row, column=1).font = Font(name='Arial', size=11, bold=False)
        ws.cell(row=row, column=2).font = Font(name='Arial', size=11, bold=(source_col == 'Số tiền thực nhận'))
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='right', vertical='center')
        ws.cell(row=row, column=2).number_format = '#,##0 "VNĐ"'
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2).border = border
        ws.row_dimensions[row].height = 23
        row += 1

    net = float(_money_to_float(employee_row.get('Số tiền thực nhận', 0)))
    net_row = row + 1
    ws.merge_cells(start_row=net_row, start_column=1, end_row=net_row, end_column=2)
    ws.cell(net_row, 1, f'Số tiền thực nhận: {net:,.0f} VNĐ')
    ws.cell(net_row, 1).font = Font(name='Arial', size=13, bold=True, color='000000')
    ws.cell(net_row, 1).alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[net_row].height = 28

    note_row = net_row + 2
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=2)
    ws.cell(note_row, 1, 'Vui lòng kiểm tra và phản hồi nếu có sai sót.')
    ws.cell(note_row, 1).font = Font(name='Arial', size=11)
    ws.cell(note_row, 1).alignment = Alignment(horizontal='left', vertical='center')

    sign_row = note_row + 2
    ws.merge_cells(start_row=sign_row, start_column=1, end_row=sign_row, end_column=2)
    ws.cell(sign_row, 1, 'Trân trọng,')
    ws.cell(sign_row, 1).font = Font(name='Arial', size=11)

    ws.merge_cells(start_row=sign_row + 1, start_column=1, end_row=sign_row + 1, end_column=2)
    ws.cell(sign_row + 1, 1, 'VERA SPA')
    ws.cell(sign_row + 1, 1).font = Font(name='Arial', size=12, bold=True)

    # Căn vừa trang và dễ đọc trên điện thoại/PC khi mở file.
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 22
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins = PageMargins(left=0.45, right=0.45, top=0.5, bottom=0.5, header=0.2, footer=0.2)
    ws.print_options.horizontalCentered = True
    ws.print_area = f'A1:B{sign_row + 1}'
    ws.sheet_view.zoomScale = 90

    # Sheet chi tiết vi phạm để nhân viên đối chiếu đúng kỳ lương.
    ws_vp = wb.create_sheet('Chi tiết vi phạm')
    ws_vp.sheet_view.showGridLines = False
    vp_headers = ['Ngày', 'Lý do nghỉ', 'Chi tiết', 'Phạt vi phạm']
    for j, h in enumerate(vp_headers, start=1):
        c = ws_vp.cell(1, j, h)
        c.font = Font(name='Arial', size=10, bold=True, color='000000')
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
    vp_df = violation_details.copy() if isinstance(violation_details, pd.DataFrame) else pd.DataFrame(columns=vp_headers)
    for c in vp_headers:
        if c not in vp_df.columns: vp_df[c] = ''
    if vp_df.empty:
        ws_vp.merge_cells('A2:D2')
        ws_vp['A2'] = 'Không có vi phạm bị phạt trong kỳ lương này.'
        ws_vp['A2'].font = Font(name='Arial', size=10, italic=True)
    else:
        for i, (_, vr) in enumerate(vp_df[vp_headers].iterrows(), start=2):
            for j, h in enumerate(vp_headers, start=1):
                val = float(_money_to_float(vr.get(h,0))) if h == 'Phạt vi phạm' else str(vr.get(h,'') or '')
                cell = ws_vp.cell(i, j, val)
                cell.border = border
                cell.font = Font(name='Arial', size=10)
                cell.alignment = Alignment(vertical='top', wrap_text=(h in {'Lý do nghỉ','Chi tiết'}))
                if h == 'Phạt vi phạm':
                    cell.number_format = '#,##0 "VNĐ"'
                    cell.alignment = Alignment(horizontal='right', vertical='top')
        total_r = len(vp_df) + 2
        ws_vp.cell(total_r, 3, 'TỔNG VI PHẠM').font = Font(name='Arial', size=10, bold=True)
        ws_vp.cell(total_r, 4, float(vp_df['Phạt vi phạm'].apply(_money_to_float).sum()))
        ws_vp.cell(total_r, 4).font = Font(name='Arial', size=10, bold=True)
        ws_vp.cell(total_r, 4).number_format = '#,##0 "VNĐ"'
    ws_vp.column_dimensions['A'].width = 14
    ws_vp.column_dimensions['B'].width = 24
    ws_vp.column_dimensions['C'].width = 42
    ws_vp.column_dimensions['D'].width = 18
    ws_vp.freeze_panes = 'A2'
    ws_vp.page_setup.orientation = ws_vp.ORIENTATION_LANDSCAPE
    ws_vp.page_setup.paperSize = ws_vp.PAPERSIZE_A4
    ws_vp.sheet_properties.pageSetUpPr.fitToPage = True
    ws_vp.page_setup.fitToWidth = 1
    ws_vp.page_setup.fitToHeight = 0

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

def send_payroll_email(sender_email, sender_password, to_email, employee_row, start_date, end_date, violation_details=None):
    try:
        emp = str(employee_row.get('Tên Hệ thống',''))
        full = str(employee_row.get('Họ và tên',''))
        subject = f"Bảng lương {emp} - {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}"
        money_fields = [
            ("Tiền Lương", "Tiền Lương"),
            ("Tiền Hỗ Trợ Hoàn Lại", "Hỗ trợ hoặc hoàn lại"),
            ("Tích lũy", "Tích lũy"),
            ("Chi Phí Sinh Hoạt", "Phí sinh hoạt"),
            ("Tiền phạt trong tháng", "Vi phạm trong kỳ"),
            ("Vi phạm kỳ trước", "Vi phạm kỳ trước"),
            ("Tiền ứng lương", "Tiền ứng lương"),
            ("Tiền hỗ trợ Locker", "Tiền hỗ trợ Locker"),
            ("Số tiền thực nhận", "Số tiền thực nhận"),
        ]
        html_rows = "".join(
            f"<tr><td style='padding:6px;border:1px solid #D9D9D9'>{label}</td><td style='padding:6px;border:1px solid #D9D9D9;text-align:right'>{_money_to_float(employee_row.get(field,0)):,.0f} VNĐ</td></tr>"
            for field, label in money_fields
        )
        vp_df = violation_details.copy() if isinstance(violation_details, pd.DataFrame) else pd.DataFrame()
        if not vp_df.empty:
            vp_rows = "".join(
                f"<tr><td style='padding:5px;border:1px solid #D9D9D9'>{str(vr.get('Ngày',''))}</td>"
                f"<td style='padding:5px;border:1px solid #D9D9D9'>{str(vr.get('Lý do nghỉ',''))}</td>"
                f"<td style='padding:5px;border:1px solid #D9D9D9'>{str(vr.get('Chi tiết',''))}</td>"
                f"<td style='padding:5px;border:1px solid #D9D9D9;text-align:right'>{_money_to_float(vr.get('Phạt vi phạm',0)):,.0f} VNĐ</td></tr>"
                for _, vr in vp_df.iterrows()
            )
            vp_total = vp_df['Phạt vi phạm'].apply(_money_to_float).sum() if 'Phạt vi phạm' in vp_df.columns else 0
            violation_html = f"""
            <p><b>Chi tiết vi phạm trong kỳ:</b></p>
            <table style='border-collapse:collapse;min-width:620px'>
            <tr><th style='padding:6px;border:1px solid #D9D9D9;background:#A1948C;white-space:normal;overflow-wrap:anywhere;word-break:break-word'>Ngày</th><th style='padding:6px;border:1px solid #D9D9D9;background:#A1948C;white-space:normal;overflow-wrap:anywhere;word-break:break-word'>Lý do</th><th style='padding:6px;border:1px solid #D9D9D9;background:#A1948C;white-space:normal;overflow-wrap:anywhere;word-break:break-word'>Chi tiết</th><th style='padding:6px;border:1px solid #D9D9D9;background:#A1948C;white-space:normal;overflow-wrap:anywhere;word-break:break-word'>Phạt</th></tr>
            {vp_rows}
            </table><p><b>Tổng vi phạm: {vp_total:,.0f} VNĐ</b></p>
            """
        else:
            violation_html = "<p><b>Chi tiết vi phạm trong kỳ:</b> Không có vi phạm bị phạt.</p>"
        html = f"""
        <html><body style='font-family:Arial,sans-serif'>
        <p>Chào <b>{full or emp}</b>,</p>
        <p>VERA SPA gửi bảng lương kỳ từ <b>{start_date.strftime('%d/%m/%Y')}</b> đến <b>{end_date.strftime('%d/%m/%Y')}</b>.</p>
        <table style='border-collapse:collapse;min-width:520px'>
        <tr><th style='padding:7px;border:1px solid #D9D9D9;background:#A1948C;color:#000;white-space:normal;overflow-wrap:anywhere;word-break:break-word'>Khoản mục</th><th style='padding:7px;border:1px solid #D9D9D9;background:#A1948C;color:#000;white-space:normal;overflow-wrap:anywhere;word-break:break-word'>Số tiền</th></tr>
        {html_rows}
        </table>
        <p><b>Số tiền thực nhận: {_money_to_float(employee_row.get('Số tiền thực nhận',0)):,.0f} VNĐ</b></p>
        {violation_html}
        <p>Vui lòng kiểm tra và phản hồi nếu có sai sót.</p><p>Trân trọng,<br><b>VERA SPA</b></p>
        </body></html>
        """
        msg = MIMEMultipart()
        msg['From'] = f"Vera Spa <{sender_email}>"
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(html, 'html'))
        attachment = build_employee_payroll_excel_bytes(employee_row, start_date, end_date, violation_details)
        part = MIMEApplication(attachment, _subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.add_header('Content-Disposition', 'attachment', filename=f"BangLuong_{normalize_login_name(emp).replace(' ','_')}_{start_date.strftime('%d%m%Y')}_{end_date.strftime('%d%m%Y')}.xlsx")
        msg.attach(part)
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(msg)
        server.quit()
        return True, "Thành công"
    except Exception as e:
        return False, str(e)

def send_payroll_summary_email(sender_email, sender_password, to_email, recipient_name, payroll_df, start_date, end_date):
    """Gửi file bảng lương TỔNG HỢP cho đúng một Lễ tân được Admin chỉ định."""
    try:
        subject = f"Bảng lương tổng hợp - {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')}"
        total_salary = recalculate_payroll_net(payroll_df)['Tiền Lương'].apply(_money_to_float).sum()
        total_net = recalculate_payroll_net(payroll_df)['Số tiền thực nhận'].apply(_money_to_float).sum()
        html = f"""
        <html><body style='font-family:Arial,sans-serif'>
        <p>Chào <b>{recipient_name}</b>,</p>
        <p>VERA SPA gửi file <b>bảng lương tổng hợp</b> kỳ <b>{start_date.strftime('%d/%m/%Y')}</b> đến <b>{end_date.strftime('%d/%m/%Y')}</b>.</p>
        <p>Số nhân viên: <b>{len(payroll_df)}</b><br>
        Tổng tiền lương: <b>{total_salary:,.0f} VNĐ</b><br>
        Tổng thực nhận: <b>{total_net:,.0f} VNĐ</b></p>
        <p>File Excel đầy đủ được đính kèm email này.</p>
        <p>Trân trọng,<br><b>VERA SPA</b></p>
        </body></html>
        """
        msg = MIMEMultipart()
        msg['From'] = f"Vera Spa <{sender_email}>"
        msg['To'] = to_email
        msg['Subject'] = subject
        msg.attach(MIMEText(html, 'html'))
        # V46: dùng đúng bộ export chuẩn nên file gửi Lễ tân cũng:
        # - bỏ Email, thay bằng Họ và Tên từ Sheet1 cột E
        # - tô vàng toàn bộ dòng có Thực nhận <= 0
        attachment = build_payroll_excel_bytes(payroll_df, start_date, end_date)
        part = MIMEApplication(attachment, _subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.add_header('Content-Disposition', 'attachment', filename=f"BangLuong_TongHop_{start_date.strftime('%d%m%Y')}_{end_date.strftime('%d%m%Y')}.xlsx")
        msg.attach(part)
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(msg)
        server.quit()
        return True, "Đã gửi bảng lương tổng hợp thành công."
    except Exception as e:
        return False, str(e)


# Tải dữ liệu
ensure_credential_control_columns()
df_credentials = load_credentials_recent()
if isinstance(df_credentials, pd.DataFrame) and not df_credentials.empty and 'Tên nhân viên' in df_credentials.columns:
    df_credentials = df_credentials.assign(
        __employee_sort=df_credentials['Tên nhân viên'].astype(str).apply(normalize_login_name)
    ).sort_values('__employee_sort', kind='stable').drop(columns='__employee_sort').reset_index(drop=True)
df_backup = load_backup_sheet_data()
df_leave_secondary = load_secondary_leave_sheet_data()
df_loai_nghi_gsheet = load_loai_nghi_from_gsheet()
GDRIVE_LINK = "https://drive.google.com/file/d/1xTjmi6BaQFSqsgn9-EM7MjVS2n2FNuxT/view?usp=sharing"

with st.spinner("Đang tải dữ liệu hệ thống..."):
    df_lich, df_nv_excel, df_loai_nghi_excel = load_lich_nghi(GDRIVE_LINK) 

df_loai_nghi = df_loai_nghi_gsheet if not df_loai_nghi_gsheet.empty else df_loai_nghi_excel

# V71: áp dụng giao diện mặc định đã lưu cho cả màn hình đăng nhập và toàn bộ trang chức năng.
_ui_theme_cfg, _ui_theme_err = load_ui_theme_config()
render_global_ui_theme_css(_ui_theme_cfg)

if df_lich.empty or df_nv_excel.empty:
    st.warning("Hệ thống chưa tìm thấy dữ liệu.")
    st.stop()

# --- ĐĂNG NHẬP ---
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.current_user = ""
    st.session_state.current_role = ""
if "birthday_login_event" not in st.session_state:
    st.session_state.birthday_login_event = False

# Tự đăng nhập bằng token đã nhớ (không lưu mật khẩu ở localStorage).
if not st.session_state.logged_in:
    try:
        remembered_token = st.query_params.get('remember_token', '')
        if _is_valid_fallback_admin_token(remembered_token):
            st.session_state.logged_in = True
            st.session_state.current_user = "Quản Trị Viên"
            st.session_state.current_role = "admin"
            st.session_state.birthday_login_event = True
        else:
            remembered_row = validate_remember_token(remembered_token, df_credentials) if remembered_token else None
            if remembered_row is not None:
                st.session_state.logged_in = True
                st.session_state.current_user = str(remembered_row['Tên nhân viên']).strip()
                st.session_state.current_role = str(remembered_row.get('Phân quyền', 'nhanvien')).strip().lower()
                st.session_state.birthday_login_event = True
                _set_default_page_after_login(st.session_state.current_role)
            elif remembered_token:
                # Token bị khóa/sai/đã thu hồi -> xóa khỏi trình duyệt.
                st.query_params['forget_login'] = '1'
                try: del st.query_params['remember_token']
                except Exception: pass
    except Exception:
        pass

if not st.session_state.logged_in:
    st.title("🔐 Đăng Nhập Hệ Thống")
    # Trên điện thoại đưa nút Đăng nhập sang phải. Khối CSS này chỉ tồn tại ở màn hình login.
    st.markdown("""
    <style>
    @media (max-width: 768px) {
        div[data-testid="stFormSubmitButton"] { display:flex !important; justify-content:flex-end !important; }
        div[data-testid="stFormSubmitButton"] > button { width:auto !important; min-width:145px !important; }
    }
    </style>
    """, unsafe_allow_html=True)
    with st.form("login_form"):
        username_input = st.text_input("Tên đăng nhập", autocomplete="username").strip()
        password_input = st.text_input("Mật khẩu", type="password", autocomplete="current-password")
        st.caption("🔐 Thiết bị này sẽ duy trì đăng nhập cho tới khi bạn bấm Đăng xuất.")

        # Nút lưu đăng nhập dùng token bảo mật hiện có; không lưu mật khẩu dạng chữ thường.
        c_save_login, c_login = st.columns([2, 1])
        with c_save_login:
            save_credentials_submit = st.form_submit_button("💾 Lưu tên đăng nhập và mật khẩu")
        with c_login:
            login_submit = st.form_submit_button("Đăng Nhập")

        if login_submit or save_credentials_submit:
            input_name_norm = normalize_login_name(username_input)

            # Tài khoản quản trị dự phòng cũ: vẫn chấp nhận HOA/thường ở tên đăng nhập.
            if input_name_norm == normalize_login_name('admin') and password_matches(password_input, '32531235'):
                st.session_state.logged_in = True
                st.session_state.current_user = "Quản Trị Viên"
                st.session_state.current_role = "admin"
                st.session_state.birthday_login_event = True
                # Admin dự phòng cũng được duy trì đăng nhập cho tới khi bấm Đăng xuất.
                st.query_params['remember_token'] = _fallback_admin_remember_token()
                st.rerun()
            else:
                user_found = False
                locked_account = False
                matched_row = None

                for _, row in df_credentials.iterrows():
                    db_name = str(row['Tên nhân viên']).strip()
                    if input_name_norm == normalize_login_name(db_name):
                        matched_row = row
                        if is_locked_value(row.get('Khóa đăng nhập', '')):
                            locked_account = True
                            break
                        if password_matches(password_input, row.get('Mật khẩu', '')):
                            st.session_state.logged_in = True
                            st.session_state.current_user = db_name
                            st.session_state.current_role = str(row.get('Phân quyền', 'nhanvien')).strip().lower()
                            st.session_state.birthday_login_event = True
                            _set_default_page_after_login(st.session_state.current_role)
                            user_found = True
                            break

                if locked_account:
                    st.error("🔒 Tài khoản này đang bị khóa đăng nhập tạm thời. Vui lòng liên hệ Admin.")
                elif user_found:
                    token = create_remember_token(st.session_state.current_user)
                    if token:
                        st.query_params['remember_token'] = token
                    st.rerun()
                else:
                    st.error("❌ Sai tên đăng nhập hoặc mật khẩu!")
    st.stop()


# V86.4: mọi tài khoản (kể cả tạp vụ) đều dùng điều hướng chung; quyền nghiệp vụ vẫn được kiểm soát riêng.

# ==========================================
# ĐIỀU HƯỚNG THEO TỪNG TRANG CHỨC NĂNG
# ==========================================
is_admin_letan = st.session_state.current_role in ["admin", "letan", "quanly"]

PAGE_SLUGS = {
    "🧭 Bảng tour": "bang-tour",
    "💰 Bảng lương": "bang-luong",
    "📅 Đăng ký nghỉ phép": "dang-ky-thong-ke-nghi-phep",
    "✏️ Quản lý lịch nghỉ": "quan-ly-lich-nghi",
    "⏰ Thiết lập ca làm việc": "thiet-lap-ca",
    "👥 Danh sách nhân sự": "danh-sach-nhan-su",
    "➕ Thêm nhân viên": "them-nhan-vien",
    "✏️ Sửa / Xóa nhân viên": "sua-xoa-nhan-vien",
    "🔒 Khóa đăng nhập": "khoa-dang-nhap",
    "🔐 Khóa quyền đăng ký": "khoa-quyen-dang-ky",
    "⏸️ Auto Update phạt": "auto-update-phat",
    "📦 Snapshot nền hôm nay": "snapshot-nen-hom-nay",
    "📘 Hướng dẫn sử dụng": "huong-dan-su-dung",
    "🔄 Đồng bộ dữ liệu": "dong-bo-du-lieu",
    "⚙️ Giao diện tùy chỉnh": "cau-hinh-cot",
    "🔐 Phân quyền chức năng": "phan-quyen-chuc-nang",
    "🎂 Sinh nhật nhân sự": "sinh-nhat-nhan-su",
    "👤 Hồ sơ cá nhân": "ho-so-ca-nhan",
}
SLUG_TO_PAGE = {v: k for k, v in PAGE_SLUGS.items()}

payroll_letan_enabled = get_payroll_letan_enabled()

PAGE_FEATURE_KEYS = {
    "🧭 Bảng tour": "tour",
    "💰 Bảng lương": "payroll",
    "📅 Đăng ký nghỉ phép": "leave",
    "✏️ Quản lý lịch nghỉ": "leave_manage",
    "⏰ Thiết lập ca làm việc": "shift",
    "👥 Danh sách nhân sự": "staff_list",
    "➕ Thêm nhân viên": "employee_add",
    "✏️ Sửa / Xóa nhân viên": "employee_edit",
    "🔒 Khóa đăng nhập": "account_lock",
    "🔐 Khóa quyền đăng ký": "registration_lock",
    "⏸️ Auto Update phạt": "auto_penalty",
    "📦 Snapshot nền hôm nay": "snapshot_today",
    "📘 Hướng dẫn sử dụng": "guide",
    "🔄 Đồng bộ dữ liệu": "sync",
    "⚙️ Giao diện tùy chỉnh": "column_config",
    "🔐 Phân quyền chức năng": "permission_admin",
    "🎂 Sinh nhật nhân sự": "birthday",
    "👤 Hồ sơ cá nhân": "profile",
}
PAGE_FEATURE_GROUPS = {
    # Một trang có thể chứa nhiều chức năng độc lập. Chỉ cần có ít nhất một quyền
    # trong nhóm thì trang được hiện; từng section bên trong vẫn kiểm tra quyền riêng.
    "💰 Bảng lương": {"payroll", "payroll_history"},
    "✏️ Sửa / Xóa nhân viên": {"employee_edit", "employment_status", "employee_delete"},
}

def has_page_access(page_name):
    if page_name == "📘 Hướng dẫn sử dụng":
        return True
    # Các trang hệ thống dưới đây tuyệt đối chỉ dành cho Admin.
    if page_name in {"⚙️ Giao diện tùy chỉnh", "⏸️ Auto Update phạt", "📦 Snapshot nền hôm nay"}:
        return st.session_state.get('current_role') == 'admin'
    features = PAGE_FEATURE_GROUPS.get(page_name)
    if features:
        return any(has_feature_access(key) for key in features)
    key = PAGE_FEATURE_KEYS.get(page_name)
    return bool(key and has_feature_access(key))

DEFAULT_PAGE_ORDER = list(PAGE_FEATURE_KEYS.keys())
PAGE_ORDER = admin_menu_order_for_pages(DEFAULT_PAGE_ORDER)
allowed_pages = [p for p in PAGE_ORDER if has_page_access(p)]
# Admin luôn giữ trang phân quyền để tránh tự khóa hệ thống.
if st.session_state.current_role == 'admin' and "🔐 Phân quyền chức năng" not in allowed_pages:
    allowed_pages.append("🔐 Phân quyền chức năng")
# Nếu tài khoản không có trang nghiệp vụ nào, giữ Hồ sơ cá nhân khi được cấp quyền; còn không sẽ chỉ thấy header/logout.
# Đọc trang từ URL để nút Back/Forward và swipe trên điện thoại hoạt động.
requested_slug = str(st.query_params.get("page", "")).strip()
requested_page = SLUG_TO_PAGE.get(requested_slug)
if requested_page in allowed_pages:
    st.session_state.app_page = requested_page
elif st.session_state.get("app_page") not in allowed_pages:
    if allowed_pages:
        preferred_default = DEFAULT_LEAVE_PAGE if st.session_state.current_role in {"letan", "quanly", "nhanvien", "leader"} and DEFAULT_LEAVE_PAGE in allowed_pages else allowed_pages[0]
        st.session_state.app_page = preferred_default
    else:
        st.session_state.app_page = ""
selected_page = st.session_state.get("app_page", "")


def render_global_unsaved_changes_guard(page_labels):
    """V70: cảnh báo khi rời trang trong lúc form/data editor còn thay đổi chưa lưu.

    Chạy phía trình duyệt nên vẫn phát hiện được thay đổi bên trong st.form trước khi
    Streamlit gửi dữ liệu về Python. Áp dụng chung cho các vùng chỉnh dữ liệu của hệ thống.
    """
    labels_json = json.dumps([str(x) for x in (page_labels or [])], ensure_ascii=False)
    components.html(f"""
<script>
(function() {{
  try {{
    const W = window.parent, D = W.document;
    const NAV_LABELS = {labels_json};
    if (W.__veraUnsavedGuardV70) {{
      W.__veraUnsavedGuardV70.setNavLabels(NAV_LABELS);
      return;
    }}

    let dirty = false;
    let navLabels = NAV_LABELS.slice();
    const bannerId = 'vera-unsaved-banner-v70';

    function ensureBanner() {{
      let b = D.getElementById(bannerId);
      if (!b) {{
        b = D.createElement('div');
        b.id = bannerId;
        b.style.cssText = 'position:fixed;left:50%;top:8px;transform:translateX(-50%);z-index:2147483647;background:#fff3cd;color:#664d03;border:1px solid #D9D9D9;border-radius:8px;padding:8px 14px;font:600 13px Arial,sans-serif;box-shadow:0 2px 10px rgba(0,0,0,.12);display:none;max-width:92vw;text-align:center;';
        b.textContent = '⚠️ Có thay đổi chưa lưu. Hãy bấm Lưu trước khi chuyển trang hoặc thoát.';
        D.body.appendChild(b);
      }}
      return b;
    }}
    function setDirty() {{ dirty = true; ensureBanner().style.display = 'block'; }}
    function clearDirty() {{ dirty = false; const b = D.getElementById(bannerId); if (b) b.style.display='none'; }}
    function isEditableArea(t) {{
      if (!t || !t.closest) return false;
      if (t.closest('[data-testid="stDataEditor"]')) return true;

      const form = t.closest('[data-testid="stForm"]');
      if (form) {{
        const txt = (form.innerText || '').toLowerCase();
        return /lưu|cập nhật|ghi đè|xóa|xoá|thêm nhân viên|chỉnh sửa|thiết lập|save/.test(txt);
      }}

      // Các màn hình cũ chưa dùng st.form: chỉ đánh dấu dirty khi input nằm trong
      // một khối có nút ghi dữ liệu, đồng thời bỏ qua các ô chỉ dùng để lọc/tìm kiếm.
      const widget = t.closest('[data-testid="stTextInput"],[data-testid="stNumberInput"],[data-testid="stSelectbox"],[data-testid="stCheckbox"],[data-testid="stDateInput"],[data-testid="stTextArea"],[data-testid="stFileUploader"]');
      if (!widget) return false;
      const widgetText = (widget.innerText || '').toLowerCase();
      if (/lọc|tìm kiếm|tìm theo|search|chọn khoảng thời gian|chọn ngày:|lọc thời gian|chọn bảng cần tùy chỉnh/.test(widgetText)) return false;
      let block = widget.closest('[data-testid="stVerticalBlock"]');
      let hops = 0;
      while (block && hops < 4) {{
        const txt = (block.innerText || '').toLowerCase();
        if (/💾|lưu|cập nhật|ghi đè|xóa|xoá|thêm nhân viên|chỉnh sửa hồ sơ|thiết lập|save/.test(txt)) return true;
        block = block.parentElement ? block.parentElement.closest('[data-testid="stVerticalBlock"]') : null;
        hops += 1;
      }}
      return false;
    }}
    function isEditEventTarget(t) {{
      if (!isEditableArea(t)) return false;
      const tag = (t.tagName || '').toLowerCase();
      return ['input','textarea','select'].includes(tag) || (t.getAttribute && (t.getAttribute('role') === 'checkbox' || t.getAttribute('contenteditable') === 'true'));
    }}
    function markFromEvent(e) {{ if (isEditEventTarget(e.target)) setDirty(); }}
    D.addEventListener('input', markFromEvent, true);
    D.addEventListener('change', markFromEvent, true);
    D.addEventListener('keydown', function(e) {{
      if (isEditableArea(e.target) && !['Shift','Control','Alt','Meta','Tab','Escape','ArrowUp','ArrowDown','ArrowLeft','ArrowRight'].includes(e.key)) setDirty();
    }}, true);

    D.addEventListener('click', function(e) {{
      const btn = e.target && e.target.closest ? e.target.closest('button') : null;
      if (!btn) return;
      const text = (btn.innerText || btn.textContent || '').trim();
      const low = text.toLowerCase();

      // Các nút thực hiện ghi/xóa dữ liệu: form sẽ gửi ngay sau click nên bỏ cờ cảnh báo.
      if (/lưu|ghi đè|xóa|xoá|cập nhật hồ sơ|cập nhật nhân viên|thêm nhân viên|save/.test(low)) {{
        clearDirty();
        return;
      }}

      const inSidebar = !!btn.closest('[data-testid="stSidebar"]');
      const isNav = inSidebar || navLabels.some(x => text === x) || /đăng xuất/i.test(text);
      if (dirty && isNav) {{
        const ok = W.confirm('Bạn đang có thay đổi CHƯA LƯU.\n\nNhấn HỦY để quay lại và bấm LƯU bản cập nhật.\nNhấn OK nếu muốn rời trang và BỎ các thay đổi chưa lưu.');
        if (!ok) {{
          e.preventDefault(); e.stopPropagation(); e.stopImmediatePropagation();
          return false;
        }}
        clearDirty();
      }}
    }}, true);

    W.addEventListener('beforeunload', function(e) {{
      if (!dirty) return;
      e.preventDefault();
      e.returnValue = '';
    }});

    W.__veraUnsavedGuardV70 = {{
      setDirty, clearDirty,
      isDirty: () => dirty,
      setNavLabels: (xs) => {{ navLabels = Array.isArray(xs) ? xs.slice() : []; }}
    }};
  }} catch (err) {{}}
}})();
</script>
""", height=0, width=0)


render_global_unsaved_changes_guard(allowed_pages)

# V72 - NÚT LƯU NỔI TOÀN HỆ THỐNG.
# Khi trang có nút Lưu/Save đang khả dụng, một nút nổi cố định sẽ luôn hiện ở góc dưới.
# Nó tự nhắm tới nút Lưu gần vùng đang xem nhất, nên dùng được cho cả form và data_editor.
components.html(r"""
<script>
(function () {
  try {
    const win = window.parent;
    const doc = win.document;
    const FLOAT_ID = 'vera-global-floating-save-v72';
    const STYLE_ID = 'vera-global-floating-save-style-v72';

    if (!doc.getElementById(STYLE_ID)) {
      const style = doc.createElement('style');
      style.id = STYLE_ID;
      style.textContent = `
        #${FLOAT_ID} {
          position: fixed; right: 20px; bottom: 18px; z-index: 2147483000;
          display: none; align-items: center; justify-content: center;
          max-width: min(420px, calc(100vw - 32px)); min-height: 46px;
          padding: 10px 18px; border-radius: 12px; border: 1px solid #D9D9D9;
          background: #ffffff; color: #262730; font: 700 16px Roboto, Arial, sans-serif;
          box-shadow: 0 8px 28px rgba(0,0,0,.18); cursor: pointer;
          white-space: nowrap; overflow: hidden; text-overflow: ellipsis;
        }
        #${FLOAT_ID}:hover { transform: translateY(-2px); box-shadow: 0 10px 32px rgba(0,0,0,.22); }
        @media (max-width: 768px) {
          #${FLOAT_ID} { right: 10px; left: 10px; bottom: 12px; max-width: none; width: calc(100vw - 20px); font-size: 15px; }
        }
      `;
      doc.head.appendChild(style);
    }

    let floater = doc.getElementById(FLOAT_ID);
    if (!floater) {
      floater = doc.createElement('button');
      floater.id = FLOAT_ID;
      floater.type = 'button';
      floater.setAttribute('aria-label', 'Lưu thay đổi');
      doc.body.appendChild(floater);
    }

    let currentTarget = null;
    function isSaveButton(btn) {
      if (!btn || btn.id === FLOAT_ID || btn.disabled) return false;
      const txt = (btn.innerText || btn.textContent || '').trim();
      const low = txt.toLocaleLowerCase('vi-VN');
      if (!(low.includes('lưu') || low.includes('save') || txt.includes('💾'))) return false;
      const style = win.getComputedStyle(btn);
      if (style.display === 'none' || style.visibility === 'hidden') return false;
      if (!btn.offsetParent) return false;
      return true;
    }

    function refreshTarget() {
      const buttons = Array.from(doc.querySelectorAll('button')).filter(isSaveButton);
      if (!buttons.length) {
        currentTarget = null; floater.style.display = 'none'; return;
      }
      const center = win.innerHeight / 2;
      let best = null, bestScore = Number.POSITIVE_INFINITY;
      for (const btn of buttons) {
        const r = btn.getBoundingClientRect();
        const btnCenter = (r.top + r.bottom) / 2;
        let score = Math.abs(btnCenter - center);
        // Ưu tiên nút đang nằm trong viewport hiện tại.
        if (r.bottom < 0 || r.top > win.innerHeight) score += win.innerHeight * 0.75;
        if (score < bestScore) { bestScore = score; best = btn; }
      }
      currentTarget = best;
      if (!best) { floater.style.display = 'none'; return; }
      let label = (best.innerText || best.textContent || '💾 Lưu thay đổi').trim();
      if (label.length > 52) label = '💾 Lưu thay đổi';
      floater.textContent = label || '💾 Lưu thay đổi';
      floater.disabled = !!best.disabled;
      floater.style.display = 'flex';
    }

    floater.onclick = function () {
      refreshTarget();
      if (currentTarget && !currentTarget.disabled) currentTarget.click();
    };

    if (!win.__veraFloatingSaveBoundV72) {
      win.__veraFloatingSaveBoundV72 = true;
      let timer = null;
      const schedule = () => { clearTimeout(timer); timer = setTimeout(refreshTarget, 70); };
      win.addEventListener('scroll', schedule, {passive:true});
      win.addEventListener('resize', schedule, {passive:true});
      const observer = new MutationObserver(schedule);
      observer.observe(doc.body, {childList:true, subtree:true});
      win.__veraFloatingSaveObserverV72 = observer;
      win.__veraFloatingSaveIntervalV72 = win.setInterval(refreshTarget, 800);
    }
    refreshTarget();
  } catch (e) { console.debug('Vera floating save:', e); }
})();
</script>
""", height=0, width=0)


def open_app_page(page_name):
    if page_name not in allowed_pages:
        return
    st.session_state.app_page = page_name
    st.query_params["page"] = PAGE_SLUGS[page_name]
    st.session_state["_vera_collapse_sidebar_once"] = True
    st.rerun()


def collapse_sidebar_after_navigation_once():
    """Sau khi chọn xong một mục MENU CHỨC NĂNG, tự thu gọn sidebar một lần."""
    if not st.session_state.pop("_vera_collapse_sidebar_once", False):
        return
    components.html(r"""
    <script>
    (function(){
      try {
        const doc = window.parent.document;
        const clickCollapse = () => {
          const collapsedControl = doc.querySelector('[data-testid="collapsedControl"]');
          if (collapsedControl && collapsedControl.offsetParent !== null) return;
          const sidebar = doc.querySelector('[data-testid="stSidebar"]');
          if (!sidebar) return;
          const candidates = [
            doc.querySelector('[data-testid="stSidebarCollapseButton"] button'),
            doc.querySelector('[data-testid="stSidebarCollapseButton"]'),
            sidebar.querySelector('button[aria-label="Close sidebar"]'),
            sidebar.querySelector('button[aria-label*="sidebar" i]'),
            sidebar.querySelector('button[kind="header"]')
          ].filter(Boolean);
          const btn = candidates.find(x => !x.disabled);
          if (btn) btn.click();
        };
        setTimeout(clickCollapse, 80);
        setTimeout(clickCollapse, 260);
      } catch (e) {}
    })();
    </script>
    """, height=0, width=0)

# V86.4: mọi tài khoản dùng MENU CHỨC NĂNG dạng dọc trong sidebar.
st.sidebar.markdown(
    "<div style='font-size:18px;font-weight:800;line-height:1.15;margin:2px 0 10px 0;'>📌 MENU CHỨC NĂNG</div>",
    unsafe_allow_html=True
)
for page_name in allowed_pages:
    if st.sidebar.button(page_name, key=f"nav_{PAGE_SLUGS[page_name]}", use_container_width=True,
                         type="primary" if selected_page == page_name else "secondary"):
        open_app_page(page_name)
if st.session_state.current_role == "admin":
    st.sidebar.markdown("---")
    st.sidebar.caption("🔐 Quyền nghiệp vụ vẫn được quản lý tại trang Phân quyền chức năng. Hướng dẫn sử dụng luôn hiển thị cho mọi tài khoản.")
collapse_sidebar_after_navigation_once()

# --- GIAO DIỆN HEADER ---
st.write("")
col_title, col_logout = st.columns([8, 2])
with col_title:
    st.markdown("""
        <div class='custom-main-title'>VERA SPA TAM HIỆP ĐỒNG NAI</div>
    """, unsafe_allow_html=True)
with col_logout:
    if st.button("🚪 Đăng xuất", use_container_width=True):
        if st.session_state.current_user and st.session_state.current_user != "Quản Trị Viên":
            revoke_remember_token(st.session_state.current_user)
        st.session_state.logged_in = False
        st.session_state.current_user = ""
        st.session_state.current_role = ""
        st.session_state.birthday_login_event = False
        st.session_state.pop("birthday_notice_count_today", None)
        st.session_state.pop("birthday_notice_muted_today", None)
        st.session_state.pop("app_page", None)
        st.query_params['forget_login'] = '1'
        try: del st.query_params['remember_token']
        except Exception: pass
        try: del st.query_params['page']
        except Exception: pass
        st.rerun()

# V48: Thông báo sinh nhật đầu tháng cho Admin/Quản lý/Lễ tân.
render_birthday_login_notice(df_credentials)
# Nút xem sinh nhật chủ động đã được chuyển vào menu chức năng riêng.

# V86.4: điện thoại dùng menu dọc sidebar + Swipe; không lặp menu dạng lưới trong nội dung.

# V86.4: Swipe trái/phải giữa các trang chức năng trên điện thoại cho MỌI tài khoản. Nếu lịch sử có sẵn,
# ưu tiên back/forward; URL page giúp trạng thái được phục hồi chính xác.
components.html(f"""
<script>
(function() {{
    try {{
        const parentWin = window.parent, doc = parentWin.document;
        if (!parentWin.matchMedia('(max-width: 768px)').matches) return;
        const pages = {json.dumps([PAGE_SLUGS[p] for p in allowed_pages], ensure_ascii=False)};
        const current = {json.dumps(PAGE_SLUGS[selected_page])};
        let x0=null, y0=null, target0=null;
        doc.addEventListener('touchstart', function(e) {{
            if (!e.touches || e.touches.length !== 1) return;
            x0=e.touches[0].clientX; y0=e.touches[0].clientY; target0=e.target;
        }}, {{passive:true}});
        doc.addEventListener('touchend', function(e) {{
            if (x0===null || !e.changedTouches || e.changedTouches.length!==1) return;
            const t=target0; target0=null;
            if (t && t.closest && t.closest('input,textarea,button,a,[data-baseweb="select"],[data-testid="stDataFrame"],[data-testid="stDataEditor"]')) {{x0=y0=null;return;}}
            const dx=e.changedTouches[0].clientX-x0, dy=e.changedTouches[0].clientY-y0; x0=y0=null;
            if (Math.abs(dx)<90 || Math.abs(dx)<Math.abs(dy)*1.4) return;
            const i=pages.indexOf(current); if(i<0) return;
            const ni=dx<0 ? i+1 : i-1;
            if(ni<0 || ni>=pages.length) return;
            const url=new URL(parentWin.location.href); url.searchParams.set('page', pages[ni]);
            parentWin.location.href=url.toString();
        }}, {{passive:true}});
    }} catch(e) {{ console.debug('Vera swipe:',e); }}
}})();
</script>
""", height=0, width=0)

# Hồ sơ cá nhân là một trang riêng và KHÔNG hiển thị cho Admin.
@st.dialog("📋 Danh sách nghỉ trong ngày")
def show_daily_leave_popup(day_label, category_label, rows_df):
    st.markdown(f"**Ngày:** {day_label}")
    st.markdown(f"**Nhóm:** {category_label}")
    if not isinstance(rows_df, pd.DataFrame) or rows_df.empty:
        st.info("Không có nhân viên trong nhóm này.")
        return
    cols = [c for c in ["Tên nhân viên", "Lý do nghỉ"] if c in rows_df.columns]
    popup_df = rows_df[cols].copy()
    if "Lý do nghỉ" in popup_df.columns:
        popup_df["Lý do nghỉ"] = popup_df["Lý do nghỉ"].astype(str).apply(clean_leave_reason_display)
    popup_df = popup_df.drop_duplicates().reset_index(drop=True)
    st.dataframe(popup_df, width="stretch", hide_index=True, height="content")


if selected_page == "👤 Hồ sơ cá nhân" and st.session_state.current_role != "admin":
    st.subheader(f"👤 Cập nhật hồ sơ cá nhân: {st.session_state.current_user}")
    cred_row = df_credentials[df_credentials['Tên nhân viên'].apply(normalize_login_name) == normalize_login_name(st.session_state.current_user)]
    curr_fullname = str(cred_row.iloc[0].get('Họ và tên đầy đủ', '')).strip() if not cred_row.empty else ""
    curr_dob = str(cred_row.iloc[0].get('Ngày sinh', '')).strip() if not cred_row.empty else ""
    curr_phone = str(cred_row.iloc[0].get('Điện thoại', '')).strip().replace("'", "") if not cred_row.empty else ""
    curr_email = str(cred_row.iloc[0].get('Email', '')).strip() if not cred_row.empty else ""
    curr_address = str(cred_row.iloc[0].get('Địa chỉ', '')).strip() if not cred_row.empty else ""
    curr_bank_account = str(cred_row.iloc[0].get('Số tài khoản ngân hàng', '')).strip().replace("'", "") if not cred_row.empty else ""
    curr_bank_name = str(cred_row.iloc[0].get('Tên ngân hàng', '')).strip() if not cred_row.empty else ""

    # Không dùng st.form ở khu vực địa chỉ để Tỉnh/Thành -> Phường/Xã cập nhật ngay khi chọn.
    old_pass = st.text_input("Mật khẩu hiện tại (🔴 Bắt buộc để lưu)", type="password", key="profile_old_pass")
    new_pass = st.text_input("Mật khẩu mới (Bỏ trống nếu không đổi)", type="password", key="profile_new_pass")
    c1, c2 = st.columns(2)
    with c1:
        in_fullname = st.text_input("Họ và tên đầy đủ", value=curr_fullname, key="profile_fullname")
        in_dob = st.text_input("Ngày sinh (VD: 15/08/1990)", value=curr_dob, key="profile_dob")
        in_phone = st.text_input("Số điện thoại", value=curr_phone, key="profile_phone")
        in_email = st.text_input("Email", value=curr_email, key="profile_email")
    with c2:
        st.markdown("**📍 Địa chỉ**")
        in_address = vietnam_address_inputs("profile_address", curr_address)
        in_bank_account = st.text_input("Số tài khoản ngân hàng", value=curr_bank_account, key="profile_bank_account")
        in_bank_name = bank_selectbox("Tên ngân hàng", key="profile_bank_name", current_value=curr_bank_name)
    if st.button("💾 Lưu thay đổi", use_container_width=True, key="profile_save_button"):
        db_old_pass = str(cred_row.iloc[0]['Mật khẩu']) if not cred_row.empty else "123456"
        if not password_matches(old_pass, db_old_pass):
            st.error("❌ Mật khẩu hiện tại không chính xác!")
        elif new_pass and len(str(new_pass)) < 4:
            st.error("❌ Mật khẩu mới quá ngắn.")
        else:
            ok, msg = update_user_profile(
                st.session_state.current_user, new_pass, in_fullname.strip(), in_dob.strip(),
                in_phone.strip(), in_email.strip(), in_address.strip(),
                in_bank_account.strip(), in_bank_name.strip()
            )
            (st.success if ok else st.error)(msg)
            if ok: st.rerun()

# ==========================================
# CÁC TRANG CHỨC NĂNG ĐỘC LẬP
# ==========================================
if selected_page == "📘 Hướng dẫn sử dụng":
    is_admin_guide = str(st.session_state.get("current_role", "")).strip().lower() == "admin"
    meta_guide, raw_guide, guide_err = load_usage_guide_document()
    if guide_err:
        st.error(guide_err)

    if is_admin_guide:
        with st.expander("🛠️ Quản trị Hướng dẫn sử dụng", expanded=(meta_guide is None)):
            st.caption(
                "Admin có thể tải PDF/ảnh mới, sửa tiêu đề/phiên bản/ghi chú hoặc xóa tài liệu. "
                f"Giới hạn mỗi file {USAGE_GUIDE_MAX_BYTES // (1024*1024)} MB."
            )
            current_title = str((meta_guide or {}).get("Tên tài liệu", "Hướng dẫn sử dụng VERA SPA"))
            current_version = str((meta_guide or {}).get("Phiên bản", ""))
            current_note = str((meta_guide or {}).get("Ghi chú", ""))
            guide_file = st.file_uploader(
                "Tải bản hướng dẫn mới (PDF/PNG/JPG/WEBP)",
                type=["pdf", "png", "jpg", "jpeg", "webp"], key="usage_guide_upload_v864"
            )
            cgt1, cgt2 = st.columns(2)
            with cgt1:
                guide_title = st.text_input("Tên tài liệu", value=current_title, key="usage_guide_title_v864")
            with cgt2:
                guide_version = st.text_input("Phiên bản", value=current_version, key="usage_guide_version_v864")
            guide_note = st.text_area("Ghi chú hiển thị", value=current_note, key="usage_guide_note_v864", height=90)
            csave, cmeta, cdel = st.columns(3)
            with csave:
                if st.button("⬆️ Tải lên / Thay thế", use_container_width=True, key="usage_guide_replace_v864"):
                    ok, msg = save_usage_guide_document(guide_file, guide_title, guide_version, guide_note, st.session_state.get("current_user", "Admin"))
                    (st.success if ok else st.error)(msg)
                    if ok: st.rerun()
            with cmeta:
                if st.button("💾 Sửa thông tin", use_container_width=True, disabled=not bool(meta_guide), key="usage_guide_meta_v864"):
                    ok, msg = update_usage_guide_metadata(guide_title, guide_version, guide_note, st.session_state.get("current_user", "Admin"))
                    (st.success if ok else st.error)(msg)
                    if ok: st.rerun()
            with cdel:
                confirm_delete = st.checkbox("Xác nhận xóa", key="usage_guide_confirm_delete_v864")
                if st.button("🗑️ Xóa tài liệu", use_container_width=True, disabled=not (bool(meta_guide) and confirm_delete), key="usage_guide_delete_v864"):
                    ok, msg = delete_usage_guide_document(st.session_state.get("current_user", "Admin"))
                    (st.success if ok else st.error)(msg)
                    if ok: st.rerun()
        meta_guide, raw_guide, guide_err = load_usage_guide_document()

    if not meta_guide or raw_guide is None:
        if not guide_err:
            st.info("Admin chưa tải Hướng dẫn sử dụng lên hệ thống.")
    else:
        if not is_admin_guide:
            st.caption("🔒 Tài liệu chỉ dành để đọc trong hệ thống. Không có nút tải xuống, chia sẻ hoặc in.")
        render_protected_usage_guide(meta_guide, raw_guide)

elif selected_page == "🎂 Sinh nhật nhân sự" and has_feature_access("birthday"):
    st.subheader("🎂 Sinh nhật nhân sự trong tháng")
    st.caption("Danh sách được lấy từ hồ sơ mới nhất và sắp theo ngày sinh trong tháng.")
    render_manual_birthday_check(
        load_credentials_recent(),
        key_prefix=f"birthday_menu_{normalize_login_name(st.session_state.current_user) or 'user'}"
    )
elif selected_page == "👤 Hồ sơ cá nhân":
    pass  # Nội dung hồ sơ đã hiển thị ở phía trên.
elif selected_page == "⏰ Thiết lập ca làm việc" and has_feature_access("shift"):
    st.subheader("⏰ Thiết lập ca làm việc")
    st.info(
        "Chỉ hiển thị tài khoản role `nhanvien`. Bảng được đặt trong Form nên khi chỉnh nhiều ô sẽ không tải lại dữ liệu nguồn; "
        "hệ thống chỉ ghi Google Sheet khi bấm **Lưu Toàn Bộ Cấu Hình Ca**."
    )

    shift_base = get_nhanvien_shift_dataframe(df_credentials)
    shift_seed_key = "shift_schedule_working_df"
    shift_seed_signature_key = "shift_schedule_seed_signature"
    shift_signature = "|".join(
        f"{normalize_login_name(r.get('Tên nhân viên',''))}:{r.get('Ca làm việc','')}:{r.get('Ngày bắt đầu ca','')}:{r.get('Chu kỳ','')}"
        for _, r in shift_base.iterrows()
    )
    if shift_seed_key not in st.session_state or st.session_state.get(shift_seed_signature_key) != shift_signature:
        st.session_state[shift_seed_key] = shift_base.copy()
        st.session_state[shift_seed_signature_key] = shift_signature

    c_export_shift, c_import_shift = st.columns(2)
    with c_export_shift:
        st.download_button(
            "📥 Export Excel template phân ca",
            data=build_shift_template_excel_bytes(df_credentials),
            file_name=f"Vera-Spa_Template_Ca_{get_vn_today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help="Template có dropdown sẵn ở cột Ca làm việc và Chu kỳ luân phiên."
        )
    with c_import_shift:
        uploaded_shift = st.file_uploader(
            "📤 Import Excel template", type=["xlsx", "xlsm"], key="shift_template_upload",
            help="Dùng đúng template Export từ hệ thống để giữ dropdown và tên cột chuẩn."
        )

    if uploaded_shift is not None:
        imported_shift, import_err = read_shift_template_excel(uploaded_shift, df_credentials)
        if import_err:
            st.error(import_err)
        elif not imported_shift.empty:
            if st.button("✅ Nạp dữ liệu Import vào bảng chỉnh sửa", use_container_width=True, key="apply_shift_import"):
                base = get_nhanvien_shift_dataframe(df_credentials).copy()
                imp_map = {normalize_login_name(r['Tên nhân viên']): r for _, r in imported_shift.iterrows()}
                for idx, rr in base.iterrows():
                    key = normalize_login_name(rr.get('Tên nhân viên',''))
                    if key in imp_map:
                        ir = imp_map[key]
                        base.at[idx, 'Ca làm việc'] = str(ir.get('Ca làm việc','')).strip()
                        base.at[idx, 'Ngày bắt đầu ca'] = str(ir.get('Ngày bắt đầu ca','')).strip()
                        base.at[idx, 'Chu kỳ'] = str(ir.get('Chu kỳ','')).strip()
                st.session_state[shift_seed_key] = base
                st.success(f"Đã nạp {len(imported_shift)} dòng từ file. Kiểm tra lại rồi bấm Lưu.")
                st.rerun()

    working_shift = st.session_state.get(shift_seed_key, shift_base).copy()
    st.markdown("#### 🔎 Bộ lọc")
    f1, f2, f3 = st.columns([2.2, 1.5, 1.5])
    with f1:
        filter_shift_name = st.text_input("Tìm theo tên nhân viên", key="shift_filter_name", placeholder="Gõ tên để lọc...")
    with f2:
        shift_filter_value = st.selectbox("Lọc theo ca", ["- Tất cả ca -"] + SHIFT_OPTIONS, key="shift_filter_shift")
    with f3:
        cycle_filter_value = st.selectbox("Lọc theo luân phiên", ["- Tất cả luân phiên -"] + SHIFT_CYCLE_OPTIONS, key="shift_filter_cycle")

    filtered_shift = working_shift.copy()
    if str(filter_shift_name).strip():
        keyq = normalize_login_name(filter_shift_name)
        filtered_shift = filtered_shift[filtered_shift['Tên nhân viên'].astype(str).apply(normalize_login_name).str.contains(keyq, regex=False)]
    if shift_filter_value != "- Tất cả ca -":
        filtered_shift = filtered_shift[filtered_shift['Ca làm việc'].astype(str).eq(shift_filter_value)]
    if cycle_filter_value != "- Tất cả luân phiên -":
        filtered_shift = filtered_shift[filtered_shift['Chu kỳ'].astype(str).eq(cycle_filter_value)]

    st.caption(f"Đang hiển thị {len(filtered_shift)}/{len(working_shift)} nhân viên.")
    if filtered_shift.empty:
        st.info("Không có nhân viên phù hợp bộ lọc.")
    else:
        with st.form("shift_schedule_batch_form", clear_on_submit=False):
            edited_df = st.data_editor(
                filtered_shift,
                height=min(900, max(160, (len(filtered_shift) * 36) + 42)),
                column_config={
                    "Tên nhân viên": st.column_config.TextColumn("Tên nhân viên", disabled=True),
                    "Ca làm việc": st.column_config.SelectboxColumn(
                        "Ca làm việc", options=SHIFT_OPTIONS, width="large"
                    ),
                    "Ngày bắt đầu ca": st.column_config.TextColumn("Ngày bắt đầu (DD/MM/YYYY)"),
                    "Chu kỳ": st.column_config.SelectboxColumn(
                        "Chu kỳ luân phiên", options=SHIFT_CYCLE_OPTIONS, width="medium"
                    )
                },
                hide_index=True, use_container_width=True, key="shift_schedule_editor_form"
            )
            submit_shift = st.form_submit_button("💾 Lưu Toàn Bộ Cấu Hình Ca", use_container_width=True)

        if submit_shift:
            # Ghép các dòng đang lọc/chỉnh vào toàn bộ danh sách rồi mới ghi 1 lần.
            updated_all = working_shift.copy()
            edit_map = {normalize_login_name(r['Tên nhân viên']): r for _, r in edited_df.iterrows()}
            for idx, rr in updated_all.iterrows():
                k = normalize_login_name(rr.get('Tên nhân viên',''))
                if k in edit_map:
                    er = edit_map[k]
                    updated_all.at[idx, 'Ca làm việc'] = str(er.get('Ca làm việc','')).strip()
                    updated_all.at[idx, 'Ngày bắt đầu ca'] = str(er.get('Ngày bắt đầu ca','')).strip()
                    updated_all.at[idx, 'Chu kỳ'] = str(er.get('Chu kỳ','')).strip()
            with st.spinner("Đang lưu đồng loạt vào hệ thống..."):
                res, msg = batch_update_shift_schedule(updated_all)
            (st.success if res else st.error)(msg)
            if res:
                st.session_state.pop(shift_seed_key, None)
                st.session_state.pop(shift_seed_signature_key, None)
                st.rerun()

elif selected_page == "👥 Danh sách nhân sự" and has_feature_access("staff_list"):
    st.subheader("👥 Danh sách nhân sự")
    staff_source_df = build_staff_list_dataframe(df_credentials)

    c_staff_export, c_staff_import = st.columns([1, 1])
    with c_staff_export:
        st.download_button(
            "📥 Export danh sách nhân viên hiện tại",
            data=staff_list_to_excel(staff_source_df),
            file_name=f"VeraSpa_DanhSachNhanSu_{get_vn_today().strftime('%d%m%Y')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="export_current_staff_list"
        )
    with c_staff_import:
        if st.session_state.current_role in {'admin', 'letan', 'quanly'}:
            staff_import_file = st.file_uploader(
                "Import lại hệ thống", type=['xlsx'], key="import_staff_list_file",
                help="Dùng file đã Export. Tên nhân viên là khóa; import không tạo/xóa tài khoản và không ghi đè mật khẩu."
            )
        else:
            staff_import_file = None
            st.button("📤 Import lại hệ thống", disabled=True, use_container_width=True, key="staff_import_not_allowed")

    if staff_import_file is not None and st.session_state.current_role in {'admin', 'letan', 'quanly'}:
        imported_staff_df, import_staff_err = read_staff_list_import(staff_import_file)
        if import_staff_err:
            st.error(import_staff_err)
        else:
            preview_cols = [c for c in STAFF_EXPORT_COLUMNS if c in imported_staff_df.columns]
            st.caption(f"Xem trước file import: {len(imported_staff_df)} nhân viên. Chỉ cập nhật các tài khoản đã tồn tại.")
            st.dataframe(imported_staff_df[preview_cols], width='stretch', height=min(480, 70 + len(imported_staff_df) * 35), hide_index=True)
            if st.button("⬆️ Xác nhận Import danh sách nhân sự", use_container_width=True, type="primary", key="confirm_import_staff_list"):
                ok, msg = batch_import_staff_list(
                    imported_staff_df, st.session_state.current_user, st.session_state.current_role
                )
                (st.success if ok else st.error)(msg)
                if ok:
                    st.rerun()

    cols_staff = [c for c in STAFF_EXPORT_COLUMNS if c in staff_source_df.columns]
    staff_df, staff_widths = apply_table_layout_df(staff_source_df[cols_staff], "staff_list")
    st.dataframe(
        apply_table_visual_styler(staff_df, "staff_list", list(staff_df.columns)),
        width='stretch', height='content', hide_index=True,
        row_height=layout_row_height("staff_list"),
        column_config=table_layout_column_config("staff_list", list(staff_df.columns))
    )
    render_admin_quick_layout_default("staff_list", list(staff_df.columns), "staff_list_page")

elif selected_page == "⚙️ Giao diện tùy chỉnh" and st.session_state.current_role == "admin":
    st.subheader("⚙️ Giao diện tùy chỉnh")
    render_admin_theme_config_panel()

    with st.expander("📌 Sắp xếp MENU CHỨC NĂNG", expanded=False):
        st.caption(
            "Chỉ áp dụng cho tài khoản Admin. Chọn một nút rồi di chuyển lên/xuống "
            "hoặc chuyển thẳng đến vị trí mong muốn. Thứ tự được lưu trên Google Sheet."
        )

        _menu_default = list(PAGE_FEATURE_KEYS.keys())
        _menu_current = admin_menu_order_for_pages(_menu_default)
        _menu_table = pd.DataFrame({
            "Vị trí": list(range(1, len(_menu_current) + 1)),
            "Nút chức năng": _menu_current,
        })
        st.dataframe(_menu_table, width="stretch", hide_index=True, height="content")

        _menu_item = st.selectbox(
            "Chọn nút cần di chuyển",
            _menu_current,
            key="admin_menu_move_item",
            filter_mode="contains",
        )
        _current_pos = _menu_current.index(_menu_item) + 1 if _menu_item in _menu_current else 1
        _target_pos = st.number_input(
            "Chuyển đến vị trí",
            min_value=1,
            max_value=max(1, len(_menu_current)),
            value=int(_current_pos),
            step=1,
            key=f"admin_menu_target_{PAGE_SLUGS.get(_menu_item, 'page')}",
        )

        _m1, _m2, _m3 = st.columns(3)
        with _m1:
            _move_up = st.button("⬆️ Lên 1 vị trí", use_container_width=True, key="admin_menu_up")
        with _m2:
            _move_down = st.button("⬇️ Xuống 1 vị trí", use_container_width=True, key="admin_menu_down")
        with _m3:
            _move_to = st.button("↕️ Chuyển đến vị trí", use_container_width=True, key="admin_menu_move_to")

        _new_menu_order = None
        if _move_up and _menu_item in _menu_current:
            _idx = _menu_current.index(_menu_item)
            _new_menu_order = _move_menu_item(_menu_current, _menu_item, max(0, _idx - 1))
        elif _move_down and _menu_item in _menu_current:
            _idx = _menu_current.index(_menu_item)
            _new_menu_order = _move_menu_item(_menu_current, _menu_item, min(len(_menu_current) - 1, _idx + 1))
        elif _move_to and _menu_item in _menu_current:
            _new_menu_order = _move_menu_item(_menu_current, _menu_item, int(_target_pos) - 1)

        if _new_menu_order is not None:
            _ok, _msg = save_admin_menu_order(_new_menu_order, st.session_state.current_user)
            (st.success if _ok else st.error)(_msg)
            if _ok:
                st.rerun()

        if st.button("♻️ Khôi phục thứ tự MENU mặc định", use_container_width=True, key="admin_menu_reset_default"):
            _ok, _msg = save_admin_menu_order(_menu_default, st.session_state.current_user)
            (st.success if _ok else st.error)(_msg)
            if _ok:
                st.rerun()

    with st.expander("⚡ Hạ tầng & hiệu năng (Cloud Run + PostgreSQL)", expanded=False):
        if vpg is None or not vpg.is_enabled():
            st.warning("PostgreSQL chưa được bật. Khi deploy Cloud Run hãy đặt VERA_DB_ENABLED=1 và cấu hình DB/Cloud SQL.")
        else:
            _pg_ok, _pg_msg = get_postgres_runtime_status()
            (st.success if _pg_ok else st.error)(_pg_msg)
            if _pg_ok:
                _pg_status = vpg.get_status()
                if isinstance(_pg_status, pd.DataFrame) and not _pg_status.empty:
                    st.dataframe(_pg_status, width="stretch", hide_index=True, height="content")
                if st.button("⚡ Đồng bộ dữ liệu nặng Google Sheets → PostgreSQL", use_container_width=True, key="admin_pg_prewarm_v75"):
                    _sync_jobs = [
                        ("credentials", _load_credentials_from_sheets, 30),
                        ("leave_primary", _load_backup_sheet_data_from_sheets, 45),
                        ("leave_secondary", _load_secondary_leave_sheet_data_from_sheets, 90),
                        ("tichluy", _load_tichluy_tracking_from_sheets, 90),
                        ("violation_debt", _load_violation_debt_ledger_from_sheets, 60),
                        ("payroll_history", _load_payroll_history_from_sheets, 90),
                    ]
                    _sync_errors = []
                    _progress = st.progress(0)
                    for _i, (_key, _loader, _ttl) in enumerate(_sync_jobs, start=1):
                        try:
                            vpg.load_dataset(_key, _loader, ttl_seconds=_ttl, force_refresh=True)
                        except Exception as _exc:
                            _sync_errors.append(f"{_key}: {_exc}")
                        _progress.progress(_i / len(_sync_jobs))
                    if _sync_errors:
                        st.error("Một số nhóm chưa đồng bộ: " + " | ".join(_sync_errors))
                    else:
                        st.success("Đã đưa các nhóm dữ liệu đọc nhiều vào PostgreSQL dùng chung cho mọi Cloud Run instance.")

    st.info(
        "Admin có thể tùy chỉnh thứ tự, độ rộng cột, độ cao dòng, font, cỡ chữ, kiểu chữ, "
        "căn lề và Wrap Text. Sau khi lưu, cấu hình được dùng chung cho toàn bộ tài khoản."
    )

    st.markdown(
        "<div style='background:#FFF2CC;border:1px solid #E6C95C;border-radius:8px;padding:8px 12px;"
        "font-weight:700;margin:4px 0 6px 0;'>🎨 Chọn bảng cần tùy chỉnh</div>",
        unsafe_allow_html=True
    )
    table_key = st.selectbox(
        "Chọn bảng cần tùy chỉnh",
        options=list(TABLE_LAYOUT_LABELS.keys()),
        format_func=lambda x: TABLE_LAYOUT_LABELS.get(x, x),
        key="ui_layout_table_selector", label_visibility="collapsed"
    )
    available_cols = get_table_columns_for_settings(table_key)
    if not available_cols:
        st.warning("Chưa xác định được danh sách cột của bảng này.")
    else:
        rows_state_key = f"layout_rows_state_{table_key}"
        version_state_key = f"layout_editor_version_{table_key}"
        row_height_state_key = f"layout_row_height_state_{table_key}"
        init_flag_key = f"layout_init_signature_{table_key}"
        signature = "|".join(map(str, available_cols))

        # Chỉ khởi tạo từ dữ liệu đã lưu khi mới mở bảng hoặc sau khi Save/Reset.
        if st.session_state.get(init_flag_key) != signature or rows_state_key not in st.session_state:
            st.session_state[rows_state_key] = _layout_editor_rows_from_saved(table_key, available_cols)
            saved_row_height, _ = get_table_visual_settings(table_key, available_cols)
            st.session_state[row_height_state_key] = int(saved_row_height)
            st.session_state[version_state_key] = int(st.session_state.get(version_state_key, 0) or 0) + 1
            st.session_state[init_flag_key] = signature

        st.markdown("#### 🧱 Kích thước dòng")
        st.number_input(
            "Độ cao dòng (px)", min_value=24, max_value=120, step=2,
            key=row_height_state_key,
            help="Áp dụng cho các dòng dữ liệu của bảng. Giá trị gợi ý: 32–48 px."
        )

        st.markdown("#### 🎨 Thiết lập từng cột")
        st.caption(
            "Cột **Vị trí** dùng cơ chế chèn tự động. Ví dụ: đổi một cột từ vị trí 6 → 3 thì "
            "các cột đang ở 3, 4, 5 sẽ tự chuyển thành 4, 5, 6."
        )

        editor_version = int(st.session_state.get(version_state_key, 0) or 0)
        editor_key = f"layout_editor_{table_key}_{editor_version}"
        cfg_df = pd.DataFrame(st.session_state.get(rows_state_key, []))

        st.data_editor(
            cfg_df,
            key=editor_key,
            width="stretch", height="content", hide_index=True, num_rows="fixed",
            disabled=["Tên cột"],
            row_height=42,
            on_change=_layout_editor_on_change,
            args=(table_key, editor_key, rows_state_key, version_state_key),
            column_config={
                "Tên cột": st.column_config.TextColumn("Tên cột", disabled=True, width=210),
                "Vị trí": st.column_config.NumberColumn(
                    "Vị trí", min_value=1, max_value=max(1, len(cfg_df)), step=1, format="%d", width=75
                ),
                "Độ rộng (px)": st.column_config.NumberColumn(
                    "Độ rộng (px)", min_value=50, max_value=800, step=10, format="%d", width=110
                ),
                "Font chữ": st.column_config.SelectboxColumn(
                    "Font chữ", options=TABLE_LAYOUT_FONT_OPTIONS, width=135
                ),
                "Cỡ chữ": st.column_config.NumberColumn(
                    "Cỡ chữ", min_value=8, max_value=30, step=1, format="%d", width=85
                ),
                "Kiểu chữ": st.column_config.SelectboxColumn(
                    "Kiểu chữ", options=TABLE_LAYOUT_FONT_STYLE_OPTIONS, width=135
                ),
                "Căn lề": st.column_config.SelectboxColumn(
                    "Căn lề", options=TABLE_LAYOUT_ALIGN_OPTIONS, width=95,
                    help="left = trái, center = giữa, right = phải"
                ),
                "Wrap text": st.column_config.CheckboxColumn(
                    "Wrap text", width=95,
                    help="Bật để nội dung được phép xuống dòng khi chiều rộng cột nhỏ."
                ),
            }
        )

        st.caption(
            "Các tiêu đề cột vẫn luôn được phép xuống dòng để tránh mất chữ. Font/căn lề/Wrap Text "
            "được áp dụng cho bảng hiển thị; độ rộng và độ cao dòng áp dụng cả bảng hiển thị và bảng chỉnh sửa."
        )

        c_save_layout, c_default_layout, c_reset_layout = st.columns(3)
        with c_save_layout:
            _save_layout_clicked = st.button("💾 Lưu & áp dụng", use_container_width=True, key=f"save_layout_{table_key}")
        with c_default_layout:
            _default_layout_clicked = st.button("⭐ Lưu làm mặc định", use_container_width=True, key=f"save_default_layout_{table_key}")

        if _save_layout_clicked or _default_layout_clicked:
            rows = [dict(r) for r in st.session_state.get(rows_state_key, [])]
            rows = sorted(rows, key=lambda r: int(float(r.get("Vị trí", 9999))))
            for pos, r in enumerate(rows, start=1):
                r["Vị trí"] = pos
            new_order = [str(r.get("Tên cột", "")) for r in rows if str(r.get("Tên cột", ""))]
            new_widths = {
                str(r["Tên cột"]): max(50, min(800, int(float(r.get("Độ rộng (px)", 140)))))
                for r in rows if str(r.get("Tên cột", ""))
            }
            visual_cols = {}
            for r in rows:
                col = str(r.get("Tên cột", ""))
                if not col:
                    continue
                visual_cols[col] = {
                    "font_family": str(r.get("Font chữ", "Roboto")),
                    "font_size": max(8, min(30, int(float(r.get("Cỡ chữ", TABLE_LAYOUT_DEFAULT_FONT_SIZE))))),
                    "font_style": str(r.get("Kiểu chữ", "Thường")),
                    "align": str(r.get("Căn lề", _default_column_alignment(col))).lower(),
                    "wrap": bool(r.get("Wrap text", True)),
                }
            visual = {
                "row_height": max(24, min(120, int(st.session_state.get(row_height_state_key, TABLE_LAYOUT_DEFAULT_ROW_HEIGHT)))),
                "columns": visual_cols,
            }
            ok, msg = save_table_layout_config(
                table_key, new_order, new_widths, st.session_state.current_user, visual=visual
            )
            if ok and _default_layout_clicked:
                msg = "Đã lưu bố cục này làm mặc định và áp dụng toàn hệ thống."
            (st.success if ok else st.error)(msg)
            if ok:
                st.session_state.pop(rows_state_key, None)
                st.session_state.pop(init_flag_key, None)
                st.rerun()

        with c_reset_layout:
            if st.button("♻️ Khôi phục mặc định", use_container_width=True, key=f"reset_layout_{table_key}"):
                default_order = list(available_cols)
                default_widths = {c: _default_column_width(c) for c in default_order}
                default_visual = {
                    "row_height": TABLE_LAYOUT_DEFAULT_ROW_HEIGHT,
                    "columns": {c: _default_column_visual_style(c) for c in default_order},
                }
                ok, msg = save_table_layout_config(
                    table_key, default_order, default_widths, st.session_state.current_user, visual=default_visual
                )
                (st.success if ok else st.error)(msg)
                if ok:
                    st.session_state.pop(rows_state_key, None)
                    st.session_state.pop(init_flag_key, None)
                    st.session_state.pop(row_height_state_key, None)
                    st.rerun()

elif selected_page == "🔐 Phân quyền chức năng" and st.session_state.current_role == "admin":
    st.subheader("🔐 Phân quyền chức năng")
    st.caption("Cấu hình theo vai trò trước; quyền riêng từng tài khoản sẽ ghi đè quyền của vai trò. Admin luôn giữ toàn quyền để tránh tự khóa hệ thống.")
    feature_keys = list(FEATURE_DEFINITIONS.keys())
    feature_label_to_key = {label: key for key, label in FEATURE_DEFINITIONS.items()}
    feature_labels = [FEATURE_DEFINITIONS[k] for k in feature_keys]

    with st.expander("👥 Quyền mặc định theo loại tài khoản", expanded=True):
        role_choice = st.selectbox("Chọn vai trò", [r for r in ALL_ACCOUNT_ROLES if r != 'admin'], key="perm_role_choice")
        current_role_allowed = [
            FEATURE_DEFINITIONS[k] for k in feature_keys
            if has_feature_access(k, role=role_choice, username='')
        ]
        role_allowed_labels = st.multiselect(
            "Các chức năng được phép", feature_labels, default=current_role_allowed,
            filter_mode="contains", key=f"perm_role_features_{role_choice}"
        )
        if st.button("💾 Lưu quyền cho vai trò", use_container_width=True, key=f"save_role_perm_{role_choice}"):
            ok, msg = save_role_feature_permissions(
                role_choice, [feature_label_to_key[x] for x in role_allowed_labels], st.session_state.current_user
            )
            (st.success if ok else st.error)(msg)
            if ok: st.rerun()

    with st.expander("👤 Quyền riêng theo từng tài khoản", expanded=False):
        account_options = sort_employee_names(df_credentials['Tên nhân viên'].dropna().astype(str).tolist()) if not df_credentials.empty else []
        account_choice = st.selectbox("Chọn tài khoản", account_options, filter_mode="contains", key="perm_account_choice") if account_options else ""
        if account_choice:
            account_row = latest_credential_row_from_credentials(df_credentials, account_choice)
            account_role = str(account_row.get('Phân quyền', 'nhanvien')).strip().lower() if account_row is not None else 'nhanvien'
            _role_cfg, _account_cfg = load_feature_permissions()
            has_account_override = any(k[0] == normalize_login_name(account_choice) for k in _account_cfg)
            if has_account_override:
                current_account_allowed = [FEATURE_DEFINITIONS[k] for k in feature_keys if _account_cfg.get((normalize_login_name(account_choice), k), False)]
            else:
                current_account_allowed = [FEATURE_DEFINITIONS[k] for k in feature_keys if has_feature_access(k, role=account_role, username='')]
            st.caption(f"Vai trò hiện tại: {account_role} · {'đang dùng quyền riêng' if has_account_override else 'đang kế thừa quyền theo vai trò'}")
            account_allowed_labels = st.multiselect(
                "Các chức năng cho tài khoản này", feature_labels, default=current_account_allowed,
                filter_mode="contains", key=f"perm_account_features_{normalize_login_name(account_choice)}"
            )
            cpa, cpb = st.columns(2)
            with cpa:
                if st.button("💾 Lưu quyền riêng", use_container_width=True, key="save_account_permissions"):
                    ok, msg = save_account_feature_permissions(
                        account_choice, [feature_label_to_key[x] for x in account_allowed_labels], st.session_state.current_user, inherit=False
                    )
                    (st.success if ok else st.error)(msg)
                    if ok: st.rerun()
            with cpb:
                if st.button("♻️ Dùng quyền theo vai trò", use_container_width=True, key="inherit_account_permissions"):
                    ok, msg = save_account_feature_permissions(account_choice, [], st.session_state.current_user, inherit=True)
                    (st.success if ok else st.error)(msg)
                    if ok: st.rerun()

elif selected_page == "➕ Thêm nhân viên" and has_feature_access("employee_add"):
    st.subheader("➕ Thêm nhân viên")
    st.write("Nhập thông tin nhân viên mới:")
    st.caption("📍 Địa chỉ dùng danh mục hành chính Việt Nam sau 01/07/2025; khi lưu sẽ tự ghép vào duy nhất cột Địa chỉ.")
    col1, col2 = st.columns(2)
    with col1:
        new_usr = st.text_input("Tên đăng nhập (Bắt buộc)", key="new_emp_username")
        new_pwd = st.text_input("Mật khẩu", value="123456", key="new_emp_password")
        _new_role_options = ALL_ACCOUNT_ROLES if st.session_state.current_role == 'admin' else ["nhanvien", "locker", "tapvu"]
        new_role = st.selectbox("Phân quyền", _new_role_options, filter_mode="contains", key="new_emp_role")
        new_fn = st.text_input("Họ và tên đầy đủ", key="new_emp_fullname")
        new_phone = st.text_input("Số điện thoại", key="new_emp_phone")
        new_email = st.text_input("Email", key="new_emp_email")
    with col2:
        st.markdown("**📍 Địa chỉ**")
        new_address = vietnam_address_inputs("new_emp_address", "")
        new_bank_account = st.text_input("Số tài khoản ngân hàng", key="new_emp_bank_account")
        new_bank_name = bank_selectbox("Tên ngân hàng", key="new_employee_bank_name", current_value="")

    if st.button("💾 Lưu Nhân Viên Mới", use_container_width=True, key="save_new_employee"):
        if new_usr:
            try:
                client = get_gspread_client()
                sheet_mk = client.open_by_key(SHEET_MAT_KHAU_ID).get_worksheet(0)
                all_emps = _gs_call_with_backoff(sheet_mk.col_values, 2)

                if normalize_login_name(new_usr) in {normalize_login_name(x) for x in all_emps}:
                    st.error("Tên đăng nhập đã tồn tại (hệ thống không phân biệt dấu và HOA/thường)!")
                else:
                    stt_new = len(all_emps)
                    row_data = [
                        stt_new, new_usr, str(new_pwd), new_role, new_fn, "", new_phone, new_email, new_address,
                        new_bank_account, new_bank_name, "0", "0", "0", "", "", "", "", "", ""
                    ]
                    _gs_call_with_backoff(sheet_mk.append_row, row_data, value_input_option='USER_ENTERED')
                    start_work_date = get_vn_today()
                    role_new = str(new_role).strip().lower()
                    # Nhân viên và Leader tham gia TichLuy. Quanly/letan/locker/tapvu/admin không xuất hiện ở TichLuy.
                    if role_new not in TICHLUY_EXCLUDED_ROLES:
                        tl_ok, tl_msg = ensure_employee_in_tichluy(new_usr, start_work_date)
                    else:
                        tl_ok, tl_msg = True, 'Vai trò này không tham gia TichLuy.'

                    # KHÔNG đồng bộ nhân viên mới sang file 1Kz0... theo yêu cầu mới.
                    # Sau khi thêm phải đánh lại STT cột A của Sheet1 và TichLuy.
                    stt_ok, stt_msg = renumber_credential_sheet_stt(sheet_mk)
                    try:
                        load_credentials.clear()
                    except Exception:
                        pass
                    tl_sync_ok, tl_sync_msg = sync_tichluy_roles_and_stt()
                    _clear_dynamic_data_caches()

                    if tl_ok and stt_ok and tl_sync_ok:
                        extra = f" · Ngày bắt đầu làm {start_work_date.strftime('%d/%m/%Y')}" if role_new not in TICHLUY_EXCLUDED_ROLES else ""
                        st.success(f"Đã thêm thành công: {new_usr}{extra} · đã sắp xếp lại STT Sheet1/TichLuy.")
                    else:
                        st.warning(
                            f"Đã tạo tài khoản {new_usr}, nhưng có bước phụ chưa hoàn tất: "
                            f"{tl_msg} | {stt_msg} | {tl_sync_msg}"
                        )
            except Exception as e:
                st.error(f"Lỗi: {e}")
        else:
            st.error("Vui lòng nhập Tên đăng nhập.")


elif selected_page == "✏️ Sửa / Xóa nhân viên" and has_page_access("✏️ Sửa / Xóa nhân viên"):
    st.subheader("✏️ Sửa / Xóa nhân viên")

    _current_role = str(st.session_state.current_role).strip().lower()
    _all_staff = df_credentials.copy()
    if _current_role != 'admin' and 'Phân quyền' in _all_staff.columns:
        _all_staff = _all_staff[_all_staff['Phân quyền'].astype(str).str.strip().str.lower().isin(FRONTDESK_MANAGEABLE_ROLES)].copy()
    _manageable_names = sort_employee_names(_all_staff['Tên nhân viên'].dropna().astype(str).tolist()) if not _all_staff.empty else []

    # 1) Chỉnh sửa hồ sơ
    if has_feature_access('employee_edit'):
        st.markdown("#### ✏️ Chỉnh sửa hồ sơ")
        edit_usr = st.selectbox(
            "Chọn nhân viên cần sửa:", _manageable_names, index=None, placeholder="Chọn nhân viên",
            key='sb_edit_employee', filter_mode="contains"
        )
        if edit_usr:
            usr_data = df_credentials[df_credentials['Tên nhân viên'].apply(normalize_login_name) == normalize_login_name(edit_usr)].iloc[-1]
            edit_key = re.sub(r"[^a-zA-Z0-9_]+", "_", normalize_login_name(edit_usr)) or "employee"
            current_target_role = str(usr_data.get('Phân quyền', 'nhanvien')).strip().lower()
            allowed_edit_roles = ALL_ACCOUNT_ROLES if _current_role == 'admin' else ["nhanvien", "locker", "tapvu"]
            if current_target_role not in allowed_edit_roles and _current_role != 'admin':
                st.error("Lễ tân/Quản lý chỉ được chỉnh sửa tài khoản nhanvien, locker hoặc tapvu.")
            else:
                e_role = st.selectbox(
                    "Phân quyền", allowed_edit_roles,
                    index=allowed_edit_roles.index(current_target_role) if current_target_role in allowed_edit_roles else 0,
                    key=f"edit_role_{edit_key}"
                )
                e_pass = st.text_input("Mật khẩu", value=str(usr_data.get('Mật khẩu', '')), key=f"edit_password_{edit_key}")
                e_fn = st.text_input("Họ tên", value=str(usr_data.get('Họ và tên đầy đủ', '')), key=f"edit_fullname_{edit_key}")
                e_dob = st.text_input("Ngày sinh", value=str(usr_data.get('Ngày sinh', '')), key=f"edit_dob_{edit_key}")
                e_phone = st.text_input("SĐT", value=str(usr_data.get('Điện thoại', '')).replace("'", ""), key=f"edit_phone_{edit_key}")
                e_email = st.text_input("Email", value=str(usr_data.get('Email', '')), key=f"edit_email_{edit_key}")
                st.markdown("**📍 Địa chỉ**")
                e_address = vietnam_address_inputs(f"edit_address_{edit_key}", str(usr_data.get('Địa chỉ', '')))
                e_bank_account = st.text_input("Số tài khoản ngân hàng", value=str(usr_data.get('Số tài khoản ngân hàng', '')).replace("'", ""), key=f"edit_bank_account_{edit_key}")
                e_bank_name = bank_selectbox("Tên ngân hàng", key=f"edit_bank_name_{edit_key}", current_value=str(usr_data.get('Tên ngân hàng', '')))
                if st.button("💾 Cập nhật dữ liệu", use_container_width=True, key=f"edit_save_{edit_key}"):
                    ok, msg = update_user_profile(
                        edit_usr, e_pass, e_fn, e_dob, e_phone, e_email, e_address,
                        e_bank_account, e_bank_name, new_role=e_role
                    )
                    (st.success if ok else st.error)(msg)
                    if ok: st.rerun()
        st.markdown("---")

    # 2) Trạng thái làm việc
    if has_feature_access('employment_status'):
        st.markdown("#### 🏷️ Trạng thái làm việc của nhân viên")
        if 'Phân quyền' in df_credentials.columns:
            _status_roles = EMPLOYMENT_STATUS_MANAGEABLE_ROLES if _current_role == 'admin' else FRONTDESK_MANAGEABLE_ROLES
            nhanvien_df_status = df_credentials[
                df_credentials['Phân quyền'].astype(str).str.strip().str.lower().isin(_status_roles)
            ].copy()
        else:
            nhanvien_df_status = pd.DataFrame(columns=df_credentials.columns)
        status_emp_options = sort_employee_names(nhanvien_df_status['Tên nhân viên'].dropna().astype(str).tolist()) if not nhanvien_df_status.empty else []
        if not status_emp_options:
            st.info("Hiện không có tài khoản phù hợp để cập nhật trạng thái.")
        else:
            c_status_emp, c_status_value, c_status_save = st.columns([2.2, 1.6, 1.2])
            with c_status_emp:
                status_emp = st.selectbox(
                    "Chọn nhân viên", options=status_emp_options, index=None, placeholder="Chọn nhân viên",
                    filter_mode="contains", key="employment_status_employee"
                )
            current_status_map = load_employment_status_map()
            if status_emp:
                current_status = current_status_map.get(normalize_login_name(status_emp), EMPLOYMENT_STATUS_ACTIVE)
                with c_status_value:
                    status_value = st.selectbox(
                        "Trạng thái", options=EMPLOYMENT_STATUS_OPTIONS,
                        index=EMPLOYMENT_STATUS_OPTIONS.index(current_status) if current_status in EMPLOYMENT_STATUS_OPTIONS else 0,
                        key=f"employment_status_value_{normalize_login_name(status_emp)}"
                    )
                with c_status_save:
                    st.write(""); st.write("")
                    if st.button("💾 Lưu trạng thái", use_container_width=True, key="save_employment_status"):
                        ok, msg = set_employee_employment_status(status_emp, status_value, st.session_state.current_user)
                        (st.success if ok else st.error)(msg)
                        if ok: st.rerun()
            else:
                with c_status_value:
                    st.selectbox("Trạng thái", options=EMPLOYMENT_STATUS_OPTIONS, index=None, placeholder="Chọn trạng thái", disabled=True, key="employment_status_value_empty")
                with c_status_save:
                    st.write(""); st.write("")
                    st.button("💾 Lưu trạng thái", use_container_width=True, disabled=True, key="save_employment_status_empty")
        st.markdown("---")

    # 3) Xóa nhân viên
    if has_feature_access('employee_delete'):
        st.markdown("#### 🗑️ Xóa nhân viên")
        del_usr = st.selectbox(
            "Chọn nhân viên cần xóa:", _manageable_names, index=None, placeholder="Chọn nhân viên",
            filter_mode="contains", key="delete_employee_select"
        )
        confirm_del = st.checkbox("Tôi xác nhận xóa tài khoản đã chọn", key="confirm_delete_employee")
        if st.button("Xác nhận xóa", use_container_width=True, disabled=not bool(del_usr and confirm_del)):
            if del_usr:
                try:
                    client = get_gspread_client()
                    sheet_mk = client.open_by_key(SHEET_MAT_KHAU_ID).get_worksheet(0)
                    cells = sheet_mk.findall(del_usr, in_column=2)
                    if cells:
                        sheet_mk.delete_rows(cells[0].row)
                        renumber_credential_sheet_stt(sheet_mk)
                        try: load_credentials.clear()
                        except Exception: pass
                        sync_tichluy_roles_and_stt()
                        _clear_dynamic_data_caches()
                        st.success(f"Đã xóa nhân viên: {del_usr} · đã sắp xếp lại STT Sheet1/TichLuy.")
                        st.rerun()
                except Exception as e:
                    st.error(f"Lỗi xóa: {e}")

elif selected_page == "🔒 Khóa đăng nhập" and has_feature_access("account_lock"):
    st.markdown("### 🔒 Khóa / mở khóa đăng nhập")
    lockable_df = df_credentials[df_credentials['Tên nhân viên'].apply(normalize_login_name) != normalize_login_name(st.session_state.current_user)].copy()
    lockable_users = sort_employee_names(lockable_df['Tên nhân viên'].dropna().astype(str).tolist())
    selected_lock_users = st.multiselect(
        "Chọn một hoặc nhiều tài khoản:",
        options=lockable_users,
        default=[],
        filter_mode="contains",
        placeholder="Gõ để tìm tài khoản..."
    )
    c_lock1, c_lock2, c_lock3, c_lock4 = st.columns(4)
    with c_lock1:
        if st.button("🔒 Khóa tài khoản đã chọn", use_container_width=True):
            if not selected_lock_users:
                st.warning("Vui lòng chọn ít nhất 1 tài khoản.")
            else:
                ok, msg = set_accounts_login_lock(selected_lock_users, True)
                (st.success if ok else st.error)(msg)
                if ok: st.rerun()
    with c_lock2:
        if st.button("🔓 Mở khóa tài khoản đã chọn", use_container_width=True):
            if not selected_lock_users:
                st.warning("Vui lòng chọn ít nhất 1 tài khoản.")
            else:
                ok, msg = set_accounts_login_lock(selected_lock_users, False)
                (st.success if ok else st.error)(msg)
                if ok: st.rerun()
    with c_lock3:
        if st.button("⛔ Khóa TOÀN BỘ", use_container_width=True):
            ok, msg = set_accounts_login_lock(lockable_users, True)
            (st.success if ok else st.error)(msg)
            if ok: st.rerun()
    with c_lock4:
        if st.button("✅ Mở TOÀN BỘ", use_container_width=True):
            ok, msg = set_accounts_login_lock(lockable_users, False)
            (st.success if ok else st.error)(msg)
            if ok: st.rerun()

    locked_now = lockable_df[lockable_df['Khóa đăng nhập'].apply(is_locked_value)]
    st.caption(f"Đang khóa: {len(locked_now)} / {len(lockable_df)} tài khoản. Tài khoản Admin đang sử dụng được loại khỏi danh sách để tránh tự khóa chính mình.")
    if not locked_now.empty:
        st.dataframe(locked_now[['Tên nhân viên', 'Phân quyền', 'Khóa đăng nhập']], width='stretch', height='content', hide_index=True)
elif selected_page == "🔐 Khóa quyền đăng ký" and has_feature_access("registration_lock"):
    st.subheader("🔐 Khóa quyền đăng ký theo từng vai trò")
    st.caption(
        "Khóa độc lập từng vai trò. Trạng thái được lưu trên Google Sheet nên vẫn giữ nguyên sau khi "
        "deploy/restart. Tài khoản Admin luôn được mở quyền và không thể bị khóa."
    )

    _role_locks = load_registration_role_locks()
    for _role in REGISTRATION_LOCK_ROLES:
        _label = REGISTRATION_LOCK_LABELS.get(_role, _role)
        _locked = bool(_role_locks.get(_role, False))
        _c1, _c2, _c3 = st.columns([3, 2, 2])
        with _c1:
            if _locked:
                st.warning(f"🔴 {_label}: ĐANG KHÓA")
            else:
                st.success(f"🟢 {_label}: ĐANG MỞ")
        with _c2:
            if st.button(
                f"🔒 Khóa {_label}",
                key=f"reglock_lock_{_role}",
                use_container_width=True,
                disabled=_locked,
            ):
                ok, msg = set_registration_role_lock(
                    _role, True, st.session_state.get("current_user", "Admin")
                )
                (st.success if ok else st.error)(msg)
                if ok:
                    st.rerun()
        with _c3:
            if st.button(
                f"🔓 Mở {_label}",
                key=f"reglock_unlock_{_role}",
                use_container_width=True,
                disabled=not _locked,
            ):
                ok, msg = set_registration_role_lock(
                    _role, False, st.session_state.get("current_user", "Admin")
                )
                (st.success if ok else st.error)(msg)
                if ok:
                    st.rerun()

    st.info("🛡️ Admin: luôn MỞ quyền đăng ký / nhập / sửa / xóa lịch nghỉ và miễn các giới hạn nghiệp vụ.")

elif selected_page == "⏸️ Auto Update phạt" and st.session_state.current_role == "admin":
    st.subheader("⏸️ Điều khiển Auto Update phạt")
    cfg = load_auto_penalty_config()
    paused = bool(cfg.get('paused'))
    if paused:
        st.warning(
            f"🔴 Auto Update phạt đang TẠM DỪNG. Không tự ghi Đi trễ TimeSoft hoặc Ra ngoài vào muộn từ Bảng tour. "
            f"Ngưỡng khi hoạt động: từ {AUTO_PENALTY_MINUTES} phút."
        )
    else:
        st.success(
            f"🟢 Auto Update phạt đang HOẠT ĐỘNG. Lịch tự động: 15:00 và 20:00 hằng ngày; ngưỡng từ {AUTO_PENALTY_MINUTES} phút."
        )
    st.caption(
        f"Cập nhật gần nhất: {cfg.get('updated_date','')} {cfg.get('updated_time','')} · "
        f"{cfg.get('updated_by','') or 'Hệ thống'}"
    )
    if cfg.get('error'):
        st.warning(f"Không đọc được cấu hình Auto Update đầy đủ: {cfg.get('error')}")

    c_auto1, c_auto2 = st.columns(2)
    with c_auto1:
        if paused:
            if st.button("▶️ Mở lại Auto Update phạt", use_container_width=True, type="primary", key="resume_auto_penalty_v84"):
                ok, msg = set_auto_penalty_paused(False, st.session_state.current_user)
                (st.success if ok else st.error)(msg)
                if ok:
                    st.rerun()
        else:
            if st.button("⏸️ Tạm dừng Auto Update phạt", use_container_width=True, type="primary", key="pause_auto_penalty_v84"):
                ok, msg = set_auto_penalty_paused(True, st.session_state.current_user)
                (st.success if ok else st.error)(msg)
                if ok:
                    st.rerun()
    with c_auto2:
        run_now = st.button(
            "▶️ Chạy Auto Update ngay", use_container_width=True,
            disabled=paused, key="run_auto_penalty_now_v84",
            help="Đọc Bảng tour hiện tại và dữ liệu TimeSoft gần nhất đang có trong phiên/snapshot."
        )

    if run_now:
        load_bang_tour_input.clear()
        _tour_now, _tour_err = load_bang_tour_input()
        _checkin_now = None
        _direct = st.session_state.get('timesoft_direct_result_v81') or {}
        if isinstance(_direct.get('employee_checkin_df'), pd.DataFrame):
            _checkin_now = _direct.get('employee_checkin_df')
        elif vpg is not None and vpg.is_enabled():
            _snap = _timesoft_read_background_snapshot(get_vn_today())
            if isinstance(_snap.get('employee_checkin'), pd.DataFrame):
                _checkin_now = _snap.get('employee_checkin')
        with st.spinner("Đang kiểm tra Bảng tour và TimeSoft theo ngưỡng 5 phút..."):
            _auto_res = run_auto_penalty_now(
                tour_df=_tour_now, checkin_df=_checkin_now, actor=f"AUTO UPDATE - {st.session_state.current_user}"
            )
        rt, rs, rc = _auto_res['tour'], _auto_res['timesoft'], _auto_res['ca1']
        m1, m2, m3, m4, m5, m6 = st.columns(6)
        m1.metric("Tour đủ ĐK", rt.get('eligible',0))
        m2.metric("Tour đã thêm", rt.get('added',0))
        m3.metric("TimeSoft đủ ĐK", rs.get('eligible',0))
        m4.metric("TimeSoft đã thêm", rs.get('added',0))
        m5.metric("Ca 1 đủ ĐK", rc.get('eligible',0))
        m6.metric("Ca 1 đã thêm", rc.get('added',0))
        all_msgs = (rt.get('messages') or []) + (rs.get('messages') or []) + (rc.get('messages') or [])
        if all_msgs:
            st.caption(" | ".join(all_msgs[:10]))

    st.markdown("#### Quy tắc V84.7")
    st.write(
        "• Auto Update tự động chỉ chạy theo Cloud Scheduler lúc **15:00 và 20:00**; mở Bảng tour hoặc lấy TimeSoft thủ công không tự ghi phạt.  \n"
        "• Ra ngoài vào muộn: chỉ Auto Update khi cột **Vào trễ >= 5 phút**. "
        "Tên như **Cẩm Nhung *** được đối chiếu như **Cẩm Nhung**.  \n"
        "• TimeSoft: check-in được so trực tiếp với giờ bắt đầu ca. **Hỗ trợ Ca 1 2 tiếng = 120 phút; Ca 1 sau 0:0H 3 tiếng = 180 phút; Ca 2 sau 0:0H 1 tiếng = 60 phút**. Chỉ Auto phạt khi vượt mức Hỗ trợ; nếu không có Hỗ trợ thì ngưỡng là **>= 5 phút**.  \n"
        "• **KHÔNG dọn vệ sinh ca 1**: chỉ áp dụng cho role `nhanvien` đang làm **Ca 1 trong tuần hiện tại**, "
        "không có **Hỗ trợ Ca 1 đi trễ 2 tiếng / Hỗ trợ Ca 1 đi trễ 3 tiếng / Hỗ trợ Ca 2 đi trễ 1 tiếng**, "
        "và hôm đó có **Đi trễ <=30 / <=60 / >60 đến <=120 phút** theo đúng loại nghỉ đã cấu hình.  \n"
        "• Tiền phạt và Số ngày tính lấy trực tiếp từ sheet **LoaiNghi**; dữ liệu phạt ghi vào Sheet1 A:J và không tạo trùng cùng Ngày + Nhân viên + Lý do.  \n"
        "• **Email bắt buộc cho mọi Auto Update có Phạt vi phạm > 0**: gửi từ `veraspabienhoa@gmail.com` đến nhân viên bị phạt; CC `veraspabienhoa@gmail.com + quanly + letan`. Email lỗi sẽ được Job kế tiếp tự thử gửi lại."
    )

elif selected_page == "📦 Snapshot nền hôm nay" and st.session_state.current_role == "admin":
    st.subheader("📦 Snapshot nền hôm nay")
    st.caption(
        "Trang riêng dành cho Admin. Dữ liệu được đọc từ snapshot TimeSoft nền trong PostgreSQL; "
        "không gọi TimeSoft trực tiếp khi chỉ mở trang này."
    )
    if st.button("🔄 Làm mới Snapshot", use_container_width=True, key="refresh_snapshot_today_admin"):
        try:
            if vpg is not None:
                # Dataset đọc theo allow_stale; rerun để lấy lại trạng thái/dữ liệu mới nhất.
                pass
        except Exception:
            pass
        st.rerun()
    render_timesoft_background_snapshot_today(show_status=True)


elif selected_page == "🔄 Đồng bộ dữ liệu" and has_feature_access("sync"):
    st.subheader("🔄 Đồng bộ dữ liệu")
    st.info("Các công cụ đồng bộ chỉ dành cho tài khoản Admin.")

    tab_timesoft, tab_gsheet = st.tabs(["🌐 TimeSoft", "📄 Excel / Google Sheets"])

    with tab_timesoft:
        st.markdown("### 🌐 TimeSoft · API trực tiếp + Đồng bộ nền Cloud Run")
        st.caption(
            "V83 giữ chế độ lấy dữ liệu trực tiếp và hiển thị snapshot nền. "
            f"Cloud Scheduler gọi Cloud Run Job mỗi {TIMESOFT_BACKGROUND_INTERVAL_MINUTES} phút; "
            "snapshot được lưu PostgreSQL để mọi instance cùng dùng dữ liệu mới nhất."
        )

        if timesoft_is_configured():
            st.success(f"✅ Đã cấu hình TimeSoft: {TIMESOFT_BASE_URL} · tài khoản đã được nạp từ Secrets.")
        else:
            st.error(
                "❌ Chưa đủ cấu hình TimeSoft. Hãy cấu hình Secret Manager trên Cloud Run hoặc [TIMESOFT] trong Streamlit Secrets."
            )

        st.markdown(
            f"#### ☁️ Đồng bộ TimeSoft tự động 24/7 · mỗi {TIMESOFT_BACKGROUND_INTERVAL_MINUTES} phút"
        )
        if vpg is not None and vpg.is_enabled():
            bg_status = _timesoft_background_status_row()
            if bg_status:
                bg_ok = str(bg_status.get("status", "")).lower() == "success"
                last_sync = str(bg_status.get("synced_at_vn", "") or bg_status.get("synced_at", ""))
                (st.success if bg_ok else st.warning)(
                    f"{'✅' if bg_ok else '⚠️'} Cloud Run Job: {bg_status.get('status', 'unknown')} · "
                    f"lần chạy gần nhất {last_sync or 'chưa rõ'}"
                )
                b1, b2, b3, b4 = st.columns(4)
                b1.metric("Chu kỳ", f"{TIMESOFT_BACKGROUND_INTERVAL_MINUTES} phút")
                b2.metric("Doanh thu · dòng", int(float(bg_status.get("invoice_rows", 0) or 0)))
                b3.metric("Chấm công · dòng", int(float(bg_status.get("checkin_rows", 0) or 0)))
                b4.metric("Thời gian chạy", f"{float(bg_status.get('duration_seconds', 0) or 0):.1f}s")
                # V85.2: vẫn giữ Snapshot hiện tại trong trang Đồng bộ dữ liệu,
                # nhưng tuyệt đối chỉ hiển thị với tài khoản Admin.
                if st.session_state.current_role == "admin":
                    with st.expander("📦 Xem snapshot nền hôm nay", expanded=False):
                        render_timesoft_background_snapshot_today(show_status=False)
            else:
                st.info("PostgreSQL đã bật nhưng chưa có snapshot TimeSoft nền. Sau khi Cloud Scheduler gọi Cloud Run Job thành công lần đầu, trạng thái sẽ xuất hiện tại đây.")
        else:
            st.caption("Máy local không bật PostgreSQL nên chỉ hiển thị chế độ trực tiếp. Trên Cloud Run, snapshot nền sẽ được đọc từ PostgreSQL.")

        st.markdown("#### 📥 Lấy dữ liệu TimeSoft thủ công")
        ts_today = get_vn_today()
        ts_default_start = st.session_state.get("timesoft_start_date_v81", ts_today)
        ts_default_end = st.session_state.get("timesoft_end_date_v81", ts_today)
        c_date1, c_date2 = st.columns(2)
        with c_date1:
            ts_start_date = st.date_input(
                "Từ ngày",
                value=ts_default_start,
                key="timesoft_start_date_v81",
                format="DD/MM/YYYY",
            )
        with c_date2:
            ts_end_date = st.date_input(
                "Đến ngày",
                value=ts_default_end,
                key="timesoft_end_date_v81",
                format="DD/MM/YYYY",
            )

        if ts_start_date > ts_end_date:
            st.warning("Từ ngày đang lớn hơn Đến ngày. Hệ thống sẽ tự đảo khoảng ngày khi lấy dữ liệu.")

        c_sync1, c_sync2, c_sync3 = st.columns([2.2, 2.2, 1.4])
        with c_sync1:
            ts_sync_now = st.button(
                "📥 Lấy dữ liệu TimeSoft",
                use_container_width=True,
                disabled=not timesoft_is_configured(),
                key="timesoft_direct_sync_v81",
                type="primary",
            )
        with c_sync2:
            ts_force_sync = st.button(
                "🔐 Đăng nhập lại & lấy dữ liệu",
                use_container_width=True,
                disabled=not timesoft_is_configured(),
                key="timesoft_force_sync_v81",
            )
        with c_sync3:
            if st.button(
                "🧹 Xóa dữ liệu",
                use_container_width=True,
                disabled=not bool(st.session_state.get("timesoft_direct_result_v81")),
                key="timesoft_clear_direct_v81",
            ):
                st.session_state.pop("timesoft_direct_result_v81", None)
                st.session_state.pop("timesoft_direct_msg_v81", None)
                st.rerun()

        if ts_sync_now or ts_force_sync:
            with st.spinner("Đang đăng nhập/kiểm tra session và lấy dữ liệu trực tiếp từ 2 API TimeSoft..."):
                ok_direct, msg_direct, result_direct = timesoft_direct_sync(
                    ts_start_date,
                    ts_end_date,
                    force_login=bool(ts_force_sync),
                )
                st.session_state["timesoft_direct_result_v81"] = result_direct
                st.session_state["timesoft_direct_msg_v81"] = (ok_direct, msg_direct)
                # V84.7: lấy TimeSoft thủ công CHỈ tải/xem dữ liệu.
                # Auto Update tự động chỉ do Cloud Scheduler chạy lúc 15:00 và 20:00.
                st.session_state.pop("timesoft_auto_penalty_result_v84", None)

        direct_msg = st.session_state.get("timesoft_direct_msg_v81")
        if direct_msg:
            ok_direct, msg_direct = direct_msg
            (st.success if ok_direct else st.error)(msg_direct)

        direct_result = st.session_state.get("timesoft_direct_result_v81") or {}
        if direct_result:
            synced_at = direct_result.get("synced_at")
            if isinstance(synced_at, datetime):
                st.caption(f"Lần đồng bộ: {synced_at.strftime('%d/%m/%Y %H:%M:%S')} · {direct_result.get('session_message', '')}")

            inv_meta = direct_result.get("summary_invoice_meta") or {}
            inv_df = direct_result.get("summary_invoice_df")
            chk_meta = direct_result.get("employee_checkin_meta") or {}
            chk_df = direct_result.get("employee_checkin_df")

            st.markdown("##### 💰 Báo cáo tổng hợp doanh thu")
            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Số dòng chi tiết", len(inv_df) if isinstance(inv_df, pd.DataFrame) else 0)
            m2.metric("Tổng tiền", f"{float(inv_meta.get('TotalMoney') or 0):,.0f} đ".replace(",", "."))
            m3.metric("Tổng giảm giá", f"{float(inv_meta.get('TotalDiscount') or 0):,.0f} đ".replace(",", "."))
            m4.metric("Doanh thu thực", f"{float(inv_meta.get('TotalActualRevenu') or 0):,.0f} đ".replace(",", "."))
            if isinstance(inv_df, pd.DataFrame) and not inv_df.empty:
                st.dataframe(inv_df, width="stretch", hide_index=True, height=360)
            else:
                st.info("Khoảng ngày đã chọn không có dòng doanh thu chi tiết hoặc TimeSoft trả Data rỗng.")

            st.markdown("##### 🕒 Báo cáo chấm công nhân viên")
            chk_display = _timesoft_checkin_display_df(chk_df)
            cc1, cc2 = st.columns(2)
            cc1.metric("Tổng bản ghi", int(chk_meta.get("Total") or (len(chk_df) if isinstance(chk_df, pd.DataFrame) else 0)))
            cc2.metric("Đã tải", len(chk_df) if isinstance(chk_df, pd.DataFrame) else 0)
            if isinstance(chk_display, pd.DataFrame) and not chk_display.empty:
                st.dataframe(chk_display, width="stretch", hide_index=True, height=420)
            else:
                st.info("Khoảng ngày đã chọn chưa có dữ liệu chấm công.")

            try:
                export_bytes = _timesoft_export_workbook(direct_result)
                start_txt = direct_result.get("start_date").strftime("%Y%m%d") if isinstance(direct_result.get("start_date"), date) else "from"
                end_txt = direct_result.get("end_date").strftime("%Y%m%d") if isinstance(direct_result.get("end_date"), date) else "to"
                st.download_button(
                    "📥 Tải Excel dữ liệu TimeSoft",
                    data=export_bytes,
                    file_name=f"TimeSoft_{start_txt}_{end_txt}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="timesoft_download_excel_v81",
                )
            except Exception as e:
                st.warning(f"Không tạo được file Excel TimeSoft: {e}")

        st.markdown("---")
        with st.expander("🧪 Công cụ kỹ thuật · Phát hiện lại API", expanded=False):
            st.caption(
                "Chỉ dùng khi TimeSoft thay đổi hệ thống. V81 đã biết 2 endpoint hiện tại nên bình thường không cần chạy discovery."
            )
            c_ts1, c_ts2 = st.columns([3, 2])
            with c_ts1:
                discover_now = st.button(
                    "🔍 Phát hiện lại API TimeSoft",
                    use_container_width=True,
                    disabled=not timesoft_is_configured(),
                    key="timesoft_auto_discover_v81",
                )
            with c_ts2:
                if st.button(
                    "🧹 Xóa kết quả phát hiện",
                    use_container_width=True,
                    disabled=not bool(st.session_state.get("timesoft_api_discovery_v81")),
                    key="timesoft_clear_discovery_v81",
                ):
                    st.session_state.pop("timesoft_api_discovery_v81", None)
                    st.session_state.pop("timesoft_api_discovery_msg_v81", None)
                    st.rerun()

            if discover_now:
                with st.spinner("Đang đăng nhập TimeSoft và bắt request thật của 2 báo cáo..."):
                    ok_ts, msg_ts, discovery_ts = timesoft_auto_discover_apis()
                    st.session_state["timesoft_api_discovery_v81"] = discovery_ts
                    st.session_state["timesoft_api_discovery_msg_v81"] = (ok_ts, msg_ts)

            discovery_ts = st.session_state.get("timesoft_api_discovery_v81", {})
            discovery_msg = st.session_state.get("timesoft_api_discovery_msg_v81")
            if discovery_msg:
                ok_ts, msg_ts = discovery_msg
                (st.success if ok_ts else st.warning)(msg_ts)

            if discovery_ts:
                best_df = _timesoft_discovery_rows(discovery_ts)
                st.dataframe(best_df, width="stretch", hide_index=True, height="content")
                safe_json = json.dumps(
                    _timesoft_sanitized_discovery(discovery_ts),
                    ensure_ascii=False,
                    indent=2,
                ).encode("utf-8")
                st.download_button(
                    "📥 Tải cấu hình API kỹ thuật (đã loại response/PII/token)",
                    data=safe_json,
                    file_name="timesoft_api_discovery_safe.json",
                    mime="application/json",
                    use_container_width=True,
                    key="download_timesoft_discovery_v81",
                )

        st.caption(
            "API V81: ReportSummaryInvoice/SearchFullText và ReportEmployeeCheckin/SearchElastic. "
            "Password, Cookie và Authorization không được hiển thị trên giao diện."
        )

    with tab_gsheet:
        st.markdown("### ⬆️ Đồng bộ Excel → Google Sheet1 A:J")
        st.caption(
            f"Đích: Google Sheet {SHEET_DU_PHONG_ID} · Sheet1. "
            "Tên nhân viên khi đối chiếu sẽ bỏ dấu * ở cuối: Cẩm Nhung * = Cẩm Nhung."
        )
        v1, v2 = st.columns(2)
        with v1:
            st.markdown("#### Phiên bản 1 · Ghi đè")
            st.warning("Sẽ xóa dữ liệu hiện có trong A2:J rồi paste toàn bộ Excel bắt đầu từ A2. Header hàng 1 được giữ nguyên.")
            confirm_v1 = st.checkbox("Tôi xác nhận cho phép ghi đè A2:J", key="confirm_excel_overwrite_v84")
            if st.button(
                "⚠️ V1 · Ghi đè Excel → Google", use_container_width=True,
                disabled=not confirm_v1, key="sync_excel_google_overwrite_v84"
            ):
                with st.spinner("Đang ghi đè A2:J từ Excel..."):
                    res, msg = admin_sync_excel_to_gsheet_overwrite()
                    (st.success if res else st.error)(msg)
        with v2:
            st.markdown("#### Phiên bản 2 · Không ghi đè")
            st.success("Giữ nguyên dữ liệu hiện có. Chỉ thêm dòng chưa tồn tại vào đúng last row và luôn ghi trong cột A:J.")
            if st.button(
                "✅ V2 · Thêm mới Excel → Google", use_container_width=True,
                key="sync_excel_google_append_v84"
            ):
                with st.spinner("Đang tìm last row A:J và thêm dữ liệu mới..."):
                    res, msg = admin_sync_excel_to_gsheet_append()
                    (st.success if res else st.error)(msg)

        st.markdown("---")
        if st.button("⬇️ Tạo Excel mới từ Google Sheets", help="Gộp dữ liệu mới từ Sheet vào file Excel gốc", use_container_width=True):
            with st.spinner("Đang tạo file..."):
                df_merged, has_new = admin_sync_gsheet_to_excel(df_backup, df_lich)
                if has_new:
                    st.download_button("📥 Tải file Excel cập nhật", data=to_excel(df_merged), file_name="LichNghi_CapNhat.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                else:
                    st.info("Excel gốc đã có đủ dữ liệu, không có dòng mới.")

elif selected_page == "💰 Bảng lương" and has_page_access("💰 Bảng lương"):
    st.subheader("💰 Bảng lương nhân viên")
    st.caption("Tiền Lương được tính theo đúng quy tắc: cột F bắt đầu bằng 'Tip' → cộng cột G theo tên nhân viên ở cột I.")

    tab_calc, tab_history = st.tabs(["🧮 Tính lương nhân viên", "🗂 Lịch sử bảng lương đã lưu"])
    with tab_calc:
        if not has_feature_access("payroll"):
            st.info("🔒 Bạn chưa được cấp quyền Tính lương nhân viên.")
        else:
            # V56: Admin kiểm tra toàn bộ nghĩa vụ Vi phạm còn mở, tách riêng 2 nguồn:
            # 1) Nợ do Thực nhận âm tự phát sinh khi lưu lương.
            # 2) Khoản Admin chủ động Tạm hoãn Vi phạm sang kỳ kế tiếp.
            if st.session_state.current_role == "admin":
                _debt_btn_col, _debt_hide_col = st.columns([3, 2])
                with _debt_btn_col:
                    if st.button(
                        "💳 Kiểm tra nghĩa vụ Vi phạm còn mở",
                        use_container_width=True,
                        key="admin_check_open_negative_payroll_debts"
                    ):
                        _clear_violation_debt_cache()
                        st.session_state.show_open_negative_payroll_debts = True
                with _debt_hide_col:
                    if st.button(
                        "✖️ Ẩn danh sách nghĩa vụ",
                        use_container_width=True,
                        key="admin_hide_open_negative_payroll_debts",
                        disabled=not bool(st.session_state.get('show_open_negative_payroll_debts', False))
                    ):
                        st.session_state.show_open_negative_payroll_debts = False
                        st.rerun()

                if st.session_state.get('show_open_negative_payroll_debts', False):
                    _negative_debt_summary, _negative_debt_detail = get_open_negative_payroll_debts()
                    _deferred_debt_summary, _deferred_debt_detail = get_open_admin_deferred_violation_debts()

                    if _negative_debt_summary.empty and _deferred_debt_summary.empty:
                        st.success("✅ Hiện không có Nghĩa vụ Vi phạm nào còn Chưa hoàn thành.")
                    else:
                        _negative_debt_total = int(round(_negative_debt_summary['Tổng còn nợ'].apply(_money_to_float).sum())) if not _negative_debt_summary.empty else 0
                        _deferred_debt_total = int(round(_deferred_debt_summary['Tổng tạm hoãn'].apply(_money_to_float).sum())) if not _deferred_debt_summary.empty else 0
                        _m1, _m2, _m3, _m4 = st.columns(4)
                        _m1.metric("NV nợ Thực nhận âm", len(_negative_debt_summary))
                        _m2.metric("Tổng nợ Thực nhận âm", f"{_negative_debt_total:,.0f} đ".replace(',', '.'))
                        _m3.metric("NV Admin tạm hoãn", len(_deferred_debt_summary))
                        _m4.metric("Tổng Admin tạm hoãn", f"{_deferred_debt_total:,.0f} đ".replace(',', '.'))
                        st.caption("Hiển thị riêng hai nhóm nghĩa vụ còn mở: Thực nhận âm tự chuyển kỳ và Vi phạm do Admin chủ động tạm hoãn.")

                        st.markdown("#### 🔴 Nợ do Thực nhận âm")
                        if _negative_debt_summary.empty:
                            st.info("Không có nhân viên nào còn nợ do Thực nhận âm.")
                        else:
                            _summary_show = _negative_debt_summary.copy()
                            _summary_show['Tổng còn nợ'] = _summary_show['Tổng còn nợ'].apply(lambda x: f"{_money_to_float(x):,.0f}".replace(',', '.'))
                            st.dataframe(
                                _summary_show,
                                hide_index=True,
                                width="stretch",
                                height="content"
                            )
                            with st.expander("🔎 Xem chi tiết từng kỳ nợ Thực nhận âm", expanded=False):
                                _detail_show = _negative_debt_detail.copy()
                                if 'Số tiền' in _detail_show.columns:
                                    _detail_show['Số tiền'] = _detail_show['Số tiền'].apply(lambda x: f"{_money_to_float(x):,.0f}".replace(',', '.'))
                                st.dataframe(
                                    _detail_show,
                                    hide_index=True,
                                    width="stretch",
                                    height="content"
                                )

                        st.markdown("#### ⏭️ Nghĩa vụ Vi phạm Admin chủ động tạm hoãn")
                        if _deferred_debt_summary.empty:
                            st.info("Không có khoản Vi phạm nào do Admin chủ động tạm hoãn đang mở.")
                        else:
                            _deferred_summary_show = _deferred_debt_summary.copy()
                            _deferred_summary_show['Tổng tạm hoãn'] = _deferred_summary_show['Tổng tạm hoãn'].apply(lambda x: f"{_money_to_float(x):,.0f}".replace(',', '.'))
                            st.dataframe(
                                _deferred_summary_show,
                                hide_index=True,
                                width="stretch",
                                height="content"
                            )
                            with st.expander("🔎 Xem chi tiết từng kỳ Admin đã tạm hoãn", expanded=False):
                                _deferred_detail_show = _deferred_debt_detail.copy()
                                if 'Số tiền' in _deferred_detail_show.columns:
                                    _deferred_detail_show['Số tiền'] = _deferred_detail_show['Số tiền'].apply(lambda x: f"{_money_to_float(x):,.0f}".replace(',', '.'))
                                st.dataframe(
                                    _deferred_detail_show,
                                    hide_index=True,
                                    width="stretch",
                                    height="content"
                                )

            default_living_db, default_locker_db = get_payroll_default_amounts()
            leader_allowance_db = get_leader_responsibility_allowance()
            payroll_default_living = float(default_living_db)
            payroll_default_locker = float(default_locker_db)
            payroll_leader_allowance = float(leader_allowance_db)

            # Chỉ Admin được thay đổi cấu hình tiền mặc định / mức riêng / tiền trách nhiệm Leader.
            if st.session_state.current_role == 'admin':
                with st.expander("⚙️ Mức khấu trừ mặc định & tiền trách nhiệm Leader", expanded=False):
                    cfg1, cfg2, cfg3, cfg4 = st.columns([3, 3, 3, 2])
                    with cfg1:
                        payroll_default_living = st.number_input(
                            "Chi phí sinh hoạt / nhân viên", min_value=0.0, step=10000.0, format="%.0f",
                            value=float(default_living_db), key="payroll_default_living"
                        )
                    with cfg2:
                        payroll_default_locker = st.number_input(
                            "Hỗ trợ Locker / nhân viên", min_value=0.0, step=10000.0, format="%.0f",
                            value=float(default_locker_db), key="payroll_default_locker"
                        )
                    with cfg3:
                        payroll_leader_allowance = st.number_input(
                            "Tiền trách nhiệm Leader / Kỳ 2", min_value=0.0, step=50000.0, format="%.0f",
                            value=float(leader_allowance_db), key="payroll_leader_allowance",
                            help="Khoản này tự động cộng vào cột Hỗ Trợ Hoàn Lại của tài khoản Leader khi tính lương."
                        )
                    with cfg4:
                        st.write("")
                        if st.button("💾 Lưu cấu hình", use_container_width=True, key="save_payroll_defaults"):
                            ok1, msg1 = set_payroll_default_amounts(payroll_default_living, payroll_default_locker)
                            ok2, msg2 = set_leader_responsibility_allowance(payroll_leader_allowance)
                            if ok1 and ok2:
                                st.success(f"{msg1} {msg2}")
                            else:
                                st.error(" | ".join([m for o, m in [(ok1,msg1),(ok2,msg2)] if not o]))
                    st.caption(
                        "Nhân viên/Leader dùng Phí Sinh Hoạt và Locker theo mức mặc định hoặc mức riêng. "
                        "Leader chỉ được cộng Tiền trách nhiệm vào Hỗ Trợ Hoàn Lại ở Kỳ 2 (ngày 16 đến cuối tháng)."
                    )

                    st.markdown("#### 👥 Mức riêng theo Nhân viên / Leader")
                    payroll_emp_choices_df = df_credentials.copy() if isinstance(df_credentials, pd.DataFrame) else pd.DataFrame()
                    if not payroll_emp_choices_df.empty and 'Tên nhân viên' in payroll_emp_choices_df.columns:
                        payroll_emp_choices_df = payroll_emp_choices_df[payroll_emp_choices_df['Tên nhân viên'].astype(str).str.strip() != ''].copy()
                        payroll_emp_choices_df = payroll_emp_choices_df[~payroll_emp_choices_df['Tên nhân viên'].astype(str).apply(normalize_login_name).isin({
                            'ten nhan vien', 'ten he thong', 'username', 'user name'
                        })]
                        if 'Phân quyền' in payroll_emp_choices_df.columns:
                            _pay_roles = payroll_emp_choices_df['Phân quyền'].astype(str).str.strip().str.lower()
                            payroll_emp_choices_df = payroll_emp_choices_df[_pay_roles.isin(PAYROLL_ELIGIBLE_ROLES)].copy()
                        payroll_employee_options = sort_employee_names(
                            payroll_emp_choices_df['Tên nhân viên'].astype(str).str.strip().tolist()
                        )
                    else:
                        payroll_employee_options = []

                    selected_payroll_override_emps = st.multiselect(
                        "Chọn 1 hoặc nhiều nhân viên cần đặt mức riêng:",
                        options=payroll_employee_options,
                        key="payroll_override_employees",
                        filter_mode="contains",
                        help="Các tài khoản được chọn sẽ dùng mức riêng thay cho mức mặc định chung khi tạo bảng lương mới."
                    )

                    existing_payroll_overrides = get_payroll_employee_overrides()
                    _selected_keys = [normalize_login_name(x) for x in selected_payroll_override_emps]
                    _living_values = [existing_payroll_overrides[k]['living'] for k in _selected_keys if k in existing_payroll_overrides]
                    _locker_values = [existing_payroll_overrides[k]['locker'] for k in _selected_keys if k in existing_payroll_overrides]
                    _living_initial = _living_values[0] if _living_values and len(set(_living_values)) == 1 else float(payroll_default_living)
                    _locker_initial = _locker_values[0] if _locker_values and len(set(_locker_values)) == 1 else float(payroll_default_locker)
                    _override_sig = hashlib.md5("|".join(sorted(_selected_keys)).encode('utf-8')).hexdigest()[:10] if _selected_keys else "none"

                    ov1, ov2, ov3, ov4 = st.columns([3, 3, 2, 2])
                    with ov1:
                        payroll_override_living = st.number_input(
                            "Chi phí sinh hoạt riêng / nhân viên", min_value=0.0, step=10000.0, format="%.0f",
                            value=float(_living_initial), key=f"payroll_override_living_{_override_sig}",
                            disabled=not bool(selected_payroll_override_emps)
                        )
                    with ov2:
                        payroll_override_locker = st.number_input(
                            "Hỗ trợ Locker riêng / nhân viên", min_value=0.0, step=10000.0, format="%.0f",
                            value=float(_locker_initial), key=f"payroll_override_locker_{_override_sig}",
                            disabled=not bool(selected_payroll_override_emps)
                        )
                    with ov3:
                        st.write("")
                        if st.button(
                            "💾 Áp dụng mức riêng", use_container_width=True, key="save_payroll_employee_overrides",
                            disabled=not bool(selected_payroll_override_emps)
                        ):
                            ok, msg = set_payroll_employee_overrides(
                                selected_payroll_override_emps, payroll_override_living, payroll_override_locker
                            )
                            if ok:
                                _apply_payroll_override_to_current_session(
                                    selected_payroll_override_emps, payroll_override_living, payroll_override_locker
                                )
                                st.success(msg)
                            else:
                                st.error(msg)
                    with ov4:
                        st.write("")
                        if st.button(
                            "♻️ Dùng lại mặc định", use_container_width=True, key="clear_payroll_employee_overrides",
                            disabled=not bool(selected_payroll_override_emps)
                        ):
                            ok, msg = clear_payroll_employee_overrides(selected_payroll_override_emps)
                            if ok:
                                _apply_payroll_override_to_current_session(
                                    selected_payroll_override_emps, payroll_default_living, payroll_default_locker
                                )
                                st.success(msg)
                            else:
                                st.error(msg)

                    if existing_payroll_overrides:
                        _override_rows = []
                        for _k, _v in existing_payroll_overrides.items():
                            _override_rows.append({
                                "Tên Hệ thống": _v.get("name", _k),
                                "Phí Sinh Hoạt riêng": int(round(_money_to_float(_v.get("living", 0)))),
                                "Hỗ trợ Locker riêng": int(round(_money_to_float(_v.get("locker", 0)))),
                            })
                        _override_df = pd.DataFrame(_override_rows)
                        if not _override_df.empty:
                            _override_df['__sort'] = _override_df['Tên Hệ thống'].apply(normalize_login_name)
                            _override_df = _override_df.sort_values('__sort').drop(columns='__sort')
                        with st.expander(f"📋 Danh sách mức riêng đang lưu ({len(_override_df)} nhân viên)", expanded=False):
                            st.dataframe(
                                _override_df, width="stretch", height="content", hide_index=True,
                                column_config={
                                    "Tên Hệ thống": st.column_config.TextColumn("Tên Hệ thống"),
                                    "Phí Sinh Hoạt riêng": st.column_config.NumberColumn("Phí Sinh Hoạt riêng", format="%,d"),
                                    "Hỗ trợ Locker riêng": st.column_config.NumberColumn("Hỗ trợ Locker riêng", format="%,d"),
                                }
                            )

            c_period, c_source = st.columns(2)
            with c_period:
                preset = st.selectbox(
                    "Chọn kỳ tính lương:",
                    ["Kỳ 1 - Tháng này", "Kỳ 2 - Tháng này", "Kỳ 1 - Tháng trước", "Kỳ 2 - Tháng trước"],
                    key="payroll_period_preset", filter_mode="contains"
                )
                p_start, p_end, period_err = resolve_payroll_period(preset, get_vn_today())
                if period_err:
                    st.error(period_err)
                elif p_start and p_end:
                    st.info(f"Kỳ đang chọn: **{p_start.strftime('%d/%m/%Y')} → {p_end.strftime('%d/%m/%Y')}**")
            with c_source:
                source_mode = st.selectbox(
                    "Nguồn dữ liệu lương:",
                    ["TimeSoft", "Upload file Excel", "Google Sheet mặc định"],
                    index=0,
                    key="payroll_source_mode_v855", filter_mode="contains"
                )
                payroll_upload = None
                if source_mode == "TimeSoft":
                    st.caption(
                        "⭐ Nguồn mặc định: TimeSoft. Khi bấm Tính lương, hệ thống tự lấy dữ liệu "
                        "đúng kỳ đang chọn và chỉ cộng các dòng có loại bắt đầu bằng 'Tip'."
                    )
                    if not timesoft_is_configured():
                        st.warning("⚠️ TimeSoft chưa được cấu hình đầy đủ trong Secrets.")
                elif source_mode == "Upload file Excel":
                    payroll_upload = st.file_uploader(
                        "Upload file dulieuluong (.xlsx/.xlsm)", type=["xlsx", "xlsm"], key="payroll_upload_file",
                        help=f"File phải có sheet '{PAYROLL_SOURCE_WORKSHEET}'."
                    )
                else:
                    st.caption("Nguồn phụ: Google Sheet 1WtYsbEAlifL1PZ-nSGBojgL4Bnur-1vF")

            # Trạng thái trực quan cho quy trình tải dữ liệu & tính lương.
            if "payroll_process_message" not in st.session_state:
                st.session_state.payroll_process_message = "⏸️ Sẵn sàng tính lương nhân viên."
            if "payroll_process_state" not in st.session_state:
                st.session_state.payroll_process_state = "idle"

            state_icon = {"idle": "⚪", "running": "🔵", "complete": "🟢", "error": "🔴"}.get(
                st.session_state.payroll_process_state, "⚪"
            )
            st.markdown(
                f"<div style='padding:8px 12px;border:1px solid #D9D9D9;border-radius:8px;"
                f"background:#fafafa;margin:4px 0 8px 0;font-weight:600;'>"
                f"{state_icon} {st.session_state.payroll_process_message}</div>",
                unsafe_allow_html=True
            )

            calc_col, recalc_col = st.columns(2)
            with calc_col:
                payroll_calc_clicked = st.button(
                    "🧮 Tính lương nhân viên", use_container_width=True, disabled=bool(period_err),
                    key="payroll_calc_button"
                )
            with recalc_col:
                payroll_clear_recalc_clicked = st.button(
                    "🧹 Xóa dữ liệu bảng lương & tính lại", use_container_width=True, disabled=bool(period_err),
                    key="payroll_clear_recalc_button",
                    help="Xóa bảng lương đang nằm trong phiên làm việc, tải lại hồ sơ/role mới nhất rồi tính lại từ đầu."
                )

            if payroll_clear_recalc_clicked:
                # Chỉ xóa dữ liệu bảng lương đang tính trong phiên, KHÔNG xóa lịch sử đã lưu.
                for _k in [
                    "payroll_current_df", "payroll_current_start", "payroll_current_end",
                    "payroll_current_source", "payroll_unmatched", "payroll_adjustment_editor"
                ]:
                    st.session_state.pop(_k, None)
                _clear_violation_debt_cache()
                st.session_state.payroll_process_state = "idle"
                st.session_state.payroll_process_message = "Đã xóa dữ liệu cũ · Đang tính lại từ đầu..."

            if payroll_calc_clicked or payroll_clear_recalc_clicked:
                progress = st.progress(0, text="0% - Bắt đầu xử lý...")
                status = st.status("🧮 Đang tính lương nhân viên...", expanded=True, state="running")
                try:
                    st.session_state.payroll_process_state = "running"
                    st.session_state.payroll_process_message = "Đang kiểm tra nguồn dữ liệu..."
                    status.write("1/5 · Kiểm tra nguồn dữ liệu và kỳ lương")
                    progress.progress(10, text="10% - Kiểm tra nguồn dữ liệu")

                    if source_mode == "TimeSoft":
                        status.write(
                            f"2/5 · Đang lấy TimeSoft: {p_start.strftime('%d/%m/%Y')} → {p_end.strftime('%d/%m/%Y')}"
                        )
                        progress.progress(20, text="20% - Đang đăng nhập và lấy dữ liệu TimeSoft")
                        src_df, src_err, _payroll_ts_result = load_payroll_source_from_timesoft(p_start, p_end)
                        st.session_state["payroll_timesoft_last_result_v855"] = _payroll_ts_result
                        src_label = f"TimeSoft {p_start.strftime('%d/%m/%Y')} - {p_end.strftime('%d/%m/%Y')}"
                        _ts_map = st.session_state.get("payroll_timesoft_mapping_v855") or {}
                        if _ts_map:
                            status.write(
                                f"✅ TimeSoft đã nhận diện {_ts_map.get('tip_count', 0)} dòng Tip "
                                f"· cấu trúc {_ts_map.get('frame', 'Data')}"
                            )
                    elif source_mode == "Upload file Excel":
                        if payroll_upload is None:
                            raise ValueError("Vui lòng upload file Excel dữ liệu lương trước khi tính.")
                        status.write(f"2/5 · Đang đọc file: {getattr(payroll_upload, 'name', 'Upload Excel')}")
                        progress.progress(25, text="25% - Đang đọc file Excel")
                        src_df, src_err = load_payroll_source_from_uploaded_excel(payroll_upload)
                        src_label = getattr(payroll_upload, 'name', 'Upload Excel')
                    else:
                        status.write("2/5 · Đang tải dữ liệu từ Google Sheet mặc định")
                        progress.progress(25, text="25% - Đang tải Google Sheet")
                        src_df, src_err = load_payroll_source_from_google_sheet()
                        src_label = f"Google Sheet {PAYROLL_SOURCE_SHEET_ID}"

                    if src_err:
                        raise ValueError(src_err)

                    row_count = len(src_df) if isinstance(src_df, pd.DataFrame) else 0
                    status.write(f"✅ Đã đọc {row_count:,} dòng dữ liệu nguồn".replace(",", "."))
                    if source_mode == "TimeSoft" and st.session_state.current_role == "admin":
                        _map_info = st.session_state.get("payroll_timesoft_mapping_v855") or {}
                        if _map_info.get("mapping"):
                            status.write(
                                "🔎 Mapping TimeSoft → Lương: "
                                + " | ".join(f"{k} ← {v}" for k, v in _map_info["mapping"].items())
                            )
                    progress.progress(45, text="45% - Đã đọc dữ liệu nguồn")

                    status.write("3/5 · Đang tải dữ liệu tiền phạt từ hệ thống")
                    st.session_state.payroll_process_message = "Đang tải dữ liệu tiền phạt..."
                    progress.progress(60, text="60% - Đang tải tiền phạt")
                    leave_primary = load_backup_sheet_data()
                    penalty_rows = len(leave_primary) if isinstance(leave_primary, pd.DataFrame) else 0
                    status.write(f"✅ Đã tải {penalty_rows:,} dòng lịch nghỉ/vi phạm".replace(",", "."))

                    status.write("4/5 · Đang tải lại vai trò nhân viên, đồng bộ TichLuy và tính lương")
                    st.session_state.payroll_process_message = "Đang tải lại vai trò nhân viên và tính lương..."
                    progress.progress(75, text="75% - Đang tính lương")

                    # V44: luôn lấy hồ sơ/Phân quyền MỚI NHẤT trước mỗi lần tính lương.
                    # Điều này bảo đảm người vừa đổi từ nhanvien -> letan/quanly/locker/tapvu
                    # bị loại ngay khỏi bảng lương, không chờ cache load_credentials hết hạn.
                    credentials_live = load_credentials_fresh()
                    df_credentials = credentials_live
                    nhanvien_live_count = 0
                    if isinstance(credentials_live, pd.DataFrame) and not credentials_live.empty and 'Phân quyền' in credentials_live.columns:
                        nhanvien_live_count = int(
                            credentials_live['Phân quyền'].astype(str).str.strip().str.lower().isin(PAYROLL_ELIGIBLE_ROLES).sum()
                        )
                    status.write(f"✅ Hồ sơ mới nhất: {nhanvien_live_count} tài khoản Nhân viên/Leader")

                    # Bảo đảm mọi nhân viên đủ điều kiện đều có một dòng trong TichLuy trước khi tính.
                    # Đồng bộ chỉ thêm người thiếu + đánh STT; không sửa D/E/F của người đã có.
                    tl_sync_ok, tl_sync_msg = sync_tichluy_roles_and_stt(credentials_live)
                    if tl_sync_ok:
                        status.write(f"✅ {tl_sync_msg}")
                    else:
                        status.write(f"⚠️ {tl_sync_msg}")
                    # Tiền phạt chỉ dùng dữ liệu ở Google Sheet 1Kz0...; không lấy nguồn lịch nghỉ thứ hai.
                    payroll_df, unmatched_names = build_payroll_table(
                        src_df, credentials_live, p_start, p_end,
                        leave_primary=leave_primary, leave_secondary=None,
                        default_living_expense=payroll_default_living,
                        default_locker_support=payroll_default_locker,
                        leader_responsibility_allowance=payroll_leader_allowance
                    )

                    status.write("5/5 · Đang hoàn tất và lưu kết quả vào phiên làm việc")
                    st.session_state.payroll_process_message = "Đang hoàn tất bảng lương..."
                    progress.progress(92, text="92% - Đang hoàn tất")
                    st.session_state.payroll_current_df = payroll_df
                    st.session_state.payroll_current_start = p_start.isoformat()
                    st.session_state.payroll_current_end = p_end.isoformat()
                    st.session_state.payroll_current_source = src_label
                    st.session_state.payroll_unmatched = unmatched_names

                    progress.progress(100, text="100% - Hoàn tất")
                    st.session_state.payroll_process_state = "complete"
                    st.session_state.payroll_process_message = f"Hoàn tất · Đã tính lương cho {len(payroll_df)} nhân viên."
                    status.update(
                        label=f"✅ Hoàn tất - Đã tính lương cho {len(payroll_df)} nhân viên",
                        state="complete", expanded=False
                    )
                    st.success(f"✅ Đã tính lương cho {len(payroll_df)} tài khoản Nhân viên/Leader.")
                    if unmatched_names:
                        status.write(f"⚠️ Có {len(unmatched_names)} tên trong dữ liệu Tip chưa khớp tài khoản hệ thống.")
                except Exception as e:
                    progress.empty()
                    st.session_state.payroll_process_state = "error"
                    st.session_state.payroll_process_message = f"Lỗi: {e}"
                    status.update(label=f"❌ Không thể tính lương: {e}", state="error", expanded=True)
                    status.write(f"❌ {e}")
                    st.error(f"❌ {e}")

            current = st.session_state.get('payroll_current_df')
            if isinstance(current, pd.DataFrame) and not current.empty:
                # Dọn cả dữ liệu đang nằm trong session từ bản cũ để không còn dòng header giả.
                current = _filter_real_payroll_rows(current)
                st.session_state.payroll_current_df = current
                current_start = date.fromisoformat(st.session_state.get('payroll_current_start'))
                current_end = date.fromisoformat(st.session_state.get('payroll_current_end'))
                unmatched = st.session_state.get('payroll_unmatched', [])
                if unmatched:
                    st.warning("Có tên ở dữ liệu Tip nhưng không khớp tài khoản hệ thống: " + ", ".join(map(str, unmatched)))

                # V49: bảng điều chỉnh được chuyển xuống CUỐI trang và luôn đóng mặc định.
                # Ở phần nội dung phía trên, dùng dữ liệu bảng lương hiện có trong session.
                # Khi Admin mở bảng điều chỉnh ở cuối trang và thay đổi số liệu, ứng dụng lưu lại
                # vào session rồi rerun để toàn bộ thống kê/export/email phía trên cập nhật đồng bộ.
                final_df = recalculate_payroll_net(current.copy())
                final_df = _filter_real_payroll_rows(final_df)
                # V60: đồng bộ mọi trường hồ sơ theo Sheet1 nguồn và áp dụng role hiện tại.
                final_df = apply_latest_profile_fields_to_payroll(
                    final_df, load_credentials_recent(), only_current_nhanvien=True
                )
                st.session_state.payroll_current_df = final_df

                # V47: Admin có thể chủ động tạm hoãn tiền Vi phạm của kỳ hiện tại.
                # Khoản đã hoãn được lưu vào sheet NoViPham và chỉ bắt đầu trừ từ kỳ kế tiếp.
                if st.session_state.current_role == "admin":
                    # V55: giữ section Tạm hoãn Vi phạm luôn mở trong lúc thao tác.
                    # Selectbox có filter_mode="contains" sẽ rerun khi gõ/chọn; nếu expanded=False
                    # thì expander tự đóng sau mỗi rerun. Để expanded=True giúp phần này không bị
                    # ẩn khi Admin đang tìm/chọn nhân viên hoặc nhập số tiền tạm hoãn.
                    with st.expander("⏭️ Tạm hoãn Vi phạm sang kỳ kế tiếp", expanded=True):
                        _leave_for_defer = load_backup_sheet_data()
                        _raw_penalty_map = _period_penalty_by_employee(current_start, current_end, _leave_for_defer, None)
                        _due_debt_map, _deferred_map, _active_debts = get_violation_debt_state(
                            current_start, current_end, final_df['Tên Hệ thống'].astype(str).tolist()
                        )
                        _defer_options = []
                        _available_by_emp = {}
                        for _emp in final_df['Tên Hệ thống'].astype(str).tolist():
                            _ek = normalize_login_name(_emp)
                            _raw_current = max(0.0, float(_money_to_float(_raw_penalty_map.get(_ek, 0))))
                            _already_deferred = max(0.0, float(_money_to_float(_deferred_map.get(_ek, 0))))
                            _available = max(0.0, _raw_current - _already_deferred)
                            if _available > 0:
                                _defer_options.append(_emp)
                                _available_by_emp[_emp] = (_raw_current, _already_deferred, _available)

                        _defer_options = sort_employee_names(_defer_options)
                        if _defer_options:
                            _defer_emp = st.selectbox(
                                "Nhân viên", _defer_options, filter_mode="contains", key="payroll_defer_violation_emp"
                            )
                            _raw_current, _already_deferred, _available = _available_by_emp[_defer_emp]
                            c_a, c_b, c_c = st.columns(3)
                            c_a.metric("Vi phạm gốc kỳ này", f"{_raw_current:,.0f} đ".replace(',', '.'))
                            c_b.metric("Đã tạm hoãn", f"{_already_deferred:,.0f} đ".replace(',', '.'))
                            c_c.metric("Còn có thể hoãn", f"{_available:,.0f} đ".replace(',', '.'))
                            _step = 50000.0 if _available >= 50000 else max(1.0, _available)
                            _defer_amount = st.number_input(
                                "Số tiền Vi phạm muốn chuyển sang kỳ kế tiếp",
                                min_value=0.0, max_value=float(_available), value=float(_available), step=float(_step),
                                format="%.0f", key=f"payroll_defer_violation_amount_{normalize_login_name(_defer_emp)}"
                            )
                            if st.button("💾 Lưu nghĩa vụ Vi phạm sang kỳ kế tiếp", use_container_width=True, key="save_deferred_violation"):
                                ok, msg = defer_violation_to_next_period(
                                    _defer_emp, _defer_amount, current_start, current_end, st.session_state.current_user
                                )
                                if ok:
                                    # Cập nhật ngay bảng đang mở; lần Tính lại sau sẽ đọc lại ledger và cho cùng kết quả.
                                    _tmp = st.session_state.payroll_current_df.copy()
                                    _mask = _tmp['Tên Hệ thống'].apply(normalize_login_name).eq(normalize_login_name(_defer_emp))
                                    if _mask.any():
                                        _idx = _tmp.index[_mask][0]
                                        _tmp.at[_idx, 'Tiền phạt trong tháng'] = max(
                                            0.0,
                                            float(_money_to_float(_tmp.at[_idx, 'Tiền phạt trong tháng'])) - float(_money_to_float(_defer_amount))
                                        )
                                        _tmp = recalculate_payroll_net(_tmp)
                                        st.session_state.payroll_current_df = _tmp
                                    st.session_state.pop('payroll_adjustment_editor', None)
                                    st.success(msg)
                                    st.rerun()
                                else:
                                    st.error(msg)
                        else:
                            st.info("Không còn khoản Vi phạm của kỳ hiện tại có thể tạm hoãn.")

                        # V54: Admin xem + sửa/xóa toàn bộ Nghĩa vụ Vi phạm đang mở.
                        # Việc sửa/xóa có hiệu lực ngay với kỳ đang mở và mọi kỳ tính sau.
                        _all_active = get_all_open_violation_debts_for_admin()
                        if isinstance(_all_active, pd.DataFrame) and not _all_active.empty:
                            _show_cols = [c for c in [
                                'Tên nhân viên','Số tiền','Nội dung','Loại','Kỳ phát sinh từ','Kỳ phát sinh đến','Bắt đầu trừ từ','Trạng thái'
                            ] if c in _all_active.columns]
                            _debt_show = _all_active[_show_cols].copy()
                            if 'Số tiền' in _debt_show.columns:
                                _debt_show['Số tiền'] = _debt_show['Số tiền'].apply(lambda x: f"{_money_to_float(x):,.0f}".replace(',', '.'))
                            st.caption("Nghĩa vụ Vi phạm đang mở")
                            st.dataframe(_debt_show, hide_index=True, width="stretch", height="content")

                            with st.expander("✏️ Sửa / Xóa Nghĩa vụ Vi phạm", expanded=False):
                                _manage_rows = _all_active.reset_index(drop=True).copy()
                                _manage_options = []
                                _manage_lookup = {}
                                for _i, _r in _manage_rows.iterrows():
                                    _sheet_row = int(_r.get('__sheet_row', 0) or 0)
                                    _emp = str(_r.get('Tên nhân viên', '')).strip()
                                    _amt = float(_money_to_float(_r.get('Số tiền', 0)))
                                    _typ = str(_r.get('Loại', '')).strip()
                                    _p1 = str(_r.get('Kỳ phát sinh từ', '')).strip()
                                    _p2 = str(_r.get('Kỳ phát sinh đến', '')).strip()
                                    _label = f"{_emp} · {_amt:,.0f} đ · {_typ} · {_p1} → {_p2}".replace(',', '.')
                                    # Bảo đảm label duy nhất nếu có nhiều dòng giống nhau.
                                    if _label in _manage_lookup:
                                        _label = f"{_label} · dòng {_sheet_row}"
                                    _manage_options.append(_label)
                                    _manage_lookup[_label] = _r.to_dict()

                                _selected_label = st.selectbox(
                                    "Chọn nghĩa vụ cần chỉnh sửa",
                                    _manage_options,
                                    filter_mode="contains",
                                    key="admin_manage_violation_debt_select",
                                )
                                _selected = _manage_lookup.get(_selected_label, {})
                                _selected_row = int(_selected.get('__sheet_row', 0) or 0)
                                _selected_amount = max(0.0, float(_money_to_float(_selected.get('Số tiền', 0))))
                                _selected_due = _parse_vn_date(_selected.get('Bắt đầu trừ từ', '')) or current_start
                                _edit_amount = st.number_input(
                                    "Số tiền nghĩa vụ",
                                    min_value=1.0,
                                    value=float(max(1.0, _selected_amount)),
                                    step=50000.0,
                                    format="%.0f",
                                    key=f"edit_violation_debt_amount_{_selected_row}",
                                )
                                _edit_content = st.text_input(
                                    "Nội dung",
                                    value=str(_selected.get('Nội dung', '')).strip() or VIOLATION_DEBT_CONTENT,
                                    key=f"edit_violation_debt_content_{_selected_row}",
                                )
                                _edit_due = st.date_input(
                                    "Bắt đầu trừ từ kỳ/ngày",
                                    value=_selected_due,
                                    format="DD/MM/YYYY",
                                    key=f"edit_violation_debt_due_{_selected_row}",
                                )
                                _btn_edit, _btn_delete = st.columns(2)
                                if _btn_edit.button(
                                    "💾 Lưu chỉnh sửa nghĩa vụ",
                                    use_container_width=True,
                                    key=f"save_violation_debt_edit_{_selected_row}",
                                ):
                                    _ok, _msg = update_violation_debt_obligation(
                                        _selected_row,
                                        _edit_amount,
                                        _edit_content,
                                        _edit_due,
                                        st.session_state.current_user,
                                    )
                                    if _ok:
                                        _clear_violation_debt_cache()
                                        try:
                                            st.session_state.payroll_current_df = refresh_current_payroll_violation_debt(
                                                st.session_state.payroll_current_df,
                                                current_start,
                                                current_end,
                                            )
                                        except Exception:
                                            pass
                                        st.session_state.pop('payroll_adjustment_editor', None)
                                        st.success(_msg + " Kỳ đang mở đã được tính lại; các kỳ tiếp theo sẽ dùng số mới.")
                                        st.rerun()
                                    else:
                                        st.error(_msg)

                                _confirm_delete = st.checkbox(
                                    "Tôi xác nhận xóa vĩnh viễn nghĩa vụ này",
                                    key=f"confirm_delete_violation_debt_{_selected_row}",
                                )
                                if _btn_delete.button(
                                    "🗑️ Xóa nghĩa vụ",
                                    use_container_width=True,
                                    disabled=not _confirm_delete,
                                    key=f"delete_violation_debt_{_selected_row}",
                                ):
                                    _ok, _msg = delete_violation_debt_obligation(
                                        _selected_row,
                                        st.session_state.current_user,
                                    )
                                    if _ok:
                                        _clear_violation_debt_cache()
                                        try:
                                            st.session_state.payroll_current_df = refresh_current_payroll_violation_debt(
                                                st.session_state.payroll_current_df,
                                                current_start,
                                                current_end,
                                            )
                                        except Exception:
                                            pass
                                        st.session_state.pop('payroll_adjustment_editor', None)
                                        st.success(_msg + " Kỳ đang mở đã được tính lại; các kỳ tiếp theo sẽ không còn khoản đã xóa.")
                                        st.rerun()
                                    else:
                                        st.error(_msg)

                                st.caption(
                                    "Thay đổi áp dụng ngay cho kỳ lương đang mở và mọi kỳ tính sau. "
                                    "Các bản lương lịch sử đã lưu không bị ghi đè tự động; khi cần cập nhật bản cũ, mở bản đó và bấm Cập nhật bảng lương từ hệ thống."
                                )
                        else:
                            st.caption("Hiện không có Nghĩa vụ Vi phạm nào đang mở.")

                # Thông báo rõ các dòng âm: khi Admin lưu bảng lương, phần âm sẽ chuyển thành nghĩa vụ Vi phạm kỳ sau.
                _negative_rows = final_df[final_df['Số tiền thực nhận'].apply(_money_to_float) < 0]
                if not _negative_rows.empty:
                    _negative_total = abs(_negative_rows['Số tiền thực nhận'].apply(_money_to_float).sum())
                    _negative_details = [
                        f"{str(_r.get('Tên Hệ thống','')).strip()}: {_money_to_float(_r.get('Số tiền thực nhận',0)):,.0f} đ".replace(',', '.')
                        for _, _r in _negative_rows.iterrows()
                    ]
                    st.warning(
                        f"Có {len(_negative_rows)} nhân viên Thực nhận âm, tổng phần chưa hoàn thành "
                        f"{_negative_total:,.0f} đ. Khi lưu bảng lương, những nhân viên Admin KHÔNG chủ động tạm hoãn "
                        f"sẽ được tự lưu thành '{VIOLATION_DEBT_CONTENT}' và trừ từ kỳ lương kế tiếp.".replace(',', '.')
                    )
                    st.markdown("**Nhân viên Thực nhận âm:** " + " · ".join(_negative_details))

                total_salary = final_df['Tiền Lương'].sum()
                total_penalty = final_df['Tiền phạt trong tháng'].apply(_money_to_float).sum() + final_df.get('Vi phạm kỳ trước', pd.Series(0, index=final_df.index)).apply(_money_to_float).sum()
                total_net = final_df['Số tiền thực nhận'].sum()
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Nhân viên", len(final_df))
                c2.metric("Tổng Tiền Lương", f"{total_salary:,.0f} đ".replace(',', '.'))
                c3.metric("Tổng tiền phạt", f"{total_penalty:,.0f} đ".replace(',', '.'))
                c4.metric("Tổng thực nhận", f"{total_net:,.0f} đ".replace(',', '.'))

                display_cols = [
                    "TT", "Tên Hệ thống", "Tiền Lương", "Tiền Hỗ Trợ Hoàn Lại",
                    "Tích lũy", "Chi Phí Sinh Hoạt", "Tiền phạt trong tháng", "Vi phạm kỳ trước", "Tiền ứng lương",
                    "Tiền hỗ trợ Locker", "Số tiền thực nhận"
                ]
                st.markdown("### 📋 Bảng lương tổng hợp")
                # HTML table dùng width:100% + table-layout:fixed để không tạo thanh cuộn ngang/dọc.
                web_df = final_df[display_cols].copy()
                web_df, payroll_web_widths = apply_table_layout_df(web_df, "payroll_current")
                payroll_internal_order = list(web_df.columns)
                money_web_cols = [c for c in payroll_internal_order if c.startswith('Tiền') or c in {'Tích lũy','Chi Phí Sinh Hoạt','Vi phạm kỳ trước','Số tiền thực nhận'}]
                for c in money_web_cols:
                    web_df[c] = web_df[c].apply(lambda v: f"{_money_to_float(v):,.0f}".replace(',', '.'))

                # V53: nếu Admin đã tạm hoãn Vi phạm của kỳ này, ghi chú ngay TRONG ô Vi phạm.
                # Dùng token rồi thay bằng HTML sau khi pandas escape bảng, để không phải tắt escaping
                # cho các dữ liệu tên nhân viên lấy từ Google Sheet.
                _violation_note_tokens = {}
                try:
                    _, _web_deferred_map, _ = get_violation_debt_state(
                        current_start, current_end, final_df['Tên Hệ thống'].astype(str).tolist()
                    )
                    if 'Tiền phạt trong tháng' in web_df.columns:
                        for _row_pos, _row_idx in enumerate(web_df.index):
                            _emp_name = str(final_df.loc[_row_idx, 'Tên Hệ thống']).strip() if _row_idx in final_df.index else ''
                            _defer_amt = max(0.0, float(_money_to_float(_web_deferred_map.get(normalize_login_name(_emp_name), 0))))
                            if _defer_amt <= 0:
                                continue
                            _base_violation = str(web_df.at[_row_idx, 'Tiền phạt trong tháng'])
                            _token = f"__VERA_DEFER_NOTE_{_row_pos}__"
                            _note_amount = f"{_defer_amt:,.0f}".replace(',', '.')
                            _violation_note_tokens[_token] = (
                                f"<div class='payroll-violation-value'>{_base_violation}</div>"
                                f"<div class='payroll-violation-note'>Trừ kỳ lương kế tiếp: {_note_amount} đ</div>"
                            )
                            web_df.at[_row_idx, 'Tiền phạt trong tháng'] = _token
                except Exception:
                    _violation_note_tokens = {}

                # V50: bảng tổng hợp trên website không hiển thị thông tin ngân hàng hoặc Email.
                # V46: ghi nhớ dòng Thực nhận <= 0 trước khi đổi định dạng/đổi tên cột để
                # có thể tô vàng toàn bộ dòng trên Bảng lương nhân viên.
                _web_non_positive_mask = final_df['Số tiền thực nhận'].apply(_money_to_float).le(0).tolist()

                # Chỉ đổi tên cột lúc hiển thị; dữ liệu nội bộ vẫn giữ tên chuẩn để tính toán/lưu lịch sử.
                web_df = web_df.rename(columns={c: PAYROLL_DISPLAY_LABELS.get(c, c) for c in web_df.columns})
                payroll_html = web_df.to_html(index=False, escape=True, classes='vera-payroll-table')
                # Chỉ mở HTML đối với token do hệ thống tự tạo; mọi dữ liệu Sheet khác vẫn được escape an toàn.
                for _token, _fragment in _violation_note_tokens.items():
                    payroll_html = payroll_html.replace(_token, _fragment)

                # Gắn class cho từng <tr> trong <tbody> theo đúng thứ tự dòng.
                # Không phụ thuộc alternating row color nên dòng <= 0 luôn nổi bật màu vàng.
                try:
                    _head_html, _body_tail = payroll_html.split('<tbody>', 1)
                    _body_html, _tail_html = _body_tail.split('</tbody>', 1)
                    _row_counter = {'i': 0}
                    def _mark_payroll_non_positive_row(_match):
                        _i = _row_counter['i']
                        _row_counter['i'] += 1
                        if _i < len(_web_non_positive_mask) and _web_non_positive_mask[_i]:
                            return '<tr class="payroll-nonpositive">'
                        return '<tr>'
                    _body_html = re.sub(r'<tr>', _mark_payroll_non_positive_row, _body_html)
                    payroll_html = _head_html + '<tbody>' + _body_html + '</tbody>' + _tail_html
                except Exception:
                    pass

                width_total = max(1, sum(int(payroll_web_widths.get(c, 140)) for c in payroll_internal_order))
                colgroup = '<colgroup>' + ''.join(
                    f'<col style="width:{(int(payroll_web_widths.get(c, 140)) / width_total) * 100:.3f}%">'
                    for c in payroll_internal_order
                ) + '</colgroup>'
                payroll_html = payroll_html.replace('>', '>' + colgroup, 1)
                payroll_layout_css = table_layout_html_css(
                    "payroll_current", payroll_internal_order, "table.vera-payroll-table"
                )
                st.markdown(
                    f"""<style>
                    {payroll_layout_css}
                    .vera-payroll-wrap{{width:100%;overflow:visible;}}
                    table.vera-payroll-table{{width:100%;table-layout:fixed;border-collapse:collapse;font-size:clamp(8px,.68vw,12px);}}
                    table.vera-payroll-table th{{background:#A1948C!important;color:#000!important;font-weight:700!important;padding:5px 3px;border:1px solid #D9D9D9;white-space:nowrap!important;overflow-wrap:normal!important;word-break:normal!important;line-height:1.15!important;vertical-align:middle!important;}}
                    table.vera-payroll-table td{{padding:4px 3px;border:1px solid #D9D9D9;white-space:nowrap!important;word-break:normal!important;vertical-align:middle;}}
                    table.vera-payroll-table .payroll-violation-value{{line-height:1.05;}}
                    table.vera-payroll-table .payroll-violation-note{{font-size:6px!important;line-height:1.05!important;margin-top:2px;white-space:normal!important;font-weight:400;}}
                    table.vera-payroll-table tbody tr:nth-child(even){{background:#fafafa;}}
                    table.vera-payroll-table tbody tr.payroll-nonpositive td{{background:#FFF2CC!important;}}
                    @media(max-width:800px){{table.vera-payroll-table{{font-size:7px;}}table.vera-payroll-table th,table.vera-payroll-table td{{padding:3px 1px;}}}}
                    </style>""" + f"<div class='vera-payroll-wrap'>{payroll_html}</div>",
                    unsafe_allow_html=True
                )
                render_admin_quick_layout_default("payroll_current", payroll_internal_order, "payroll_current_summary")

                c_save, c_export = st.columns(2)
                with c_save:
                    if st.button("💾 Lưu bảng lương kỳ này vào hệ thống", use_container_width=True):
                        ok, msg, batch_id = save_payroll_snapshot(
                            final_df, current_start, current_end,
                            st.session_state.get('payroll_current_source', ''), st.session_state.current_user
                        )
                        (st.success if ok else st.error)(msg)
                        if ok:
                            load_payroll_history.clear()
                            if vpg is not None and vpg.is_enabled():
                                try:
                                    vpg.invalidate_dataset("payroll_history")
                                except Exception:
                                    pass
                            st.caption(f"Mã bản lưu: {batch_id}")
                with c_export:
                    excel_bytes = build_payroll_excel_bytes(final_df, current_start, current_end)
                    st.download_button(
                        "📥 Export toàn bộ Bảng lương Excel",
                        data=excel_bytes,
                        file_name=f"BangLuong_{current_start.strftime('%d%m%Y')}_{current_end.strftime('%d%m%Y')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

                if has_feature_access("payroll_email"):
                    with st.expander("📧 GỬI BẢNG LƯƠNG QUA EMAIL"):
                        # V59: danh sách gửi chỉ phụ thuộc Thực nhận > 0. Email KHÔNG lấy từ bảng lương
                        # vì có thể đã cũ; địa chỉ mới nhất sẽ được đọc trực tiếp từ Sheet1 ngay khi bấm Gửi.
                        _email_net = final_df['Số tiền thực nhận'].apply(_money_to_float) if 'Số tiền thực nhận' in final_df.columns else pd.Series(0, index=final_df.index)
                        emailable = final_df[_email_net > 0].copy()
                        employees_email = sort_employee_names(emailable['Tên Hệ thống'].astype(str).tolist())
                        selected_email_emps = st.multiselect(
                            "Chọn 1, nhiều hoặc tất cả nhân viên:", employees_email, default=employees_email,
                            filter_mode="contains", key="payroll_email_recipients"
                        )
                        st.caption(
                            f"Có {len(employees_email)} nhân viên có Thực nhận > 0. "
                            "Email sẽ luôn được đọc lại mới nhất từ Sheet1 ngay trước khi gửi; "
                            "nhân viên có Thực nhận ≤ 0 sẽ không được gửi mail."
                        )
                        if st.button("🚀 Gửi bảng lương cho nhân viên đã chọn", use_container_width=True):
                            if not selected_email_emps:
                                st.warning("Vui lòng chọn ít nhất 1 nhân viên.")
                            else:
                                sender_email, sender_pass = get_smtp_sender_credentials()
                                progress = st.progress(0)
                                ok_count, errors = 0, []
                                # V59: đọc hồ sơ nhân sự MỚI NHẤT đúng 1 lần cho cả lượt gửi.
                                # Không dùng Email đang nằm trong final_df vì bảng lương có thể đã được tính/lưu trước khi Email được sửa.
                                live_email_creds = load_credentials_fresh_for_email()
                                # Chỉ đọc Sheet1 lịch nghỉ một lần rồi lọc theo từng nhân viên, tránh quota 429.
                                email_leave_df = load_backup_sheet_data()
                                for idx, emp in enumerate(selected_email_emps):
                                    matched_pay = emailable[emailable['Tên Hệ thống'].astype(str) == str(emp)]
                                    if matched_pay.empty:
                                        errors.append(f"{emp}: Không tìm thấy dữ liệu bảng lương.")
                                        progress.progress((idx + 1) / len(selected_email_emps))
                                        continue
                                    row = matched_pay.iloc[0].copy()
                                    _live_row_df = apply_latest_profile_fields_to_payroll(
                                        pd.DataFrame([row]), live_email_creds, only_current_nhanvien=False
                                    )
                                    if not _live_row_df.empty:
                                        row = _live_row_df.iloc[0].copy()
                                    to_email = str(row.get('Email', '')).strip() or latest_email_from_credentials(live_email_creds, emp)
                                    if not to_email or '@' not in to_email:
                                        errors.append(f"{emp}: Email mới nhất trong Sheet1 không hợp lệ hoặc đang để trống.")
                                        progress.progress((idx + 1) / len(selected_email_emps))
                                        continue
                                    # Đồng bộ Email trên row chỉ để nội dung/file nội bộ luôn phản ánh dữ liệu mới nhất.
                                    row['Email'] = to_email
                                    emp_violations = get_employee_violation_details(emp, current_start, current_end, email_leave_df)
                                    ok, msg = send_payroll_email(
                                        sender_email, sender_pass, to_email, row,
                                        current_start, current_end, emp_violations
                                    )
                                    if ok: ok_count += 1
                                    else: errors.append(f"{emp}: {msg}")
                                    progress.progress((idx + 1) / len(selected_email_emps))
                                    time.sleep(0.35)
                                if ok_count: st.success(f"Đã gửi thành công {ok_count}/{len(selected_email_emps)} email bảng lương.")
                                for e in errors: st.error(e)

                    if st.session_state.current_role == "admin":
                        with st.expander("📨 GỬI BẢNG LƯƠNG TỔNG HỢP CHO LỄ TÂN"):
                            # Hiển thị TẤT CẢ tài khoản Lễ tân trước; chỉ sau khi Admin check tên
                            # mới lấy Email tương ứng từ hồ sơ hệ thống. Không lọc Email từ đầu.
                            letan_df = df_credentials.copy()
                            if not letan_df.empty and 'Phân quyền' in letan_df.columns:
                                letan_df = letan_df[
                                    letan_df['Phân quyền'].astype(str).str.strip().str.lower().isin(['letan', 'quanly'])
                                ].copy()
                                if 'Tên nhân viên' in letan_df.columns:
                                    letan_df = letan_df[
                                        ~letan_df['Tên nhân viên'].astype(str).apply(normalize_login_name).isin({
                                            'ten nhan vien', 'ten he thong', 'username', 'user name'
                                        })
                                    ].copy()
                            if not letan_df.empty and 'Tên nhân viên' in letan_df.columns:
                                letan_df = letan_df.assign(__sort=letan_df['Tên nhân viên'].apply(normalize_login_name)).sort_values('__sort').drop(columns='__sort')
                            if letan_df.empty:
                                st.info("Không có tài khoản Lễ tân trong hồ sơ hệ thống.")
                            else:
                                st.write("**Check đúng 1 Lễ tân để hệ thống lấy Email từ hồ sơ:**")
                                checked_letan = []
                                for i, (_, lr) in enumerate(letan_df.iterrows()):
                                    lname = str(lr.get('Tên nhân viên', '')).strip()
                                    if not lname:
                                        continue
                                    if st.checkbox(lname, key=f"payroll_letan_check_{i}_{normalize_login_name(lname)}"):
                                        checked_letan.append(lname)

                                if len(checked_letan) > 1:
                                    st.warning("⚠️ Chỉ được check 1 Lễ tân cho mỗi lần gửi.")
                                elif len(checked_letan) == 1:
                                    selected_letan = checked_letan[0]
                                    matched = letan_df[
                                        letan_df['Tên nhân viên'].astype(str).apply(normalize_login_name)
                                        == normalize_login_name(selected_letan)
                                    ]
                                    if matched.empty:
                                        st.error("Không tìm thấy hồ sơ Lễ tân đã chọn.")
                                    else:
                                        st.caption("📧 Email người nhận sẽ được kiểm tra lại trực tiếp từ Sheet1 ngay khi bấm Gửi.")
                                        if st.button(
                                            "📤 Gửi bảng lương tổng hợp cho Lễ tân đã check",
                                            use_container_width=True,
                                            key="send_payroll_summary_letan"
                                        ):
                                            live_letan_creds = load_credentials_fresh_for_email()
                                            letan_email = latest_email_from_credentials(live_letan_creds, selected_letan)
                                            if not letan_email or '@' not in letan_email:
                                                st.error(f"⚠️ Tài khoản {selected_letan} chưa có Email hợp lệ trong Sheet1 mới nhất.")
                                            else:
                                                sender_email, sender_pass = get_smtp_sender_credentials()
                                                ok, msg = send_payroll_summary_email(
                                                    sender_email, sender_pass, letan_email,
                                                    selected_letan, final_df, current_start, current_end
                                                )
                                                if ok:
                                                    st.success(f"{msg} Email mới nhất: {letan_email}")
                                                else:
                                                    st.error(msg)
                                else:
                                    st.caption("Chưa chọn Lễ tân nhận bảng lương tổng hợp.")

                # V49: BẢNG ĐIỀU CHỈNH BẢNG LƯƠNG LUÔN ẨN VÀ NẰM CUỐI CÙNG CỦA TRANG TÍNH LƯƠNG.
                # Đây chính là bảng TT / Tên Hệ thống / Tiền Lương / Hỗ Trợ Hoàn Lại / Tích lũy / ...
                # Người dùng chỉ mở khi thật sự cần điều chỉnh.
                with st.expander("✏️ Mở bảng điều chỉnh bảng lương", expanded=False):
                    st.caption("Bảng này được ẩn mặc định. Chỉ mở khi cần chỉnh các khoản cho bảng lương hiện tại.")
                    editor_cols = [
                        "TT", "Tên Hệ thống", "Tiền Lương", "Tiền Hỗ Trợ Hoàn Lại",
                        "Tích lũy", "Chi Phí Sinh Hoạt", "Tiền phạt trong tháng", "Vi phạm kỳ trước", "Tiền ứng lương", "Tiền hỗ trợ Locker"
                    ]
                    editor_source = st.session_state.get('payroll_current_df')
                    if isinstance(editor_source, pd.DataFrame) and not editor_source.empty:
                        editor_source = _filter_real_payroll_rows(editor_source.copy())
                        editor_df = editor_source[editor_cols].copy()
                        editor_df, _ = apply_table_layout_df(editor_df, "payroll_current")
                        col_cfg = {
                            "TT": st.column_config.NumberColumn(PAYROLL_DISPLAY_LABELS["TT"], format="%d", disabled=True, width=layout_width("payroll_current", "TT", "small")),
                            "Tên Hệ thống": st.column_config.TextColumn(PAYROLL_DISPLAY_LABELS["Tên Hệ thống"], disabled=True, width=layout_width("payroll_current", "Tên Hệ thống", "small")),
                            "Tiền Lương": st.column_config.NumberColumn(PAYROLL_DISPLAY_LABELS["Tiền Lương"], format="%,d", disabled=True, width=layout_width("payroll_current", "Tiền Lương", "small")),
                            "Tiền phạt trong tháng": st.column_config.NumberColumn(PAYROLL_DISPLAY_LABELS["Tiền phạt trong tháng"], format="%,d", disabled=True, width=layout_width("payroll_current", "Tiền phạt trong tháng", "small")),
                            "Vi phạm kỳ trước": st.column_config.NumberColumn(PAYROLL_DISPLAY_LABELS["Vi phạm kỳ trước"], format="%,d", disabled=True, width=layout_width("payroll_current", "Vi phạm kỳ trước", "small")),
                            "Tích lũy": st.column_config.NumberColumn(PAYROLL_DISPLAY_LABELS["Tích lũy"], format="%,d", disabled=True, width=layout_width("payroll_current", "Tích lũy", "small")),
                        }
                        for c in [x for x in PAYROLL_ADJUSTMENT_COLUMNS if x != "Tích lũy"]:
                            col_cfg[c] = st.column_config.NumberColumn(
                                PAYROLL_DISPLAY_LABELS.get(c, c), min_value=0.0, step=50000.0,
                                format="%,d", width=layout_width("payroll_current", c, "small")
                            )
                        edited = st.data_editor(
                            editor_df, key="payroll_adjustment_editor", width="stretch", height="content", hide_index=True,
                            row_height=layout_row_height("payroll_current"),
                            column_config=col_cfg,
                            disabled=["TT", "Tên Hệ thống", "Tiền Lương", "Tích lũy", "Tiền phạt trong tháng", "Vi phạm kỳ trước"]
                        )

                        adjusted_df = editor_source.copy()
                        for c in editor_cols:
                            if c in edited.columns:
                                adjusted_df[c] = edited[c].values
                        adjusted_df = recalculate_payroll_net(adjusted_df)
                        adjusted_df = _filter_real_payroll_rows(adjusted_df)

                        compare_cols = [c for c in editor_cols + ["Số tiền thực nhận"] if c in adjusted_df.columns and c in editor_source.columns]
                        _changed = False
                        try:
                            _left = adjusted_df[compare_cols].reset_index(drop=True).fillna(0)
                            _right = editor_source[compare_cols].reset_index(drop=True).fillna(0)
                            _changed = not _left.equals(_right)
                        except Exception:
                            _changed = True
                        if _changed:
                            st.session_state.payroll_current_df = adjusted_df
                            st.rerun()
                    else:
                        st.info("Chưa có dữ liệu bảng lương để điều chỉnh.")
    with tab_history:
        if not has_feature_access("payroll_history"):
            st.info("🔒 Bạn chưa được cấp quyền xem Lịch sử bảng lương.")
        else:
            delete_flash = st.session_state.pop("payroll_history_delete_flash", None)
            if delete_flash:
                flash_type, flash_text = delete_flash
                if flash_type == "success":
                    st.success(flash_text)
                else:
                    st.warning(flash_text)

            history = load_payroll_history()
            if history.empty or 'Mã bản lưu' not in history.columns:
                st.info("Chưa có bảng lương nào được lưu trong hệ thống.")
            else:
                batches = [x for x in history['Mã bản lưu'].dropna().astype(str).unique().tolist() if x.strip()]
                # Bản mới nhất nằm cuối Sheet nên đảo lên đầu.
                batches = list(reversed(batches))

                # Admin có thể xóa bớt một hoặc nhiều bản lịch sử. Chức năng này chỉ xóa
                # snapshot bảng lương, tuyệt đối không hoàn tác Tích lũy / Vi phạm / hồ sơ.
                if st.session_state.current_role == "admin":
                    batch_labels = {}
                    for _batch_id in batches:
                        _g = history[history['Mã bản lưu'].astype(str) == str(_batch_id)]
                        if _g.empty:
                            batch_labels[_batch_id] = str(_batch_id)
                        else:
                            _r = _g.iloc[0]
                            _period = f"{_r.get('Từ ngày','')} → {_r.get('Đến ngày','')}"
                            _saved_at = f"{_r.get('Ngày lưu','')} {_r.get('Giờ lưu','')}".strip()
                            batch_labels[_batch_id] = f"{_batch_id} | {_period} | lưu {_saved_at}"

                    with st.expander("🗑 Xóa bớt lịch sử bảng lương (Admin)", expanded=False):
                        st.caption(
                            "Có thể chọn 1 hoặc nhiều bản lương để xóa. Việc này chỉ xóa Lịch sử bảng lương đã lưu; "
                            "không thay đổi sheet TichLuy, dữ liệu vi phạm, lịch nghỉ hoặc hồ sơ nhân viên."
                        )
                        delete_batches = st.multiselect(
                            "Chọn bản lương cần xóa:",
                            options=batches,
                            default=[],
                            format_func=lambda x: batch_labels.get(x, str(x)),
                            filter_mode="contains",
                            key="payroll_history_delete_batches",
                        )
                        if delete_batches:
                            st.warning(
                                f"Bạn đang chọn xóa {len(delete_batches)} bản lương. Thao tác này không có nút hoàn tác trong ứng dụng."
                            )
                        confirm_delete_history = st.checkbox(
                            "Tôi xác nhận xóa vĩnh viễn các bản lương đã chọn khỏi lịch sử",
                            key="confirm_delete_payroll_history",
                        )
                        if st.button(
                            "🗑 Xóa các bản lương đã chọn",
                            use_container_width=True,
                            type="primary",
                            key="delete_selected_payroll_history",
                            disabled=not (delete_batches and confirm_delete_history),
                        ):
                            ok_delete, msg_delete, deleted_batches = delete_payroll_snapshots(delete_batches)
                            if ok_delete:
                                # Dọn state liên quan đến các batch vừa xóa để lần rerun sau không giữ editor cũ.
                                for _deleted_batch in deleted_batches:
                                    for _state_key in (
                                        f"payroll_history_system_refresh_{_deleted_batch}",
                                        f"payroll_history_editor_version_{_deleted_batch}",
                                    ):
                                        st.session_state.pop(_state_key, None)
                                st.session_state.pop("payroll_history_batch", None)
                                st.session_state.pop("payroll_history_delete_batches", None)
                                st.session_state.pop("confirm_delete_payroll_history", None)
                                st.session_state["payroll_history_delete_flash"] = ("success", msg_delete)
                                st.rerun()
                            else:
                                st.error(msg_delete)

                batch = st.selectbox("Chọn bản lương đã lưu:", batches, filter_mode="contains", key="payroll_history_batch")
                saved = history[history['Mã bản lưu'].astype(str) == str(batch)].copy()
                if not saved.empty:
                    st.info(
                        f"Kỳ {saved.iloc[0].get('Từ ngày','')} → {saved.iloc[0].get('Đến ngày','')} | "
                        f"Lưu bởi {saved.iloc[0].get('Người lưu','')} lúc {saved.iloc[0].get('Giờ lưu','')} ngày {saved.iloc[0].get('Ngày lưu','')}"
                    )
                    saved_table = payroll_history_to_table(saved)
                    saved_table = _filter_real_payroll_rows(saved_table)
                    # V60: hồ sơ của bản lịch sử lấy từ Sheet1 gần nhất ngay khi mở.
                    _history_live_creds = load_credentials_recent()
                    saved_table = apply_latest_profile_fields_to_payroll(
                        saved_table, _history_live_creds, only_current_nhanvien=False
                    )

                    # Nếu Admin vừa bấm "Cập nhật bảng lương từ hệ thống", dùng bản đã làm mới
                    # làm dữ liệu nền cho editor ở lần rerun kế tiếp. Mỗi batch có state riêng.
                    hist_refresh_key = f"payroll_history_system_refresh_{batch}"
                    hist_editor_version_key = f"payroll_history_editor_version_{batch}"
                    if hist_refresh_key in st.session_state:
                        try:
                            refreshed_state_df = st.session_state.get(hist_refresh_key)
                            if isinstance(refreshed_state_df, pd.DataFrame) and not refreshed_state_df.empty:
                                saved_table = _filter_real_payroll_rows(refreshed_state_df.copy())
                                saved_table = apply_latest_profile_fields_to_payroll(
                                    saved_table, _history_live_creds, only_current_nhanvien=False
                                )
                        except Exception:
                            pass

                    # V60: role cũng lấy theo Sheet1 gần nhất; bảng lương áp dụng `nhanvien` và `leader`.
                    try:
                        _current_role_map = _credential_role_map(_history_live_creds)
                        if 'Tên Hệ thống' in saved_table.columns and _current_role_map:
                            saved_table = saved_table[
                                saved_table['Tên Hệ thống'].astype(str).apply(
                                    lambda _n: _current_role_map.get(normalize_login_name(_n), 'nhanvien') in PAYROLL_ELIGIBLE_ROLES
                                )
                            ].copy()
                            saved_table = _filter_real_payroll_rows(saved_table)
                    except Exception:
                        pass

                    try:
                        hs = pd.to_datetime(saved.iloc[0]['Từ ngày'], dayfirst=True).date()
                        he = pd.to_datetime(saved.iloc[0]['Đến ngày'], dayfirst=True).date()
                    except Exception:
                        hs, he = get_vn_today(), get_vn_today()

                    # Cập nhật các dữ liệu hệ thống có thể thay đổi sau khi bản lương đã được lưu.
                    # Nút được đặt ngay phía trên tiêu đề Mở lại và chỉnh sửa bản lương theo yêu cầu.
                    st.caption(
                        "Nút cập nhật hệ thống sẽ làm mới: Tích lũy, Vi phạm, Phí Sinh Hoạt, Tiền hỗ trợ Locker, "
                        "Họ và Tên, Tài khoản ngân hàng, Tên ngân hàng, Email và vai trò hiện tại. Tiền Lương không bị thay đổi."
                    )
                    if st.button(
                        "🔄 Cập nhật bảng lương từ hệ thống",
                        use_container_width=True,
                        key=f"refresh_payroll_from_system_{batch}"
                    ):
                        progress_refresh = st.progress(0)
                        status_refresh = st.empty()
                        try:
                            status_refresh.info("⏳ Đang tải hồ sơ nhân viên mới nhất...")
                            progress_refresh.progress(20)
                            credentials_live = load_credentials_fresh()

                            status_refresh.info("⏳ Đang tải tiền phạt trong kỳ từ hệ thống lịch nghỉ...")
                            progress_refresh.progress(45)
                            try:
                                load_backup_sheet_data.clear()
                            except Exception:
                                pass
                            leave_live = load_backup_sheet_data()

                            status_refresh.info("⏳ Đang tải Tích lũy, Phí sinh hoạt / Locker và cập nhật bảng lương...")
                            progress_refresh.progress(70)
                            _clear_payroll_config_cache()
                            # Bắt buộc làm mới TichLuy để nút cập nhật không dùng snapshot cache 120 giây cũ.
                            try:
                                load_tichluy_tracking.clear()
                                if vpg is not None and vpg.is_enabled():
                                    try:
                                        vpg.invalidate_dataset("tichluy")
                                    except Exception:
                                        pass
                            except Exception:
                                pass
                            refreshed_df, refresh_meta = refresh_saved_payroll_from_system(
                                saved_table, hs, he,
                                credentials_df=credentials_live,
                                leave_primary=leave_live
                            )

                            current_hist_version = int(st.session_state.get(hist_editor_version_key, 0) or 0)
                            st.session_state[hist_refresh_key] = refreshed_df
                            st.session_state[hist_editor_version_key] = current_hist_version + 1
                            saved_table = _filter_real_payroll_rows(refreshed_df.copy())

                            progress_refresh.progress(100)
                            status_refresh.success(
                                f"✅ Đã cập nhật dữ liệu hệ thống cho {refresh_meta.get('updated', len(refreshed_df))} nhân viên; "
                                f"Tích lũy kỳ này có số tiền ở {refresh_meta.get('tichluy_updated', 0)} nhân viên. "
                                "Tiền Lương và các khoản nhập tay được giữ nguyên; Thực nhận đã tính lại."
                            )
                            missing_profiles = refresh_meta.get('missing', [])
                            if missing_profiles:
                                st.warning(
                                    "⚠️ Không tìm thấy hồ sơ hệ thống của: " + ", ".join(missing_profiles)
                                    + ". Các thông tin ngân hàng/email cũ của những người này được giữ nguyên."
                                )
                        except Exception as e:
                            progress_refresh.empty()
                            status_refresh.error(f"❌ Không cập nhật được bảng lương từ hệ thống: {e}")

                    # V61: giữ vị trí email ngay dưới nút cập nhật hệ thống, nhưng nội dung
                    # được tạo sau khi editor đã tính xong edited_saved_table. Streamlit container
                    # cho phép render đúng vị trí mà không nhân đôi logic email.
                    history_email_actions = st.container()

                    st.markdown("#### ✏️ Mở lại và chỉnh sửa bản lương")
                    st.caption(
                        "Bạn có thể sửa trực tiếp các khoản tiền bên dưới. Cột Thực nhận được hệ thống tự tính lại. "
                        "Khi bấm Ghi đè, Mã bản lưu được giữ nguyên và bản cũ sẽ được thay bằng dữ liệu mới."
                    )

                    history_edit_cols = [c for c in [
                        "TT", "Tên Hệ thống", "Tiền Lương", "Tiền Hỗ Trợ Hoàn Lại", "Tích lũy",
                        "Chi Phí Sinh Hoạt", "Tiền phạt trong tháng", "Vi phạm kỳ trước", "Tiền ứng lương", "Tiền hỗ trợ Locker",
                        "Số tiền thực nhận", "Số tài khoản ngân hàng", "Tên ngân hàng", "Email"
                    ] if c in saved_table.columns]
                    hist_editor_df = saved_table[history_edit_cols].copy()
                    hist_editor_df, _ = apply_table_layout_df(hist_editor_df, "payroll_history")

                    hist_col_cfg = {
                        "TT": st.column_config.NumberColumn(PAYROLL_DISPLAY_LABELS.get("TT", "TT"), format="%d", disabled=True, width=layout_width("payroll_history", "TT", "small")),
                        "Tên Hệ thống": st.column_config.TextColumn(PAYROLL_DISPLAY_LABELS.get("Tên Hệ thống", "Tên Hệ thống"), disabled=True, width=layout_width("payroll_history", "Tên Hệ thống", "small")),
                        "Số tiền thực nhận": st.column_config.NumberColumn(PAYROLL_DISPLAY_LABELS.get("Số tiền thực nhận", "Thực nhận"), format="%,d", disabled=True, width=layout_width("payroll_history", "Số tiền thực nhận", "small")),
                        "Số tài khoản ngân hàng": st.column_config.TextColumn(PAYROLL_DISPLAY_LABELS.get("Số tài khoản ngân hàng", "Tài khoản ngân hàng"), disabled=True, width=layout_width("payroll_history", "Số tài khoản ngân hàng", "small")),
                        "Tên ngân hàng": st.column_config.TextColumn(PAYROLL_DISPLAY_LABELS.get("Tên ngân hàng", "Tên ngân hàng"), disabled=True, width=layout_width("payroll_history", "Tên ngân hàng", "small")),
                        "Email": st.column_config.TextColumn(PAYROLL_DISPLAY_LABELS.get("Email", "Email"), disabled=True, width=layout_width("payroll_history", "Email", "small")),
                    }
                    for c in [
                        "Tiền Lương", "Tiền Hỗ Trợ Hoàn Lại", "Tích lũy", "Chi Phí Sinh Hoạt",
                        "Tiền phạt trong tháng", "Vi phạm kỳ trước", "Tiền ứng lương", "Tiền hỗ trợ Locker"
                    ]:
                        if c in hist_editor_df.columns:
                            hist_col_cfg[c] = st.column_config.NumberColumn(
                                PAYROLL_DISPLAY_LABELS.get(c, c), min_value=0.0, step=50000.0, format="%,d", width=layout_width("payroll_history", c, "small"),
                                disabled=(c in {"Tích lũy", "Vi phạm kỳ trước"})
                            )

                    # V58: Streamlit chỉ áp dụng Pandas Styler cho các cột KHÔNG chỉnh sửa
                    # trong st.data_editor. Ở V57 việc ép màu bằng !important khiến một số ô
                    # disabled bị render thành màu đen, trong khi các ô editable vẫn không thể
                    # nhận màu theo hàng. Vì vậy editor dùng nền chuẩn để không còn ô đen;
                    # ngay bên dưới sẽ có bảng xem trước tô vàng TOÀN BỘ hàng Thực nhận <= 0.
                    hist_editor_version = int(st.session_state.get(hist_editor_version_key, 0) or 0)
                    edited_hist = st.data_editor(
                        hist_editor_df,
                        key=f"payroll_history_editor_{batch}_{hist_editor_version}",
                        width="stretch", height="content", hide_index=True,
                        row_height=layout_row_height("payroll_history"),
                        column_config=hist_col_cfg,
                        disabled=[c for c in ["TT", "Tên Hệ thống", "Tích lũy", "Vi phạm kỳ trước", "Số tiền thực nhận", "Số tài khoản ngân hàng", "Tên ngân hàng", "Email"] if c in hist_editor_df.columns]
                    )

                    edited_saved_table = saved_table.copy()
                    for c in edited_hist.columns:
                        if c in edited_saved_table.columns:
                            edited_saved_table[c] = edited_hist[c].values
                    edited_saved_table = recalculate_payroll_net(edited_saved_table)
                    edited_saved_table = _filter_real_payroll_rows(edited_saved_table)

                    # V58: Bảng xem trước sau chỉnh sửa. Dùng HTML để bảo đảm TẤT CẢ ô trong
                    # hàng có Thực nhận <= 0 đều nền vàng, không phụ thuộc giới hạn Styler của
                    # st.data_editor và không bị theme tối ghi đè thành màu đen.
                    try:
                        _hist_preview_cols = [c for c in history_edit_cols if c in edited_saved_table.columns]
                        _hist_preview = edited_saved_table[_hist_preview_cols].copy()
                        _hist_non_positive = _hist_preview.get(
                            "Số tiền thực nhận", pd.Series([0] * len(_hist_preview), index=_hist_preview.index)
                        ).apply(_money_to_float).le(0).tolist()

                        _hist_money_cols = {
                            "Tiền Lương", "Tiền Hỗ Trợ Hoàn Lại", "Tích lũy", "Chi Phí Sinh Hoạt",
                            "Tiền phạt trong tháng", "Vi phạm kỳ trước", "Tiền ứng lương",
                            "Tiền hỗ trợ Locker", "Số tiền thực nhận"
                        }
                        for _c in _hist_preview.columns:
                            if _c in _hist_money_cols:
                                _hist_preview[_c] = _hist_preview[_c].apply(
                                    lambda _v: f"{_money_to_float(_v):,.0f}" if _money_to_float(_v) != 0 else "0"
                                )
                            elif _c == "TT":
                                _hist_preview[_c] = pd.to_numeric(_hist_preview[_c], errors="coerce").fillna(0).astype(int)
                            else:
                                _hist_preview[_c] = _hist_preview[_c].fillna("").astype(str)

                        _hist_preview = _hist_preview.rename(
                            columns={c: PAYROLL_DISPLAY_LABELS.get(c, c) for c in _hist_preview.columns}
                        )
                        _hist_html = _hist_preview.to_html(
                            index=False, escape=True, classes="vera-history-payroll-preview"
                        )
                        try:
                            _hh, _hbt = _hist_html.split("<tbody>", 1)
                            _hb, _ht = _hbt.split("</tbody>", 1)
                            _hc = {"i": 0}
                            def _mark_history_nonpositive(_m):
                                _i = _hc["i"]
                                _hc["i"] += 1
                                if _i < len(_hist_non_positive) and _hist_non_positive[_i]:
                                    return '<tr class="history-nonpositive">'
                                return '<tr>'
                            _hb = re.sub(r"<tr>", _mark_history_nonpositive, _hb)
                            _hist_html = _hh + "<tbody>" + _hb + "</tbody>" + _ht
                        except Exception:
                            pass

                        st.markdown("**👁 Bảng xem trước sau chỉnh sửa**")
                        # V60: không thụt lề HTML; Markdown sẽ không còn hiểu thành code block.
                        _hist_preview_css = (
                            "<style>"
                            ".vera-history-preview-wrap{width:100%;overflow-x:auto;margin:4px 0 10px 0;}"
                            "table.vera-history-payroll-preview{width:max-content;min-width:100%;border-collapse:collapse;table-layout:auto;font-size:12px;}"
                            "table.vera-history-payroll-preview th{background:#A1948C!important;color:#000!important;font-weight:700!important;padding:5px 4px;border:1px solid #D9D9D9;white-space:nowrap!important;overflow-wrap:normal!important;word-break:normal!important;}"
                            "table.vera-history-payroll-preview td{background:#fff!important;color:#000!important;padding:4px;border:1px solid #D9D9D9;white-space:nowrap!important;word-break:normal!important;}"
                            "table.vera-history-payroll-preview tbody tr:nth-child(even) td{background:#fafafa!important;}"
                            "table.vera-history-payroll-preview tbody tr.history-nonpositive td{background:#FFF2CC!important;color:#000!important;}"
                            "</style>"
                        )
                        st.markdown(
                            _hist_preview_css + f"<div class='vera-history-preview-wrap'>{_hist_html}</div>",
                            unsafe_allow_html=True,
                        )
                    except Exception:
                        pass
                    render_admin_quick_layout_default("payroll_history", history_edit_cols, f"history_{batch}")

                    # Hiển thị nhanh tổng sau khi sửa để kiểm tra trước khi ghi đè.
                    # V85.4: riêng Admin có thêm "Tổng phạt vi phạm"
                    # = Vi phạm trong kỳ + Vi phạm kỳ trước.
                    _hist_total_salary = edited_saved_table.get(
                        "Tiền Lương", pd.Series(0, index=edited_saved_table.index)
                    ).apply(_money_to_float).sum()
                    _hist_total_net = edited_saved_table.get(
                        "Số tiền thực nhận", pd.Series(0, index=edited_saved_table.index)
                    ).apply(_money_to_float).sum()

                    if str(st.session_state.get("current_role", "")).strip().lower() == "admin":
                        _hist_total_violation = (
                            edited_saved_table.get(
                                "Tiền phạt trong tháng", pd.Series(0, index=edited_saved_table.index)
                            ).apply(_money_to_float).sum()
                            +
                            edited_saved_table.get(
                                "Vi phạm kỳ trước", pd.Series(0, index=edited_saved_table.index)
                            ).apply(_money_to_float).sum()
                        )
                        h1, h2, h3, h4 = st.columns(4)
                        h1.metric("Nhân viên", len(edited_saved_table))
                        h2.metric("Tổng Tiền Lương", f"{_hist_total_salary:,.0f} đ".replace(',', '.'))
                        h3.metric("Tổng phạt vi phạm", f"{_hist_total_violation:,.0f} đ".replace(',', '.'))
                        h4.metric("Tổng Thực nhận", f"{_hist_total_net:,.0f} đ".replace(',', '.'))
                    else:
                        h1, h2, h3 = st.columns(3)
                        h1.metric("Nhân viên", len(edited_saved_table))
                        h2.metric("Tổng Tiền Lương", f"{_hist_total_salary:,.0f} đ".replace(',', '.'))
                        h3.metric("Tổng Thực nhận", f"{_hist_total_net:,.0f} đ".replace(',', '.'))

                    confirm_overwrite = st.checkbox(
                        f"Tôi xác nhận ghi đè bản lương {batch}",
                        key=f"confirm_payroll_overwrite_{batch}"
                    )
                    c_overwrite, c_export_hist = st.columns(2)
                    with c_overwrite:
                        if st.button(
                            "💾 Ghi đè cập nhật bản lương này",
                            use_container_width=True,
                            key=f"overwrite_payroll_{batch}",
                            disabled=not confirm_overwrite
                        ):
                            source_label = str(saved.iloc[0].get('Nguồn dữ liệu', '')).strip()
                            ok, msg = overwrite_payroll_snapshot(
                                batch, edited_saved_table, hs, he, source_label, st.session_state.current_user
                            )
                            if ok:
                                load_payroll_history.clear()
                                if vpg is not None and vpg.is_enabled():
                                    try:
                                        vpg.invalidate_dataset("payroll_history")
                                    except Exception:
                                        pass
                                try:
                                    st.session_state.pop(hist_refresh_key, None)
                                    st.session_state[hist_editor_version_key] = hist_editor_version + 1
                                except Exception:
                                    pass
                                st.success(msg)
                                st.info(
                                    f"Bản {batch} đã được cập nhật lúc {datetime.now(VN_TZ).strftime('%H:%M:%S %d/%m/%Y')} "
                                    f"bởi {st.session_state.current_user}."
                                )
                            else:
                                st.error(msg)

                    with c_export_hist:
                        try:
                            hist_excel = build_payroll_excel_bytes(edited_saved_table, hs, he)
                            st.download_button(
                                "📥 Export bản đang chỉnh sửa",
                                data=hist_excel,
                                file_name=f"BangLuong_DaLuu_{hs.strftime('%d%m%Y')}_{he.strftime('%d%m%Y')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key=f"export_payroll_history_{batch}"
                            )
                        except Exception as e:
                            st.warning(f"Không tạo được file export lịch sử: {e}")

                    if has_feature_access("payroll_email"):
                        with history_email_actions:
                            # --- GỬI EMAIL TỪ BẢN LƯƠNG ĐANG MỞ / ĐANG CHỈNH SỬA ---
                            # Dùng trực tiếp edited_saved_table để email phản ánh đúng số liệu Admin/Lễ tân/Quản lý
                            # đang nhìn thấy trên màn hình, kể cả trước khi bấm Ghi đè.
                            with st.expander("📧 GỬI BẢNG LƯƠNG QUA EMAIL (BẢN ĐANG CHỈNH SỬA)"):
                                st.caption(
                                    "Email và file đính kèm sẽ dùng số liệu của bản đang chỉnh sửa hiện tại. "
                                    "Nếu cần lưu các thay đổi này vào hệ thống, hãy bấm Ghi đè cập nhật bản lương."
                                )
                                hist_emailable = edited_saved_table.copy()
                                _hist_email_net = (
                                    hist_emailable['Số tiền thực nhận'].apply(_money_to_float)
                                    if 'Số tiền thực nhận' in hist_emailable.columns
                                    else pd.Series(0, index=hist_emailable.index)
                                )
                                hist_emailable = hist_emailable[_hist_email_net > 0].copy()

                                hist_employee_names = (
                                    sort_employee_names(hist_emailable['Tên Hệ thống'].astype(str).tolist())
                                    if not hist_emailable.empty and 'Tên Hệ thống' in hist_emailable.columns else []
                                )
                                hist_selected_email_emps = st.multiselect(
                                    "Chọn 1, nhiều hoặc tất cả nhân viên:",
                                    hist_employee_names,
                                    default=hist_employee_names,
                                    filter_mode="contains",
                                    key=f"payroll_history_email_recipients_{batch}"
                                )
                                st.caption(
                                    f"Có {len(hist_employee_names)} nhân viên có Thực nhận > 0 trong bản lương này. "
                                    "Email được lấy lại mới nhất từ Sheet1 khi bấm Gửi; nhân viên có Thực nhận ≤ 0 không được gửi mail."
                                )

                                if st.button(
                                    "🚀 Gửi bảng lương cho nhân viên đã chọn",
                                    use_container_width=True,
                                    key=f"send_payroll_history_employees_{batch}"
                                ):
                                    if not hist_selected_email_emps:
                                        st.warning("Vui lòng chọn ít nhất 1 nhân viên.")
                                    else:
                                        sender_email, sender_pass = get_smtp_sender_credentials()
                                        progress_hist_email = st.progress(0)
                                        hist_ok_count, hist_errors = 0, []
                                        # V59: đọc Email MỚI NHẤT trực tiếp từ Sheet1 một lần cho cả lượt gửi lịch sử.
                                        hist_live_email_creds = load_credentials_fresh_for_email()
                                        # Một snapshot lịch vi phạm dùng chung cho toàn bộ email trong lần gửi.
                                        hist_email_leave_df = load_backup_sheet_data()
                                        for idx, emp in enumerate(hist_selected_email_emps):
                                            matched_emp = hist_emailable[
                                                hist_emailable['Tên Hệ thống'].astype(str) == str(emp)
                                            ]
                                            if matched_emp.empty:
                                                hist_errors.append(f"{emp}: Không tìm thấy dữ liệu bảng lương.")
                                            else:
                                                row = matched_emp.iloc[0].copy()
                                                _hist_live_row_df = apply_latest_profile_fields_to_payroll(
                                                    pd.DataFrame([row]), hist_live_email_creds, only_current_nhanvien=False
                                                )
                                                if not _hist_live_row_df.empty:
                                                    row = _hist_live_row_df.iloc[0].copy()
                                                to_email = str(row.get('Email', '')).strip() or latest_email_from_credentials(hist_live_email_creds, emp)
                                                if not to_email or '@' not in to_email:
                                                    hist_errors.append(f"{emp}: Email mới nhất trong Sheet1 không hợp lệ hoặc đang để trống.")
                                                else:
                                                    row['Email'] = to_email
                                                    emp_violations = get_employee_violation_details(emp, hs, he, hist_email_leave_df)
                                                    ok, msg = send_payroll_email(
                                                        sender_email, sender_pass, to_email, row, hs, he, emp_violations
                                                    )
                                                    if ok:
                                                        hist_ok_count += 1
                                                    else:
                                                        hist_errors.append(f"{emp}: {msg}")
                                            progress_hist_email.progress((idx + 1) / len(hist_selected_email_emps))
                                            time.sleep(0.35)

                                        if hist_ok_count:
                                            st.success(
                                                f"Đã gửi thành công {hist_ok_count}/{len(hist_selected_email_emps)} "
                                                "email bảng lương từ bản đang chỉnh sửa."
                                            )
                                        for err in hist_errors:
                                            st.error(err)

                            if st.session_state.current_role == "admin":
                                with st.expander("📨 GỬI BẢNG LƯƠNG TỔNG HỢP CHO LỄ TÂN (BẢN ĐANG CHỈNH SỬA)"):
                                    hist_letan_df = df_credentials.copy()
                                    if not hist_letan_df.empty and 'Phân quyền' in hist_letan_df.columns:
                                        hist_letan_df = hist_letan_df[
                                            hist_letan_df['Phân quyền'].astype(str).str.strip().str.lower().isin(['letan', 'quanly'])
                                        ].copy()
                                        if 'Tên nhân viên' in hist_letan_df.columns:
                                            hist_letan_df = hist_letan_df[
                                                ~hist_letan_df['Tên nhân viên'].astype(str).apply(normalize_login_name).isin({
                                                    'ten nhan vien', 'ten he thong', 'username', 'user name'
                                                })
                                            ].copy()

                                    if not hist_letan_df.empty and 'Tên nhân viên' in hist_letan_df.columns:
                                        hist_letan_df = hist_letan_df.assign(__sort=hist_letan_df['Tên nhân viên'].apply(normalize_login_name)).sort_values('__sort').drop(columns='__sort')
                                    if hist_letan_df.empty:
                                        st.info("Không có tài khoản Lễ tân trong hồ sơ hệ thống.")
                                    else:
                                        st.write("**Check đúng 1 Lễ tân để hệ thống lấy Email từ hồ sơ:**")
                                        hist_checked_letan = []
                                        for i, (_, lr) in enumerate(hist_letan_df.iterrows()):
                                            lname = str(lr.get('Tên nhân viên', '')).strip()
                                            if not lname:
                                                continue
                                            if st.checkbox(
                                                lname,
                                                key=f"payroll_history_letan_check_{batch}_{i}_{normalize_login_name(lname)}"
                                            ):
                                                hist_checked_letan.append(lname)

                                        if len(hist_checked_letan) > 1:
                                            st.warning("⚠️ Chỉ được check 1 Lễ tân cho mỗi lần gửi.")
                                        elif len(hist_checked_letan) == 1:
                                            hist_selected_letan = hist_checked_letan[0]
                                            hist_matched_letan = hist_letan_df[
                                                hist_letan_df['Tên nhân viên'].astype(str).apply(normalize_login_name)
                                                == normalize_login_name(hist_selected_letan)
                                            ]
                                            if hist_matched_letan.empty:
                                                st.error("Không tìm thấy hồ sơ Lễ tân đã chọn.")
                                            else:
                                                st.caption("📧 Email người nhận sẽ được kiểm tra lại trực tiếp từ Sheet1 ngay khi bấm Gửi.")
                                                if st.button(
                                                    "📤 Gửi bảng lương tổng hợp cho Lễ tân đã check",
                                                    use_container_width=True,
                                                    key=f"send_payroll_history_summary_letan_{batch}"
                                                ):
                                                    hist_live_letan_creds = load_credentials_fresh_for_email()
                                                    hist_letan_email = latest_email_from_credentials(hist_live_letan_creds, hist_selected_letan)
                                                    if not hist_letan_email or '@' not in hist_letan_email:
                                                        st.error(
                                                            f"⚠️ Tài khoản {hist_selected_letan} chưa có Email hợp lệ trong Sheet1 mới nhất."
                                                        )
                                                    else:
                                                        sender_email, sender_pass = get_smtp_sender_credentials()
                                                        ok, msg = send_payroll_summary_email(
                                                            sender_email, sender_pass, hist_letan_email,
                                                            hist_selected_letan, edited_saved_table, hs, he
                                                        )
                                                        if ok:
                                                            st.success(
                                                                f"✅ Đã gửi bảng lương tổng hợp của bản {batch} "
                                                                f"cho {hist_selected_letan} ({hist_letan_email})."
                                                            )
                                                        else:
                                                            st.error(msg)
                                        else:
                                            st.caption("Chưa chọn Lễ tân nhận bảng lương tổng hợp.")
elif selected_page == "🧭 Bảng tour":
    st.subheader("🧭 Bảng tour")

    c_refresh, _ = st.columns([2, 8])
    with c_refresh:
        if st.button("🔄 Làm mới Bảng tour", use_container_width=True):
            load_bang_tour_input.clear()
            st.rerun()

    df_tour, tour_err = load_bang_tour_input()
    if tour_err:
        st.error(tour_err)
    elif df_tour.empty:
        st.info("Không có dữ liệu trong sheet Input.")
    else:
        # V84.7: mở/làm mới Bảng tour chỉ để xem dữ liệu.
        # Không tự ghi phạt tại đây; Auto Update tự động chạy theo Scheduler 15:00 và 20:00.

        # Chỉ Admin/Lễ tân/Quản lý được xem Thống kê Bảng tour.
        if str(st.session_state.current_role).strip().lower() in {"admin", "letan", "quanly"}:
            # Bảng thống kê dùng dữ liệu GỐC vừa đọc, trước khi làm trống thời gian <= -15.
            tour_stats_df = calculate_bang_tour_stats(df_tour)
            with st.expander("📊 Thống kê Bảng tour", expanded=False):

                            def style_tour_stats_row(row):

                                if str(row.get("Chỉ số", "")).strip() == "Có thể lên tour":

                                    return ["background-color:#92D050;color:#000000;font-weight:700;"] * len(row)

                                return [""] * len(row)


                            tour_stats_styled = (

                                tour_stats_df.style

                                .apply(style_tour_stats_row, axis=1)

                                .set_table_styles([

                                    {

                                        "selector": "th",

                                        "props": [

                                            ("background-color", "#A1948C"),

                                            ("color", "#000000"),

                                            ("font-weight", "700"),

                                            ("text-align", "center"),

                                            ("white-space", "normal"),

                                            ("overflow-wrap", "anywhere"),

                                            ("word-break", "break-word"),

                                            ("line-height", "1.15"),

                                        ],

                                    }

                                ])

                            )

                            st.dataframe(

                                tour_stats_styled,

                                use_container_width=True,

                                hide_index=True,

                                height="content"

                            )


        # Sau khi lấy dữ liệu: sắp cột + định dạng các cột thời gian.
        # Thời gian còn lại <= -15 và Thời gian < -180 được làm trống trên giao diện.
        df_tour_display = prepare_bang_tour_display(df_tour)
        df_tour_display, _tour_widths = apply_table_layout_df(df_tour_display, "tour_main")

        # Auto-fit toàn bộ chiều cao: hiển thị đủ dòng, bỏ thanh cuộn dọc.
        status_col_display = _find_tour_col(df_tour_display, "Trạng thái")
        remain_col_display = _find_tour_col(df_tour_display, "Thời gian còn lại")
        tour_column_config = table_layout_column_config("tour_main", list(df_tour_display.columns))
        if status_col_display is not None:
            tour_column_config[status_col_display] = st.column_config.TextColumn(
                status_col_display, width=layout_width("tour_main", status_col_display, "medium")
            )
        if remain_col_display is not None:
            tour_column_config[remain_col_display] = st.column_config.TextColumn(
                remain_col_display, width=layout_width("tour_main", remain_col_display, "small")
            )

        st.dataframe(
            apply_table_visual_styler(style_bang_tour(df_tour_display), "tour_main", list(df_tour_display.columns)),
            use_container_width=True,
            hide_index=True,
            height="content",
            row_height=layout_row_height("tour_main"),
            column_config=tour_column_config
        )
        render_admin_quick_layout_default("tour_main", list(df_tour_display.columns), "tour_main_page")
        st.caption(
            "Màu dòng: Nghỉ phép = chữ mờ/nền trắng; Đi làm = chữ đen/nền trắng; "
            "≥15 phút = xanh; 0–<15 = vàng; -15–<0 = đỏ; ≤-15 làm trống thời gian; Break = cam."
        )

elif selected_page == "📅 Đăng ký nghỉ phép":
    is_admin_leave_registration = str(st.session_state.get("current_role", "")).strip().lower() == "admin"
    if is_admin_leave_registration:
        leave_registration_area = st.expander("➕ Đăng ký lịch nghỉ", expanded=False)
    else:
        leave_registration_area = st.container()

    with leave_registration_area:
        if not is_admin_leave_registration:
            st.subheader("➕ Đăng ký lịch nghỉ")
        all_users = get_leave_eligible_employee_names(df_credentials, df_nv_excel)
        _current_registration_role = str(st.session_state.get("current_role", "")).strip().lower()
        _registration_locked = is_registration_role_locked(_current_registration_role)
        if _registration_locked and not is_admin_leave_registration:
            st.error(
                f"🔒 Quyền đăng ký lịch nghỉ của vai trò "
                f"{REGISTRATION_LOCK_LABELS.get(_current_registration_role, _current_registration_role)} "
                "đang bị Admin tạm khóa."
            )
        else:
            if is_admin_letan:
                list_nv_input = ["-- Chọn nhân viên --"] + all_users
                chosen_dates = st.date_input("Chọn ngày nghỉ (Khoảng thời gian nếu là Phép năm):", value=(get_vn_today(), get_vn_today()), key="sb_chosen_date")
            else:
                list_nv_input = [st.session_state.current_user]
                emp_min_date, emp_max_date = employee_registration_window()
                chosen_dates = st.date_input(
                    "Chọn ngày nghỉ (Nhân viên chọn 1 ngày):",
                    get_vn_today(),
                    min_value=emp_min_date,
                    max_value=emp_max_date,
                    key="sb_chosen_date"
                )
                st.caption(f"Nhân viên được đăng ký từ {emp_min_date.strftime('%d/%m/%Y')} đến hết {emp_max_date.strftime('%d/%m/%Y')}.")

            if isinstance(chosen_dates, tuple):
                if len(chosen_dates) == 2: start_date, end_date = chosen_dates
                elif len(chosen_dates) == 1: start_date = end_date = chosen_dates[0]
                else: start_date = end_date = get_vn_today()
            else:
                start_date = end_date = chosen_dates

            chosen_nv = st.selectbox("Chọn nhân viên:", list_nv_input, key="sb_chosen_nv", filter_mode="contains")

            # --- BỘ LỌC ĐỘNG CHO LÝ DO NGHỈ ---
            list_loai_nghi = []
            loai_nghi_dict = {}
            current_role = st.session_state.current_role.lower()
            role_for_leave_rules = "letan" if current_role == "quanly" else current_role

            if not df_loai_nghi.empty:
                for idx, row in df_loai_nghi.iterrows():
                    row_vals = row.tolist()
                    l_name = str(row_vals[1]).strip() if len(row_vals) > 1 else ""
                    if not l_name or l_name.lower() in ["nan", "none"]:
                        l_name = str(row.get('Lý do nghỉ', row.get('Loại nghỉ', ''))).strip()

                    if l_name and l_name.lower() not in ["nan", "loại nghỉ", "lý do nghỉ", "none", ""]:
                        dk_ngay = str(row_vals[6]).strip().lower() if len(row_vals) > 6 else ""
                        dk_role = str(row_vals[7]).strip().lower() if len(row_vals) > 7 else ""

                        role_allowed = True
                        if dk_role and dk_role not in ["nan", "none", "tất cả", "all", ""]:
                            if role_for_leave_rules not in dk_role: role_allowed = False

                        day_allowed = True
                        special_day_exempt = is_special_day_rule_exempt(current_role, l_name)
                        if (not special_day_exempt) and dk_ngay and dk_ngay not in ["nan", "none", "tất cả", "all", ""]:
                            wd = start_date.weekday()
                            wd_map = {
                                0: ["hai", "t2"], 1: ["ba", "t3"], 2: ["tư", "tu", "t4"],
                                3: ["năm", "nam", "t5"], 4: ["sáu", "sau", "t6"],
                                5: ["bảy", "bẩy", "t7", "cuối tuần"], 6: ["chủ nhật", "chu nhat", "cn", "cuối tuần"]
                            }
                            day_allowed = any(k in dk_ngay for k in wd_map[wd])

                        if day_allowed and role_allowed:
                            if "không phép" in l_name.lower(): l_name = f"🔴 {l_name}"
                            list_loai_nghi.append(l_name)
                            try:
                                s_ngay_str = str(row_vals[4]).replace(',', '').strip() if len(row_vals) > 4 else ""
                                s_ngay = float(s_ngay_str) if s_ngay_str != "" else 0.0
                            except: s_ngay = 0.0

                            try:
                                p_str = str(row_vals[5] if len(row_vals)>5 else "0").replace('.', '').replace(',', '').replace(' ', '').replace('đ', '').replace('VNĐ', '').replace('VND', '')
                                p_val = 0.0 if p_str.lower() in ["", "-", "nan", "none"] else float(p_str)
                            except: p_val = 0.0

                            # V86.9: lưu thêm cột C = Loại nghỉ để xác định trường hợp
                            # Có phép nhưng Số ngày tính = 0 (không chiếm suất nghỉ trong ngày).
                            loai_type = str(row_vals[2]).strip() if len(row_vals) > 2 else ""
                            loai_nghi_dict[l_name.lower()] = [s_ngay, p_val, loai_type]

            if not list_loai_nghi:
                list_loai_nghi = ["Nghỉ phép", "🔴 Nghỉ không phép", "Nghỉ phát sinh", "🔴 Đi trễ không phép", "🔴 Về sớm không phép"]
                loai_nghi_dict = {l.lower(): [0.0, 0.0, ""] for l in list_loai_nghi}

            chosen_loai = st.selectbox("Lý do nghỉ:", ["-- Chọn lý do nghỉ --"] + list_loai_nghi, key="sb_loai_nghi_live", filter_mode="contains")

            default_songay = 0.0
            default_phat = 0.0
            chosen_leave_type = ""
            if chosen_loai and chosen_loai != "-- Chọn lý do nghỉ --" and chosen_loai.lower() in loai_nghi_dict:
                default_songay = loai_nghi_dict[chosen_loai.lower()][0]
                default_phat = loai_nghi_dict[chosen_loai.lower()][1]
                chosen_leave_type = loai_nghi_dict[chosen_loai.lower()][2] if len(loai_nghi_dict[chosen_loai.lower()]) > 2 else ""

            # V86.9:
            # Nếu danh mục LoaiNghi có Cột C = "Có phép" và Số ngày tính = 0,
            # lý do này KHÔNG chiếm suất người nghỉ trong ngày.
            # Vì vậy không cảnh báo và không chặn bởi giới hạn 5 người ngày thường / 3 người cuối tuần.
            is_zero_day_co_phep = (
                normalize_login_name(chosen_leave_type) == "co phep"
                and abs(float(default_songay or 0)) < 1e-9
            )

            is_loi_vi_pham = "lỗi vi phạm khác" in chosen_loai.lower() if chosen_loai else False
            is_nghi_ly_do_khac = "nghỉ lý do khác" in chosen_loai.lower() if chosen_loai else False
            if is_loi_vi_pham: default_songay = 0.0

            # --- CẢNH BÁO SỚM SỐ NGƯỜI NGHỈ ---
            early_warning = ""
            norm_loai_temp = chosen_loai.strip().lower() if chosen_loai else ""
            is_video_leave_temp = is_video_leave_reason(chosen_loai)
            is_special_day_exempt_temp = is_special_day_rule_exempt(current_role, chosen_loai)
            if chosen_loai and chosen_loai != "-- Chọn lý do nghỉ --":
                num_days_temp = (end_date - start_date).days + 1
                if num_days_temp > 1 and "phép năm" not in norm_loai_temp:
                    early_warning = "❌ Chọn Khoảng thời gian nhiều ngày chỉ áp dụng cho 'Nghỉ Phép năm'."
                elif (not is_zero_day_co_phep
                      and not is_special_day_exempt_temp and not is_nghi_ly_do_khac and default_phat <= 0
                      and "phép năm" not in norm_loai_temp and not is_loi_vi_pham):
                    for i in range(num_days_temp):
                        chk_d = start_date + timedelta(days=i)
                        chk_is_we = chk_d.weekday() >= 5
                        if norm_loai_temp == "nghỉ phát sinh":
                            # Cảnh báo NGAY khi vừa chọn Nghỉ phát sinh, chưa cần bấm Lưu.
                            current_hour = datetime.now(VN_TZ).hour
                            if current_hour < 9 or current_hour >= 17:
                                early_warning = "❌ Khung giờ đăng ký 'Nghỉ phát sinh' chỉ cho phép từ 09:00 đến 17:00!"
                                break
                            if chk_is_we:
                                early_warning = f"❌ Ngày {chk_d.strftime('%d/%m/%Y')} là cuối tuần, không được phép 'Nghỉ phát sinh'!"
                                break
                            c_ps = len(df_lich[(df_lich['Ngày'] == chk_d) & (df_lich['Lý do nghỉ'].astype(str).str.strip().str.lower() == "nghỉ phát sinh")]) if not df_lich.empty else 0
                            if c_ps >= 2:
                                early_warning = f"❌ Ngày {chk_d.strftime('%d/%m/%Y')} đã đạt giới hạn 2 người 'Nghỉ phát sinh'!"
                                break
                            # Đồng thời kiểm tra hạn mức tổng số người nghỉ trong ngày.
                            m_ppl = 5 if not chk_is_we else 3
                            _quota_df = _leave_rows_counting_toward_quota(df_lich)
                            c_nghi = len(_quota_df[(_quota_df['Ngày'] == chk_d) & (_quota_df['Số ngày tính'] > 0)]) if not _quota_df.empty else 0
                            if c_nghi >= m_ppl:
                                early_warning = f"❌ Ngày {chk_d.strftime('%d/%m/%Y')} đã đủ hạn mức {m_ppl} người nghỉ trong ngày!"
                                break
                        else:
                            m_ppl = 5 if not chk_is_we else 3
                            _quota_df = _leave_rows_counting_toward_quota(df_lich)
                            c_nghi = len(_quota_df[(_quota_df['Ngày'] == chk_d) & (_quota_df['Số ngày tính'] > 0)]) if not _quota_df.empty else 0
                            if c_nghi >= m_ppl:
                                early_warning = f"❌ Ngày {chk_d.strftime('%d/%m/%Y')} đã đạt giới hạn {m_ppl} người nghỉ chung/ngày."
                                break

            if early_warning:
                st.error(early_warning)

            # Không hiện cảnh báo hạn mức người nghỉ cho Có phép + Số ngày tính = 0.
            # Trường hợp này vẫn được phép ghi dù ngày đã đủ 5/3 người.
            if is_zero_day_co_phep:
                st.caption("ℹ️ Lý do này là Có phép và Số ngày tính = 0 nên không tính vào giới hạn người nghỉ trong ngày.")

            # Kiểm tra lịch đã có từ CẢ HAI nguồn để chặn đăng ký trùng ngay trên giao diện.
            registration_all_df = combine_leave_sources_for_daily_stats(df_lich, df_leave_secondary, df_backup)
            existing_today = []
            if not registration_all_df.empty and chosen_nv != "-- Chọn nhân viên --":
                ex_df = registration_all_df[
                    (registration_all_df['Tên nhân viên'].astype(str).apply(normalize_login_name) == normalize_login_name(chosen_nv)) &
                    (registration_all_df['Ngày'] == start_date)
                ]
                existing_today = ex_df['Lý do nghỉ'].astype(str).str.strip().tolist()

            dyn_key_suffix = f"{chosen_loai}_{start_date}_{chosen_nv}"

            # Hiển thị trước thứ tự và mức cộng phạt cho 3 nhóm vi phạm lũy tiến.
            progressive_preview_reason = get_progressive_penalty_reason(chosen_loai)
            if progressive_preview_reason:
                preview_ordinal, preview_extra = _progressive_ordinal_and_bonus(
                    registration_all_df, start_date, chosen_loai
                )
                preview_total = float(default_phat) + float(preview_extra)
                st.warning(
                    f"⚠️ {progressive_preview_reason} ngày {start_date.strftime('%d/%m/%Y')}: Người Thứ {preview_ordinal}. "
                    f"Phạt theo quy định {float(default_phat):,.0f} VNĐ"
                    + (f" + lũy tiến {preview_extra:,.0f} VNĐ" if preview_extra > 0 else "")
                    + f" = {preview_total:,.0f} VNĐ."
                )

            st.info(
                "📌 Quy tắc trong 1 ngày / 1 nhân viên: "
                "chỉ được có tối đa 1 dòng có Số ngày tính 0.5 hoặc 1 "
                "(không cho 0.5 + 0.5); đồng thời không được có 2 lần cùng nhóm "
                "CÓ phép, KHÔNG phép hoặc PHÁT SINH."
            )

            with st.form("form_nhap_lich_inner"):
                txt_chitiet_label = "Chi tiết vi phạm / Ghi chú (🔴 **Bắt buộc**):" if (is_loi_vi_pham or is_nghi_ly_do_khac) else "Chi tiết vi phạm / Ghi chú (nếu có):"
                input_chitiet = st.text_input(txt_chitiet_label).strip()

                col_p1, col_p2 = st.columns(2)
                with col_p1:
                    if is_admin_leave_registration:
                        val_songay = st.number_input(
                            "Số ngày tính:",
                            value=float(default_songay),
                            step=0.5,
                            key=f"num_songay_{dyn_key_suffix}",
                            disabled=False,
                            help="Admin toàn quyền: không áp dụng giới hạn Số ngày tính.",
                        )
                    else:
                        val_songay = st.number_input(
                            "Số ngày tính:",
                            min_value=0.0,
                            max_value=1.0,
                            value=min(1.0, max(0.0, float(default_songay))),
                            step=0.5,
                            key=f"num_songay_{dyn_key_suffix}",
                            disabled=(is_loi_vi_pham or is_zero_day_co_phep),
                        help=(
                            "Mỗi dòng chỉ được 0, 0.5 hoặc 1 ngày. Trong cùng ngày không được có thêm "
                            "một dòng Số ngày tính > 0 khác."
                            if not is_zero_day_co_phep
                            else "Theo cấu hình LoaiNghi: Có phép + Số ngày tính = 0 nên không chiếm suất nghỉ trong ngày."
                        ),
                    )

                # HIỂN THỊ Ô MỨC PHẠT CHO TẤT CẢ TÀI KHOẢN (ĐÃ MỞ LẠI)
                with col_p2:
                    txt_phat_label = "Mức phạt vi phạm VNĐ (🔴 **Bắt buộc**):" if is_loi_vi_pham else "Mức phạt vi phạm (VNĐ):"
                    val_phat = st.number_input(
                        txt_phat_label,
                        value=float(default_phat),
                        step=50000.0,
                        key=f"num_phat_{dyn_key_suffix}",
                        disabled=is_progressive_penalty_reason(chosen_loai)
                    )

                confirm_multiple = True
                if existing_today:
                    normalized_existing = {normalize_leave_reason(x) for x in existing_today}
                    if is_admin_leave_registration:
                        st.info(
                            f"🛡️ Admin toàn quyền: {chosen_nv} đã có lịch trong ngày "
                            f"{start_date.strftime('%d/%m/%Y')}: {', '.join(existing_today)}. "
                            "Admin vẫn được phép ghi thêm."
                        )
                    elif normalize_leave_reason(chosen_loai) in normalized_existing:
                        st.error(f"❌ Nhân viên này đã có Lý do nghỉ: '{clean_leave_reason_display(chosen_loai)}' vào ngày này rồi. KHÔNG THỂ trùng cùng 1 loại nghỉ!")
                        confirm_multiple = False
                    else:
                        st.warning(f"⚠️ CẢNH BÁO: Nhân viên '{chosen_nv}' đã có các lịch sau trong ngày {start_date.strftime('%d/%m/%Y')}: {', '.join(existing_today)}")
                        confirm_multiple = st.checkbox("Tôi xác nhận đăng ký này là ĐÚNG và MỚI.")

                submit_lich = st.form_submit_button("💾 Xác Nhận Ghi Lịch Nghỉ")

                if submit_lich:
                    today = get_vn_today()
                    can_proceed = True

                    if current_role in EMPLOYEE_LIKE_ROLES:
                        special_emp_reason = (
                            is_video_leave_reason(chosen_loai)
                            or (current_role == "leader" and is_leader_policy_leave_reason(chosen_loai))
                        )
                        if not special_emp_reason:
                            emp_min_date, emp_max_date = employee_registration_window(today)
                            if start_date < emp_min_date or end_date > emp_max_date:
                                st.error(f"❌ Tài khoản NHÂN VIÊN chỉ được đăng ký từ hôm nay đến hết ngày {emp_max_date.strftime('%d/%m/%Y')} (tháng hiện tại và 1 tháng kế tiếp).")
                                can_proceed = False
                    elif (current_role in ["letan", "quanly"] and start_date < today
                          and not (is_video_leave_reason(chosen_loai) or is_bereavement_leave_reason(chosen_loai))):
                        st.error("❌ Lỗi: Tài khoản LỄ TÂN/QUẢN LÝ không được đăng ký lịch trong **QUÁ KHỨ**. Muốn sửa lịch cũ, vui lòng liên hệ Admin.")
                        can_proceed = False

                    if can_proceed:
                        if not confirm_multiple:
                            st.error("❌ Vui lòng tick Xác nhận cảnh báo bên trên trước khi lưu.")
                        elif chosen_nv == "-- Chọn nhân viên --" or not chosen_nv:
                            st.error("❌ Vui lòng chọn nhân viên cần nhập lịch nghỉ!")
                        elif chosen_loai == "-- Chọn lý do nghỉ --" or not chosen_loai:
                            st.error("❌ Vui lòng chọn lý do nghỉ!")
                        elif early_warning and not is_admin_leave_registration:
                            st.error(f"❌ Không thể lưu: {early_warning}")
                        else:
                            norm_loai = normalize_leave_reason(chosen_loai)
                            is_video_leave = is_video_leave_reason(chosen_loai)
                            num_days_selected = (end_date - start_date).days + 1

                            if is_loi_vi_pham:
                                val_songay = 0.0 
                                if not input_chitiet:
                                    st.error("❌ Bắt buộc nhập Chi tiết vi phạm / Ghi chú đối với 'Lỗi vi phạm khác'.")
                                    can_proceed = False
                                if val_phat <= 0 and st.session_state.current_role == "admin":
                                    st.error("❌ Bắt buộc nhập số tiền Phạt vi phạm > 0 đối với 'Lỗi vi phạm khác'.")
                                    can_proceed = False

                            if is_nghi_ly_do_khac and not input_chitiet:
                                st.error("❌ Bắt buộc nhập Chi tiết vi phạm / Ghi chú đối với 'Nghỉ lý do khác'.")
                                can_proceed = False

                            # KIỂM TRA GIỚI HẠN NHÂN SỰ CÁ NHÂN
                            # V86.13: Admin đi qua khối để lưu, nhưng bỏ qua toàn bộ phép/giới hạn bên dưới.
                            if can_proceed:
                                nv_info = df_credentials[df_credentials['Tên nhân viên'].str.lower() == chosen_nv.lower()]
                                limit_ps = pd.to_numeric(nv_info.iloc[0].get('Phát sinh tháng', 0), errors='coerce') if not nv_info.empty else 0
                                limit_cp = pd.to_numeric(nv_info.iloc[0].get('Có phép tháng', 0), errors='coerce') if not nv_info.empty else 0
                                limit_pn = pd.to_numeric(nv_info.iloc[0].get('Phép năm', 0), errors='coerce') if not nv_info.empty else 0

                                if pd.isna(limit_ps): limit_ps = 0
                                if pd.isna(limit_cp): limit_cp = 0
                                if pd.isna(limit_pn): limit_pn = 0

                                user_hist = df_lich[df_lich['Tên nhân viên'] == chosen_nv] if not df_lich.empty else pd.DataFrame(columns=['Ngày', 'Lý do nghỉ', 'Số ngày tính'])
                                # Nghỉ phép quay video không tiêu hao bất kỳ hạn mức nghỉ nào.
                                user_hist_quota = _leave_rows_counting_toward_quota(user_hist)
                                user_hist['Ngày_DT'] = pd.to_datetime(user_hist['Ngày'], errors='coerce')
                                user_hist_quota['Ngày_DT'] = pd.to_datetime(user_hist_quota['Ngày'], errors='coerce')
                                user_hist['M'] = user_hist['Ngày_DT'].dt.month
                                user_hist['Y'] = user_hist['Ngày_DT'].dt.year
                                user_hist_quota['M'] = user_hist_quota['Ngày_DT'].dt.month
                                user_hist_quota['Y'] = user_hist_quota['Ngày_DT'].dt.year

                                curr_m = start_date.month
                                curr_y = start_date.year

                                total_phep_required = val_songay * num_days_selected
                                accumulated_month = user_hist_quota[(user_hist_quota['M'] == curr_m) & (user_hist_quota['Y'] == curr_y)]['Số ngày tính'].sum()

                                if is_admin_leave_registration or is_video_leave:
                                    # Admin luôn miễn mọi giới hạn; Nghỉ quay video cũng miễn hạn mức cá nhân.
                                    pass
                                elif "phép năm" in norm_loai:
                                    used_pn = user_hist_quota[(user_hist_quota['Y'] == curr_y) & (user_hist_quota['Lý do nghỉ'].str.lower().str.contains("phép năm", na=False))]['Số ngày tính'].sum()
                                    if limit_pn > 0 and (used_pn + total_phep_required > limit_pn):
                                        st.error(f"❌ Vượt quá số ngày Phép năm! Bạn cần {total_phep_required} ngày nhưng quỹ phép chỉ còn {limit_pn - used_pn} ngày trong năm {curr_y}.")
                                        can_proceed = False

                                elif "phát sinh" in norm_loai:
                                    used_ps = len(user_hist_quota[(user_hist['M'] == curr_m) & (user_hist['Y'] == curr_y) & (user_hist['Lý do nghỉ'].str.lower().str.contains("phát sinh", na=False))])
                                    if limit_ps > 0 and (used_ps >= limit_ps):
                                        st.error(f"❌ Vượt giới hạn Phát sinh! Nhân viên này chỉ được đăng ký {limit_ps} lần phát sinh/tháng.")
                                        can_proceed = False

                                elif not is_nghi_ly_do_khac and "không phép" not in norm_loai and val_songay > 0:
                                    used_cp = user_hist_quota[(user_hist['M'] == curr_m) & (user_hist['Y'] == curr_y) & (~user_hist['Lý do nghỉ'].str.lower().str.contains("không phép|phát sinh|lý do khác", na=False, regex=True))]['Số ngày tính'].sum()
                                    if limit_cp > 0 and (used_cp + total_phep_required > limit_cp):
                                        st.error(f"❌ Vượt số ngày Có phép trong tháng! Nhân viên này chỉ được nghỉ tối đa {limit_cp} ngày/tháng.")
                                        can_proceed = False

                                if can_proceed:
                                    all_saved = True
                                    save_success_notes = []
                                    for i in range(num_days_selected):
                                        curr_date_iter = start_date + timedelta(days=i)
                                        is_weekend_iter = curr_date_iter.weekday() >= 5

                                        if val_songay is None:
                                            val_songay = 0.0
                                        if not is_video_leave:
                                            accumulated_month += val_songay

                                        # V86.11: ngoài chống trùng đúng Lý do, còn áp quy tắc:
                                        # - chỉ 1 dòng có Số ngày tính > 0 mỗi nhân viên/ngày;
                                        # - không 0.5 + 0.5;
                                        # - không 2 lần CÓ phép / KHÔNG phép / PHÁT SINH.
                                        latest_registration_df = combine_leave_sources_for_daily_stats(
                                            df_lich, df_leave_secondary, df_backup
                                        )

                                        if not is_admin_leave_registration:
                                            daily_rule_ok, daily_rule_msg = _validate_daily_employee_registration_rule(
                                                latest_registration_df,
                                                curr_date_iter,
                                                chosen_nv,
                                                chosen_loai,
                                                val_songay,
                                            )
                                            if not daily_rule_ok:
                                                st.error(
                                                    f"❌ {chosen_nv} · {curr_date_iter.strftime('%d/%m/%Y')}: "
                                                    f"{daily_rule_msg}"
                                                )
                                                all_saved = False
                                                break

                                            if _leave_exists_in_sources(latest_registration_df, curr_date_iter, chosen_nv, chosen_loai):
                                                st.error(
                                                    f"❌ {chosen_nv} đã có đúng lý do '{clean_leave_reason_display(chosen_loai)}' ngày "
                                                    f"{curr_date_iter.strftime('%d/%m/%Y')}. Hãy chọn lý do khác nếu cần ghi thêm vi phạm."
                                                )
                                                all_saved = False
                                                break

                                        special_day_exempt_save = is_special_day_rule_exempt(current_role, chosen_loai)
                                        if (not is_admin_leave_registration
                                            and not is_zero_day_co_phep
                                            and not special_day_exempt_save and not is_nghi_ly_do_khac and val_phat <= 0
                                            and "phép năm" not in norm_loai and not is_loi_vi_pham):
                                            if norm_loai == "nghỉ phát sinh":
                                                current_hour = datetime.now(VN_TZ).hour
                                                if current_hour < 9 or current_hour >= 17:
                                                    st.error("❌ Khung giờ đăng ký 'Nghỉ phát sinh' chỉ cho phép từ 09:00 đến 17:00!")
                                                    all_saved = False
                                                    break
                                                elif is_weekend_iter:
                                                    st.error(f"❌ Ngày {curr_date_iter.strftime('%d/%m/%Y')} là cuối tuần, không được phép 'Nghỉ phát sinh'!")
                                                    all_saved = False
                                                    break
                                                else:
                                                    count_ps = len(df_lich[(df_lich['Ngày'] == curr_date_iter) & (df_lich['Lý do nghỉ'].astype(str).str.strip().str.lower() == "nghỉ phát sinh")]) if not df_lich.empty else 0
                                                    if count_ps >= 2:
                                                        st.error(f"❌ Ngày {curr_date_iter.strftime('%d/%m/%Y')} đã đạt giới hạn 2 người 'Nghỉ phát sinh'!")
                                                        all_saved = False
                                                        break
                                            else:
                                                max_people = 5 if not is_weekend_iter else 3
                                                _quota_df = _leave_rows_counting_toward_quota(df_lich)
                                                today_total_nghi = len(_quota_df[(_quota_df['Ngày'] == curr_date_iter) & (_quota_df['Số ngày tính'] > 0)]) if not _quota_df.empty else 0
                                                if today_total_nghi >= max_people:
                                                    st.error(f"❌ Ngày {curr_date_iter.strftime('%d/%m/%Y')} đã đạt giới hạn {max_people} người nghỉ chung/ngày.")
                                                    all_saved = False
                                                    break

                                        # GỌI HÀM LƯU LÊN GOOGLE SHEETS
                                        penalty_to_save = float(default_phat) if is_progressive_penalty_reason(chosen_loai) else float(val_phat)
                                        success_bk, msg_bk = save_lich_nghi_to_backup_sheet(
                                            curr_date_iter.strftime('%d/%m/%Y'), chosen_nv, clean_leave_reason_display(chosen_loai),
                                            input_chitiet, val_songay, accumulated_month, penalty_to_save, st.session_state.current_user,
                                            df_main_source=df_lich
                                        )

                                        if not success_bk:
                                            st.error(f"❌ LỖI GOOGLE SHEETS: {msg_bk}")
                                            all_saved = False
                                            break
                                        if msg_bk:
                                            save_success_notes.append(msg_bk)

                                    # CHỈ IN THÀNH CÔNG NẾU API THỰC SỰ TRẢ VỀ SUCCESS
                                    if all_saved:
                                        st.success(f"✅ Đã ghi nhận lịch nghỉ thành công cho {num_days_selected} ngày!")
                                        for note in save_success_notes:
                                            if "Người Thứ" in note:
                                                st.info(note)
                                        _clear_dynamic_data_caches()



    st.markdown("---")

    # Bộ lọc thời gian & nhân viên
    col_date, col_name, col_refresh = st.columns([5, 4, 2])

    with col_date:
        render_leave_filter_label_css()
        today = get_vn_today() 
        col_d1, col_d2 = st.columns(2)
        with col_d1:
            filter_type = st.selectbox(
                "Lọc thời gian:", 
                ["Hôm nay", "Hôm qua", "Ngày mai", "Chọn ngày", "Khoảng thời gian", "Tuần này", "Tuần trước", "Tuần sau", "Tháng này", "Tháng trước", "Tháng sau"],
                index=0, filter_mode="contains", key="leave_stats_time_filter"
            )
        with col_d2:
            if filter_type == "Hôm nay": start_date = end_date = today
            elif filter_type == "Hôm qua": start_date = end_date = today - timedelta(days=1)
            elif filter_type == "Ngày mai": start_date = end_date = today + timedelta(days=1)
            elif filter_type == "Tuần này":
                start_date = today - timedelta(days=today.weekday())
                end_date = start_date + timedelta(days=6)
            elif filter_type == "Tuần trước":
                start_date = today - timedelta(days=today.weekday() + 7)
                end_date = start_date + timedelta(days=6)
            elif filter_type == "Tuần sau":
                start_date = today - timedelta(days=today.weekday()) + timedelta(days=7)
                end_date = start_date + timedelta(days=6)
            elif filter_type == "Tháng này":
                start_date = today.replace(day=1)
                end_date = today.replace(day=calendar.monthrange(today.year, today.month)[1])
            elif filter_type == "Tháng trước":
                end_date = today.replace(day=1) - timedelta(days=1)
                start_date = end_date.replace(day=1)
            elif filter_type == "Tháng sau":
                start_date = today.replace(year=today.year + 1, month=1, day=1) if today.month == 12 else today.replace(month=today.month + 1, day=1)
                end_date = start_date.replace(day=calendar.monthrange(start_date.year, start_date.month)[1])
            elif filter_type == "Chọn ngày":
                start_date = end_date = st.date_input("Chọn ngày:", today)
            elif filter_type == "Khoảng thời gian":
                date_range = st.date_input("Chọn khoảng thời gian:", [today, today])
                start_date, end_date = (date_range[0], date_range[1]) if len(date_range) == 2 else (date_range[0], date_range[0])
            else: start_date = end_date = today

    with col_name:
        list_nv = ["- Tất cả nhân viên -"] + get_leave_eligible_employee_names(df_credentials, df_nv_excel)
        selected_nv = st.selectbox("👤 Tìm kiếm nhân viên:", list_nv, filter_mode="contains")

    with col_refresh:
        st.write("") 
        if st.button("🔄 Cập Nhật Dữ Liệu", use_container_width=True):
            _clear_dynamic_data_caches()
            st.rerun()

    # Lọc dữ liệu: phần thống kê/Chi tiết danh sách dùng ĐÚNG 2 Google Sheet:
    # 1) SHEET_DU_PHONG_ID (nơi nhập liệu hiện tại)
    # 2) SHEET_LICH_NGHI_2_ID
    # Nếu trùng Ngày + Tên nhân viên + Lý do nghỉ thì ưu tiên Sheet dự phòng.
    detail_all_df = combine_leave_sources_for_daily_stats(df_leave_secondary, df_backup)
    # Nếu cache nguồn từng trả về rỗng do lỗi API tạm thời, chủ động đọc lại đúng một lần
    # để phần Chi tiết danh sách không bị trắng dù Google Sheet đang có dữ liệu.
    if detail_all_df.empty:
        try:
            load_backup_sheet_data.clear(); load_secondary_leave_sheet_data.clear()
            df_backup = load_backup_sheet_data()
            df_leave_secondary = load_secondary_leave_sheet_data()
            detail_all_df = combine_leave_sources_for_daily_stats(df_leave_secondary, df_backup)
        except Exception:
            pass
    if not detail_all_df.empty:
        mask_date = (detail_all_df['Ngày'] >= start_date) & (detail_all_df['Ngày'] <= end_date)
        filtered_df = detail_all_df[mask_date].copy()
        if selected_nv != "- Tất cả nhân viên -":
            filtered_df = filtered_df[
                filtered_df['Tên nhân viên'].astype(str).str.strip().str.casefold() == selected_nv.strip().casefold()
            ]
    else:
        filtered_df = detail_all_df.copy()

    # --- THỐNG KÊ ---
    excluded_keywords = ["đi trễ", "di tre", "không dọn vệ sinh", "khong don ve sinh", "lỗi vi phạm", "loi vi pham", "qua tour", "xuống phòng", "xuong phong", "ra sớm", "ra som", "vào muộn", "vao muon", "đi tua", "di tua", "ngưng nhận", "ngung nhan", "hỗ trợ ca", "ho tro ca"]
    def is_excluded(r): return any(kw in str(r).lower() for kw in excluded_keywords)

    # Nguồn riêng cho "Thống kê chi tiết theo từng ngày": chỉ nhân sự Đang làm việc.
    daily_all_df = _filter_active_employees_for_leave_stats(detail_all_df.copy())
    if not daily_all_df.empty:
        daily_mask = (daily_all_df['Ngày'] >= start_date) & (daily_all_df['Ngày'] <= end_date)
        daily_filtered_df = daily_all_df[daily_mask].copy()
        if selected_nv != "- Tất cả nhân viên -":
            daily_filtered_df = daily_filtered_df[
                daily_filtered_df['Tên nhân viên'].astype(str).apply(normalize_login_name) == normalize_login_name(selected_nv)
            ]
    else:
        daily_filtered_df = daily_all_df.copy()

    daily_thuc_nghi = (
        daily_filtered_df[~daily_filtered_df['Lý do nghỉ'].apply(is_excluded)].copy()
        if not daily_filtered_df.empty else pd.DataFrame(columns=daily_all_df.columns)
    )

    stats_filtered_df = _filter_active_employees_for_leave_stats(filtered_df.copy())
    if stats_filtered_df.empty:
        df_thuc_nghi = phat_sinh_df = khong_phep_df = co_phep_df = pd.DataFrame(columns=stats_filtered_df.columns if hasattr(stats_filtered_df, 'columns') else [])
        tong_phat = 0.0
    else:
        df_thuc_nghi = stats_filtered_df[~stats_filtered_df['Lý do nghỉ'].apply(is_excluded)].copy()
        if df_thuc_nghi.empty:
            phat_sinh_df = khong_phep_df = co_phep_df = pd.DataFrame(columns=stats_filtered_df.columns)
        else:
            # V86.7: thống kê PHÂN NHÓM TRỰC TIẾP theo nội dung cột `Lý do nghỉ`.
            # - chứa "có phép"    -> CÓ phép
            # - chứa "phát sinh"  -> PHÁT SINH
            # - chứa "không phép" -> KHÔNG phép
            _reason_norm = df_thuc_nghi['Lý do nghỉ'].astype(str).apply(normalize_login_name)

            _is_phat_sinh = _reason_norm.str.contains('phat sinh', na=False)
            _is_khong_phep = _reason_norm.str.contains('khong phep', na=False)

            # V86.8: nhóm CÓ phép nếu Lý do nghỉ chứa một trong:
            # CÓ phép, CP, Nghỉ phép, Nghỉ đám hiếu.
            # Loại trừ rõ các dòng đã thuộc PHÁT SINH hoặc KHÔNG phép để tránh đếm chồng.
            _is_co_phep = (
                _reason_norm.str.contains('co phep', na=False)
                | _reason_norm.str.contains(r'(^|\s)cp($|\s)', na=False, regex=True)
                | _reason_norm.str.contains('nghi phep', na=False)
                | _reason_norm.str.contains('nghi dam hieu', na=False)
            ) & (~_is_phat_sinh) & (~_is_khong_phep)

            phat_sinh_df = df_thuc_nghi[_is_phat_sinh].copy()
            khong_phep_df = df_thuc_nghi[_is_khong_phep].copy()
            co_phep_df = df_thuc_nghi[_is_co_phep].copy()
        tong_phat = stats_filtered_df['Phạt vi phạm'].sum()

    # Chỉ tài khoản admin được xem Tổng tiền phạt.
    # Ẩn các metric Tổng số người nghỉ / CÓ phép / PHÁT SINH / KHÔNG phép.
    is_admin = str(st.session_state.get("current_role", "")).strip().lower() == "admin"
    if is_admin:
        st.write("")
        st.metric("💰 Tổng tiền phạt", f"{tong_phat:,.0f} đ".replace(",", "."))
        cols_to_hide = []
    else:
        cols_to_hide = ['Phạt vi phạm']

    st.markdown("### 📅 Thống kê chi tiết theo từng ngày")
    st.caption("Nhân viên/Leader: Nghỉ CÓ phép chỉ được Sửa/Hủy trước ít nhất 3 ngày; Nghỉ KHÔNG phép được Sửa/Hủy không giới hạn 3 ngày nhưng chỉ với ngày tương lai; Nghỉ phép quay video không chịu các giới hạn này.")
    if not daily_thuc_nghi.empty:
        daily_stats = []
        daily_limit_flags = []
        for d in sorted(daily_filtered_df['Ngày'].dropna().unique()):
            day_df = daily_filtered_df[daily_filtered_df['Ngày'] == d]
            day_thuc_nghi = day_df[~day_df['Lý do nghỉ'].apply(is_excluded)]
            # V86.7: đếm theo nội dung cột `Lý do nghỉ` đúng như quy định nghiệp vụ.
            d_reason = day_thuc_nghi['Lý do nghỉ'].astype(str).apply(normalize_login_name)

            d_is_phat_sinh = d_reason.str.contains('phat sinh', na=False)
            d_is_khong_phep = d_reason.str.contains('khong phep', na=False)
            d_is_co_phep = (
                d_reason.str.contains('co phep', na=False)
                | d_reason.str.contains(r'(^|\s)cp($|\s)', na=False, regex=True)
                | d_reason.str.contains('nghi phep', na=False)
                | d_reason.str.contains('nghi dam hieu', na=False)
            ) & (~d_is_phat_sinh) & (~d_is_khong_phep)

            count_co_phep = int(d_is_co_phep.sum())
            count_phat_sinh = int(d_is_phat_sinh.sum())
            count_khong_phep = int(d_is_khong_phep.sum())
            is_weekend = d.weekday() >= 5
            max_people = 3 if is_weekend else 5
            day_quota_df = _leave_rows_counting_toward_quota(day_df)
            total_count_for_limit = len(day_quota_df[pd.to_numeric(day_quota_df['Số ngày tính'], errors='coerce').fillna(0) > 0])

            stat_row = {
                "Ngày": d.strftime('%d/%m/%Y'),
                "Thứ ngày": _vn_weekday_label(d),
                "Tổng nghỉ": int(count_co_phep + count_phat_sinh + count_khong_phep),
                "Tổng số người nghỉ": len(day_thuc_nghi),
                "✅ CÓ phép": count_co_phep,
                "⚠️ PHÁT SINH": count_phat_sinh,
                "❌ KHÔNG phép": count_khong_phep
            }
            if st.session_state.current_role == "admin":
                stat_row["💰 Tổng tiền phạt"] = f"{pd.to_numeric(day_df['Phạt vi phạm'], errors='coerce').fillna(0).sum():,.0f} đ".replace(",", ".")

            daily_stats.append(stat_row)
            daily_limit_flags.append({
                'co_phep_full': total_count_for_limit >= max_people,
                'phat_sinh_full': (count_phat_sinh >= 2) or (is_weekend and count_phat_sinh > 0)
            })

        daily_stats_df = pd.DataFrame(daily_stats)

        def highlight_daily_limits(row):
            styles = [''] * len(row)
            flags = daily_limit_flags[row.name]
            for idx, col in enumerate(row.index):
                if col == '✅ CÓ phép' and flags['co_phep_full']:
                    styles[idx] = 'background-color: #ffcdd2; color: #b71c1c; font-weight: 700;'
                elif col == '⚠️ PHÁT SINH' and flags['phat_sinh_full']:
                    styles[idx] = 'background-color: #ffcdd2; color: #b71c1c; font-weight: 700;'
            return styles

        st.caption("💡 CÓ phép = Lý do nghỉ chứa CÓ phép / CP / Nghỉ phép / Nghỉ đám hiếu. PHÁT SINH và KHÔNG phép vẫn phân theo từ khóa tương ứng. Bấm vào số để xem chi tiết.")
        _stat_widths = [1.35, 0.9, 1.0, 1.15, 1.15, 1.15]
        if st.session_state.current_role == "admin":
            _stat_widths.append(1.25)

        _heads = st.columns(_stat_widths)
        _titles = ["Ngày", "Thứ ngày", "Tổng nghỉ", "✅ CÓ phép", "⚠️ PHÁT SINH", "❌ KHÔNG phép"]
        if st.session_state.current_role == "admin":
            _titles.append("💰 Tổng tiền phạt")
        for _hc, _title in zip(_heads, _titles):
            _hc.markdown(f"**{_title}**")

        for _idx, _stat in daily_stats_df.iterrows():
            _day_label = str(_stat.get("Ngày", ""))
            _day_obj = _parse_vn_date(_day_label)
            _day_df = daily_filtered_df[daily_filtered_df["Ngày"] == _day_obj].copy() if _day_obj is not None else pd.DataFrame()
            _day_thuc = _day_df[~_day_df["Lý do nghỉ"].apply(is_excluded)].copy() if not _day_df.empty else pd.DataFrame()
            if not _day_thuc.empty:
                _rn = _day_thuc["Lý do nghỉ"].astype(str).apply(normalize_login_name)

                _rn_ps = _rn.str.contains('phat sinh', na=False)
                _rn_kp = _rn.str.contains('khong phep', na=False)
                _rn_cp = (
                    _rn.str.contains('co phep', na=False)
                    | _rn.str.contains(r'(^|\s)cp($|\s)', na=False, regex=True)
                    | _rn.str.contains('nghi phep', na=False)
                    | _rn.str.contains('nghi dam hieu', na=False)
                ) & (~_rn_ps) & (~_rn_kp)

                _co_rows = _day_thuc[_rn_cp].copy()
                _ps_rows = _day_thuc[_rn_ps].copy()
                _kp_rows = _day_thuc[_rn_kp].copy()
            else:
                _co_rows = _ps_rows = _kp_rows = pd.DataFrame(columns=daily_filtered_df.columns)

            _rc = st.columns(_stat_widths)
            _rc[0].markdown(_day_label)
            _rc[1].markdown(str(_stat.get("Thứ ngày", "")))

            _total_n = int(_stat.get("Tổng nghỉ", 0) or 0)
            _co_n = int(_stat.get("✅ CÓ phép", 0) or 0)
            _ps_n = int(_stat.get("⚠️ PHÁT SINH", 0) or 0)
            _kp_n = int(_stat.get("❌ KHÔNG phép", 0) or 0)

            _rc[2].markdown(f"**{_total_n}**")
            if _rc[3].button(str(_co_n), key=f"daily_co_{_idx}_{_day_label}", use_container_width=True, disabled=_co_n <= 0):
                show_daily_leave_popup(_day_label, "CÓ phép", _co_rows)
            if _rc[4].button(str(_ps_n), key=f"daily_ps_{_idx}_{_day_label}", use_container_width=True, disabled=_ps_n <= 0):
                show_daily_leave_popup(_day_label, "PHÁT SINH", _ps_rows)
            if _rc[5].button(str(_kp_n), key=f"daily_kp_{_idx}_{_day_label}", use_container_width=True, disabled=_kp_n <= 0):
                show_daily_leave_popup(_day_label, "KHÔNG phép", _kp_rows)
            if st.session_state.current_role == "admin":
                _rc[6].markdown(str(_stat.get("💰 Tổng tiền phạt", "")))

    else:
        st.info("Không có dữ liệu báo nghỉ trong khoảng thời gian đã chọn ở cả hai nguồn.")

    st.markdown("---")

    export_source_df = filtered_df.drop(columns=cols_to_hide + ['__source_sheet_id', '__source_row'], errors='ignore').copy()
    export_source_df = add_source_leave_type_column(export_source_df)
    export_df = format_display_df(export_source_df)
    df_for_excel = export_df.copy()
    if st.session_state.current_role == "admin" and not df_for_excel.empty:
        tong_cong_row = pd.Series(index=df_for_excel.columns, dtype=object)
        tong_cong_row['Tên nhân viên'] = "TỔNG TIỀN PHẠT:"
        tong_cong_row['Phạt vi phạm'] = tong_phat
        df_for_excel = pd.concat([df_for_excel, tong_cong_row.to_frame().T], ignore_index=True)

    if st.session_state.current_role in EMPLOYEE_LIKE_ROLES:
        st.subheader(f"Chi tiết danh sách (Từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')})")
    else:
        col_header, col_download = st.columns([7, 3])
        with col_header:
            st.subheader(f"Chi tiết danh sách (Từ {start_date.strftime('%d/%m/%Y')} đến {end_date.strftime('%d/%m/%Y')})")
        with col_download:
            st.write("")
            if not export_df.empty:
                st.download_button(
                    "📥 Tải Dữ Liệu Lọc Xuống (Excel)",
                    data=to_excel(df_for_excel),
                    file_name=f"Vera-Spa_{start_date.strftime('%d%m%Y')}_to_{end_date.strftime('%d%m%Y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            else:
                st.button("📥 Tải Dữ Liệu Lọc Xuống (Excel)", disabled=True, use_container_width=True)

    # --- KHU VỰC CHỈ DÀNH CHO ADMIN: GỬI EMAIL BÁO CÁO ---
    if st.session_state.current_role == "admin" and not filtered_df.empty:
        with st.expander("📧 GỬI BÁO CÁO QUA EMAIL CHO NHÂN VIÊN"):
            st.info("Hệ thống sẽ tự động tách dữ liệu của từng nhân viên và gửi đến đúng Email của họ. Bạn có thể chọn gửi cho 1 người, nhiều người hoặc tất cả.")

            unique_employees_in_filter = sort_employee_names(filtered_df['Tên nhân viên'].dropna().astype(str).tolist())

            with st.form("form_send_email"):
                # Thêm multiselect cho phép chọn người nhận
                selected_to_send = st.multiselect(
                    "Chọn nhân viên nhận báo cáo:", 
                    options=unique_employees_in_filter, 
                    default=unique_employees_in_filter,
                    filter_mode="contains",
                    help="Có thể xóa bớt hoặc chọn lại. Mặc định là gửi cho tất cả những người có trong danh sách lọc bên trên."
                )

                # Đã lưu cứng thông tin Email và Mật khẩu ứng dụng vào code
                sender_email, sender_pass = get_smtp_sender_credentials()

                st.write(f"📧 **Email gửi đi mặc định:** `{sender_email}`")

                if st.form_submit_button("🚀 Xác Nhận Gửi Email"):
                    if not sender_email or not sender_pass:
                        st.error("❌ Vui lòng nhập đầy đủ Email và Mật khẩu ứng dụng!")
                    elif not selected_to_send:
                        st.warning("⚠️ Vui lòng chọn ít nhất 1 nhân viên để gửi!")
                    else:
                        success_count = 0
                        error_messages = []

                        progress_bar = st.progress(0)
                        # V59: mọi email báo cáo cũng luôn lấy địa chỉ mới nhất trực tiếp từ Sheet1.
                        live_report_email_creds = load_credentials_fresh_for_email()

                        for i, emp in enumerate(selected_to_send):
                            df_emp = filtered_df[filtered_df['Tên nhân viên'] == emp]
                            total_phat = df_emp['Phạt vi phạm'].sum()

                            emp_email = latest_email_from_credentials(live_report_email_creds, emp)

                            if not emp_email or "@" not in emp_email:
                                error_messages.append(f"⚠️ Bỏ qua {emp}: Không có Email hợp lệ.")
                            else:
                                res, msg = send_email_report(
                                    sender_email, sender_pass, emp_email, emp, df_emp, 
                                    total_phat, start_date.strftime('%d/%m/%Y'), end_date.strftime('%d/%m/%Y')
                                )
                                if res:
                                    success_count += 1
                                else:
                                    error_messages.append(f"❌ Lỗi gửi {emp}: {msg}")

                            progress_bar.progress((i + 1) / len(selected_to_send))
                            time.sleep(0.5) # Chờ nửa giây để tránh bị Google chặn Spam

                        if success_count > 0:
                            st.success(f"✅ Đã gửi thành công {success_count} email báo cáo!")
                        if error_messages:
                            for err in error_messages:
                                st.error(err)

    # V86.3: Chi tiết danh sách KHÔNG dùng bất kỳ conditional formatting nào.
    # Mọi dòng dùng nền/chữ mặc định của Streamlit để hiển thị ổn định trên desktop và mobile.

    tab1, tab2, tab3, tab4 = st.tabs(["Tất cả danh sách", "Danh sách Nghỉ CÓ phép", "Danh sách Nghỉ PHÁT SINH", "Danh sách Nghỉ KHÔNG phép"])

    with tab1:
        if export_df.empty:
            st.info("Trống.")
        elif st.session_state.current_role in ["admin", "letan", "quanly"]:
            # Admin/Lễ tân/Quản lý: checkbox chọn 1 hoặc nhiều dòng và sửa trực tiếp tại bảng.
            raw_detail_full = filtered_df.copy().reset_index(drop=True)
            raw_detail = raw_detail_full.drop(columns=cols_to_hide + ['__source_sheet_id', '__source_row'], errors='ignore').copy()
            raw_detail = add_source_leave_type_column(raw_detail)
            if 'Lý do nghỉ' in raw_detail.columns:
                raw_detail['Lý do nghỉ'] = raw_detail['Lý do nghỉ'].apply(clean_leave_reason_display)

            # Danh mục Lý do nghỉ dùng trực tiếp trong bảng sửa.
            reason_options = get_leave_reason_options(
                globals().get('df_loai_nghi', pd.DataFrame()),
                raw_detail['Lý do nghỉ'].tolist() if 'Lý do nghỉ' in raw_detail.columns else []
            )

            # V70: toàn bộ chỉnh sửa lịch trong Chi tiết danh sách nằm trong st.form.
            # Vì vậy thay đổi nhiều ô KHÔNG reload/re-run trang; chỉ khi bấm Lưu/Xóa mới gửi dữ liệu.
            fingerprint_parts = []
            for _, _r in raw_detail_full.iterrows():
                fingerprint_parts.append(
                    f"{_r.get('__source_sheet_id','')}|{_r.get('__source_row','')}|{schedule_key(_r)}|"
                    f"{_r.get('Ngày cập nhật','')}|{_r.get('Giờ cập nhật','')}"
                )
            detail_fp = "||".join(fingerprint_parts)
            if st.session_state.get('_detail_editor_fingerprint') != detail_fp:
                seed_df = raw_detail.copy()
                seed_df.insert(0, "Chọn", False)
                st.session_state['_detail_editor_seed'] = seed_df
                st.session_state['_detail_editor_fingerprint'] = detail_fp
                st.session_state['_detail_editor_version'] = int(st.session_state.get('_detail_editor_version', 0)) + 1

            editor_df = st.session_state.get('_detail_editor_seed', raw_detail.copy()).copy()
            if 'Chọn' not in editor_df.columns:
                editor_df.insert(0, "Chọn", False)
            editor_df, _ = apply_table_layout_df(editor_df, "leave_detail")
            # V74: ép đúng dtype trước st.data_editor để tương thích Streamlit mới.
            editor_df = prepare_leave_editor_types(editor_df)

            derived_cols = [
                "Số ngày tính", "Số ngày phép cộng dồn", "Phạt vi phạm",
                "Ngày cập nhật", "Giờ cập nhật", "Người cập nhật"
            ]
            disabled_cols = [c for c in derived_cols if c in editor_df.columns]
            if "Loại nghỉ" in editor_df.columns:
                disabled_cols.append("Loại nghỉ")
            if st.session_state.current_role in EMPLOYEE_LIKE_ROLES and "Tên nhân viên" in editor_df.columns:
                disabled_cols.append("Tên nhân viên")

            editor_version = int(st.session_state.get('_detail_editor_version', 1))
            editor_key = f"detail_schedule_editor_batch_v{editor_version}"
            detail_col_config = table_layout_column_config("leave_detail", list(editor_df.columns))
            if "Chọn" in editor_df.columns:
                detail_col_config["Chọn"] = st.column_config.CheckboxColumn(
                    "Chọn", help="Chỉ cần tick khi muốn XÓA. Khi sửa, hệ thống tự nhận biết dòng đã thay đổi.",
                    default=False, width=layout_width("leave_detail", "Chọn", "small")
                )
            if "Ngày" in editor_df.columns:
                detail_col_config["Ngày"] = st.column_config.DateColumn(
                    "Ngày", format="DD/MM/YYYY", width=layout_width("leave_detail", "Ngày", "small")
                )
            if "Thứ ngày" in editor_df.columns:
                detail_col_config["Thứ ngày"] = st.column_config.TextColumn(
                    "Thứ ngày", disabled=True, width=layout_width("leave_detail", "Thứ ngày", "small")
                )
            if "Lý do nghỉ" in editor_df.columns:
                detail_col_config["Lý do nghỉ"] = st.column_config.SelectboxColumn(
                    "Lý do nghỉ", options=reason_options, required=True,
                    width=layout_width("leave_detail", "Lý do nghỉ", "medium"),
                    help="Bấm vào ô để mở dropdown rồi gõ tên loại nghỉ để tìm nhanh. Danh sách lấy từ sheet LoaiNghi."
                )
            if "Loại nghỉ" in editor_df.columns:
                detail_col_config["Loại nghỉ"] = st.column_config.TextColumn(
                    "Loại nghỉ", disabled=True,
                    width=layout_width("leave_detail", "Loại nghỉ", "medium"),
                    help="Giá trị nguồn trực tiếp từ cột C của Google Sheet lịch nghỉ."
                )
            if "Số ngày tính" in editor_df.columns:
                detail_col_config["Số ngày tính"] = st.column_config.NumberColumn(
                    "Số ngày tính", step=0.5, format="%.1f", disabled=True,
                    width=layout_width("leave_detail", "Số ngày tính", "small")
                )
            if "Số ngày phép cộng dồn" in editor_df.columns:
                detail_col_config["Số ngày phép cộng dồn"] = st.column_config.NumberColumn(
                    "Số ngày phép cộng dồn", step=0.5, format="%.1f", disabled=True,
                    width=layout_width("leave_detail", "Số ngày phép cộng dồn", "small")
                )
            if "Phạt vi phạm" in editor_df.columns:
                detail_col_config["Phạt vi phạm"] = st.column_config.NumberColumn(
                    "Phạt vi phạm", step=50000, format="%.0f", disabled=True,
                    width=layout_width("leave_detail", "Phạt vi phạm", "small")
                )
            for _c in ["Ngày cập nhật", "Giờ cập nhật", "Người cập nhật"]:
                if _c in editor_df.columns:
                    detail_col_config[_c] = st.column_config.TextColumn(
                        _c, disabled=True, width=layout_width("leave_detail", _c, "small")
                    )

            with st.form(f"detail_schedule_batch_form_v{editor_version}", clear_on_submit=False):
                detail_editor = st.data_editor(
                    editor_df,
                    width="stretch", height="content", hide_index=True,
                    row_height=layout_row_height("leave_detail"),
                    num_rows="fixed", disabled=disabled_cols,
                    column_config=detail_col_config, key=editor_key
                )
                st.caption(
                    "Sửa trực tiếp nhiều dòng rồi bấm Lưu một lần. Cột Lý do nghỉ là dropdown có thể gõ để tìm; cột Loại nghỉ lấy từ cột C của sheet LoaiNghi và chỉ đọc. "
                    "Checkbox Chọn chỉ dùng khi muốn xóa dòng."
                )
                _d_save, _d_delete = st.columns(2)
                with _d_save:
                    submit_detail_save = st.form_submit_button("💾 Lưu tất cả thay đổi", use_container_width=True)
                with _d_delete:
                    submit_detail_delete = st.form_submit_button("🗑️ Xóa các dòng đã chọn", use_container_width=True)

            detail_edit_only = detail_editor.drop(columns=['Chọn'], errors='ignore').copy()
            original_compare = raw_detail.copy().reset_index(drop=True)
            detail_compare = detail_edit_only.copy().reset_index(drop=True)
            changed_positions = get_changed_schedule_positions(original_compare, detail_compare)
            selected_positions = detail_editor.index[detail_editor.get('Chọn', False) == True].tolist() if 'Chọn' in detail_editor.columns else []

            if submit_detail_save:
                if not changed_positions:
                    st.info("Không có thay đổi nào cần lưu.")
                else:
                    all_ok = True
                    messages = []
                    for pos in changed_positions:
                        if pos >= len(raw_detail_full) or pos >= len(detail_compare):
                            continue
                        original = raw_detail_full.iloc[pos].copy()
                        edited = detail_compare.iloc[pos].copy()
                        permitted, perm_msg = validate_schedule_edit_permission(
                            original, edited, st.session_state.current_role, get_vn_today(),
                            current_user=st.session_state.current_user
                        )
                        if not permitted:
                            st.error(f"❌ {original.get('Tên nhân viên','')}: {perm_msg}")
                            all_ok = False
                            break
                        ok, msg = update_schedule_record(original, edited, st.session_state.current_user)
                        messages.append((ok, msg))
                        if not ok:
                            all_ok = False
                            break
                    for ok, msg in messages:
                        (st.success if ok else st.error)(msg)
                    if all_ok:
                        st.session_state.pop('_detail_editor_seed', None)
                        st.session_state.pop('_detail_editor_fingerprint', None)
                        _clear_dynamic_data_caches()
                        st.rerun()

            if submit_detail_delete:
                if not selected_positions:
                    st.warning("Vui lòng tick ít nhất 1 dòng cần xóa.")
                else:
                    originals = [raw_detail_full.iloc[pos].copy() for pos in selected_positions if pos < len(raw_detail_full)]
                    today_del = get_vn_today()
                    can_delete = True
                    for r in originals:
                        permitted, perm_msg = validate_schedule_delete_permission(
                            r, st.session_state.current_role,
                            current_user=st.session_state.current_user, today=today_del
                        )
                        if not permitted:
                            st.error(f"❌ {r.get('Tên nhân viên','')}: {perm_msg}")
                            can_delete = False
                            break
                    if can_delete:
                        ok, msg = delete_schedule_records(originals, st.session_state.current_user)
                        (st.success if ok else st.error)(msg)
                        if ok:
                            st.session_state.pop('_detail_editor_seed', None)
                            st.session_state.pop('_detail_editor_fingerprint', None)
                            _clear_dynamic_data_caches()
                            st.rerun()
        else:
            # Nhân viên: chỉ xem, không có checkbox sửa/xóa và không có Export Excel.
            export_view_df, _ = apply_table_layout_df(export_df.copy(), "leave_detail")
            st.dataframe(
                export_view_df,
                width="stretch",
                height="content",
                hide_index=True,
                row_height=layout_row_height("leave_detail"),
                column_config=table_layout_column_config("leave_detail", list(export_view_df.columns))
            )

    with tab2:
        if co_phep_df.empty:
            st.info("Trống.")
        else:
            co_display_source = add_source_leave_type_column(co_phep_df.drop(columns=cols_to_hide + ['__source_sheet_id', '__source_row'], errors='ignore').copy())
            co_display = format_display_df(co_display_source)
            co_display, _ = apply_table_layout_df(co_display, "leave_detail")
            st.dataframe(
                co_display,
                width="stretch", height="content", hide_index=True, row_height=layout_row_height("leave_detail"),
                column_config=table_layout_column_config("leave_detail", list(co_display.columns))
            )

    with tab3:
        if phat_sinh_df.empty:
            st.info("Trống.")
        else:
            ps_display_source = add_source_leave_type_column(phat_sinh_df.drop(columns=cols_to_hide + ['__source_sheet_id', '__source_row'], errors='ignore').copy())
            ps_display = format_display_df(ps_display_source)
            ps_display, _ = apply_table_layout_df(ps_display, "leave_detail")
            st.dataframe(
                ps_display,
                width="stretch", height="content", hide_index=True, row_height=layout_row_height("leave_detail"),
                column_config=table_layout_column_config("leave_detail", list(ps_display.columns))
            )

    with tab4:
        if khong_phep_df.empty:
            st.success("Không có ai!")
        else:
            kp_display_source = add_source_leave_type_column(khong_phep_df.drop(columns=cols_to_hide + ['__source_sheet_id', '__source_row'], errors='ignore').copy())
            kp_display = format_display_df(kp_display_source)
            kp_display, _ = apply_table_layout_df(kp_display, "leave_detail")
            st.dataframe(
                kp_display,
                width="stretch", height="content", hide_index=True, row_height=layout_row_height("leave_detail"),
                column_config=table_layout_column_config("leave_detail", list(kp_display.columns))
            )

    # Nút lưu bố cục ngay tại nhóm bảng Chi tiết danh sách, chỉ Admin.
    render_admin_quick_layout_default(
        "leave_detail", get_table_columns_for_settings("leave_detail"), "leave_detail_tabs"
    )


elif selected_page == "✏️ Quản lý lịch nghỉ":
    st.subheader("✏️ Quản lý lịch nghỉ")
    st.markdown("### 🗑️ Xóa / Quản lý lịch nghỉ đã đăng ký")
    st.markdown("""
### 3. Quy định Sửa / Hủy lịch nghỉ

**Đối với Nhân viên/Leader:**

- Chỉ được phép Sửa/Hủy lịch đối với loại Nghỉ CÓ phép của chính mình.
- Chỉ được phép Sửa/Hủy lịch đối với loại Nghỉ KHÔNG phép của chính mình. Nghỉ KHÔNG phép không bị giới hạn 3 ngày, nhưng không được sửa/hủy ngày hiện tại và ngày trong quá khứ.
- Nghỉ CÓ phép phải thực hiện Sửa/Hủy trước ít nhất **3 ngày**.
- Tuyệt đối không được đổi lịch sang tên người khác.

**Đối với Lễ tân / Quản lý:**

- Không bị giới hạn 3 ngày, nhưng tuyệt đối không được Sửa/Xóa lịch trong quá khứ (các ngày trước ngày hôm nay) để bảo vệ dữ liệu tính lương.

**Ngoại lệ — Nghỉ phép quay video:**

- Không bị ràng buộc bởi các giới hạn thời gian, hạn mức ngày/tháng, giới hạn ngày thường/Chủ nhật.
- Không chiếm hạn mức nghỉ trong ngày và không làm giảm suất nghỉ của nhân viên khác.

**Ngoại lệ — Leader theo chính sách:**

- `Leader nghỉ phép theo chính sách`
- `Leader về sớm về sớm theo chính sách`
- `Leader đi trễ sớm theo chính sách`

Leader được đăng ký/Sửa/Hủy các lý do trên không bị giới hạn 3 ngày và không bị giới hạn theo thứ/ngày trong tuần.

**Ngoại lệ — Nghỉ đám hiếu:**

- Admin/Lễ tân/Quản lý được cập nhật hoặc hủy `Nghỉ đám hiếu` không bị giới hạn 3 ngày và không bị giới hạn theo thứ/ngày trong tuần.
""")

    render_leave_filter_label_css()
    manage_today = get_vn_today()
    mf_date, mf_name, mf_refresh = st.columns([5, 4, 2])
    with mf_date:
        md1, md2 = st.columns(2)
        with md1:
            manage_filter_type = st.selectbox(
                "Lọc thời gian:",
                ["Hôm nay", "Hôm qua", "Ngày mai", "Chọn ngày", "Khoảng thời gian", "Tuần này", "Tuần trước", "Tuần sau", "Tháng này", "Tháng trước", "Tháng sau", "Tất cả"],
                index=0, key="leave_manage_time_filter", filter_mode="contains"
            )
        with md2:
            if manage_filter_type == "Hôm nay": manage_start = manage_end = manage_today
            elif manage_filter_type == "Hôm qua": manage_start = manage_end = manage_today - timedelta(days=1)
            elif manage_filter_type == "Ngày mai": manage_start = manage_end = manage_today + timedelta(days=1)
            elif manage_filter_type == "Tuần này":
                manage_start = manage_today - timedelta(days=manage_today.weekday()); manage_end = manage_start + timedelta(days=6)
            elif manage_filter_type == "Tuần trước":
                manage_start = manage_today - timedelta(days=manage_today.weekday() + 7); manage_end = manage_start + timedelta(days=6)
            elif manage_filter_type == "Tuần sau":
                manage_start = manage_today - timedelta(days=manage_today.weekday()) + timedelta(days=7); manage_end = manage_start + timedelta(days=6)
            elif manage_filter_type == "Tháng này":
                manage_start = manage_today.replace(day=1); manage_end = manage_today.replace(day=calendar.monthrange(manage_today.year, manage_today.month)[1])
            elif manage_filter_type == "Tháng trước":
                manage_end = manage_today.replace(day=1) - timedelta(days=1); manage_start = manage_end.replace(day=1)
            elif manage_filter_type == "Tháng sau":
                manage_start = manage_today.replace(year=manage_today.year + 1, month=1, day=1) if manage_today.month == 12 else manage_today.replace(month=manage_today.month + 1, day=1)
                manage_end = manage_start.replace(day=calendar.monthrange(manage_start.year, manage_start.month)[1])
            elif manage_filter_type == "Chọn ngày":
                manage_start = manage_end = st.date_input("Chọn ngày:", manage_today, key="leave_manage_single_date")
            elif manage_filter_type == "Khoảng thời gian":
                _manage_range = st.date_input("Chọn khoảng thời gian:", [manage_today, manage_today], key="leave_manage_date_range")
                manage_start, manage_end = (_manage_range[0], _manage_range[1]) if len(_manage_range) == 2 else (_manage_range[0], _manage_range[0])
            else:
                manage_start = manage_end = None

    df_backup_view = df_backup.copy()
    if st.session_state.current_role in EMPLOYEE_LIKE_ROLES:
        df_backup_view = df_backup_view[
            df_backup_view['Tên nhân viên'].astype(str).apply(normalize_login_name).eq(normalize_login_name(st.session_state.current_user))
        ].copy()
        manage_selected_nv = st.session_state.current_user
        with mf_name:
            st.text_input("👤 Nhân viên:", value=st.session_state.current_user, disabled=True, key="leave_manage_emp_locked")
    else:
        with mf_name:
            manage_employee_options = ["- Tất cả nhân viên -"] + sort_employee_names(df_backup_view.get('Tên nhân viên', pd.Series(dtype=str)).dropna().astype(str).tolist())
            manage_selected_nv = st.selectbox("👤 Tìm kiếm nhân viên:", manage_employee_options, key="leave_manage_employee_filter", filter_mode="contains")

    with mf_refresh:
        st.write("")
        if st.button("🔄 Cập Nhật Dữ Liệu", use_container_width=True, key="refresh_leave_manage"):
            load_backup_sheet_data.clear()
            _clear_dynamic_data_caches()
            st.rerun()

    if not df_backup_view.empty:
        df_backup_view = df_backup_view.copy()
        df_backup_view['__filter_date'] = df_backup_view['Ngày'].apply(_parse_vn_date)
        if manage_start is not None and manage_end is not None:
            df_backup_view = df_backup_view[
                df_backup_view['__filter_date'].apply(lambda d: d is not None and manage_start <= d <= manage_end)
            ].copy()
        if st.session_state.current_role not in EMPLOYEE_LIKE_ROLES and manage_selected_nv != "- Tất cả nhân viên -":
            df_backup_view = df_backup_view[
                df_backup_view['Tên nhân viên'].astype(str).apply(normalize_login_name).eq(normalize_login_name(manage_selected_nv))
            ].copy()
        df_backup_view = df_backup_view.drop(columns=['__filter_date'], errors='ignore')

    if df_backup_view.empty:
        st.info("Chưa có lịch nghỉ nào được đăng ký trong bộ lọc hiện tại.")
    else:
        manage_raw_full = df_backup_view.reset_index(drop=True).copy()
        manage_visible = manage_raw_full.drop(columns=['__source_sheet_id', '__source_row'], errors='ignore').copy()
        manage_visible = add_source_leave_type_column(manage_visible)
        if st.session_state.current_role != "admin" and "Phạt vi phạm" in manage_visible.columns:
            manage_visible = manage_visible.drop(columns=["Phạt vi phạm"])
        if 'Lý do nghỉ' in manage_visible.columns:
            manage_visible['Lý do nghỉ'] = manage_visible['Lý do nghỉ'].apply(clean_leave_reason_display)
        manage_visible.insert(0, 'Chọn', False)
        manage_visible, _ = apply_table_layout_df(manage_visible, "leave_manage")
        # V74: ép đúng dtype trước st.data_editor để tương thích Streamlit mới.
        manage_visible = prepare_leave_editor_types(manage_visible)

        manage_reason_options = get_leave_reason_options(
            globals().get('df_loai_nghi', pd.DataFrame()),
            manage_visible['Lý do nghỉ'].tolist() if 'Lý do nghỉ' in manage_visible.columns else []
        )

        manage_col_config = table_layout_column_config("leave_manage", list(manage_visible.columns))
        if 'Chọn' in manage_visible.columns:
            manage_col_config['Chọn'] = st.column_config.CheckboxColumn(
                'Chọn', default=False, width=70,
                help='Tick khi muốn XÓA. Không cần tick khi sửa; hệ thống tự nhận biết dòng đã thay đổi.'
            )
        if 'Ngày' in manage_visible.columns:
            manage_col_config['Ngày'] = st.column_config.DateColumn('Ngày', format='DD/MM/YYYY', width=layout_width('leave_manage', 'Ngày', 'small'))
        if 'Thứ ngày' in manage_visible.columns:
            manage_col_config['Thứ ngày'] = st.column_config.TextColumn(
                'Thứ ngày', disabled=True, width=layout_width('leave_manage', 'Thứ ngày', 'small')
            )
        if 'Lý do nghỉ' in manage_visible.columns:
            manage_col_config['Lý do nghỉ'] = st.column_config.SelectboxColumn(
                'Lý do nghỉ', options=manage_reason_options, required=True,
                width=layout_width('leave_manage', 'Lý do nghỉ', 'medium'),
                help='Click vào ô để mở dropdown. Có thể gõ trực tiếp tên loại nghỉ để tìm trong danh sách.'
            )
        if 'Loại nghỉ' in manage_visible.columns:
            manage_col_config['Loại nghỉ'] = st.column_config.TextColumn(
                'Loại nghỉ', disabled=True, width=layout_width('leave_manage', 'Loại nghỉ', 'medium'),
                help='Loại nghỉ lấy từ cột C của sheet LoaiNghi, đối chiếu theo Lý do nghỉ ở cột B.'
            )
        if 'Số ngày tính' in manage_visible.columns:
            manage_col_config['Số ngày tính'] = st.column_config.NumberColumn('Số ngày tính', format='%.1f', disabled=True)
        if 'Số ngày phép cộng dồn' in manage_visible.columns:
            manage_col_config['Số ngày phép cộng dồn'] = st.column_config.NumberColumn('Số ngày phép cộng dồn', format='%.1f', disabled=True)
        if 'Phạt vi phạm' in manage_visible.columns:
            manage_col_config['Phạt vi phạm'] = st.column_config.NumberColumn('Phạt vi phạm', format='%.0f', disabled=True)
        for _c in ['Ngày cập nhật', 'Giờ cập nhật', 'Người cập nhật']:
            if _c in manage_visible.columns:
                manage_col_config[_c] = st.column_config.TextColumn(_c, disabled=True)

        manage_derived = [c for c in ['Thứ ngày','Loại nghỉ','Số ngày tính','Số ngày phép cộng dồn','Phạt vi phạm','Ngày cập nhật','Giờ cập nhật','Người cập nhật'] if c in manage_visible.columns]
        if st.session_state.current_role in EMPLOYEE_LIKE_ROLES and 'Tên nhân viên' in manage_visible.columns:
            manage_derived.append('Tên nhân viên')
        manage_locked = st.session_state.current_role in EMPLOYEE_LIKE_ROLES and system_status['lock_nv']
        if manage_locked:
            st.error("🔒 Admin đang khóa quyền thay đổi lịch nghỉ của nhân viên. Bảng chỉ được xem cho đến khi mở khóa.")
            manage_derived = [c for c in manage_visible.columns if c != 'Chọn']

        with st.form('leave_manage_batch_edit_form_v70', clear_on_submit=False):
            manage_editor = st.data_editor(
                manage_visible,
                width='stretch', height='content', hide_index=True, num_rows='fixed',
                row_height=layout_row_height('leave_manage'),
                disabled=manage_derived,
                column_config=manage_col_config,
                key='leave_manage_batch_editor_v70'
            )
            st.caption(
                "Sửa trực tiếp nhiều dòng rồi bấm Lưu một lần. Dropdown Lý do nghỉ hỗ trợ gõ để tìm; cột Loại nghỉ lấy từ cột C của sheet LoaiNghi và chỉ đọc. "
                "Checkbox Chọn chỉ dùng cho thao tác Xóa."
            )
            _m_save, _m_delete = st.columns(2)
            with _m_save:
                manage_submit_save = st.form_submit_button('💾 Lưu tất cả thay đổi', use_container_width=True, disabled=manage_locked)
            with _m_delete:
                manage_submit_delete = st.form_submit_button('🗑️ Xóa các dòng đã chọn', use_container_width=True, disabled=manage_locked)

        # V86.3: Quản lý lịch nghỉ không dùng bất kỳ conditional formatting nào.
        # Chỉ dùng st.data_editor mặc định + cấu hình kiểu cột, không tô màu theo Lý do/Loại nghỉ.
        render_admin_quick_layout_default("leave_manage", [c for c in manage_visible.columns if c != 'Chọn'], "leave_manage_page")

        manage_edit_only = manage_editor.drop(columns=['Chọn'], errors='ignore').reset_index(drop=True)
        manage_original_visible = manage_raw_full.drop(columns=['__source_sheet_id','__source_row'], errors='ignore').copy()
        manage_original_visible = add_source_leave_type_column(manage_original_visible)
        if st.session_state.current_role != 'admin' and 'Phạt vi phạm' in manage_original_visible.columns:
            manage_original_visible = manage_original_visible.drop(columns=['Phạt vi phạm'])
        if 'Lý do nghỉ' in manage_original_visible.columns:
            manage_original_visible['Lý do nghỉ'] = manage_original_visible['Lý do nghỉ'].apply(clean_leave_reason_display)
        manage_original_visible = manage_original_visible.reset_index(drop=True)
        manage_original_visible = prepare_leave_editor_types(manage_original_visible)
        manage_changed = get_changed_schedule_positions(manage_original_visible, manage_edit_only)
        manage_selected = manage_editor.index[manage_editor.get('Chọn', False) == True].tolist() if 'Chọn' in manage_editor.columns else []

        if manage_submit_save:
            if not manage_changed:
                st.info('Không có thay đổi nào cần lưu.')
            else:
                all_ok = True
                save_messages = []
                for pos in manage_changed:
                    if pos >= len(manage_raw_full) or pos >= len(manage_edit_only):
                        continue
                    original = manage_raw_full.iloc[pos].copy()
                    edited = manage_edit_only.iloc[pos].copy()
                    permitted, perm_msg = validate_schedule_edit_permission(
                        original, edited, st.session_state.current_role, manage_today,
                        current_user=st.session_state.current_user
                    )
                    if not permitted:
                        st.error(f"❌ {original.get('Tên nhân viên','')}: {perm_msg}")
                        all_ok = False
                        break
                    ok, msg = update_schedule_record(original, edited, st.session_state.current_user)
                    save_messages.append((ok, msg))
                    if not ok:
                        all_ok = False
                        break
                for ok, msg in save_messages:
                    (st.success if ok else st.error)(msg)
                if all_ok:
                    _clear_dynamic_data_caches()
                    st.rerun()

        if manage_submit_delete:
            if not manage_selected:
                st.warning('Vui lòng tick ít nhất 1 dòng cần xóa.')
            else:
                originals = [manage_raw_full.iloc[pos].copy() for pos in manage_selected if pos < len(manage_raw_full)]
                can_delete = True
                for r in originals:
                    permitted, perm_msg = validate_schedule_delete_permission(
                        r, st.session_state.current_role,
                        current_user=st.session_state.current_user, today=manage_today
                    )
                    if not permitted:
                        st.error(f"❌ {r.get('Tên nhân viên','')}: {perm_msg}")
                        can_delete = False
                        break
                if can_delete:
                    ok, msg = delete_schedule_records(originals, st.session_state.current_user)
                    (st.success if ok else st.error)(msg)
                    if ok:
                        _clear_dynamic_data_caches()
                        st.rerun()
