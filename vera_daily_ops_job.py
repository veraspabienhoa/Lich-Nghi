"""Vera Spa daily operations Cloud Run Job.

Actions:
  python vera_daily_ops_job.py cleanup
      - At 08:00 ICT: clear previous-day Break/Giờ ra/Giờ vào in Input!R21:R100,
        S21:S100, U21:U100 of the Bảng tour XLSM and upload the same file back to Drive.

  python vera_daily_ops_job.py violations
      - At 20:10 ICT: analyze Bảng tour break records for today.
      - >90 minutes outside => map penalty reason from LoaiNghi and append Auto update violation.
      - Only one of Giờ ra/Giờ vào => map configured/sheet reason and append Auto update violation.
      - Email employee and CC admin/letan/quanly + veraspabienhoa@gmail.com.

This script is intentionally standalone and does not import Streamlit app.py.
Credentials/passwords are read from Cloud Run environment / Secret Manager, never hard-coded.
"""

from __future__ import annotations

import json
import math
import os
import re
import smtplib
import sys
import tempfile
import unicodedata
from datetime import date, datetime, timedelta, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import gspread
import pandas as pd
from google.auth.transport.requests import AuthorizedSession
from google.oauth2.service_account import Credentials

VN_TZ = timezone(timedelta(hours=7))

SHEET_MAT_KHAU_ID = os.getenv("SHEET_MAT_KHAU_ID", "1DGXy3kPyMPwtz-3CnG8i6BiQbXFDApasoXVFzSmUe24")
SHEET_DU_PHONG_ID = os.getenv("SHEET_DU_PHONG_ID", "1Kz0aw-JatptAN9G7YSwZ6rJO09urOPaD-rS-18eZSY0")
BANG_TOUR_FILE_ID = os.getenv("BANG_TOUR_FILE_ID", "1yA1Oog_6R-HmDFatcku-x8s-59p2dP9R")
BANG_TOUR_AUDIT_WORKSHEET = os.getenv("BANG_TOUR_AUDIT_WORKSHEET", "DoiSoatRaNgoai")
BANG_TOUR_BREAK_LIMIT_MINUTES = int(os.getenv("BANG_TOUR_BREAK_LIMIT_MINUTES", "90") or 90)
BANG_TOUR_SINGLE_SIDE_REASON = (os.getenv("BANG_TOUR_SINGLE_SIDE_REASON", "Ra ngoài thiếu giờ ra/vào") or "Ra ngoài thiếu giờ ra/vào").strip()

SMTP_SENDER_EMAIL = (os.getenv("SMTP_SENDER_EMAIL", "veraspabienhoa@gmail.com") or "").strip()
SMTP_APP_PASSWORD = (os.getenv("SMTP_APP_PASSWORD", "") or "").strip()
AUTO_CC_EMAIL = "veraspabienhoa@gmail.com"

AUDIT_HEADERS = [
    "Ngày", "Tên nhân viên", "Loại vi phạm", "Số phút", "Mức phạt",
    "Ghi phiếu phạt", "Email", "Chi tiết", "Cập nhật lúc", "Người cập nhật",
]


def vn_now() -> datetime:
    return datetime.now(VN_TZ)


def normalize_text(value) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = text.replace("đ", "d").replace("Đ", "D")
    return " ".join(text.strip().split()).casefold()


def parse_money(value) -> float:
    if value is None:
        return 0.0
    text = str(value).strip().replace("đ", "").replace("Đ", "").replace("VND", "").replace("VNĐ", "").replace(" ", "")
    if not text or text.casefold() in {"nan", "none", "-"}:
        return 0.0
    # Vietnamese money cells are usually integer with . or , as thousands separators.
    text = text.replace(".", "").replace(",", "")
    try:
        return float(text)
    except Exception:
        return 0.0


def _credentials(scopes):
    env_json = (os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON", "") or "").strip()
    if env_json:
        return Credentials.from_service_account_info(json.loads(env_json), scopes=scopes)
    import google.auth
    creds, _ = google.auth.default(scopes=scopes)
    return creds


def gspread_client():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    return gspread.authorize(_credentials(scopes))


def drive_session():
    scopes = ["https://www.googleapis.com/auth/drive"]
    return AuthorizedSession(_credentials(scopes))


def download_drive_file(file_id: str, path: str):
    session = drive_session()
    url = f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media&supportsAllDrives=true"
    r = session.get(url, timeout=120, stream=True)
    if r.status_code != 200:
        raise RuntimeError(f"Drive download HTTP {r.status_code}: {str(r.text)[:400]}")
    with open(path, "wb") as f:
        for chunk in r.iter_content(1024 * 1024):
            if chunk:
                f.write(chunk)
    if not os.path.exists(path) or os.path.getsize(path) == 0:
        raise RuntimeError("Drive returned an empty file.")


def upload_drive_file(file_id: str, path: str):
    session = drive_session()
    url = f"https://www.googleapis.com/upload/drive/v3/files/{file_id}?uploadType=media&supportsAllDrives=true"
    with open(path, "rb") as f:
        r = session.patch(
            url,
            data=f,
            headers={"Content-Type": "application/vnd.ms-excel.sheet.macroEnabled.12"},
            timeout=180,
        )
    if r.status_code not in {200, 201}:
        raise RuntimeError(f"Drive upload HTTP {r.status_code}: {str(r.text)[:500]}")


def parse_datetime(value, fallback_date: date | None = None) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, pd.Timestamp):
        dt = value.to_pydatetime()
    elif isinstance(value, datetime):
        dt = value
    elif isinstance(value, date):
        dt = datetime.combine(value, datetime.min.time())
    else:
        parsed = pd.to_datetime(str(value).strip(), dayfirst=True, errors="coerce")
        if pd.isna(parsed):
            return None
        dt = parsed.to_pydatetime() if isinstance(parsed, pd.Timestamp) else parsed
    if dt.tzinfo is not None:
        dt = dt.astimezone(VN_TZ).replace(tzinfo=None)
    if dt.year in {1899, 1900, 1970} and fallback_date:
        dt = datetime.combine(fallback_date, dt.time())
    return dt


def open_bang_tour(data_only=False):
    from openpyxl import load_workbook
    fd, path = tempfile.mkstemp(suffix=".xlsm")
    os.close(fd)
    download_drive_file(BANG_TOUR_FILE_ID, path)
    wb = load_workbook(path, keep_vba=True, data_only=data_only)
    if "Input" not in wb.sheetnames:
        try:
            os.remove(path)
        except Exception:
            pass
        raise RuntimeError("Không tìm thấy sheet Input trong Bảng tour.")
    return wb, wb["Input"], path


def cleanup_previous_day() -> str:
    today = vn_now().date()
    yesterday = today - timedelta(days=1)
    wb, ws, source_path = open_bang_tour(data_only=False)
    out_path = source_path + ".out.xlsm"
    cleared = 0
    try:
        for row in range(21, 101):
            out_dt = parse_datetime(ws[f"S{row}"].value)
            in_dt = parse_datetime(ws[f"U{row}"].value)
            out_date = out_dt.date() if out_dt else None
            in_date = in_dt.date() if in_dt else None
            if out_date == yesterday or in_date == yesterday:
                ws[f"R{row}"] = None
                ws[f"S{row}"] = None
                ws[f"U{row}"] = None
                cleared += 1
        if cleared:
            wb.save(out_path)
            upload_drive_file(BANG_TOUR_FILE_ID, out_path)
        return f"cleanup success: {cleared} row(s) cleared for {yesterday.strftime('%d/%m/%Y')}"
    finally:
        for fp in (source_path, out_path):
            try:
                if os.path.exists(fp):
                    os.remove(fp)
            except Exception:
                pass


def leave_catalog(client) -> dict:
    ss = client.open_by_key(SHEET_DU_PHONG_ID)
    try:
        ws = ss.worksheet("LoaiNghi")
    except Exception:
        return {}
    vals = ws.get_all_values()
    catalog = {}
    for row in vals:
        if len(row) < 2:
            continue
        name = str(row[1] or "").strip()
        if not name or normalize_text(name) in {"loai nghi", "ly do nghi"}:
            continue
        penalty = parse_money(row[5] if len(row) > 5 else 0)
        catalog[normalize_text(name)] = {"name": name, "penalty": penalty}
    return catalog


def pick_late_reason(catalog: dict, late_minutes: int) -> str:
    candidates = []
    for item in catalog.values():
        name = str(item.get("name", "")).strip()
        key = normalize_text(name)
        if "ra ngoai" not in key or not any(t in key for t in ("vao muon", "vao tre", "tre")):
            continue
        nums = [int(x) for x in re.findall(r"\d+", key)]
        threshold = max(nums) if nums else None
        candidates.append((threshold, name))
    for threshold, name in sorted([x for x in candidates if x[0] is not None], key=lambda x: x[0]):
        if late_minutes < threshold:
            return name
    untiered = [name for threshold, name in candidates if threshold is None]
    if untiered:
        return untiered[0]
    if late_minutes < 30:
        return "Ra ngoài vào muộn dưới 30 phút"
    if late_minutes < 60:
        return "Ra ngoài vào muộn dưới 60 phút"
    return "Ra ngoài vào muộn dưới 120 phút"


def pick_single_side_reason(catalog: dict) -> str:
    token_pairs = [
        ("ra ngoai", "thieu"),
        ("ra ngoai", "mot lan"),
        ("ra ngoai", "1 lan"),
        ("ra ngoai", "khong du"),
        ("ra ngoai", "khong co gio"),
    ]
    for item in catalog.values():
        name = str(item.get("name", "")).strip()
        key = normalize_text(name)
        if any(all(token in key for token in pair) for pair in token_pairs):
            return name
    return BANG_TOUR_SINGLE_SIDE_REASON


def penalty_for(catalog: dict, reason: str) -> float:
    return float((catalog.get(normalize_text(reason)) or {}).get("penalty", 0) or 0)


def get_or_create_ws(spreadsheet, title, rows=3000, cols=12):
    try:
        return spreadsheet.worksheet(title)
    except Exception:
        return spreadsheet.add_worksheet(title=title, rows=rows, cols=cols)


def audit_ws(client):
    ss = client.open_by_key(SHEET_MAT_KHAU_ID)
    ws = get_or_create_ws(ss, BANG_TOUR_AUDIT_WORKSHEET, rows=3000, cols=12)
    current = ws.row_values(1)
    if current[: len(AUDIT_HEADERS)] != AUDIT_HEADERS:
        ws.update("A1:J1", [AUDIT_HEADERS], value_input_option="USER_ENTERED")
    return ws


def sent_keys(client) -> set:
    ws = audit_ws(client)
    vals = ws.get_all_values()
    if len(vals) <= 1:
        return set()
    idx = {h: i for i, h in enumerate(vals[0])}
    result = set()
    for row in vals[1:]:
        def cell(name):
            i = idx.get(name, -1)
            return row[i] if i >= 0 and i < len(row) else ""
        if normalize_text(cell("Email")) not in {"1", "true", "yes", "da gui"}:
            continue
        result.add((cell("Ngày"), normalize_text(cell("Tên nhân viên")), normalize_text(cell("Loại vi phạm"))))
    return result


def employee_directory(client):
    ws = client.open_by_key(SHEET_MAT_KHAU_ID).get_worksheet(0)
    vals = ws.get_all_values()
    emails = {}
    cc = [AUTO_CC_EMAIL]
    for row in vals[1:]:
        name = str(row[1] if len(row) > 1 else "").strip()
        role = str(row[3] if len(row) > 3 else "").strip().lower()
        email = str(row[7] if len(row) > 7 else "").strip()
        if name and "@" in email:
            emails[normalize_text(name)] = email
        if role in {"admin", "letan", "quanly"} and "@" in email:
            cc.append(email)
    dedup = []
    seen = set()
    for e in cc:
        k = e.casefold()
        if e and k not in seen:
            dedup.append(e)
            seen.add(k)
    return emails, dedup


def existing_violation_keys(client) -> set:
    ws = client.open_by_key(SHEET_DU_PHONG_ID).get_worksheet(0)
    vals = ws.get_all_values()
    out = set()
    for row in vals[1:]:
        if len(row) < 3:
            continue
        out.add((str(row[0]).strip(), normalize_text(row[1]), normalize_text(row[2])))
    return out


def append_violation(client, target_date: date, employee: str, reason: str, detail: str, penalty: float) -> tuple[bool, str]:
    ws = client.open_by_key(SHEET_DU_PHONG_ID).get_worksheet(0)
    date_text = target_date.strftime("%d/%m/%Y")
    key = (date_text, normalize_text(employee), normalize_text(reason))
    if key in existing_violation_keys(client):
        return False, "Đã tồn tại"
    now = vn_now()
    ws.append_row(
        [
            date_text,
            employee,
            reason,
            detail,
            0,
            0,
            float(penalty or 0),
            now.strftime("%d/%m/%Y"),
            now.strftime("%H:%M:%S"),
            "Auto update",
        ],
        value_input_option="USER_ENTERED",
    )
    return True, "Đã ghi"


def send_email(to_email: str, cc: list[str], employee: str, target_date: date, rows: list[dict]) -> tuple[bool, str]:
    if not SMTP_SENDER_EMAIL or not SMTP_APP_PASSWORD:
        return False, "SMTP_SENDER_EMAIL/SMTP_APP_PASSWORD chưa cấu hình."
    if not to_email or "@" not in to_email:
        return False, "Nhân viên chưa có email hợp lệ."
    cc = [e for e in cc if e and "@" in e and e.casefold() != to_email.casefold()]
    detail_html = "".join(
        "<tr>"
        f"<td style='padding:6px;border:1px solid #ddd'>{r['reason']}</td>"
        f"<td style='padding:6px;border:1px solid #ddd'>{int(r.get('minutes', 0) or 0)}</td>"
        f"<td style='padding:6px;border:1px solid #ddd'>{float(r.get('penalty', 0) or 0):,.0f} VNĐ</td>"
        f"<td style='padding:6px;border:1px solid #ddd'>{r.get('detail', '')}</td>"
        "</tr>"
        for r in rows
    )
    html = f"""
    <html><body style='font-family:Arial,sans-serif;color:#222'>
      <p>Chào <b>{employee}</b>,</p>
      <p>Hệ thống Vera Spa ghi nhận vi phạm ra ngoài/Break ngày <b>{target_date.strftime('%d/%m/%Y')}</b>:</p>
      <table style='border-collapse:collapse;width:100%'>
        <thead><tr>
          <th style='padding:6px;border:1px solid #ddd'>Vi phạm</th>
          <th style='padding:6px;border:1px solid #ddd'>Số phút</th>
          <th style='padding:6px;border:1px solid #ddd'>Mức phạt</th>
          <th style='padding:6px;border:1px solid #ddd'>Chi tiết</th>
        </tr></thead><tbody>{detail_html}</tbody>
      </table>
      <p>Nếu dữ liệu chưa chính xác, vui lòng phản hồi với Lễ tân/Quản lý.</p>
      <p>Trân trọng,<br><b>VERA SPA</b><br>
      <b>Địa chỉ:</b> 193 Trương Định, Tam Hiệp, Đồng Nai<br>
      <b>Điện Thoại:</b> 0833229939</p>
    </body></html>
    """
    msg = MIMEMultipart()
    msg["From"] = f"Vera Spa <{SMTP_SENDER_EMAIL}>"
    msg["To"] = to_email
    if cc:
        msg["Cc"] = ", ".join(cc)
    msg["Subject"] = f"Thông báo vi phạm ra ngoài {target_date.strftime('%d/%m/%Y')} - {employee}"
    msg.attach(MIMEText(html, "html"))
    server = smtplib.SMTP("smtp.gmail.com", 587, timeout=30)
    try:
        server.starttls()
        server.login(SMTP_SENDER_EMAIL, SMTP_APP_PASSWORD)
        server.send_message(msg)
    finally:
        try:
            server.quit()
        except Exception:
            pass
    return True, "Đã gửi"


def collect_violations(target_date: date, catalog: dict) -> list[dict]:
    wb, ws, path = open_bang_tour(data_only=False)
    results = []
    try:
        for row in range(21, 101):
            employee = str(ws[f"B{row}"].value or "").strip()
            if not employee:
                continue
            out_dt = parse_datetime(ws[f"S{row}"].value, fallback_date=target_date)
            in_dt = parse_datetime(ws[f"U{row}"].value, fallback_date=(out_dt.date() if out_dt else target_date))
            out_matches = out_dt is not None and out_dt.date() == target_date
            in_matches = in_dt is not None and in_dt.date() == target_date
            if not out_matches and not in_matches:
                continue
            note = str(ws[f"V{row}"].value or "").strip()
            if out_dt is not None and in_dt is not None:
                end_dt = in_dt + (timedelta(days=1) if in_dt < out_dt else timedelta(0))
                elapsed = max(0.0, (end_dt - out_dt).total_seconds() / 60.0)
                late = max(0.0, elapsed - BANG_TOUR_BREAK_LIMIT_MINUTES)
                if late <= 0:
                    continue
                minutes = int(math.ceil(late))
                reason = pick_late_reason(catalog, minutes)
                results.append({
                    "employee": employee,
                    "reason": reason,
                    "minutes": minutes,
                    "penalty": penalty_for(catalog, reason),
                    "detail": f"Auto update · Bảng tour: Giờ ra {out_dt.strftime('%H:%M:%S')}, Giờ vào {in_dt.strftime('%H:%M:%S')}, vượt quy định {minutes} phút. {note}".strip(),
                })
            else:
                reason = pick_single_side_reason(catalog)
                missing = "Giờ vào" if out_dt is not None else "Giờ ra"
                results.append({
                    "employee": employee,
                    "reason": reason,
                    "minutes": 0,
                    "penalty": penalty_for(catalog, reason),
                    "detail": f"Auto update · Bảng tour chỉ có một mốc thời gian, thiếu {missing}. {note}".strip(),
                })
        return results
    finally:
        try:
            os.remove(path)
        except Exception:
            pass


def process_violations() -> str:
    client = gspread_client()
    target_date = vn_now().date()
    catalog = leave_catalog(client)
    violations = collect_violations(target_date, catalog)
    emails, cc = employee_directory(client)
    already_sent = sent_keys(client)
    audit = audit_ws(client)

    grouped = {}
    audit_rows = []
    saved_count = 0
    for item in violations:
        ok, save_msg = append_violation(
            client,
            target_date,
            item["employee"],
            item["reason"],
            item["detail"],
            item["penalty"],
        )
        if ok:
            saved_count += 1
        key = (target_date.strftime("%d/%m/%Y"), normalize_text(item["employee"]), normalize_text(item["reason"]))
        if key not in already_sent:
            grouped.setdefault(item["employee"], []).append(item)
        audit_rows.append({
            "Ngày": target_date.strftime("%d/%m/%Y"),
            "Tên nhân viên": item["employee"],
            "Loại vi phạm": item["reason"],
            "Số phút": item["minutes"],
            "Mức phạt": item["penalty"],
            "Ghi phiếu phạt": "1" if ok else "Đã tồn tại",
            "Email": "",
            "Chi tiết": save_msg,
            "Cập nhật lúc": vn_now().strftime("%d/%m/%Y %H:%M:%S"),
            "Người cập nhật": "Auto update",
        })

    mail_results = {}
    for employee, rows in grouped.items():
        to_email = emails.get(normalize_text(employee), "")
        mail_results[employee] = send_email(to_email, cc, employee, target_date, rows)

    for row in audit_rows:
        result = mail_results.get(row["Tên nhân viên"])
        if result:
            row["Email"] = "1" if result[0] else "0"
            if not result[0]:
                row["Chi tiết"] = row["Chi tiết"] + " | Email: " + result[1]

    if audit_rows:
        audit.append_rows(
            [[row.get(h, "") for h in AUDIT_HEADERS] for row in audit_rows],
            value_input_option="USER_ENTERED",
        )

    zero_penalty = sum(1 for x in violations if float(x.get("penalty", 0) or 0) <= 0)
    sent_count = sum(1 for ok, _ in mail_results.values() if ok)
    return (
        f"violations success: detected={len(violations)}, saved={saved_count}, emails={sent_count}, "
        f"zero_penalty={zero_penalty}"
    )


def main():
    action = (sys.argv[1] if len(sys.argv) > 1 else os.getenv("DAILY_OPS_ACTION", "")).strip().lower()
    if action in {"cleanup", "clean", "08:00", "0800"}:
        print(cleanup_previous_day())
        return
    if action in {"violations", "violation", "20:10", "2010"}:
        print(process_violations())
        return
    if action == "both":
        print(cleanup_previous_day())
        print(process_violations())
        return
    raise SystemExit("Usage: python vera_daily_ops_job.py cleanup|violations|both")


if __name__ == "__main__":
    main()
